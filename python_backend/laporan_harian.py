# laporan_harian.py
# Tujuan     : Replika pipeline Power Query lama ("2. To Format" + "SalesBase") di pandas.
#              Ganti refresh Excel ~35 menit -> proses ~detik. Dari 1. Paste Data.xlsx (+ 3. Stock.xlsx)
#              hasilkan: (a) baris per SPV & per SM, (b) stock per SPV, (c) agregat progress harian
#              (salesDailyProgress) untuk dashboard insentif-sales.
# Caller     : python_backend/main.py endpoint /laporan-harian/process (dipanggil Next.js
#              app/api/laporan-harian/upload). Juga dipakai parity test Tahap 0.
# Dependensi : pandas, openpyxl, laporan_harian_lookups.json, laporan_harian_targets.py.
#              python-calamine dipakai bila tersedia (baca cepat).
# Main Functions:
#   load_lookups(f_format, f_spv) -> LookupTables
#   process(paste_path, stock_path, lookups) -> dict {per_spv, per_sm, stock, progress, summary}
#   build_stock_frame(...) / write_report_files(...) -> output XLSX dan Stock per target SPV/SM/principal.
# Side Effects: Baca file sumber dan tulis XLSX hasil ke runtime; tidak mengubah file sumber/tidak kirim email.
# Catatan parity (dikonfirmasi user): sumber penjualan = sheet "Paste Acc" (export Accurate),
#   sheet "Paste Lap. Penj" lama sudah kosong -> tidak dipakai. Retur = "Paste Lap. Retur" (dinegasikan).
from __future__ import annotations
import io
from dataclasses import dataclass
from typing import Optional
import numpy as np
import pandas as pd
from laporan_harian_targets import REPORT_TARGETS

PENJ_LABEL = "1. Penjualan Bruto"
RETUR_LABEL = "2. (-) Retur Penjualan"

# Rename raw "Paste Lap. Retur" -> nama FIX
REN_RETUR = {
    "No.Retur": "NO_NOTA", "Tanggal": "TANGGAL", "Mata Uang": "MATA_UANG",
    "Kode Customer": "KODE_CUST", "Nama Customer": "CUSTOMER", "Kota Customer": "KOTA",
    "Kode Salesman": "KODE_SALESMAN", "Nama Salesman": "SALESMAN", "Kode Barang": "KODE_BARANG",
    "Nama Barang": "NAMA_BARANG", "Qty": "QTY", "Nilai Disc": "POTONGAN", "Nilai Bruto": "NILAI_JUAL",
    "Nilai Pajak": "NILAI_PAJAK", "Satuan": "SATUAN", "Nama Principal": "PRINCIPAL",
    "Kode Jenis Produk": "JENISPRODUK_KODE", "Market": "MARKET", "Region": "REGION",
    "Alamat Customer": "ALAMAT", "Keterangan": "REM",
}
# Rename raw "Paste Acc" -> nama FIX
REN_ACC = {
    "KODE PELANGGAN INDUK": "KODE_CUST", "Nama Pelanggan Faktur Penjualan": "CUSTOMER",
    "NILAI JUAL": "NILAI_JUAL", "NPWP Pelanggan Faktur Penjualan": "NPWP",
    "PRINCIPLE": "PRINCIPAL", "JENIS PRODUK": "JENISPRODUK",
}

FIX_COLS = ["NO_NOTA", "TANGGAL", "TANGGAL_DATE", "MATA_UANG", "KODE_CUST", "CUSTOMER",
            "KODE_SALESMAN", "SALESMAN", "KODE_BARANG", "NAMA_BARANG", "QTY", "HARGA",
            "POTONGAN", "NILAI_JUAL", "DPP", "NILAI_PAJAK", "SATUAN", "PRINCIPAL",
            "JENISPRODUK", "KELOMPOKPRODUK", "GOLONGAN", "MARKET", "REGION", "ALAMAT",
            "KOTA", "TAHUN", "BULAN", "JENIS_TRANSAKSI", "REM"]


@dataclass
class LookupTables:
    principal_to_spv: dict   # PRINCIPAL(name) -> NAMA SPV
    conca_to_spv: dict       # PRINCIPAL+JENISPRODUK -> NAMA SPV
    jp_map: dict             # kode jenis produk -> nama
    sm_map: dict             # PRINCIPLE -> NAMA SM
    report_targets: dict = None  # keyword email -> {group_type, values}; alias principal non-exact


# ---------- IO cepat ----------
def _read_sheet(path: str, sheet: str) -> pd.DataFrame:
    """Baca satu sheet; pakai calamine bila ada (jauh lebih cepat utk file besar)."""
    try:
        return pd.read_excel(path, sheet_name=sheet, engine="calamine", dtype=object)
    except Exception:
        return pd.read_excel(path, sheet_name=sheet, dtype=object)


def _num(s) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").fillna(0)


def load_lookups(f_format: str, f_spv: str) -> LookupTables:
    import openpyxl

    def sheet_df(path, name):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c is not None else "" for c in next(it)]
        df = pd.DataFrame(list(it), columns=hdr)
        wb.close()
        return df

    gol = sheet_df(f_format, "GOLONGAN")            # KODE SPV(=principal), NAMA SPV, JENIS PRODUK
    spv = gol[["KODE SPV", "NAMA SPV", "JENIS PRODUK"]].dropna(subset=["KODE SPV"]).copy()
    spv["KODE SPV"] = spv["KODE SPV"].astype(str).str.strip()
    spv["Conca"] = spv["KODE SPV"] + spv["JENIS PRODUK"].fillna("").astype(str)
    p2s = dict(zip(spv["KODE SPV"], spv["NAMA SPV"]))
    c2s = dict(zip(spv["Conca"], spv["NAMA SPV"]))
    jp = sheet_df(f_format, "JENIS PRODUK")
    jpm = dict(zip(jp["JENISPRODUK"].astype(str).str.strip(), jp["NAMA JENIS PRODUK"]))
    mp = sheet_df(f_spv, "Mapping")                 # PRINCIPLE, NAMA SPV, NAMA SM
    sm = dict(zip(mp["PRINCIPLE"].astype(str).str.strip(), mp["NAMA SM"]))
    return LookupTables(p2s, c2s, jpm, sm, {})


# ---------- Stage A: Paste Data -> FIX ----------
def _prep_acc(acc: pd.DataFrame, lk: LookupTables) -> pd.DataFrame:
    # CUSTOMER asli di Paste Acc = KODE; nama asli ada di "Nama Pelanggan Faktur Penjualan".
    # Buang kolom CUSTOMER(kode) agar rename bawa NAMA jadi CUSTOMER (bukan drop nama saat dedup).
    a = acc.rename(columns=REN_ACC)
    a = a.loc[:, ~a.columns.duplicated()]  # buang kolom duplikat pasca-rename
    if "KODE_CUST" in a.columns:
        a["CUSTOMER"] = a["KODE_CUST"]   # per instruksi: CUSTOMER dari kolom F (KODE PELANGGAN INDUK)
    a["QTYBONUS"] = 0
    if "MARKET" in a.columns:
        a["JENISMARKET"] = a["MARKET"]   # jenis market (TT/GT/MT) ada di kolom MARKET (AE) Paste Acc
    a = a[a["NO_NOTA"].astype("string").str.strip().fillna("") != ""].copy()
    for c in ["PRINCIPAL", "KODE_SALESMAN", "KODE_CUST", "KODE_BARANG", "SATUAN", "REM", "CUSTOMER"]:
        if c in a:
            a[c] = a[c].astype("string").str.strip()
    a["JENIS_TRANSAKSI"] = np.where(
        a["NO_NOTA"].astype("string").str.upper().str.startswith("INV"), PENJ_LABEL, RETUR_LABEL)
    a["NILAI_JUAL"] = _num(a.get("NILAI_JUAL"))
    a["DPP"] = _num(a["DPP"]) if "DPP" in a else a["NILAI_JUAL"]
    a["QTY"] = _num(a.get("QTY"))
    a["Conca"] = a["PRINCIPAL"].fillna("") + a.get("JENISPRODUK", pd.Series("", index=a.index)).fillna("").astype(str)
    a["GOLONGAN"] = a["Conca"].map(lk.conca_to_spv).fillna(a["PRINCIPAL"].map(lk.principal_to_spv)).fillna("YARMAN")
    return a


def _prep_retur(retur: pd.DataFrame, lk: LookupTables) -> pd.DataFrame:
    r = retur.rename(columns=REN_RETUR)
    r = r.loc[:, ~r.columns.duplicated()]
    r = r[r["NO_NOTA"].astype("string").str.strip().fillna("") != ""].copy()
    for c in ["PRINCIPAL", "KODE_SALESMAN", "KODE_CUST", "KODE_BARANG", "SATUAN", "REM", "CUSTOMER"]:
        if c in r:
            r[c] = r[c].astype("string").str.strip()
    r["JENIS_TRANSAKSI"] = RETUR_LABEL
    r["JENISPRODUK"] = r.get("JENISPRODUK_KODE").map(lk.jp_map).fillna(r.get("JENISPRODUK_KODE"))
    r["DPP"] = -(_num(r.get("NILAI_JUAL")) - _num(r.get("POTONGAN")))
    r["QTY"] = -_num(r.get("QTY"))
    r["NILAI_JUAL"] = -_num(r.get("NILAI_JUAL"))
    r["Conca"] = r["PRINCIPAL"].fillna("") + r["JENISPRODUK"].fillna("").astype(str)
    r["GOLONGAN"] = r["Conca"].map(lk.conca_to_spv).fillna(r["PRINCIPAL"].map(lk.principal_to_spv))
    return r


def build_fix(paste_path: str, lk: LookupTables) -> pd.DataFrame:
    acc = _read_sheet(paste_path, "Paste Acc")
    acc.columns = [str(c).strip() for c in acc.columns]
    retur = _read_sheet(paste_path, "Paste Lap. Retur")
    retur.columns = [str(c).strip() for c in retur.columns]
    a = _prep_acc(acc, lk)
    r = _prep_retur(retur, lk)
    for df in (a, r):
        df["TANGGAL"] = pd.to_datetime(df.get("TANGGAL"), errors="coerce")
        if "MATA_UANG" not in df:
            df["MATA_UANG"] = "IDR"
    keep = [c for c in ["NO_NOTA", "TANGGAL", "MATA_UANG", "KODE_CUST", "CUSTOMER", "KODE_SALESMAN",
                        "SALESMAN", "KODE_BARANG", "NAMA_BARANG", "QTY", "NILAI_JUAL", "DPP",
                        "PRINCIPAL", "JENISPRODUK", "GOLONGAN", "MARKET", "REGION", "JENIS_TRANSAKSI", "REM"]]
    for c in keep:
        if c not in a:
            a[c] = None
        if c not in r:
            r[c] = None
    fix = pd.concat([a[keep], r[keep]], ignore_index=True)
    fix = fix[fix["MATA_UANG"].astype("string").str.strip().fillna("") != ""].copy()
    fix["TANGGAL_DATE"] = fix["TANGGAL"].dt.date
    fix["TAHUN"] = fix["TANGGAL"].dt.year
    fix["BULAN"] = fix["TANGGAL"].dt.month
    return fix


# ---------- Stage B: SalesBase (flag AO/EC/IA + kolom turunan) ----------
def _flag_first(df: pd.DataFrame, mask_non_retur: pd.Series, keys: list, col: str):
    """Set col=1 HANYA di baris pertama tiap kombinasi unik (meniru List.Min RowID di M).
    Jadi SUM(col) = jumlah kombinasi unik, bukan jumlah baris."""
    df[col] = 0
    idx = df[mask_non_retur].drop_duplicates(keys).index
    df.loc[idx, col] = 1


def build_salesbase(fix: pd.DataFrame, lk: LookupTables) -> pd.DataFrame:
    df = fix.reset_index(drop=True).copy()
    non_retur = df["JENIS_TRANSAKSI"] != RETUR_LABEL

    # AO/EC/IA: flag di baris pertama kombinasi (non-retur)
    _flag_first(df, non_retur, ["KODE_SALESMAN", "KODE_CUST", "PRINCIPAL"], "AO")
    _flag_first(df, non_retur, ["KODE_SALESMAN", "KODE_CUST", "PRINCIPAL", "TANGGAL_DATE"], "EC")
    _flag_first(df, non_retur, ["KODE_CUST", "PRINCIPAL", "KODE_BARANG"], "IA")
    df["Item Aktif"] = df["IA"]

    # Item batal: REM mengandung "BATAL"
    rem = df["REM"].fillna("").astype(str).str.upper() if "REM" in df else pd.Series("", index=df.index)
    df["Item batal"] = rem.str.contains("BATAL").astype(int)

    # Value Retur Termasuk Batal (Exc.PPN): |DPP| utk baris retur
    df["Value Retur Termasuk Batal (Exc.PPN)"] = np.where(~non_retur, df["DPP"].abs(), 0.0)

    # Qty Los / Value Los butuh QTY_REF (tak ada di Paste Acc) -> 0 bila kolom absen
    qref_raw = pd.to_numeric(df["QTY_REF"], errors="coerce") if "QTY_REF" in df.columns else pd.Series(np.nan, index=df.index)
    has_ref = qref_raw.notna() & (qref_raw != 0)   # tanpa QTY_REF -> Los = 0 (bukan -QTY)
    qn = _num(df.get("QTY")); hg = _num(df.get("HARGA"))
    df["Qty Los"] = np.where(non_retur & has_ref, qref_raw.fillna(0) - qn, 0.0)
    df["Value Los"] = np.where(non_retur & has_ref, (qref_raw.fillna(0) - qn) * hg, 0.0)

    # Nota Batal (per nota, 1x): baris item-batal, unik per PRINCIPAL+SALESMAN+NO_NOTA+CUST
    keys_nota = ["PRINCIPAL", "KODE_SALESMAN", "NO_NOTA", "KODE_CUST"]
    df["Nota Batal"] = 0
    batal_idx = df[df["Item batal"] == 1].drop_duplicates(keys_nota).index
    df.loc[batal_idx, "Nota Batal"] = 1
    # Jumlah Nota Retur (per nota, 1x): baris retur unik per nota
    df["Jumlah Nota Retur"] = 0
    retur_idx = df[~non_retur].drop_duplicates(keys_nota).index
    df.loc[retur_idx, "Jumlah Nota Retur"] = 1

    # Kategori Baru: SIRUP->Cordial, Nutrijell->Nutrijell, else REGULER
    nb = df["NAMA_BARANG"].fillna("").astype(str) if "NAMA_BARANG" in df else pd.Series("", index=df.index)
    df["Kategori Baru"] = np.where(nb.str.contains("SIRUP", case=False), "Cordial",
                          np.where(nb.str.contains("Nutrijell", case=False), "Nutrijell", "REGULER"))

    # NAMA SM (mapping principal). Catatan: match nama principal perlu normalisasi (principle_matcher) — tuning go-live.
    df["Harga Sesuai Inputan"] = _num(df.get("HARGA"))          # = HARGA (per Power Query)
    df["@ Toko Retur"] = df["Jumlah Nota Retur"]                 # = Jumlah Nota Retur (per Power Query)
    df["Mapping_PIC.NAMA SM"] = df["PRINCIPAL"].map(lk.sm_map)
    df["NAMA_SM"] = df["Mapping_PIC.NAMA SM"]
    return df


# ---------- Stage C: split + agregat ----------
def split_per_group(sb: pd.DataFrame):
    per_spv = {name: g.copy() for name, g in sb.groupby("GOLONGAN", dropna=True)}
    per_sm = {name: g.copy() for name, g in sb.groupby("NAMA_SM", dropna=True)}
    return per_spv, per_sm


def aggregate_progress(sb: pd.DataFrame) -> pd.DataFrame:
    """Agregat -> salesDailyProgress: per salesCode+principle+date."""
    d = sb.copy()
    d["date"] = pd.to_datetime(d["TANGGAL"], errors="coerce").dt.strftime("%Y-%m-%d")
    grp = d.groupby(["KODE_SALESMAN", "PRINCIPAL", "GOLONGAN", "date"], dropna=False)
    out = grp.agg(achievedValueDpp=("DPP", "sum"),
                  achievedEc=("EC", "sum"),
                  achievedAo=("AO", "sum"),
                  achievedIa=("IA", "sum")).reset_index()
    out = out.rename(columns={"KODE_SALESMAN": "salesCode", "PRINCIPAL": "principle", "GOLONGAN": "branch"})
    out["periodMonth"] = pd.to_datetime(out["date"], errors="coerce").dt.month
    out["periodYear"] = pd.to_datetime(out["date"], errors="coerce").dt.year
    return out


def process(paste_path: str, stock_path: Optional[str], lookups: LookupTables) -> dict:
    fix = build_fix(paste_path, lookups)
    sb = build_salesbase(fix, lookups)
    per_spv, per_sm = split_per_group(sb)
    progress = aggregate_progress(sb)
    summary = (sb.groupby("GOLONGAN", dropna=True)
                 .agg(rows=("NO_NOTA", "size"), dpp=("DPP", "sum"),
                      ao=("AO", "sum"), ec=("EC", "sum"), ia=("IA", "sum"))
                 .reset_index().sort_values("GOLONGAN"))
    stock_per_spv = {}
    if stock_path:
        try:
            stock_per_spv = build_stock(stock_path, sb, lookups)
        except Exception as e:  # stock opsional; jangan gagalkan seluruh proses
            stock_per_spv = {"__error__": str(e)}
    return {
        "fix_rows": int(len(fix)),
        "sales_rows": int(len(sb)),
        "net_dpp": float(sb["DPP"].sum()),
        "per_spv": per_spv,
        "per_sm": per_sm,
        "stock_per_spv": stock_per_spv,
        "progress": progress,
        "summary": summary,
    }


def _read_stock(stock_path: str) -> pd.DataFrame:
    """Baca stock Accurate ("Kuantitas Barang per Gudang", header bukan di baris 1) atau
    format lama. Pakai calamine (Accurate xlsx kadang tak terbaca openpyxl). Deteksi baris header
    dari sel "Kode Barang"."""
    rows = None
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(stock_path)
        sheet = next((n for n in wb.sheet_names if "gudang" in n.lower() or "kuantitas" in n.lower()),
                     wb.sheet_names[0])
        rows = wb.get_sheet_by_name(sheet).to_python()
    except Exception:
        raw = _read_sheet(stock_path, 0)
        raw.columns = [str(c).strip() for c in raw.columns]
        return raw
    hdr_i = next((i for i, r in enumerate(rows)
                  if any(str(c).strip().lower() == "kode barang" for c in r)), 0)
    header = [str(c).strip() for c in rows[hdr_i]]
    df = pd.DataFrame(rows[hdr_i + 1:], columns=header)
    return df.loc[:, [c for c in df.columns if c != ""]]


def build_stock_frame(stock_path: str, sb: pd.DataFrame) -> pd.DataFrame:
    """Stock -> perkaya KODE_BARANG dengan PRINCIPAL, GOLONGAN, dan NAMA_SM dari sales."""
    st = _read_stock(stock_path)
    st.columns = [str(c).strip() for c in st.columns]
    kode_col = next((c for c in st.columns if c.lower() in ("kode barang", "kode") or "kode barang" in c.lower()),
                    next((c for c in st.columns if "kode" in c.lower()), st.columns[0]))
    qty_col = next((c for c in st.columns if "kuantitas in pcs" in c.lower()
                    or "saldo akhir" in c.lower() or c.lower() == "qty"), None)
    ren = {kode_col: "KODE_BARANG"}
    if qty_col:
        ren[qty_col] = "QTY AKHIR"
    st = st.rename(columns=ren)
    st = st[st["KODE_BARANG"].notna()].copy()
    st["KODE_BARANG"] = st["KODE_BARANG"].astype("string").str.strip()
    mapping_cols = ["KODE_BARANG", "PRINCIPAL", "GOLONGAN", "NAMA_SM"]
    product_map = (sb.reindex(columns=mapping_cols)
                     .dropna(subset=["KODE_BARANG"])
                     .drop_duplicates("KODE_BARANG")
                     .set_index("KODE_BARANG"))
    for column in ("PRINCIPAL", "GOLONGAN", "NAMA_SM"):
        st[column] = st["KODE_BARANG"].map(product_map[column])
    return st


def build_stock(stock_path: str, sb: pd.DataFrame, lk: LookupTables) -> dict:
    """Backward-compatible: stock dipisah per SPV untuk caller pipeline lama."""
    st = build_stock_frame(stock_path, sb)
    return {name: g.copy() for name, g in st.dropna(subset=["GOLONGAN"]).groupby("GOLONGAN")}


if __name__ == "__main__":  # self-test / parity Tahap 0
    import sys, time
    base = sys.argv[1] if len(sys.argv) > 1 else "/sessions/keen-vigilant-mccarthy/mnt/Laporan/New Format Laporan - Copy"
    t = time.time()
    lk = load_lookups(f"{base}/2. To Format Laporan.xlsx", f"{base}/2.3 To SPV dan SM New.xlsx")
    res = process(f"{base}/1. Paste Data.xlsx", f"{base}/3. Stock.xlsx", lk)
    msg = ("fix_rows={:,} sales_rows={:,} net_dpp={:,.0f} SPV={} SM={} "
           "stock_spv={} progress_rows={:,} ({:.1f}s)").format(
        res["fix_rows"], res["sales_rows"], res["net_dpp"], len(res["per_spv"]),
        len(res["per_sm"]), len(res["stock_per_spv"]), len(res["progress"]), time.time() - t)
    print(msg)
    print(res["summary"].to_string(index=False))


# ---------- Alternatif Stage A: input = FIX LAP PENJ jadi (2. To Format, Paste Acc only, retur sudah minus) ----------
def load_fix(fix_path: str, sheet: str = "FIX LAP PENJ") -> pd.DataFrame:
    """Baca FIX LAP PENJ yang sudah jadi (retur sudah termasuk & sudah minus).
    Tidak ada rekonstruksi/negasi — semua kolom native dipakai apa adanya."""
    df = _read_sheet(fix_path, sheet)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df["NO_NOTA"].astype("string").str.strip().fillna("") != ""].copy()
    df["TANGGAL"] = pd.to_datetime(df.get("TANGGAL"), errors="coerce")
    df["TANGGAL_DATE"] = df["TANGGAL"].dt.date
    if "TAHUN" not in df: df["TAHUN"] = df["TANGGAL"].dt.year
    if "BULAN" not in df: df["BULAN"] = df["TANGGAL"].dt.month
    for c in ["DPP", "QTY", "HARGA", "NILAI_JUAL", "POTONGAN", "JUMLAH"]:
        if c in df: df[c] = _num(df[c])
    return df


def process_from_fix(fix_path: str, stock_path: Optional[str], sm_map: dict) -> dict:
    """Pipeline bila sumber = FIX LAP PENJ jadi. Hanya Stage B (SalesBase) + split + agregat."""
    lk = LookupTables({}, {}, {}, sm_map, {})
    fix = load_fix(fix_path)
    sb = build_salesbase(fix, lk)
    per_spv, per_sm = split_per_group(sb)
    progress = aggregate_progress(sb)
    summary = (sb.groupby("GOLONGAN", dropna=True)
                 .agg(rows=("NO_NOTA", "size"), dpp=("DPP", "sum"),
                      ao=("AO", "sum"), ec=("EC", "sum"), ia=("Item Aktif", "sum"))
                 .reset_index().sort_values("GOLONGAN"))
    stock = {}
    if stock_path:
        try:
            stock = build_stock(stock_path, sb, lk)
        except Exception as e:
            stock = {"__error__": str(e)}
    return {"sales_rows": int(len(sb)), "net_dpp": float(sb["DPP"].sum()),
            "per_spv": per_spv, "per_sm": per_sm, "stock_per_spv": stock,
            "progress": progress, "summary": summary, "sb": sb}


# ---------- Output per-SPV (layout 50 kolom seperti 2.3) ----------
REPORT_COLUMNS = ["NO_NOTA","TANGGAL","TGL_JT","KODE_CUST","CUSTOMER","KODE_SALESMAN","SALESMAN",
    "KODE_BARANG","NAMA_BARANG","QTY","HARGA","POTONGAN","NILAI_JUAL","SATUAN","NPWP","GOLONGAN",
    "GDGNAMA","ALAMAT","KOTA","QTYBONUS","JUMLAH","PRINCIPAL","JENISPRODUK","KELOMPOKPRODUK","MARKET",
    "JENISMARKET","REGION","QTY_REF","UNIT_REF","TAHUN","BULAN","JENIS_TRANSAKSI","QTYKONV",
    "QTY_SATUANKECIL","SATUAN_KECIL","REM","Value Los","AO","EC","Item batal","Nota Batal","Qty Los",
    "@ Toko Retur","Value Retur Termasuk Batal (Exc.PPN)","DPP","Harga Sesuai Inputan","Item Aktif",
    "Jumlah Nota Retur","Mapping_PIC.NAMA SM","Kategori Baru"]


def build_report_frame(sb: pd.DataFrame) -> pd.DataFrame:
    """DataFrame siap-tulis: kolom 50 seperti 2.3, tanpa #NUM!/NaN (blank), tanggal string."""
    df = sb.loc[:, ~sb.columns.duplicated()].copy()
    if "FIX QTY_SATUAN KECIL" in df.columns and "QTY_SATUANKECIL" not in df.columns:
        df["QTY_SATUANKECIL"] = df["FIX QTY_SATUAN KECIL"]
    for c in REPORT_COLUMNS:
        if c not in df.columns:
            df[c] = np.nan
    for c in ("TANGGAL", "TGL_JT"):
        df = df.assign(**{c: pd.to_datetime(df[c], errors="coerce").dt.strftime("%Y-%m-%d")})
    numcols = ["QTY","HARGA","POTONGAN","NILAI_JUAL","JUMLAH","DPP","Harga Sesuai Inputan",
               "Value Los","Qty Los","Value Retur Termasuk Batal (Exc.PPN)","QTY_SATUANKECIL","QTY_REF"]
    # GOLONGAN sudah ada di REPORT_COLUMNS; duplikasi di sini menggeser seluruh nilai setelah kolom itu saat XLSX ditulis.
    out = df[REPORT_COLUMNS].copy()
    for c in numcols:
        if c in out:
            out = out.assign(**{
                c: pd.to_numeric(out[c], errors="coerce").replace([np.inf, -np.inf], np.nan)
            })
    return out.astype(object).where(pd.notna(out), None)


def _normal_text(value) -> str:
    return " ".join(str(value or "").strip().upper().split())


def resolve_report_groups(sb: pd.DataFrame, report_keywords: list, lk: LookupTables) -> tuple:
    """Resolve keyword penerima menjadi grup data SPV, SM, atau principal tanpa fuzzy filename matching."""
    columns = {"spv": "GOLONGAN", "sm": "NAMA_SM", "principal": "PRINCIPAL"}
    known = {
        group_type: {
            _normal_text(value): str(value)
            for value in sb[column].dropna().unique()
            if _normal_text(value)
        }
        for group_type, column in columns.items()
    }
    aliases = {
        _normal_text(keyword): target
        for keyword, target in (lk.report_targets or {}).items()
    }
    groups, unmatched = [], []
    for raw_keyword in report_keywords:
        keyword = str(raw_keyword).strip()
        normalized = _normal_text(keyword)
        target = aliases.get(normalized)
        if target:
            group_type = str(target.get("group_type", "")).lower()
            values = [str(value) for value in target.get("values", [])]
        elif normalized in known["spv"]:
            group_type, values = "spv", [known["spv"][normalized]]
        elif normalized in known["sm"]:
            group_type, values = "sm", [known["sm"][normalized]]
        else:
            unmatched.append(keyword)
            continue
        column = columns.get(group_type)
        if not column:
            unmatched.append(keyword)
            continue
        normalized_values = {_normal_text(value) for value in values}
        mask = sb[column].map(_normal_text).isin(normalized_values)
        frame = sb[mask].copy()
        if frame.empty:
            unmatched.append(keyword)
            continue
        groups.append({
            "keyword": keyword,
            "groupType": group_type,
            "column": column,
            "values": values,
            "frame": frame,
        })
    return groups, unmatched


def write_report_files(sb: pd.DataFrame, out_dir: str, report_date: str,
                       report_keywords: list, lk: LookupTables,
                       stock_frame: Optional[pd.DataFrame] = None) -> tuple:
    """Tulis file per keyword aktif; setiap file berisi data dan sheet Stock dengan cakupan yang sama."""
    import os
    from pyexcelerate import Workbook
    os.makedirs(out_dir, exist_ok=True)
    groups, unmatched = resolve_report_groups(sb, report_keywords, lk)
    written = []
    for target in groups:
        keyword = target["keyword"]
        data = target["frame"]
        frame = build_report_frame(data)
        safe = keyword.replace("/", "-").replace("\\", "-")
        file_name = f"{report_date}_{safe}.xlsx"
        path = os.path.join(out_dir, file_name)
        wb = Workbook()
        sheet_base = keyword.replace("/", "-").replace("\\", "-").replace("&", "dan") or "NA"
        wb.new_sheet(sheet_base[:31],
                     data=[REPORT_COLUMNS] + frame[REPORT_COLUMNS].values.tolist())
        stock_rows = 0
        if stock_frame is not None and not stock_frame.empty:
            normalized_values = {_normal_text(value) for value in target["values"]}
            stock_mask = stock_frame[target["column"]].map(_normal_text).isin(normalized_values)
            target_stock = stock_frame[stock_mask].loc[:, ~stock_frame.columns.duplicated()].copy()
            target_stock = target_stock.astype(object).where(pd.notna(target_stock), None)
            stock_headers = [str(column) for column in target_stock.columns]
            stock_rows = int(len(target_stock))
        else:
            target_stock = None
        if target_stock is not None and stock_rows:
            wb.new_sheet(f"{sheet_base[:25]} Stock"[:31],
                         data=[stock_headers] + target_stock.values.tolist())
        wb.save(path)
        written.append({
            "keyword": keyword,
            "groupType": target["groupType"],
            "fileName": file_name,
            "path": path,
            "rows": int(len(data)),
            "stockRows": stock_rows,
        })
    return written, unmatched


def write_per_spv_files(sb: pd.DataFrame, out_dir: str, report_date: str,
                        stock_per_spv: Optional[dict] = None) -> list:
    """Backward-compatible wrapper untuk test/caller lama yang masih meminta semua SPV."""
    keywords = [str(value) for value in sb["GOLONGAN"].dropna().unique()]
    stock_frame = None
    if stock_per_spv:
        stock_frame = pd.concat(stock_per_spv.values(), ignore_index=True)
    files, _ = write_report_files(sb, out_dir, report_date, keywords,
                                  LookupTables({}, {}, {}, {}, {}), stock_frame)
    return files


# ---------- Stage A (baru): bangun FIX dari 2 file mentah Accurate (penjualan + retur) ----------
import os as _os, json as _json

RINCIAN_SHEET = "Rincian Faktur Penjualan"
LOOKUPS_JSON = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "laporan_harian_lookups.json")


def load_lookups_json(path: str = LOOKUPS_JSON) -> "LookupTables":
    """Master lookup (GOLONGAN/JENIS PRODUK/Mapping) dari JSON kecil yang di-commit.
    Dipakai saat web membangun 2.ToFormat sendiri dari export Accurate (tanpa upload master)."""
    d = _json.load(open(path, encoding="utf-8"))
    return LookupTables(d.get("principal_to_spv", {}), d.get("conca_to_spv", {}),
                        d.get("jp_map", {}), d.get("sm_map", {}),
                        REPORT_TARGETS)


def build_fix_from_accurate(penjualan_path: str, retur_path: Optional[str], lk: LookupTables) -> pd.DataFrame:
    """Bangun FIX LAP PENJ dari export Accurate mentah:
       - penjualan (INV, nilai positif) via _prep_acc
       - retur (RJN, nilai mentah POSITIF -> dinegasikan)
    Menghasilkan tabel setara 'FIX LAP PENJ' (retur sudah minus)."""
    acc = _read_sheet(penjualan_path, RINCIAN_SHEET)
    acc.columns = [str(c).strip() for c in acc.columns]
    a = _prep_acc(acc, lk)              # JENIS by prefix INV; CUSTOMER=KODE_CUST; JENISMARKET=MARKET; QTYBONUS=0
    frames = [a]
    if retur_path:
        rr_raw = _read_sheet(retur_path, RINCIAN_SHEET)
        rr_raw.columns = [str(c).strip() for c in rr_raw.columns]
        r = _prep_acc(rr_raw, lk)
        r["JENIS_TRANSAKSI"] = RETUR_LABEL   # file retur = semua RJN
        for c in ("QTY", "DPP", "NILAI_JUAL", "POTONGAN", "JUMLAH", "NILAI_PAJAK"):
            if c in r:
                r[c] = -_num(r[c])           # nilai mentah positif -> minus
        frames.append(r)
    for df in frames:
        df["TANGGAL"] = pd.to_datetime(df.get("TANGGAL"), errors="coerce")
        if "MATA_UANG" not in df:
            df["MATA_UANG"] = "IDR"
    fix = pd.concat(frames, ignore_index=True, sort=False)
    fix = fix[fix["MATA_UANG"].astype("string").str.strip().fillna("") != ""].copy()
    fix["TANGGAL_DATE"] = fix["TANGGAL"].dt.date
    fix["TAHUN"] = fix["TANGGAL"].dt.year
    fix["BULAN"] = fix["TANGGAL"].dt.month
    return fix
