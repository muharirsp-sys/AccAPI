"""
loop_master_builder.py

Loop-engineering tool (offline, 0 API): bangun master ber-taksonomi gaya
URC/Priskila dari SEMUA workbook mentah "FIX FORM MASTER BARANG - *", lalu
cocokkan surat program yang ada (teks PDF native) ke tiap master via
urc_matcher (mesin generik). Output:
  - master_barang_principle/MASTER BARANG <NAME>.xlsx  (tidak menimpa yg ada)
  - data/loop_master_report.txt  (ringkasan utk dikoreksi user)

Sumber data per workbook: sheet "Fix Mapping" (KODE BARANG | NAMA BARANG |
ISI/CTN | SATUAN | NAMA KELOMPOK) -- kelompok sudah kurasi manusia; aroma/
gramasi/kemasan diderivasi dari NAMA BARANG (heuristik v1, draft utk koreksi).
"""
import os, re, sys, json
import openpyxl
from openpyxl import Workbook

BASE = os.path.dirname(os.path.abspath(__file__))
MD = os.path.join(BASE, "..", "master_barang_principle")
# Output rebuild ditulis ke staging (non-destruktif): tidak menimpa master
# lama & tidak bentrok bila user sedang membuka file di Excel (~$ lock).
MD_OUT = os.path.join(BASE, "data", "rebuild_master")
SP = os.path.join(BASE, "..", "reference_surat_program")
REPORT = os.path.join(BASE, "data", "loop_master_report.txt")

HEADERS = [
    "Nama Barang Principle", "Kode Barang", "Nama Barang", "Nama Pcpl",
    "kode  2 Digit                (hrf+No)", "Nama KLP",
    "kode  2 Digit                (Nomor)", "Nama Sub KLP",
    "kode  1 Digit                (Nomor)", "Nama Sub KLP2",
    "kode  1 Digit                (Nomor)5", "Nama            Aroma/             Rasa",
    "kode  2 Digit                (Nomor)2", "Nama            Gramasi atau Jumlah Pack per CTN",
    "kode  4 Digit                (Nomor)", "Nama Jenis Kemasan",
    "kode  1 Digit                (Nomor)3", "Nama            Promo",
    "kode  1 Digit                (Nomor)4",
]

GRAM_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*(KG|GR|G|ML|LTR|L|CC|SHEET|SHEETS|S|PCS|YARD|M)\b", re.I)


def principle_name(fname: str) -> str:
    n = os.path.splitext(fname)[0].upper()
    for pat in ("FIX_FORM MASTER BARANG TO WIN -", "FIX_FORM MASTER BARANG -", "FIX FORM MASTER BARANG -",
                "FIX MASTER BARANG", "FIX_2 CONTOH FORM MASTER BARANG", "FIX NEW FORM", "FIX NEW",
                "NEW FORM MASTER BARANG", "MRI NEW FORM MASTER BARANG", "NEW", "MASTER BARANG"):
        n = n.replace(pat, " ")
    n = n.replace("FORM MASTER BARANG", " ").replace("MASTER BARANG", " ")
    n = re.sub(r"\(\d+\)", " ", n)
    return " ".join(n.split()).strip() or os.path.splitext(fname)[0]


def _norm_hdr(h):
    return re.sub(r"\s+", " ", str(h or "").upper()).strip()


def formfix_colmap(hdr):
    """Peta kolom sheet 'Form Fix' (header baris 5) by NAMA header -- posisi
    bergeser antar-file (mis. GDI menyisipkan kolom), jadi jangan indeks tetap.
    Taksonomi WIN sudah terurai & terkurasi manusia -> ambil langsung, tanpa
    heuristik derive dari nama."""
    m = {}
    for i, h in enumerate(hdr):
        n = _norm_hdr(h)
        if "KODE BARANG WIN" in n and "kode" not in m: m["kode"] = i
        elif n == "NAMA WIN" and "nama" not in m: m["nama"] = i
        elif n == "NAMA KLP" and "klp" not in m: m["klp"] = i
        elif n == "NAMA SUB KLP" and "sub1" not in m: m["sub1"] = i
        elif n == "NAMA SUB KLP2" and "sub2" not in m: m["sub2"] = i
        elif n.startswith("NAMA AROMA") and "aroma" not in m: m["aroma"] = i
        elif "GRAMASI" in n and "gram" not in m: m["gram"] = i
        elif "JENIS KEMASAN" in n and "kemasan" not in m: m["kemasan"] = i
    return m


def editform_colmap(hdr):
    """Peta kolom sheet keluarga 'Edit_MASTER BARANG X' / 'Master Barang X'
    (form migrasi; header dinamis, label tunggal). Ambil taksonomi + kode/nama
    WIN final."""
    m = {}
    for i, h in enumerate(hdr):
        n = _norm_hdr(h)
        if "PENULISAN CODE" in n and "kode" not in m: m["kode"] = i
        elif "PENULISAN NAMA" in n and "nama" not in m: m["nama"] = i
        elif "KATEGORI" in n and "SUB" not in n and "klp" not in m: m["klp"] = i
        elif "SUB KATEGORI" in n and "sub1" not in m: m["sub1"] = i
        elif "AROMA" in n and "aroma" not in m: m["aroma"] = i
        elif "GRAMASI" in n and "gram" not in m: m["gram"] = i
        elif "KEMASAN" in n and "kemasan" not in m: m["kemasan"] = i
    return m


def find_editform_sheet(wb):
    """Keluarga tanpa 'Form Fix' (keputusan user): pakai sheet 'Edit_MASTER
    BARANG X' bila ada, jika tidak 'Master Barang X' (bukan template HEINZ/DATA).
    Header taksonomi ada di baris bawah -> cari baris yg memuat GRAMASI+AROMA/
    KATEGORI. Return (sheetname, ws, header_row_1based, colmap) atau None."""
    edit = [s for s in wb.sheetnames if s.upper().startswith("EDIT") and "MASTER BARANG" in s.upper()]
    plain = [s for s in wb.sheetnames if "MASTER BARANG" in s.upper()
             and "HEINZ" not in s.upper() and "DATA" not in s.upper() and not s.upper().startswith("EDIT")]
    sn = edit[0] if edit else (plain[0] if plain else None)
    if not sn:
        return None
    ws = wb[sn]
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=70, values_only=True), start=1):
        n = " | ".join(_norm_hdr(c) for c in row)
        if "GRAMASI" in n and ("AROMA" in n or "KATEGORI" in n):
            cm = editform_colmap(row)
            if all(k in cm for k in ("kode", "klp", "aroma", "gram")):
                return sn, ws, ridx, cm
    return None


def _edit_txt(row, i, offset_ok=False):
    """Nilai sel; buang error formula (#REF!/#VALUE!). Beberapa kolom (mis.
    'Penulisan Nama Barang') nilainya bergeser 1 kolom -> coba i lalu i+1."""
    for j in ([i, i + 1] if offset_ok else [i]):
        if j is not None and 0 <= j < len(row) and row[j] is not None:
            s = str(row[j]).strip()
            if s and not s.startswith("#"):
                return s
    return ""


# Noise benefit/tier di baris surat yang bukan bagian identitas produk.
LINE_NOISE_RE = re.compile(
    r"^\s*[>*\-•]+\s*|\bBELI\s+\d+\b|\bALL\s+VARIANTS?\b|\bGRATIS\b.*$|\bFREE\b.*$"
    r"|\bDISKON\s+\d+%?|\bPROMO\b|\bPERIODE?\b.*$", re.I)


def clean_letter_line(ln: str) -> str:
    out = LINE_NOISE_RE.sub(" ", ln)
    return " ".join(out.split())


def _save_master(pname, rows_out):
    out = Workbook(); ows = out.active; ows.title = "Sheet1"; ows.append(HEADERS)
    for r in rows_out:
        ows.append(r)
    out.save(os.path.join(MD_OUT, f"MASTER BARANG {pname}.xlsx"))


def _edit_pname(f, sheetname):
    """Nama principle dari nama sheet 'Edit_MASTER BARANG X' / 'Master Barang X'.
    Nama sheet lebih andal drpd nama file (yg mewariskan 'FORM')."""
    name = re.sub(r"^(EDIT[_ ]*)?MASTER BARANG", "", sheetname.upper()).strip()
    return name or principle_name(f)


def build_all():
    """Bangun master (keputusan user):
    - Utama: sheet 'Form Fix' (taksonomi WIN final terkurasi).
    - Keluarga tanpa 'Form Fix': sheet 'Edit_MASTER BARANG X' bila ada, jika
      tidak 'Master Barang X'. Tabrakan nama principle (mis. FORISA) -> versi
      Form Fix menang; URC selalu dari master buatan tangan."""
    os.makedirs(MD_OUT, exist_ok=True)
    results, built = [], set()
    files = sorted(os.listdir(MD))
    # dua fase: Form Fix dulu (menang saat tabrakan), lalu keluarga-Edit.
    for phase in ("formfix", "edit"):
        for f in files:
            if f.startswith("~") or not f.lower().endswith(".xlsx") or f.upper().startswith("MASTER BARANG"):
                continue
            try:
                wb = openpyxl.load_workbook(os.path.join(MD, f), data_only=True)
            except Exception as e:
                if phase == "formfix":
                    results.append((principle_name(f), f, f"LOAD FAIL {type(e).__name__}", 0, 0))
                continue
            has_ff = "Form Fix" in wb.sheetnames
            if phase == "formfix":
                if not has_ff:
                    continue
                pname = principle_name(f)
                if pname == "URC":
                    results.append((pname, f, "SKIP (buatan tangan, dari MD)", 0, 0)); continue
                ws = wb["Form Fix"]
                hdr = next(ws.iter_rows(min_row=5, max_row=5, values_only=True))
                cm = formfix_colmap(hdr)
                if not all(k in cm for k in ("kode", "nama", "klp")):
                    results.append((pname, f, f"FORM FIX HEADER TAK LENGKAP {sorted(cm)}", 0, 0)); continue
                rows_out, no_gram = [], 0
                for row in ws.iter_rows(min_row=6, values_only=True):
                    def g(key):
                        i = cm.get(key)
                        return str(row[i]).strip() if (i is not None and i < len(row) and row[i] is not None) else ""
                    kode, nama = g("kode"), g("nama")
                    if not (kode and nama):
                        continue
                    gramasi = g("gram").upper()
                    if not gramasi:
                        no_gram += 1
                    r = [""] * len(HEADERS)
                    r[1], r[2], r[5], r[7], r[9], r[11], r[13], r[15] = (
                        kode, nama, g("klp").upper(), g("sub1").upper(), g("sub2").upper(),
                        g("aroma").upper(), gramasi, g("kemasan").upper())
                    rows_out.append(r)
                if not rows_out:
                    results.append((pname, f, "FORM FIX: 0 ITEM VALID", 0, 0)); continue
                _save_master(pname, rows_out); built.add(pname)
                results.append((pname, f, "OK", len(rows_out), no_gram))
            else:  # phase == "edit"
                if has_ff:
                    continue
                fb = find_editform_sheet(wb)
                if fb is None:
                    results.append((principle_name(f), f, "NO EDIT/MASTER SHEET (skip)", 0, 0)); continue
                sn, ws, hrow, cm = fb
                pname = _edit_pname(f, sn)
                if pname == "URC":
                    results.append((pname, f, "SKIP (buatan tangan, dari MD)", 0, 0)); continue
                if pname in built:
                    results.append((pname, f, "SKIP (sudah dari Form Fix)", 0, 0)); continue
                rows_out, no_gram = [], 0
                for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                    kode = _edit_txt(row, cm["kode"])
                    klp = _edit_txt(row, cm["klp"]).upper()
                    if not (kode and klp) or len(klp) <= 1:  # buang baris contoh/legenda (klp 1-huruf)
                        continue
                    aroma = _edit_txt(row, cm.get("aroma")).upper() if cm.get("aroma") is not None else ""
                    gramasi = _edit_txt(row, cm.get("gram")).upper() if cm.get("gram") is not None else ""
                    kemasan = _edit_txt(row, cm.get("kemasan")).upper() if cm.get("kemasan") is not None else ""
                    sub1 = _edit_txt(row, cm.get("sub1")).upper() if cm.get("sub1") is not None else ""
                    nama = _edit_txt(row, cm["nama"], offset_ok=True) if cm.get("nama") is not None else ""
                    if not nama:
                        nama = " ".join(x for x in (klp, aroma, gramasi) if x)
                    if not gramasi:
                        no_gram += 1
                    r = [""] * len(HEADERS)
                    r[1], r[2], r[5], r[7], r[11], r[13], r[15] = kode, nama, klp, sub1, aroma, gramasi, kemasan
                    rows_out.append(r)
                if not rows_out:
                    results.append((pname, f, f"EDIT ({sn}): 0 ITEM VALID", 0, 0)); continue
                _save_master(pname, rows_out); built.add(pname)
                results.append((pname, f, f"OK (edit:{sn})", len(rows_out), no_gram))
    return results


def load_master_items(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    items = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        klp, sub1, sub2 = (r[5] or ""), (r[7] or ""), (r[9] or "")
        kel = " - ".join(str(x).strip() for x in (klp, sub1, sub2) if str(x or "").strip())
        items.append({"kode_barang": str(r[1] or "").strip(), "nama_barang": str(r[2] or "").strip(),
                      "kelompok": kel, "variant": str(r[11] or "").strip(), "gramasi": str(r[13] or "").strip()})
    return [it for it in items if it["kode_barang"]]


def letter_item_lines(pdf_path):
    import fitz
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return []
    lines = []
    for page in doc:
        for ln in page.get_text().splitlines():
            ln = " ".join(ln.split())
            if 6 < len(ln) < 90 and GRAM_RE.search(ln) and re.search(r"[A-Za-z]{3}", ln):
                lines.append(ln)
    # dedup
    seen, out = set(), []
    for ln in lines:
        k = ln.upper()
        if k in seen: continue
        seen.add(k); out.append(ln)
    return out[:80]


def _all_variant_union(ln_clean, items, um, rules):
    """Baris 'ALL VARIANT' (atau tie antar-rasa) = ambil SEMUA rasa kelompok
    itu pada gramasi tsb: kelompok yang token-nya subset dari token baris +
    gramasi sama -> match bila ada >=1 row."""
    want_g = um.extract_gramasi(ln_clean)
    if not want_g:
        return False
    ln_toks = um._tokenize(ln_clean, rules)
    for it in items:
        if um._norm_gramasi_value(it.get("gramasi", "")) != want_g:
            continue
        kel_toks = um._stem_tokens(um._canon(it.get("kelompok", ""), rules), rules)
        if not kel_toks:
            continue
        # dua arah: baris memuat seluruh nama kelompok (kel ⊆ ln), ATAU baris
        # hanya menulis sebagian nama kelompok tanpa kata asing (ln ⊆ kel,
        # mis. "BONEETO 685GR" utk kelompok "ANCHOR BONEETO").
        if kel_toks <= ln_toks or (ln_toks and ln_toks <= kel_toks):
            return True
    return False


def match_letters(masters):
    import urc_matcher as um
    rules = {"type_synonyms": {}, "token_map": {}}
    rep = []
    for f in sorted(os.listdir(SP)):
        if not f.lower().endswith(".pdf"):
            continue
        lines = letter_item_lines(os.path.join(SP, f))
        if not lines:
            rep.append((f, None, 0, 0, ["(tidak ada baris item terbaca dari teks native PDF -- kemungkinan tabel gambar, butuh OCR live)"]))
            continue
        best = None
        for pname, items in masters.items():
            ok, un = 0, []
            for ln in lines:
                lc = clean_letter_line(ln)
                try:
                    res = um.resolve_surat_line(lc, items, rules)
                    matched = not res.get("unmatched")
                    if not matched:
                        matched = _all_variant_union(lc, items, um, rules)
                except Exception:
                    matched = False
                if not matched:
                    un.append(ln)
                else:
                    ok += 1
            if best is None or ok > best[1]:
                best = (pname, ok, len(lines), un)
        # 0 match = tak ada principle cocok -> jangan sebut nama principle acak.
        pname_out = best[0] if best[1] > 0 else None
        rep.append((f, pname_out, best[1], best[2], best[3][:8]))
    return rep


if __name__ == "__main__":
    print("=== TAHAP 1: BANGUN MASTER ===")
    results = build_all()
    for pname, f, status, n, nogram in results:
        print(f"  {pname:22s} <- {f[:38]:40s} {status}" + (f" ({n} item, {nogram} tanpa-gramasi)" if n else ""))

    print("\n=== TAHAP 2: MUAT SEMUA MASTER JADI ===")
    masters = {}
    # master hasil rebuild (staging) + URC buatan tangan (dari MD)
    for src_dir, only in ((MD_OUT, None), (MD, {"URC"})):
        if not os.path.isdir(src_dir):
            continue
        for f in sorted(os.listdir(src_dir)):
            if not (f.upper().startswith("MASTER BARANG") and f.lower().endswith(".xlsx") and not f.startswith("~")):
                continue
            pname = f[len("MASTER BARANG "):-len(".xlsx")].strip()
            if only is not None and pname not in only:
                continue
            try:
                masters[pname] = load_master_items(os.path.join(src_dir, f))
                print(f"  {pname:22s} {len(masters[pname])} item")
            except Exception as e:
                print(f"  {pname:22s} LOAD FAIL {type(e).__name__}")
    # Priskila kanonik (hand-enriched, dipakai produksi) menggantikan hasil auto.
    _prisk = r"D:\AccAPI\MASTER BARANG PRISKILA.xlsx"
    if os.path.exists(_prisk):
        try:
            masters["PRISKILA"] = load_master_items(_prisk)
            print(f"  PRISKILA (kanonik)     {len(masters['PRISKILA'])} item <- D:\\AccAPI")
        except Exception as e:
            print(f"  PRISKILA kanonik LOAD FAIL {type(e).__name__}")

    print("\n=== TAHAP 3: COCOKKAN SURAT (teks native, 0 API) ===")
    rep = match_letters(masters)
    lines_out = []
    for f, pname, ok, total, un in rep:
        head = f"{f[:60]:62s} -> {pname or '-':18s} {ok}/{total}"
        print("  " + head)
        lines_out.append(head)
        for u in un:
            print("      UNMATCHED:", u[:80])
            lines_out.append("    UNMATCHED: " + u)

    os.makedirs(os.path.dirname(REPORT), exist_ok=True)
    with open(REPORT, "w", encoding="utf-8") as fo:
        fo.write("LAPORAN LOOP MASTER BUILDER (offline)\n\n")
        fo.write("== BUILD ==\n")
        for pname, f, status, n, nogram in results:
            fo.write(f"{pname} <- {f}: {status} ({n} item, {nogram} tanpa-gramasi)\n")
        fo.write("\n== MATCH SURAT ==\n")
        fo.write("\n".join(lines_out))
    print(f"\nreport: {REPORT}")
