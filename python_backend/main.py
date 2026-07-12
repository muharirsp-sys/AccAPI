# main.py (PATCH v12) — Channel lookup uploader + Internal dataset optional + Required-file guard UI/API
# Tujuan: FastAPI backend untuk validator, payments restore/SPPD, finance approval, RBAC, proof upload, dan helper export berformat.
# Caller: Next.js dashboard routes, browser uploads, dan service local AccAPI.
# Dependensi: FastAPI, pandas/openpyxl, payments.py, template DOCX SPPD, Better Auth SQLite DB, filesystem JSON/output, auth utilities.
# Main Functions: render_sppd_docx, payments_upload, payments_update, payments_clear, parse_lpb_upload, payments_template_download, payments_sppd_settings_get/save/upload, payments_finance_data, payments_finance_proof, payments_finance_update.
# Side Effects: HTTP response/download, file upload/read/write, payments.json backup/mutation, DOCX/XLSX generation with number/date formatting, audit logging.
# =======================================================================================================
# You requested:
# 1) Engine reads program by channel using lookup "Data Channel by SUB" + data penjualan.
# 2) Add uploader for:
#    - Data Channel by SUB (required)
#    - Dataset Diskon Internal (optional)
# 3) If user doesn't upload Data Penjualan / Dataset Diskon Pabrik / Data Channel:
#    - block validate
#    - show message box: "Data Penjualan masih kosong tuh, upload dulu dong adiks-adiks/kakaks-kakaks"
#    (We'll use the same message for any missing required file to keep UX simple; details shown in UI text.)
# 4) Dataset Diskon Internal may be empty/missing (allow validate anyway).
#
# Notes:
# - This patch keeps v11 "Program Lock for DISC_PCT (match MDSTRING)" + v9 satpam trigger-unit.
# - Internal dataset is loaded (if provided) but NOT applied to expected yet (as per your earlier request).
#   We keep it ready for next patch to move TanpaTuan -> Internal.
#
# Run:
#   python -m uvicorn main:app --reload --port 8000

from fastapi import FastAPI, UploadFile, File, Request, Form, Response, Cookie, BackgroundTasks
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse, ORJSONResponse
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import asyncio
import io, os, re, uuid, json, math, base64, hashlib, hmac, time, zipfile, copy, mimetypes
from dotenv import load_dotenv
load_dotenv(override=False)
import traceback
import xml.etree.ElementTree as ET
# Modul determinisme Summary Program (FASE 1-4): cache OCR, parser tier posisional,
# resolusi varian tabel, correction store stable-key. Lihat SYSTEM_MAP.md.
from variant_resolver import load_variant_mapping, resolve_variant
from correction_store import load_corrections as load_stable_corrections, apply_corrections as apply_stable_corrections, save_correction, correction_key
from tier_parser import regroup_rows_by_tier
from golden_store import canonical_signature, golden_check_and_freeze
from deterministic_output import enable_pdf_determinism, finalize_xlsx
from parse_cache import parse_cache_key, parse_cache_get, parse_cache_put
_VARIANT_MAPPING = load_variant_mapping()
from typing import List, Dict, Tuple, Optional, Set, Any
from urllib.parse import urlparse, unquote
from auth import (
    LoginRateLimiter,
    build_security_headers,
    enforce_non_default_auth_secret,
    is_production_env,
    login_rate_key,
    parse_bool_env,
)
from validator_engine import extract_pdf_text_safe, read_upload_file_limited
from payments import (
    lpb_upload_template_rows,
    validator_channel_template_rows,
    validator_promo_template_rows,
    validator_sales_template_rows,
)
from principle_matcher import (
    normalize_principle_name,
    get_principle_match_key,
    find_best_match,
    build_normalized_key_map,
    generate_import_report,
)
try:
    import bcrypt as _bcrypt  # type: ignore
except Exception:
    _bcrypt = None

try:
    from argon2 import PasswordHasher as _Argon2Hasher  # type: ignore
    _ARGON2 = _Argon2Hasher()
except Exception:
    _ARGON2 = None

PATCH_VERSION = "v15"
PATCH_TITLE = "Channel lookup required + internal discount engine (expected internal + allocation)"

REQUIRED_MISSING_MSG = "Data Penjualan masih kosong tuh, upload dulu dong adiks-adiks/kakaks-kakaks"

# Nginx X-Accel-Redirect (optional)
USE_X_ACCEL = str(os.getenv("USE_X_ACCEL", "1")).strip().lower() in ["1", "true", "yes", "y"]
X_ACCEL_PREFIX = str(os.getenv("X_ACCEL_PREFIX", "/protected-downloads")).strip()

# Auth (no storage) - set via env
AUTH_USERS = str(os.getenv("AUTH_USERS", "admin:admin")).strip()
AUTH_SECRET = str(os.getenv("AUTH_SECRET", "dev-secret")).strip()
AUTH_COOKIE = str(os.getenv("AUTH_COOKIE", "dv_auth")).strip()
AUTH_TTL_SECONDS = int(os.getenv("AUTH_TTL_SECONDS", "43200"))  # 12 hours
AUTH_COOKIE_SECURE = str(os.getenv("AUTH_COOKIE_SECURE", "1")).strip().lower() in ["1", "true", "yes", "y"]
AUTH_COOKIE_SAMESITE = str(os.getenv("AUTH_COOKIE_SAMESITE", "strict")).strip().lower()
AUTH_PASSWORD_SCHEME = str(os.getenv("AUTH_PASSWORD_SCHEME", "auto")).strip().lower()
AUTH_PBKDF2_ITERATIONS = int(os.getenv("AUTH_PBKDF2_ITERATIONS", "260000"))
AUTH_BCRYPT_ROUNDS = int(os.getenv("AUTH_BCRYPT_ROUNDS", "12"))
CSRF_COOKIE = str(os.getenv("CSRF_COOKIE", "dv_csrf")).strip()
CSRF_COOKIE_SAMESITE = str(os.getenv("CSRF_COOKIE_SAMESITE", "lax")).strip().lower()
CSRF_TTL_SECONDS = int(os.getenv("CSRF_TTL_SECONDS", "7200"))
APP_ENV = str(os.getenv("APP_ENV", "development")).strip().lower()
APP_DEBUG = parse_bool_env(os.getenv("APP_DEBUG"), default=False)
APP_IS_PRODUCTION = is_production_env(APP_ENV)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BETTER_AUTH_DB_PATH = str(os.getenv("BETTER_AUTH_DB_PATH", os.path.join(BASE_DIR, "..", "sqlite.db"))).strip()
LOGIN_BG_PATH = os.path.join(BASE_DIR, "login page", "bg.png")
AUTH_USERS_JSON = str(os.getenv("AUTH_USERS_JSON", os.path.join(BASE_DIR, "users.json"))).strip()
AUTH_ADMINS = set([u.strip() for u in str(os.getenv("AUTH_ADMINS", "admin")).split(",") if u.strip()])
AUTH_FINANCE = set([u.strip() for u in str(os.getenv("AUTH_FINANCE", "")).split(",") if u.strip()])
FAVICON_PATH = str(os.getenv("FAVICON_PATH", os.path.join(os.path.dirname(os.path.abspath(__file__)), "favicon.png"))).strip()
AUDIT_LOG_PATH = str(os.getenv("AUDIT_LOG_PATH", os.path.join(BASE_DIR, "data", "audit_log.jsonl"))).strip()
ERROR_LOG_PATH = str(os.getenv("ERROR_LOG_PATH", os.path.join(BASE_DIR, "data", "error_log.jsonl"))).strip()
PAYMENTS_DB_PATH = str(os.getenv("PAYMENTS_DB_PATH", os.path.join(BASE_DIR, "data", "payments.json"))).strip()
PAYMENTS_FILES_DIR = str(os.getenv("PAYMENTS_FILES_DIR", os.path.join(BASE_DIR, "output", "payments"))).strip()
PAYMENTS_PROOFS_DIR = str(os.getenv("PAYMENTS_PROOFS_DIR", os.path.join(PAYMENTS_FILES_DIR, "proofs"))).strip()
BANK_DATA_PATH = str(os.getenv("BANK_DATA_PATH", os.path.join(BASE_DIR, "Data No Rekening Principle.xlsx"))).strip()
SPPD_TEMPLATE_PATH = str(os.getenv("SPPD_TEMPLATE_PATH", os.path.join(BASE_DIR, "SPPD TGL 24 FEBRUARI 2026.docx"))).strip()
MAX_EXCEL_UPLOAD_BYTES = int(os.getenv("MAX_EXCEL_UPLOAD_BYTES", str(15 * 1024 * 1024)))
MAX_PROOF_UPLOAD_BYTES = int(os.getenv("MAX_PROOF_UPLOAD_BYTES", str(8 * 1024 * 1024)))
MAX_PDF_UPLOAD_BYTES = int(os.getenv("MAX_PDF_UPLOAD_BYTES", str(12 * 1024 * 1024)))
MAX_PDF_OCR_PAGES = int(os.getenv("MAX_PDF_OCR_PAGES", "8"))
MAX_PDF_OCR_TIMEOUT_SECONDS = int(os.getenv("MAX_PDF_OCR_TIMEOUT_SECONDS", "20"))
MAX_PDF_OCR_PAGE_TIMEOUT_SECONDS = int(os.getenv("MAX_PDF_OCR_PAGE_TIMEOUT_SECONDS", "4"))
LOGIN_MAX_FAILED_ATTEMPTS = int(os.getenv("LOGIN_MAX_FAILED_ATTEMPTS", "5"))
LOGIN_FAILED_WINDOW_SECONDS = int(os.getenv("LOGIN_FAILED_WINDOW_SECONDS", "300"))
LOGIN_LOCKOUT_SECONDS = int(os.getenv("LOGIN_LOCKOUT_SECONDS", "300"))

PERMISSION_MODULES = [
    "dashboard",
    "api_wrapper",
    "payments",
    "sppd",
    "finance",
    "principles",
    "summary",
    "validator",
    "users",
]
PERMISSION_ACTIONS = [
    "view",
    "create",
    "edit",
    "update",
    "delete",
    "upload",
    "export",
    "submit",
    "execute",
    "edit_settings",
    "upload_excel",
    "generate",
    "download",
    "approve",
    "transfer",
    "upload_proof",
    "post_accurate",
    "retry_post",
    "run",
    "email",
    "sync",
    "manage",
    "create_user",
    "edit_user",
    "delete_user",
    "set_role",
    "set_permission",
]

if APP_IS_PRODUCTION:
    enforce_non_default_auth_secret(AUTH_SECRET)

LOGIN_LIMITER = LoginRateLimiter(
    max_attempts=LOGIN_MAX_FAILED_ATTEMPTS,
    window_seconds=LOGIN_FAILED_WINDOW_SECONDS,
    cooldown_seconds=LOGIN_LOCKOUT_SECONDS,
)

PATCH_NOTES_HTML = r"""
<div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] shadow-lg overflow-hidden">
  <div class="p-5 border-b border-[#8A4703]">
    <div class="flex items-start justify-between gap-3">
      <div>
        <div class="text-xs uppercase tracking-widest text-[#1D1F1E]">Patch Notes</div>
        <div class="text-lg font-semibold text-[#1D1F1E]">PATCH v15</div>
        <div class="mt-1 text-sm text-[#1D1F1E]">Channel lookup + hitung diskon Internal (Expected_Internal + FixClaim_Internal).</div>
      </div>
      <div class="shrink-0 rounded-xl px-3 py-2 bg-[#BD7401] text-[#EFEFEF] text-xs font-semibold">
        Engine + UI
      </div>
    </div>
  </div>

  <div class="p-5 space-y-4">
    <div class="rounded-xl bg-[#EFEFEF] border border-[#8A4703] p-4">
      <div class="text-sm font-semibold text-[#1D1F1E]">Yang berubah di v15</div>
      <ul class="mt-2 text-sm text-[#1D1F1E] list-disc pl-5 space-y-1">
        <li><b>Internal aktif</b>: jika dataset diskon internal diupload, engine akan hitung <code>Expected_Internal</code>.</li>
        <li><b>Alokasi claim</b>: Actual discount mengikuti urutan diskon di <code>MDSTRING</code> untuk
          menentukan porsi <code>FixClaim_Pabrik</code> / <code>FixClaim_Internal</code>; sisanya <code>TanpaTuan</code>.
        </li>
        <li><b>Rule internal</b>: match by <code>CHANNEL</code> (hasil lookup SUB), optional by <code>SUB</code> dan <code>KODE_BARANG</code>, tier by <code>MIN_GROSS/MAX_GROSS</code>.</li>
        <li>Dataset internal tetap optional (kalau kosong, Expected_Internal=0).</li>
      </ul>
    </div>
  </div>
</div>
"""

app = FastAPI(title="Discount Validator API", version=f"PATCH-{PATCH_VERSION}")

# ----- Background Jobs Storage -----
# In a real enterprise app, use Redis/Celery. For now, in-memory dict mapped by job_id.
BACKGROUND_JOBS: Dict[str, Dict[str, Any]] = {}

from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1000)

# CORS allowlist — hanya origin frontend yang sah yang boleh memanggil API dengan
# credentials. Mengganti pola lama `allow_origin_regex="^https?://.*$"` yang
# mengizinkan SEMUA origin (celah pencurian data lintas-situs saat digabung
# allow_credentials=True).
# Sumber origin (prioritas): env CORS_ORIGINS (pisahkan koma) > NEXT_PUBLIC_APP_URL
# > localhost dev. Origin lokal dikunci ke http://localhost:3000 agar cookie
# auth tidak pecah karena campur localhost/127.0.0.1 atau port dev lain.
_cors_env = str(os.getenv("CORS_ORIGINS", "")).strip()
if _cors_env:
    CORS_ALLOWED_ORIGINS = [o.strip() for o in _cors_env.split(",") if o.strip()]
else:
    CORS_ALLOWED_ORIGINS = [
        o for o in [
            str(os.getenv("NEXT_PUBLIC_APP_URL", "")).strip(),
            "http://localhost:3000",
        ] if o
    ]

def is_allowed_cors_origin(origin: str) -> bool:
    try:
        parsed = urlparse(origin)
    except Exception:
        return False
    host = str(parsed.hostname or "").strip().lower()
    is_local = host in {"localhost", "127.0.0.1", "0.0.0.0"}
    return (not is_local) or origin == "http://localhost:3000"

CORS_ALLOWED_ORIGINS = [o for o in CORS_ALLOWED_ORIGINS if is_allowed_cors_origin(o)]
# Dedup sambil menjaga urutan.
CORS_ALLOWED_ORIGINS = list(dict.fromkeys(CORS_ALLOWED_ORIGINS))

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    forwarded_proto = s(request.headers.get("x-forwarded-proto", "")).lower()
    is_https = request.url.scheme == "https" or forwarded_proto == "https"
    for k, v in build_security_headers(is_https).items():
        if k not in response.headers:
            response.headers[k] = v
    return response

# ---------------------------
# Helpers
# ---------------------------
def s(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

def norm(x) -> str:
    return s(x).upper()

def ensure_col(df: pd.DataFrame, col: str, default):
    if col not in df.columns:
        df[col] = default
    return df

def to_float_series(df: pd.DataFrame, cols: List[str]):
    for c in cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    return df

def parse_number_id(x) -> float:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return 0.0
    if isinstance(x, (int, float)) and not pd.isna(x):
        return float(x)

    t = str(x).strip()
    if t == "":
        return 0.0

    t = re.sub(r"[^0-9\.\,\-]", "", t)

    # Both '.' and ',' => EU style "1.234,56"
    if "." in t and "," in t:
        t2 = t.replace(".", "").replace(",", ".")
        try:
            return float(t2)
        except:
            return 0.0

    # Only comma
    if "," in t and "." not in t:
        parts = t.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            t2 = t.replace(",", ".")
        else:
            t2 = t.replace(",", "")
        try:
            return float(t2)
        except:
            return 0.0

    # Only dot
    if "." in t and "," not in t:
        groups = t.split(".")
        if len(groups) >= 2 and all(g.isdigit() for g in groups if g != ""):
            if len(groups[-1]) == 3:
                t2 = "".join(groups)
                try:
                    return float(t2)
                except:
                    return 0.0
        try:
            return float(t)
        except:
            return 0.0

    try:
        return float(t)
    except:
        return 0.0

def packaging_to_int(packaging: str) -> int:
    packaging = s(packaging).upper()
    m = re.search(r"(\d+)", packaging)
    return int(m.group(1)) if m else 1

def qty_in_unit(row: pd.Series, unit: str) -> float:
    unit = norm(unit)
    qty = float(row.get("QTY", 0) or 0.0)
    qtypcs = float(row.get("QTYPCS", 0) or 0.0)
    u = norm(row.get("UNIT", ""))
    pack = packaging_to_int(row.get("PACKAGING", "")) or 1

    if unit == "PCS":
        if qtypcs > 0:
            return qtypcs
        if u in ["CTN", "KRT"]:
            return qty * float(pack)
        return qty

    if unit in ["CTN", "KRT"]:
        if u in ["CTN", "KRT"]:
            return qty
        pcs = qtypcs if qtypcs > 0 else qty
        return float(pcs) / float(pack)

    return qty

def parse_pct_chain(value: str) -> List[float]:
    value = s(value)
    if not value:
        return []
    parts = re.split(r"\s*\+\s*", value.replace("%", ""))
    out: List[float] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        try:
            out.append(float(p))
        except:
            pass
    return out

def parse_mdstring_pct(md: str) -> List[float]:
    md = s(md)
    if not md:
        return []
    nums = re.findall(r"(-?\d+(?:\.\d+)?)\s*%", md)
    if not nums:
        return parse_pct_chain(md)
    out = []
    for n in nums:
        try:
            out.append(float(n))
        except:
            pass
    return out

def normalize_pct_list(pcts) -> List[float]:
    if pcts is None:
        return []
    if isinstance(pcts, str):
        raw = parse_pct_chain(pcts)
    elif isinstance(pcts, (list, tuple, set)):
        raw = list(pcts)
    else:
        raw = [pcts]
    out: List[float] = []
    for p in raw:
        try:
            f = float(p)
        except:
            continue
        if abs(f) < 1e-9:
            continue
        out.append(round(f, 4))
    return out

def fmt_pct(p: float) -> str:
    if abs(p - round(p)) < 1e-6:
        return str(int(round(p)))
    return f"{p:.2f}".rstrip("0").rstrip(".")

def pct_list_str(pcts: List[float]) -> str:
    if not pcts:
        return ""
    return "+".join([fmt_pct(p) for p in pcts])

def pct_diff(md_parts: List[float], elig_parts: List[float]) -> Tuple[List[float], List[float]]:
    remaining = list(elig_parts)
    matched: List[float] = []
    missing: List[float] = []
    for p in md_parts:
        found = False
        for i, ep in enumerate(remaining):
            if abs(ep - p) < 1e-6:
                found = True
                matched.append(p)
                remaining.pop(i)
                break
        if not found:
            missing.append(p)
    return matched, missing

def fmt_metric(value: float, unit: str = "") -> str:
    v = float(value or 0.0)
    u = norm(unit)
    if u in ["PCS", "CTN", "KRT"]:
        return str(int(round(v)))
    if u == "GROSSAMOUNT":
        return f"{v:,.0f}"
    return f"{v:.2f}".rstrip("0").rstrip(".")

def find_favicon_path() -> Optional[str]:
    path = s(FAVICON_PATH)
    if not path:
        return None
    if os.path.isdir(path):
        for name in ["favicon.ico", "favicon.png", "favicon.jpg", "favicon.jpeg"]:
            p = os.path.join(path, name)
            if os.path.exists(p):
                return p
        return None
    if os.path.exists(path):
        return path
    for ext in [".ico", ".png", ".jpg", ".jpeg"]:
        p = path + ext
        if os.path.exists(p):
            return p
    return None

def favicon_media_type(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".ico":
        return "image/x-icon"
    if ext == ".png":
        return "image/png"
    if ext in [".jpg", ".jpeg"]:
        return "image/jpeg"
    return "application/octet-stream"

def extract_pdf_text(pdf_bytes: bytes) -> str:
    return extract_pdf_text_safe(
        pdf_bytes=pdf_bytes,
        max_pdf_bytes=MAX_PDF_UPLOAD_BYTES,
        max_pages=MAX_PDF_OCR_PAGES,
        total_timeout_seconds=MAX_PDF_OCR_TIMEOUT_SECONDS,
        per_page_ocr_timeout_seconds=MAX_PDF_OCR_PAGE_TIMEOUT_SECONDS,
    )

def normalize_ocr_text(text: str) -> str:
    # basic cleanup for OCR noise
    t = text.replace("\u2013", "-").replace("\u2014", "-")
    t = re.sub(r"[ \t]+", " ", t)
    return t

def find_line_with_pattern(lines: List[str], pattern: str) -> Optional[str]:
    for line in lines:
        if re.search(pattern, line, flags=re.IGNORECASE):
            return line
    return None

def parse_gumindo_program(text: str, list_mode: str) -> Dict[str, str]:
    raw = normalize_ocr_text(text)
    lines = [s(l) for l in raw.splitlines() if s(l)]

    surat_program = ""
    nama_program = ""
    for line in lines:
        m = re.search(r"\b\d{3,4}/[A-Z0-9]+/[A-Z0-9]+/[A-Z0-9]+/\d{2}\b", line, flags=re.IGNORECASE)
        if m:
            surat_program = m.group(0)
            tail = s(line.replace(surat_program, ""))
            if tail:
                nama_program = tail
            break

    periode = ""
    bulan = "(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)"
    for line in lines:
        if re.search(bulan, line, flags=re.IGNORECASE):
            periode = line
            break

    channel = ""
    for line in lines:
        if re.search(r"\b(GT|MT)\b", line, flags=re.IGNORECASE):
            channel = line
            break

    mekanisme_line = find_line_with_pattern(lines, r"Gramasi")
    if not mekanisme_line:
        mekanisme_line = find_line_with_pattern(lines, r"\bgr\b")

    mekanisme_text = ""
    mech_idx = None
    for i, line in enumerate(lines):
        if re.search(r"\bGramasi\b", line, flags=re.IGNORECASE):
            mech_idx = i
            break
    if mech_idx is None:
        for i, line in enumerate(lines):
            if re.search(r"\bgr\b", line, flags=re.IGNORECASE):
                mech_idx = i
                break

    if mech_idx is not None:
        block_lines = []
        for j in range(mech_idx, min(mech_idx + 6, len(lines))):
            line = lines[j]
            if j > mech_idx:
                if re.search(r"klaim.*mencantum|mencantum.*klaim", line, flags=re.IGNORECASE):
                    break
                if re.match(r"^\d+\s+", line):
                    if re.search(r"Surat|Cover|Klaim|Invoice|Harga|Tembusan|Rekap|Proses|Mengirim|Return", line, flags=re.IGNORECASE):
                        break
                if re.search(r"Demikian|Hormat|Tangerang|Terima kasih", line, flags=re.IGNORECASE):
                    break
            block_lines.append(line)
        mekanisme_text = " ".join(block_lines).strip()
    elif mekanisme_line:
        mekanisme_text = mekanisme_line.strip()

    gramasi = ""
    kelompok_barang = ""
    variant = ""
    ketentuan = ""
    benefit = ""

    def extract_ketentuan_benefit_lists(src: str) -> Tuple[List[str], List[str]]:
        if not src:
            return ([], [])
        t = normalize_ocr_text(src)
        t = re.sub(r"[=<>]", " ", t)
        t = re.sub(r"(ctn|krt|pcs)[\.\,]", r"\1", t, flags=re.IGNORECASE)
        t = re.sub(r"(ctn|krt|pcs)\s*u\s*p", r"\1 up", t, flags=re.IGNORECASE)
        t = re.sub(r"(ctn|krt|pcs)\s*up", r"\1 up", t, flags=re.IGNORECASE)

        unit_pat = r"(ctn|krt|pcs)"
        ketentuan_list = []
        ket_re = re.compile(
            r"(?:>=|=>|=|>|<)?\s*(\d+)\s*-\s*(\d+)\s*" + unit_pat + r"\b|(?:>=|=>|=|>|<)?\s*(\d+)\s*" + unit_pat + r"\s*(?:up|\+|ke\s*atas)\b",
            flags=re.IGNORECASE,
        )
        for m in ket_re.finditer(t):
            if m.group(1) and m.group(2) and m.group(3):
                ketentuan_list.append(f"{m.group(1)}-{m.group(2)} {m.group(3).lower()}")
            elif m.group(4) and m.group(5):
                ketentuan_list.append(f"{m.group(4)} {m.group(5).lower()} up")

        rp_vals = []
        for m in re.finditer(r"R\s*P\.?\s*([0-9][0-9\.\,]*)", t, flags=re.IGNORECASE):
            val = s(m.group(1))
            if val:
                rp_vals.append("Rp " + val)

        return (ketentuan_list, rp_vals)

    if mekanisme_text:
        mg = re.search(r"(\d+)\s*gr", mekanisme_text, flags=re.IGNORECASE)
        if mg:
            gramasi = mg.group(1) + " gr"
        after_gr = ""
        if mg:
            after_gr = mekanisme_text[mg.end():].strip()
        mword = re.search(r"\b([A-Za-z][A-Za-z0-9\-]+)\b", after_gr)
        if mword:
            kelompok_barang = mword.group(1)

        if "all variant" in s(nama_program).lower() or "all variant" in s(mekanisme_text).lower():
            variant = "All Variant"

        ket_list, ben_list = extract_ketentuan_benefit_lists(mekanisme_text)
        ketentuan = ", ".join(ket_list) if ket_list else ""
        benefit = ", ".join(ben_list) if ben_list else ""

        if not ketentuan or not benefit:
            rp_parts = re.split(r"\bRp\b", mekanisme_text, flags=re.IGNORECASE)
            if len(rp_parts) > 1:
                if not ketentuan:
                    ketentuan = s(rp_parts[0].replace("Gramasi", ""))
                if not benefit:
                    rp_vals = []
                    for part in rp_parts[1:]:
                        val = s(re.split(r"\s", part, maxsplit=1)[0])
                        if val:
                            rp_vals.append("Rp " + val)
                    if rp_vals:
                        benefit = " / ".join(rp_vals)

    mech_block = ""
    if not ketentuan or not benefit:
        m_start = re.search(r"mekan", raw, flags=re.IGNORECASE)
        if m_start:
            tail = raw[m_start.start():]
            m_end = re.search(r"klaim", tail, flags=re.IGNORECASE)
            mech_block = tail[:m_end.start()] if m_end else tail
        k2, b2 = extract_ketentuan_benefit_lists(mech_block)
        if not ketentuan:
            ketentuan = ", ".join(k2) if k2 else ""
        if not benefit:
            benefit = ", ".join(b2) if b2 else ""

    if not ketentuan or not benefit:
        k3, b3 = extract_ketentuan_benefit_lists(raw)
        if not ketentuan:
            ketentuan = ", ".join(k3) if k3 else ""
        if not benefit:
            benefit = ", ".join(b3) if b3 else ""

    ket_list_final, ben_list_final = extract_ketentuan_benefit_lists(mekanisme_text)
    if not ket_list_final and not ben_list_final:
        ket_list_final, ben_list_final = extract_ketentuan_benefit_lists(mech_block)
    if not ket_list_final and not ben_list_final:
        ket_list_final, ben_list_final = extract_ketentuan_benefit_lists(raw)

    syarat_claim = ""
    idx = None
    for i, line in enumerate(lines):
        if re.search(r"klaim.*mencantum", line, flags=re.IGNORECASE):
            idx = i
            break
    if idx is not None:
        syarat_lines = []
        for j in range(idx + 1, len(lines)):
            if re.search(r"Demikian|Hormat|Tangerang|Terima kasih", lines[j], flags=re.IGNORECASE):
                break
            if len(lines[j]) < 2:
                continue
            syarat_lines.append(lines[j])
        syarat_claim = "; ".join(syarat_lines)

    result = {
        "Surat Program": surat_program,
        "Nama Program": nama_program,
        "GT / MT": channel,
        "ADA LIST / TANPA LIST": list_mode,
        "Periode": periode,
        "Kelompok Barang": kelompok_barang,
        "Variant": variant,
        "Gramasi": gramasi,
        "Ketentuan Pengambilan": ketentuan,
        "Benefit": benefit,
        "Syarat Claim": syarat_claim,
        "Update": pd.Timestamp.today().strftime("%d-%m-%Y"),
        "Keterangan": "",
        "_KET_LIST": ket_list_final,
        "_BEN_LIST": ben_list_final,
    }
    return result

def build_summary_rows(text: str, list_mode: str, template: str) -> List[Dict[str, str]]:
    if template == "GUMINDO":
        base = parse_gumindo_program(text, list_mode)
        ket_list = base.pop("_KET_LIST", []) or []
        ben_list = base.pop("_BEN_LIST", []) or []
        if ket_list or ben_list:
            max_len = max(len(ket_list), len(ben_list))
            rows = []
            for i in range(max_len):
                row = dict(base)
                row["No."] = i + 1
                row["Ketentuan Pengambilan"] = ket_list[i] if i < len(ket_list) else ""
                row["Benefit"] = ben_list[i] if i < len(ben_list) else ""
                rows.append(row)
            return rows
        base["No."] = 1
        return [base]
    return []

def write_summary_excel(rows: List[Dict[str, str]], out_path: str):
    cols = [
        "No.",
        "Surat Program",
        "Nama Program",
        "GT / MT",
        "ADA LIST / TANPA LIST",
        "Periode",
        "Kelompok Barang",
        "Variant",
        "Gramasi",
        "Ketentuan Pengambilan",
        "Benefit",
        "Syarat Claim",
        "Update",
        "Keterangan",
    ]
    df = pd.DataFrame(rows, columns=cols)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SUMMARY")


# ---------------------------
# AI Summary (SumoPod OpenAI-compatible)
# ---------------------------
SUMOPOD_API_KEY = str(os.getenv("SUMOPOD_API_KEY", "")).strip()
SUMOPOD_BASE_URL = str(os.getenv("SUMOPOD_BASE_URL", "https://ai.sumopod.com/v1")).strip()
SUMOPOD_TIMEOUT = int(os.getenv("SUMOPOD_TIMEOUT", "120"))
DEFAULT_SUMMARY_AI_MODEL = str(os.getenv("SUMOPOD_MODEL", "kimi-k2-250905")).strip()

def _sumopod_url(path: str) -> str:
    base = SUMOPOD_BASE_URL.rstrip("/")
    return f"{base}/{path.lstrip('/')}"

def _strip_code_fences(t: str) -> str:
    t = (t or "").strip()
    # remove ```json ... ``` fences
    t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*```$", "", t)
    return t.strip()

def sumopod_chat_completion(
    model: str,
    messages: List[Dict[str, str]],
    temperature: float = 0.0,
    max_tokens: int = 1200,
) -> str:
    """
    Minimal OpenAI-compatible call to SumoPod using requests.
    """
    if not SUMOPOD_API_KEY:
        raise RuntimeError("SUMOPOD_API_KEY belum diset (export SUMOPOD_API_KEY='sk-...').")

    import requests  # requests is already used widely and usually available on VPS

    payload = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    r = requests.post(
        _sumopod_url("/chat/completions"),
        headers={
            "Authorization": f"Bearer {SUMOPOD_API_KEY}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=SUMOPOD_TIMEOUT,
    )
    if not r.ok:
        txt = (r.text or "")[:800]
        raise RuntimeError(f"SumoPod API error {r.status_code}: {txt}")

    data = r.json()
    try:
        return data["choices"][0]["message"]["content"]
    except Exception:
        raise RuntimeError("Format response SumoPod tidak sesuai (choices/message/content).")

def ai_extract_summary_rows(
    text: str,
    list_mode: str,
    template: str,
    model: Optional[str] = None,
) -> List[Dict[str, str]]:
    """
    Extract rows for write_summary_excel() from program letter text.
    Returns list of dicts with keys matching write_summary_excel columns.
    """
    model = s(model) or DEFAULT_SUMMARY_AI_MODEL
    template = s(template).upper()
    list_mode = s(list_mode).upper()

    cols = [
        "No.",
        "Surat Program",
        "Nama Program",
        "GT / MT",
        "ADA LIST / TANPA LIST",
        "Periode",
        "Kelompok Barang",
        "Variant",
        "Gramasi",
        "Ketentuan Pengambilan",
        "Benefit",
        "Syarat Claim",
        "Update",
        "Keterangan",
    ]

    sys_prompt = (
        "Kamu adalah asisten data yang mengekstrak isi surat program promo menjadi tabel ringkas. "
        "Keluaran HARUS JSON valid saja (tanpa penjelasan, tanpa markdown). "
        "Jika ragu, isi string kosong, tapi jangan mengarang angka. "
        "Jika ada beberapa ketentuan/benefit, buat beberapa baris (rows) dengan No. berurutan mulai 1."
    )

    user_prompt = f"""
Template: {template}
Mode List: {list_mode}

Ambil informasi dari teks surat program berikut dan hasilkan JSON dengan format:

{{
  "rows": [
    {{
      "No.": 1,
      "Surat Program": "...",
      "Nama Program": "...",
      "GT / MT": "GT" atau "MT" atau "",
      "ADA LIST / TANPA LIST": "{list_mode}",
      "Periode": "...",
      "Kelompok Barang": "...",
      "Variant": "...",
      "Gramasi": "...",
      "Ketentuan Pengambilan": "...",
      "Benefit": "...",
      "Syarat Claim": "...",
      "Update": "...",
      "Keterangan": "..."
    }}
  ],
  "confidence": 0-100
}}

Aturan:
- Jangan buat field tambahan selain "rows" dan "confidence".
- Setiap baris wajib punya semua kolom di atas (boleh kosong).
- "No." harus integer berurutan mulai 1.
- Jika ada bullet/daftar ketentuan/benefit, pecah menjadi beberapa baris.
- Jangan menebak tanggal/nominal kalau tidak tertulis.

Teks surat:
{text}
""".strip()

    content = sumopod_chat_completion(
        model=model,
        messages=[
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=0.0,
        max_tokens=1600,
    )

    raw = _strip_code_fences(content)
    # try to locate JSON object
    m = re.search(r"\{.*\}", raw, flags=re.S)
    if m:
        raw = m.group(0)

    data = json.loads(raw)
    rows = data.get("rows") if isinstance(data, dict) else None
    if not isinstance(rows, list):
        raise RuntimeError("AI output tidak berisi 'rows' list.")

    out: List[Dict[str, str]] = []
    for i, r in enumerate(rows, start=1):
        if not isinstance(r, dict):
            continue
        row: Dict[str, str] = {}
        for c in cols:
            row[c] = s(r.get(c, ""))
        # enforce No.
        try:
            row["No."] = int(r.get("No.", i))
        except Exception:
            row["No."] = i
        if not row.get("ADA LIST / TANPA LIST"):
            row["ADA LIST / TANPA LIST"] = list_mode
        out.append(row)

    return out


def normalize_permissions(raw) -> Dict[str, Set[str]]:
    perms: Dict[str, Set[str]] = {}
    if isinstance(raw, str):
        raw_text = s(raw)
        if not raw_text:
            return {}
        try:
            raw = json.loads(raw_text)
        except Exception:
            raw = [item.strip() for item in raw_text.replace(";", ",").split(",") if item.strip()]
    if isinstance(raw, dict) and "__custom" in raw:
        raw = raw.get("permissions", {})
    if isinstance(raw, dict):
        for mod, actions in raw.items():
            mod = s(mod).lower()
            if mod not in PERMISSION_MODULES:
                continue
            acts: List[str] = []
            if isinstance(actions, str):
                acts = [a.strip() for a in actions.replace(",", " ").split()]
            elif isinstance(actions, (list, tuple, set)):
                acts = [s(a) for a in actions]
            for a in acts:
                a = s(a).lower()
                if a in PERMISSION_ACTIONS:
                    perms.setdefault(mod, set()).add(a)
        return perms
    if isinstance(raw, (list, tuple, set)):
        for item in raw:
            item = s(item)
            if not item:
                continue
            if ":" in item:
                mod, act = item.split(":", 1)
            elif "." in item:
                mod, act = item.split(".", 1)
            else:
                continue
            mod = s(mod).lower()
            act = s(act).lower()
            if mod in PERMISSION_MODULES and act in PERMISSION_ACTIONS:
                perms.setdefault(mod, set()).add(act)
    return perms

def parse_permission_profile(raw) -> Tuple[Dict[str, Set[str]], bool]:
    if raw is None:
        return ({}, False)
    parsed = raw
    if isinstance(raw, str):
        raw_text = s(raw)
        if not raw_text or raw_text == "{}":
            return ({}, False)
        try:
            parsed = json.loads(raw_text)
        except Exception:
            return (normalize_permissions(raw_text), True)
    if isinstance(parsed, dict) and parsed.get("__custom") is True:
        return (normalize_permissions(parsed.get("permissions", {})), True)
    if isinstance(parsed, dict) and parsed.get("__custom") is False:
        return ({}, False)
    perms = normalize_permissions(parsed)
    return (perms, bool(perms))



def get_auth_user_records() -> Dict[str, Dict[str, str]]:
    # ponytail: store kredensial paralel (users.json + env AUTH_USERS) dimatikan (#7).
    # Identitas & RBAC kini hanya dari Better Auth (sqlite.db user/session). Fungsi
    # dipertahankan agar pemanggil legacy aman; selalu kosong.
    return {}

def get_auth_user_map() -> Dict[str, str]:
    records = get_auth_user_records()
    return {u: s(info.get("password", "")) for u, info in records.items() if s(info.get("password", ""))}

def get_user_role(username: str) -> str:
    username = s(username)
    if not username:
        return "user"
    if username.startswith("betterauth|"):
        parts = username.split("|", 2)
        role = s(parts[1] if len(parts) > 1 else "").lower()
        if role in {"admin", "manager", "finance", "staff", "viewer"}:
            return role
        return "viewer"
    rec = get_auth_user_records().get(username)
    if rec:
        role = s(rec.get("role", ""))
        if role:
            return role
    if username in AUTH_ADMINS:
        return "admin"
    if username in AUTH_FINANCE:
        return "finance"
    return "user"

def _b64url_encode(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).rstrip(b"=").decode("ascii")

def _b64url_decode(sv: str) -> bytes:
    pad = "=" * (-len(sv) % 4)
    return base64.urlsafe_b64decode(sv + pad)

def _normalize_samesite(value: str) -> str:
    v = s(value).lower()
    if v in ["lax", "strict", "none"]:
        return v
    return "strict"


def _pbkdf2_hash(password: str, iterations: Optional[int] = None, salt: Optional[bytes] = None) -> str:
    iters = int(iterations or AUTH_PBKDF2_ITERATIONS)
    salt_bytes = salt or os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt_bytes, iters)
    return f"pbkdf2${iters}${_b64url_encode(salt_bytes)}${_b64url_encode(dk)}"

def _pbkdf2_verify(password: str, stored: str) -> bool:
    parts = s(stored).split("$")
    if len(parts) != 4:
        return False
    try:
        iters = int(parts[1])
        salt = _b64url_decode(parts[2])
        expected = _b64url_decode(parts[3])
    except Exception:
        return False
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iters)
    return hmac.compare_digest(dk, expected)



def make_csrf_token() -> str:
    exp = int(time.time()) + int(CSRF_TTL_SECONDS)
    nonce = _b64url_encode(os.urandom(16))
    payload = f"{nonce}|{exp}"
    sig = hmac.new(AUTH_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{_b64url_encode(payload.encode('utf-8'))}.{sig}"

def validate_csrf_token(token: str) -> bool:
    token = s(token)
    if not token or "." not in token:
        return False
    payload_b64, sig = token.split(".", 1)
    try:
        payload = _b64url_decode(payload_b64).decode("utf-8")
    except Exception:
        return False
    expected_sig = hmac.new(AUTH_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sig, expected_sig):
        return False
    if "|" not in payload:
        return False
    _nonce, exp_s = payload.rsplit("|", 1)
    try:
        exp = int(exp_s)
    except Exception:
        return False
    return exp >= int(time.time())

def get_or_create_csrf_token(request: Request) -> str:
    token = s(request.cookies.get(CSRF_COOKIE, ""))
    if token and validate_csrf_token(token):
        return token
    return make_csrf_token()

def validate_csrf_request(request: Request, token: str) -> bool:
    token = s(token)
    cookie_token = s(request.cookies.get(CSRF_COOKIE, ""))
    if not token:
        return is_same_origin_request(request)
    if not validate_csrf_token(token):
        return False
    if cookie_token and hmac.compare_digest(token, cookie_token):
        return True
    # Fallback: allow signed token for same-origin requests to reduce false negatives
    # when browser cookies are stale/not yet synchronized.
    return is_same_origin_request(request)


def is_same_origin_request(request: Request) -> bool:
    host = s(request.headers.get("host", ""))
    if not host:
        return False
    host = host.lower()
    for hdr in ["origin", "referer"]:
        raw = s(request.headers.get(hdr, ""))
        if not raw:
            continue
        try:
            p = urlparse(raw)
        except Exception:
            return False
        src_host = s(p.netloc).lower()
        if not src_host:
            return False
        
        # Allow exact match
        if src_host == host:
            continue
            
        # Allow locked local frontend (localhost:3000) to call the local backend.
        src_hostname = src_host.split(":")[0] if ":" in src_host else src_host
        dst_hostname = host.split(":")[0] if ":" in host else host
        if src_host == "localhost:3000" and dst_hostname == "localhost":
            continue

        return False
    return True


def get_current_user(request: Request) -> Optional[str]:
    # ponytail: auth Python paralel dihapus (#7) — satu-satunya sumber identitas
    # adalah sesi Better Auth (cookie better-auth.session_token) divalidasi ke sqlite.db.
        
    # 2. Try Better-Auth SQLite Session
    ba_token = None
    raw_cookie = request.headers.get("cookie", "")
    if raw_cookie:
        for chunk in raw_cookie.split(";"):
            chunk = chunk.strip()
            if chunk.startswith("better-auth.session_token="):
                ba_token = unquote(chunk.split("=", 1)[1]).split(".")[0]
                break
            elif chunk.startswith("__Secure-better-auth.session_token="):
                ba_token = unquote(chunk.split("=", 1)[1]).split(".")[0]
                break
    
    # print(f"[DEBUG AUTH] Manual ba_token: {ba_token}")
    
    if ba_token:
        try:
            import sqlite3
            db_path = BETTER_AUTH_DB_PATH
            if os.path.exists(db_path):
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute('''
                    SELECT user.email, user.name, COALESCE(user.role, 'viewer'), session.expiresAt
                    FROM session 
                    JOIN user ON session.userId = user.id 
                    WHERE session.token = ?
                ''', (ba_token,))
                row = c.fetchone()
                conn.close()
                if row:
                    email, name, role, expiresAt = row
                    import time
                    now = time.time()
                    exp = expiresAt / 1000.0 if expiresAt > 20000000000 else expiresAt
                    if exp > now:
                        role = s(role).lower() or "viewer"
                        if role not in {"admin", "manager", "finance", "staff", "viewer"}:
                            role = "viewer"
                        return f"betterauth|{role}|{s(email).lower() or s(name)}"
        except Exception as e:
            print(f"Error checking Better-Auth session: {e}")

    return None

def is_admin_user(username: Optional[str]) -> bool:
    if not username:
        return False
    return get_user_role(username) == "admin"

def is_finance_user(username: Optional[str]) -> bool:
    if not username:
        return False
    return get_user_role(username) == "finance"

def get_user_permissions_info(username: str) -> Tuple[Dict[str, Set[str]], bool]:
    username = s(username)
    if not username:
        return ({}, False)
    if username.startswith("betterauth|"):
        parts = username.split("|", 2)
        email = s(parts[2] if len(parts) > 2 else "").lower()
        if not email:
            return ({}, False)
        try:
            import sqlite3
            if not os.path.exists(BETTER_AUTH_DB_PATH):
                return ({}, False)
            conn = sqlite3.connect(BETTER_AUTH_DB_PATH)
            c = conn.cursor()
            c.execute("SELECT permissions FROM user WHERE lower(email) = ? LIMIT 1", (email,))
            row = c.fetchone()
            conn.close()
            if not row:
                return ({}, False)
            return parse_permission_profile(row[0])
        except Exception as e:
            append_error_log("get_user_permissions_info_betterauth", e, {"email": email})
            return ({}, False)
    rec = get_auth_user_records().get(username)
    if not rec:
        return ({}, False)
    raw = rec.get("permissions", None)
    return parse_permission_profile(raw)

def user_has_permission(username: Optional[str], module: str, action: str) -> bool:
    if not username:
        return False
    if is_admin_user(username):
        return True
    module = s(module).lower()
    action = s(action).lower()
    role = get_user_role(username)
    role_permissions = {
        "admin": {mod: set(PERMISSION_ACTIONS) for mod in PERMISSION_MODULES},
        "manager": {
            "dashboard": {"view"},
            "api_wrapper": {"view", "execute"},
            "payments": {"view", "export", "submit", "edit", "update"},
            "sppd": {"view", "generate", "download"},
            "finance": {"view", "approve", "export"},
            "principles": {"view"},
            "summary": {"view", "export"},
            "validator": {"view", "download"},
        },
        "finance": {
            "dashboard": {"view"},
            "payments": {"view", "export"},
            "sppd": {"view", "download"},
            "finance": {"view", "approve", "transfer", "upload_proof", "post_accurate", "retry_post", "export", "update", "edit"},
            "principles": {"view"},
        },
        "staff": {
            "dashboard": {"view"},
            "payments": {"view", "create", "edit", "update", "upload", "submit"},
            "sppd": {"view", "generate", "download"},
            "principles": {"view"},
            "summary": {"view", "upload", "generate", "export", "edit", "update"},
            "validator": {"view", "upload", "run", "download", "edit"},
        },
        "viewer": {
            "dashboard": {"view"},
            "validator": {"view"},
            "summary": {"view"},
            "payments": {"view"},
            "sppd": {"view"},
            "finance": {"view"},
        },
    }
    perms, defined = get_user_permissions_info(username)
    if defined:
        allowed = perms.get(module, set())
        return action in allowed
    if role in role_permissions:
        return action in role_permissions.get(role, {}).get(module, set())
    if module == "finance":
        return is_finance_user(username)
    return True

def format_permissions(perms: Optional[Dict[str, Set[str]]], defined: bool) -> str:
    if not defined:
        return "default"
    if not perms:
        return "none"
    parts = []
    for mod in PERMISSION_MODULES:
        acts = sorted(perms.get(mod, set()))
        if acts:
            parts.append(f"{mod}:{','.join(acts)}")
    return "; ".join(parts) if parts else "none"

def append_audit_log(user: str, action: str, entity: str, details: Optional[Dict[str, Any]] = None) -> None:
    if not AUDIT_LOG_PATH:
        return
    entry = {
        "ts": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": s(user),
        "action": s(action),
        "entity": s(entity),
        "details": details or {},
    }
    try:
        os.makedirs(os.path.dirname(AUDIT_LOG_PATH), exist_ok=True)
        with open(AUDIT_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=True) + "\n")
    except Exception:
        pass


def append_error_log(where: str, err: Exception, context: Optional[Dict[str, Any]] = None) -> None:
    if not ERROR_LOG_PATH:
        return
    entry = {
        "ts": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "where": s(where),
        "error": s(str(err)),
        "context": context or {},
        "trace": traceback.format_exc(),
    }
    try:
        os.makedirs(os.path.dirname(ERROR_LOG_PATH), exist_ok=True)
        with open(ERROR_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=True) + "\n")
    except Exception:
        pass

# ---------------------------
# Payments (LPB) helpers
# ---------------------------
_PAYMENTS_DB_CACHE: Optional[Dict[str, Any]] = None
_PAYMENTS_DB_MTIME: float = 0.0
_PAYMENTS_DB_EMPTY: Dict[str, Any] = {"lpb": {}, "submissions": {}, "drafts": {}, "finance_mappings": {}, "proofs": {}, "sppd_settings": {}}
# ponytail: satu lock global cukup karena payments.json adalah satu file tunggal.
# Ceiling: serializes semua write — tidak ada parallelism untuk mutation routes.
# Upgrade path: per-resource lock jika ada multiple JSON stores.
_PAYMENTS_DB_LOCK: asyncio.Lock = asyncio.Lock()

def load_payments_db() -> Dict[str, Any]:
    global _PAYMENTS_DB_CACHE, _PAYMENTS_DB_MTIME
    if not PAYMENTS_DB_PATH or not os.path.exists(PAYMENTS_DB_PATH):
        return dict(_PAYMENTS_DB_EMPTY)
    try:
        mtime = os.path.getmtime(PAYMENTS_DB_PATH)
        if _PAYMENTS_DB_CACHE is not None and mtime == _PAYMENTS_DB_MTIME:
            return _PAYMENTS_DB_CACHE
        with open(PAYMENTS_DB_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return dict(_PAYMENTS_DB_EMPTY)
    if not isinstance(data, dict):
        return dict(_PAYMENTS_DB_EMPTY)
    data.setdefault("lpb", {})
    data.setdefault("submissions", {})
    data.setdefault("drafts", {})
    data.setdefault("finance_mappings", {})
    data.setdefault("proofs", {})
    data.setdefault("sppd_settings", {})
    _PAYMENTS_DB_CACHE = data
    _PAYMENTS_DB_MTIME = mtime
    return data

def save_payments_db(data: Dict[str, Any]) -> None:
    global _PAYMENTS_DB_CACHE, _PAYMENTS_DB_MTIME
    if not PAYMENTS_DB_PATH:
        return
    os.makedirs(os.path.dirname(PAYMENTS_DB_PATH), exist_ok=True)
    tmp_path = PAYMENTS_DB_PATH + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=True, indent=2)
    os.replace(tmp_path, PAYMENTS_DB_PATH)
    _PAYMENTS_DB_CACHE = data
    _PAYMENTS_DB_MTIME = os.path.getmtime(PAYMENTS_DB_PATH)

async def load_and_lock_payments_db():
    """Acquire _PAYMENTS_DB_LOCK lalu load payments.json.

    Gunakan sebagai async context manager di semua route yang melakukan
    read-modify-write agar tidak ada dua request concurrent yang menulis
    snapshot berbeda ke file yang sama (last-writer-wins silent data loss).

    Contoh pemakaian:
        async with load_and_lock_payments_db() as db:
            db["lpb"][key] = rec
            save_payments_db(db)
    """
    # ponytail: yield inside asyncio.Lock context — satu lock serializes semua
    # mutation. Ceiling: throughput mutation routes turun jadi sequential.
    # Upgrade: per-key optimistic locking jika contention terbukti tinggi.
    class _Ctx:
        def __init__(self):
            self._db: Dict[str, Any] = {}
        async def __aenter__(self) -> Dict[str, Any]:
            await _PAYMENTS_DB_LOCK.acquire()
            self._db = load_payments_db()
            return self._db
        async def __aexit__(self, exc_type, exc, tb):
            _PAYMENTS_DB_LOCK.release()
    return _Ctx()

def empty_payments_db_preserving_config(db: Dict[str, Any]) -> Dict[str, Any]:
    db = db if isinstance(db, dict) else {}
    cleared: Dict[str, Any] = {
        "lpb": {},
        "submissions": {},
        "drafts": {},
        "finance_mappings": db.get("finance_mappings", {}),
        "proofs": {},
        "sppd_settings": db.get("sppd_settings", {}),
    }
    if "sppd_seq" in db:
        cleared["sppd_seq"] = db.get("sppd_seq")
    return cleared

def normalize_lpb_no(no_lpb: str) -> str:
    return s(no_lpb).upper()


def normalize_pengajuan_type(raw: str) -> str:
    t = s(raw).upper()
    if t in ["LPB", "CBD", "NON_LPB"]:
        return t
    return "LPB"


def make_payment_record_id(prefix: str = "RID") -> str:
    base = re.sub(r"[^A-Z0-9_]+", "", s(prefix).upper()) or "RID"
    return f"{base}_{str(uuid.uuid4())[:8]}".upper()


def resolve_payment_record_key(db: Dict[str, Any], row_id: str) -> str:
    row_id = s(row_id)
    if not row_id:
        return ""
    if row_id in db.get("lpb", {}):
        return row_id
    row_norm = normalize_lpb_no(row_id)
    if row_norm in db.get("lpb", {}):
        return row_norm
    for k, r in db.get("lpb", {}).items():
        if normalize_lpb_no(s(r.get("no_lpb", ""))) == row_norm:
            return k
    return ""


def find_lpb_duplicate_key(db: Dict[str, Any], no_lpb: str, exclude_key: str = "") -> str:
    target = normalize_lpb_no(no_lpb)
    if not target:
        return ""
    exclude_key = s(exclude_key)
    for k, r in db.get("lpb", {}).items():
        if exclude_key and k == exclude_key:
            continue
        existing = normalize_lpb_no(s(r.get("no_lpb", "")) or s(k))
        if existing == target:
            return k
    return ""


def has_submitted_duplicate_payment(db: Dict[str, Any], current_key: str, rec: Dict[str, Any]) -> bool:
    principle = s(rec.get("principle", "")).upper()
    invoice_no = s(rec.get("invoice_no", "")).upper()
    nilai_invoice = parse_number_id(rec.get("nilai_invoice", rec.get("nilai_principle", 0)))
    for k, other in db.get("lpb", {}).items():
        if s(k) == s(current_key):
            continue
        if not s(other.get("submission_id", "")):
            continue
        if s(other.get("status_pembayaran", "")).lower() == "ajukan ulang":
            continue
        if normalize_pengajuan_type(other.get("tipe_pengajuan", "LPB")) != "CBD":
            continue
        if s(other.get("principle", "")).upper() != principle:
            continue
        other_invoice = s(other.get("invoice_no", "")).upper()
        if invoice_no and other_invoice and invoice_no == other_invoice:
            return True
        other_paid = parse_number_id(other.get("nilai_pembayaran", other.get("nilai_invoice", other.get("nilai_principle", 0))))
        if (not invoice_no or not other_invoice) and abs(float(other_paid or 0.0) - float(nilai_invoice or 0.0)) <= 1.0:
            return True
    return False

def to_date_str(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return s(val)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return s(val)

def parse_sppd_date_ddmmyyyy(value) -> str:
    """Normalkan tanggal import Excel SPPD ke YYYY-MM-DD dengan asumsi urutan
    DD/MM/YYYY untuk string. Tolak tanggal mustahil/rusak (raise ValueError).

    - datetime/Timestamp asli (sel tanggal Excel) -> dipakai apa adanya (tak ambigu).
    - "YYYY/MM/DD" (tahun 4-digit di depan); tukar ke YYYY/DD/MM hanya bila bulan>12.
    - "DD/MM/YYYY" (tahun 4-digit di belakang); default day-first. Tukar ke MM/DD
      hanya bila tak ambigu (bulan-slot >12 & hari-slot <=12).
    - Ambigu (keduanya <=12) -> tetap DD/MM (tidak ditukar).
    - Pemisah `-`, `/`, atau `.`. Format/komponen lain -> ValueError.
    """
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    from datetime import datetime as _dt, date as _date
    if isinstance(value, (_dt, _date)):
        return value.strftime("%Y-%m-%d")
    raw = s(value).strip()
    if not raw:
        return ""
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", raw)
    if m:
        y, a, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
        mo, da = a, b
        if a > 12 and b <= 12:  # YYYY/DD/MM -> tukar
            mo, da = b, a
    else:
        m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", raw)
        if not m:
            raise ValueError(f"'{raw}'")
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        da, mo = a, b  # default DD/MM
        if b > 12 and a <= 12:  # sebenarnya MM/DD -> tukar
            da, mo = b, a
    try:
        parsed = _date(y, mo, da)
    except ValueError:
        raise ValueError(f"'{raw}'")
    return parsed.strftime("%Y-%m-%d")

def parse_lpb_upload(content: bytes) -> List[Dict[str, Any]]:
    df = pd.read_excel(io.BytesIO(content))
    cols = {c.strip().upper(): c for c in df.columns}
    required = ["TGL. SETOR", "NO. LPB", "TGL. WIN", "TGL. J. TEMPO WIN", "PRINCIPLE", "NILAI WIN", "TGL TERIMA BARANG"]
    missing = [c for c in required if c not in cols]
    if missing:
        raise ValueError("Kolom wajib tidak lengkap: " + ", ".join(missing))

    out = []
    date_errors: List[str] = []

    def parse_lpb_upload_date(row, excel_row: int, *names: str) -> str:
        col = _col_lookup(cols, *names)
        if col is None:
            return ""
        try:
            return parse_sppd_date_ddmmyyyy(row[col])
        except ValueError as de:
            date_errors.append(f"kolom '{s(col)}' baris {excel_row} = {str(de)}")
            return ""

    for idx, (_, r) in enumerate(df.iterrows()):
        excel_row = idx + 2
        no_lpb = s(r[cols["NO. LPB"]])
        if not no_lpb:
            continue
        nilai_win = parse_number_id(r[cols["NILAI WIN"]])
        nilai_invoice = parse_number_id(_row_value(r, cols, "Nilai Invoice", "NILAI INVOICE", default=0))
        tgl_setor = parse_lpb_upload_date(r, excel_row, "TGL. SETOR")
        tgl_win = parse_lpb_upload_date(r, excel_row, "TGL. WIN")
        tgl_jtempo_win = parse_lpb_upload_date(r, excel_row, "TGL. J. TEMPO WIN")
        tgl_terima_barang = parse_lpb_upload_date(r, excel_row, "TGL TERIMA BARANG")
        tgl_invoice = parse_lpb_upload_date(r, excel_row, "Tgl Invoice", "TGL INVOICE", "TGL. INVOICE")
        jt_invoice = parse_lpb_upload_date(r, excel_row, "J.T Invoice", "J.T INVOICE", "JT INVOICE", "JATUH TEMPO INVOICE")
        actual_date = parse_lpb_upload_date(r, excel_row, "Actual Date", "ACTUAL DATE")
        tgl_pembayaran = parse_lpb_upload_date(r, excel_row, "Tgl Pembayaran", "TGL PEMBAYARAN")
        out.append({
            "no_lpb": no_lpb,
            "tgl_setor": tgl_setor,
            "tgl_win": tgl_win,
            "tgl_jtempo_win": tgl_jtempo_win,
            "principle": s(r[cols["PRINCIPLE"]]),
            "nilai_win": nilai_win,
            "tgl_terima_barang": tgl_terima_barang,
            "tgl_invoice": tgl_invoice,
            "invoice_no": s(_row_value(r, cols, "No Invoice", "NO INVOICE", "NO. INVOICE")),
            "nilai_invoice": nilai_invoice if nilai_invoice > 0 else "",
            "jt_invoice": jt_invoice,
            "actual_date": actual_date,
            "tgl_pembayaran": tgl_pembayaran,
            "gap_nilai": (nilai_win - nilai_invoice) if nilai_invoice > 0 else 0.0,
            "jenis_dokumen": s(_row_value(r, cols, "Jenis Dokumen", "JENIS DOKUMEN")),
            "nomor_dokumen": s(_row_value(r, cols, "Nomor Dokumen", "NOMOR DOKUMEN")),
            "keterangan": s(_row_value(r, cols, "Keterangan", "KETERANGAN")),
        })
    if date_errors:
        shown = "; ".join(date_errors[:5])
        extra = len(date_errors) - 5
        suffix = f"; dan {extra} lainnya" if extra > 0 else ""
        raise ValueError(
            f"Tanggal tidak valid (harus DD/MM/YYYY): {shown}{suffix}. Upload dibatalkan."
        )
    return out

def _col_lookup(cols: Dict[str, Any], *names: str) -> Optional[Any]:
    for name in names:
        key = str(name).strip().upper()
        if key in cols:
            return cols[key]
    return None

def to_datetime_str(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return s(val)
        if dt.hour or dt.minute or dt.second:
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return s(val)

def _row_value(row, cols: Dict[str, Any], *names: str, default: Any = "") -> Any:
    col = _col_lookup(cols, *names)
    if col is None:
        return default
    try:
        return row[col]
    except Exception:
        return default

def looks_like_payments_backup(cols: Dict[str, Any]) -> bool:
    required = ["RECORD ID", "TIPE PENGAJUAN", "PRINCIPLE", "NILAI INVOICE", "STATUS PEMBAYARAN"]
    return all(c in cols for c in required)

def parse_payments_backup_upload(content: bytes) -> List[Tuple[str, Dict[str, Any]]]:
    xl = pd.ExcelFile(io.BytesIO(content))
    sheet_name = "PAYMENTS" if "PAYMENTS" in xl.sheet_names else xl.sheet_names[0]
    df = xl.parse(sheet_name)
    cols = {str(c).strip().upper(): c for c in df.columns}
    if not looks_like_payments_backup(cols):
        raise ValueError("Format backup PAYMENTS tidak valid.")

    rows: List[Tuple[str, Dict[str, Any]]] = []
    date_errors: List[str] = []

    def parse_backup_date(r, excel_row: int, *names: str) -> str:
        col = _col_lookup(cols, *names)
        if col is None:
            return ""
        try:
            return parse_sppd_date_ddmmyyyy(r[col])
        except ValueError as de:
            date_errors.append(f"kolom '{s(col)}' baris {excel_row} = {str(de)}")
            return ""

    for idx, (_, r) in enumerate(df.iterrows()):
        excel_row = idx + 2
        raw_key = s(_row_value(r, cols, "Record ID"))
        tipe = normalize_pengajuan_type(_row_value(r, cols, "Tipe Pengajuan", default="LPB"))
        no_lpb = s(_row_value(r, cols, "No LPB"))
        principle = s(_row_value(r, cols, "Principle"))
        invoice_no = s(_row_value(r, cols, "No Invoice"))
        jenis_dokumen = s(_row_value(r, cols, "Jenis Dokumen"))
        nomor_dokumen = s(_row_value(r, cols, "Nomor Dokumen"))
        if not raw_key:
            raw_key = normalize_lpb_no(no_lpb) if no_lpb else make_payment_record_id(tipe)
        key = raw_key
        if not key:
            continue
        nilai_win = parse_number_id(_row_value(r, cols, "Nilai Sistem", "Nilai WIN", default=0))
        nilai_invoice = parse_number_id(_row_value(r, cols, "Nilai Invoice", default=0))
        potongan = parse_number_id(_row_value(r, cols, "Potongan", default=0))
        nilai_pembayaran = parse_number_id(_row_value(r, cols, "Nilai Pembayaran", default=0))
        gap_nilai = parse_number_id(_row_value(r, cols, "Gap Nilai", default=nilai_win - nilai_invoice))
        rec = {
            "record_id": key,
            "tipe_pengajuan": tipe,
            "no_lpb": no_lpb,
            "tgl_setor": parse_backup_date(r, excel_row, "Tgl Setor", "TGL. SETOR"),
            "tgl_win": parse_backup_date(r, excel_row, "Tgl Win", "TGL. WIN"),
            "tgl_jtempo_win": parse_backup_date(r, excel_row, "Tgl J.Tempo Win", "TGL. J. TEMPO WIN"),
            "principle": principle,
            "nilai_win": nilai_win,
            "tgl_terima_barang": parse_backup_date(r, excel_row, "Tgl Terima Barang", "TGL TERIMA BARANG"),
            "tgl_invoice": parse_backup_date(r, excel_row, "Tgl Invoice"),
            "jt_invoice": parse_backup_date(r, excel_row, "J.T Invoice"),
            "tgl_pembayaran": parse_backup_date(r, excel_row, "Tgl Pembayaran"),
            "actual_date": parse_backup_date(r, excel_row, "Actual Date"),
            "nilai_invoice": nilai_invoice,
            "gap_nilai": gap_nilai,
            "invoice_no": invoice_no,
            "status_pembayaran": s(_row_value(r, cols, "Status Pembayaran")),
            "payment_method": s(_row_value(r, cols, "Metode Pembayaran")),
            "submitted_at": to_datetime_str(_row_value(r, cols, "Submitted At")),
            "submitted_by": s(_row_value(r, cols, "Submitted By")),
            "submission_id": s(_row_value(r, cols, "Submission ID")),
            "draft_id": s(_row_value(r, cols, "Draft ID")),
            "potongan": potongan,
            "nilai_pembayaran": nilai_pembayaran,
            "target_payment_date": parse_backup_date(r, excel_row, "Tanggal Pengajuan Pembayaran"),
            "jenis_pembayaran": s(_row_value(r, cols, "Jenis Pembayaran")),
            "jenis_dokumen": jenis_dokumen,
            "nomor_dokumen": nomor_dokumen,
            "keterangan": s(_row_value(r, cols, "Keterangan")),
            "sppd_no": s(_row_value(r, cols, "SPPD No")),
            "created_at": to_datetime_str(_row_value(r, cols, "Created At")),
            "created_by": s(_row_value(r, cols, "Created By")),
        }
        rows.append((key, rec))
    if date_errors:
        shown = "; ".join(date_errors[:5])
        extra = len(date_errors) - 5
        suffix = f"; dan {extra} lainnya" if extra > 0 else ""
        raise ValueError(
            f"Tanggal tidak valid (harus DD/MM/YYYY): {shown}{suffix}. Restore backup dibatalkan."
        )
    return rows

def max_sppd_sequence_from_records(records: List[Dict[str, Any]]) -> int:
    max_seq = 0
    for rec in records:
        sppd_no = s(rec.get("sppd_no", ""))
        m = re.match(r"^\s*(\d+)\s*/", sppd_no)
        if not m:
            continue
        try:
            max_seq = max(max_seq, int(m.group(1)))
        except Exception:
            continue
    return max_seq

def rebuild_payment_submissions(db: Dict[str, Any]) -> None:
    submissions = dict(db.get("submissions", {}) or {})
    grouped: Dict[str, List[Tuple[str, Dict[str, Any]]]] = {}
    for key, rec in db.get("lpb", {}).items():
        submission_id = s(rec.get("submission_id", ""))
        if submission_id:
            grouped.setdefault(submission_id, []).append((s(key), rec))
    for submission_id, items in grouped.items():
        if submission_id in submissions:
            continue
        first = items[0][1]
        payment_method = s(first.get("payment_method", ""))
        method = "BANK_PANIN" if payment_method.lower() == "bank panin" else "NON_PANIN"
        submissions[submission_id] = {
            "id": submission_id,
            "created_at": s(first.get("submitted_at", "")),
            "created_by": s(first.get("submitted_by", "")),
            "draft_id": s(first.get("draft_id", "")),
            "method": method,
            "target_payment_date": s(first.get("target_payment_date", "")),
            "record_ids": [key for key, _ in items],
            "files": [],
            "sppd_file": "",
            "sppd_no": s(first.get("sppd_no", "")),
            "cart_items": {},
        }
    db["submissions"] = submissions

def validate_backup_restore_conflicts(db: Dict[str, Any], rows: List[Tuple[str, Dict[str, Any]]]) -> List[str]:
    conflicts: List[str] = []
    seen_keys: Set[str] = set()
    seen_lpb: Dict[str, str] = {}
    for key, rec in rows:
        if key in seen_keys:
            conflicts.append(f"Record ID duplikat di file: {key}")
        seen_keys.add(key)
        if key in db.get("lpb", {}):
            conflicts.append(f"Record ID sudah ada: {key}")
        no_lpb = s(rec.get("no_lpb", ""))
        if no_lpb:
            norm_no = normalize_lpb_no(no_lpb)
            if norm_no in seen_lpb and seen_lpb[norm_no] != key:
                conflicts.append(f"No LPB duplikat di file: {no_lpb}")
            seen_lpb[norm_no] = key
            dup_key = find_lpb_duplicate_key(db, no_lpb)
            if dup_key:
                conflicts.append(f"No LPB sudah ada di sistem: {no_lpb}")
        if len(conflicts) >= 10:
            break
    return conflicts

def slugify(text: str) -> str:
    t = re.sub(r"[^A-Za-z0-9]+", "-", s(text).strip())
    return t.strip("-").lower() or "item"

def format_idr(val: float) -> str:
    try:
        return f"{float(val):,.0f}".replace(",", ".")
    except Exception:
        return "0"

def format_idr_decimal(val: float) -> str:
    try:
        s = f"{float(val):,.2f}"
        return s.replace(",", "_").replace(".", ",").replace("_", ".")
    except Exception:
        return "0,00"

def format_id_date(dt: Optional[pd.Timestamp]) -> str:
    if dt is None or (isinstance(dt, float) and pd.isna(dt)):
        return ""
    try:
        d = pd.to_datetime(dt)
    except Exception:
        return s(dt)
    months = [
        "Januari","Februari","Maret","April","Mei","Juni",
        "Juli","Agustus","September","Oktober","November","Desember"
    ]
    return f"{d.day} {months[d.month-1]} {d.year}"

def roman_month(n: int) -> str:
    romans = ["I","II","III","IV","V","VI","VII","VIII","IX","X","XI","XII"]
    if n < 1 or n > 12:
        return ""
    return romans[n - 1]

def format_sppd_number(seq: int, dt: pd.Timestamp) -> str:
    return format_sppd_number_with_template(seq, dt, "{seq:03d}/SPA/PDSB/{roman_month}/{year}")

def format_sppd_number_with_template(seq: int, dt: pd.Timestamp, template: str) -> str:
    num = f"{int(seq):03d}"
    month = roman_month(int(dt.month))
    year = dt.year
    tpl = s(template) or "{seq:03d}/SPA/PDSB/{roman_month}/{year}"
    try:
        return tpl.format(
            seq=int(seq),
            seq3=num,
            roman_month=month,
            month=int(dt.month),
            year=year,
            yy=str(year)[-2:],
        )
    except Exception:
        return f"{num}/SPA/PDSB/{month}/{year}"

def default_sppd_settings(db: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    try:
        start_seq = int(os.getenv("SPPD_SEQ_START", "6"))
    except Exception:
        start_seq = 6
    legacy_last = start_seq - 1
    if isinstance(db, dict) and "sppd_seq" in db:
        try:
            legacy_last = int(db.get("sppd_seq", legacy_last))
        except Exception:
            pass
    return {
        "last_sequence": max(0, legacy_last),
        "number_template": "{seq:03d}/SPA/PDSB/{roman_month}/{year}",
        "fixed_jaminan_date": "2026-02-19",
        "maturity_months": 6,
        "items_per_page": max(1, int(os.getenv("SPPD_TRANSFER_ITEMS_PER_PAGE", "7"))),
    }

def normalize_sppd_settings(raw: Dict[str, Any], db: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    defaults = default_sppd_settings(db)
    raw = raw if isinstance(raw, dict) else {}
    settings = {**defaults, **raw}
    try:
        settings["last_sequence"] = max(0, int(settings.get("last_sequence", defaults["last_sequence"])))
    except Exception:
        settings["last_sequence"] = defaults["last_sequence"]
    settings["number_template"] = s(settings.get("number_template", defaults["number_template"])) or defaults["number_template"]
    fixed_date = _normalize_yyyy_mm_dd(s(settings.get("fixed_jaminan_date", defaults["fixed_jaminan_date"])))
    settings["fixed_jaminan_date"] = fixed_date or defaults["fixed_jaminan_date"]
    try:
        settings["maturity_months"] = max(1, min(24, int(settings.get("maturity_months", defaults["maturity_months"]))))
    except Exception:
        settings["maturity_months"] = defaults["maturity_months"]
    try:
        settings["items_per_page"] = max(1, min(20, int(settings.get("items_per_page", defaults["items_per_page"]))))
    except Exception:
        settings["items_per_page"] = defaults["items_per_page"]
    return settings

def get_sppd_settings(db: Dict[str, Any]) -> Dict[str, Any]:
    settings = normalize_sppd_settings(db.get("sppd_settings", {}), db)
    db["sppd_settings"] = settings
    return settings

def next_sppd_number(db: Dict[str, Any], dt: pd.Timestamp) -> Tuple[int, str, Dict[str, Any]]:
    settings = get_sppd_settings(db)
    next_seq = int(settings.get("last_sequence", 0)) + 1
    settings["last_sequence"] = next_seq
    db["sppd_seq"] = next_seq
    return next_seq, format_sppd_number_with_template(next_seq, dt, s(settings.get("number_template", ""))), settings

def _docx_paragraph_text(p: ET.Element, ns: Dict[str, str]) -> str:
    texts = []
    for t in p.findall(".//w:t", ns):
        texts.append(t.text or "")
    return "".join(texts)

def _docx_set_paragraph_text(p: ET.Element, text: str, ns: Dict[str, str]) -> None:
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return
    t_nodes[0].text = text
    for t in t_nodes[1:]:
        t.text = ""

def _docx_clear_highlights(element: ET.Element, ns: Dict[str, str]) -> None:
    for r_pr in element.findall(".//w:rPr", ns):
        for highlight in list(r_pr.findall("w:highlight", ns)):
            r_pr.remove(highlight)

def _docx_set_prefix_value(p: ET.Element, prefix: str, value: str, ns: Dict[str, str]) -> None:
    _docx_clear_highlights(p, ns)
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return
    if len(t_nodes) == 1:
        t_nodes[0].text = f"{prefix}{value}"
        return
    t_nodes[0].text = prefix
    t_nodes[1].text = value
    for t in t_nodes[2:]:
        t.text = ""

def _docx_set_field_value(p: ET.Element, value: str, ns: Dict[str, str], fallback_text: str) -> None:
    _docx_clear_highlights(p, ns)
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return
    colon_idx = None
    for idx, node in enumerate(t_nodes):
        if ":" in (node.text or ""):
            colon_idx = idx
    if colon_idx is None:
        _docx_set_paragraph_text(p, fallback_text, ns)
        return
    node = t_nodes[colon_idx]
    before = (node.text or "").split(":", 1)[0]
    node.text = f"{before}: {value}"
    for trailing in t_nodes[colon_idx + 1:]:
        trailing.text = ""

def _docx_replace_date_after_keyword(p: ET.Element, keyword: str, date_text: str, ns: Dict[str, str]) -> bool:
    _docx_clear_highlights(p, ns)
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return False
    chunks = []
    pos = 0
    full = ""
    for node in t_nodes:
        text = node.text or ""
        start = pos
        full += text
        pos += len(text)
        chunks.append((node, start, pos))
    keyword_idx = full.lower().find(keyword.lower())
    if keyword_idx < 0:
        return False
    match = re.search(r"\d{1,2}\s+[A-Za-zÀ-ÿ]+\s+\d{4}", full[keyword_idx:])
    if not match:
        return False
    start = keyword_idx + match.start()
    end = keyword_idx + match.end()
    first_node: Optional[ET.Element] = None
    for node, node_start, node_end in chunks:
        if node_end <= start or node_start >= end:
            continue
        before = (node.text or "")[:max(0, start - node_start)] if node_start <= start < node_end else ""
        after = (node.text or "")[max(0, end - node_start):] if node_start < end <= node_end else ""
        if first_node is None:
            node.text = f"{before}{date_text}{after}"
            first_node = node
        else:
            node.text = after
    return first_node is not None

def _docx_register_namespaces(xml_bytes: bytes) -> Dict[str, str]:
    namespace_map: Dict[str, str] = {}
    seen: Set[str] = set()
    for _, ns_def in ET.iterparse(io.BytesIO(xml_bytes), events=("start-ns",)):
        prefix, uri = ns_def
        namespace_map[prefix] = uri
        if prefix in seen:
            continue
        seen.add(prefix)
        try:
            ET.register_namespace(prefix, uri)
        except ValueError:
            continue
    return namespace_map

def _docx_preserve_namespace_declarations(xml_bytes: bytes, namespace_map: Dict[str, str]) -> bytes:
    text = xml_bytes.decode("utf-8")
    root_start = 0
    if text.startswith("<?xml"):
        root_start = text.find("<", text.find("?>") + 2)
    root_end = text.find(">", root_start)
    if root_start < 0 or root_end < 0:
        return xml_bytes
    root_open = text[root_start:root_end]
    additions = []
    for prefix, uri in namespace_map.items():
        if prefix == "xml":
            continue
        if prefix:
            marker = f"xmlns:{prefix}="
            attr = f' xmlns:{prefix}="{uri}"'
        else:
            marker = "xmlns="
            attr = f' xmlns="{uri}"'
        if marker not in root_open:
            additions.append(attr)
    if not additions:
        return xml_bytes
    return (text[:root_end] + "".join(additions) + text[root_end:]).encode("utf-8")

def _docx_page_break_paragraph(ns: Dict[str, str]) -> ET.Element:
    w_ns = ns["w"]
    p = ET.Element(f"{{{w_ns}}}p")
    r = ET.SubElement(p, f"{{{w_ns}}}r")
    br = ET.SubElement(r, f"{{{w_ns}}}br")
    br.set(f"{{{w_ns}}}type", "page")
    return p

def _label_prefix(text: str, default: str) -> str:
    if ":" in text:
        return text.split(":", 1)[0].strip()
    return default

def render_sppd_docx(
    template_path: str,
    out_path: str,
    submit_dt: pd.Timestamp,
    sppd_no: str,
    transfer_items: List[Dict[str, Any]],
    settings: Optional[Dict[str, Any]] = None,
) -> None:
    if not os.path.exists(template_path):
        raise FileNotFoundError("Template SPPD tidak ditemukan.")
    with zipfile.ZipFile(template_path) as z:
        data = {name: z.read(name) for name in z.namelist()}
    if "word/document.xml" not in data:
        raise ValueError("Template SPPD tidak valid.")

    namespace_map = _docx_register_namespaces(data["word/document.xml"])
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    root = ET.fromstring(data["word/document.xml"])
    body = root.find(".//w:body", ns)
    if body is None:
        raise ValueError("Template SPPD tidak valid.")
    settings = normalize_sppd_settings(settings or {})
    _docx_clear_highlights(root, ns)

    paras = list(body.findall("w:p", ns))
    texts = [_docx_paragraph_text(p, ns) for p in paras]

    # Update tanggal surat (Makassar, ...)
    tanggal_surat = format_id_date(submit_dt)
    for p, t in zip(paras, texts):
        if t.strip().startswith("Makassar"):
            _docx_set_prefix_value(p, "Makassar, ", tanggal_surat, ns)
            break

    # Update nomor surat
    for p, t in zip(paras, texts):
        if t.strip().startswith("Nomor") and "/SPA/PDSB/" in t:
            _docx_set_field_value(p, sppd_no, ns, f"Nomor : {sppd_no}")
            break

    # Update tanggal dokumen Jaminan dari konfigurasi SPPD.
    jaminan_date = format_id_date(pd.to_datetime(settings.get("fixed_jaminan_date", "2026-02-19")))
    for p, t in zip(paras, texts):
        if "Jaminan tertanggal" in t:
            _docx_replace_date_after_keyword(p, "Jaminan tertanggal", jaminan_date, ns)
            break

    # Update jatuh tempo Bank dari tanggal Makassar.
    maturity_months = int(settings.get("maturity_months", 6) or 6)
    jatuh_tempo = format_id_date(submit_dt + pd.DateOffset(months=maturity_months))
    for p, t in zip(paras, texts):
        if "jatuh tempo pada Bank yaitu tanggal" in t:
            _docx_replace_date_after_keyword(p, "jatuh tempo pada Bank yaitu tanggal", jatuh_tempo, ns)
            break

    # Replace daftar bank
    start_idx = None
    end_idx = None
    for i, t in enumerate(texts):
        if t.strip().startswith("Bank"):
            start_idx = i
            break
    for i, t in enumerate(texts):
        if "Kami menyatakan dan menyanggupi" in t:
            end_idx = i - 1
            break

    if start_idx is not None and end_idx is not None and end_idx >= start_idx:
        bank_tpl = paras[start_idx]
        norek_tpl = paras[start_idx + 1] if start_idx + 1 < len(paras) else bank_tpl
        atas_tpl = paras[start_idx + 2] if start_idx + 2 < len(paras) else bank_tpl
        jumlah_tpl = paras[start_idx + 3] if start_idx + 3 < len(paras) else bank_tpl

        sep_tpl = None
        for i in range(start_idx, min(start_idx + 8, len(paras))):
            if not texts[i].strip():
                sep_tpl = paras[i]
                break

        bank_label = _label_prefix(texts[start_idx], "Bank")
        norek_label = _label_prefix(texts[start_idx + 1], "Nomor Rekening")
        atas_label = _label_prefix(texts[start_idx + 2], "Atas Nama")
        jumlah_label = _label_prefix(texts[start_idx + 3], "Sejumlah")

        new_paras: List[ET.Element] = []
        items_per_page = max(1, int(settings.get("items_per_page", 7) or 7))
        for idx, item in enumerate(transfer_items):
            p_bank = copy.deepcopy(bank_tpl)
            bank_value = s(item.get("bank", ""))
            _docx_set_field_value(p_bank, bank_value, ns, f"{bank_label}: {bank_value}")
            new_paras.append(p_bank)

            p_norek = copy.deepcopy(norek_tpl)
            norek_value = s(item.get("rekening", ""))
            _docx_set_field_value(p_norek, norek_value, ns, f"{norek_label}: {norek_value}")
            new_paras.append(p_norek)

            p_atas = copy.deepcopy(atas_tpl)
            atas_value = s(item.get("penerima", ""))
            _docx_set_field_value(p_atas, atas_value, ns, f"{atas_label}: {atas_value}")
            new_paras.append(p_atas)

            p_jumlah = copy.deepcopy(jumlah_tpl)
            amount = float(item.get("amount", 0) or 0.0)
            jumlah_value = f"Rp {format_idr(amount)}"
            _docx_set_field_value(p_jumlah, jumlah_value, ns, f"{jumlah_label}: {jumlah_value}")
            new_paras.append(p_jumlah)

            if idx < len(transfer_items) - 1:
                if (idx + 1) % items_per_page == 0:
                    new_paras.append(_docx_page_break_paragraph(ns))
                elif sep_tpl is not None:
                    new_paras.append(copy.deepcopy(sep_tpl))

        for p in paras[start_idx:end_idx + 1]:
            body.remove(p)
        for offset, p in enumerate(new_paras):
            body.insert(start_idx + offset, p)

    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    new_xml = _docx_preserve_namespace_declarations(new_xml, namespace_map)
    data["word/document.xml"] = new_xml
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for name, content in data.items():
            z.writestr(name, content)

def load_bank_map() -> Dict[str, Dict[str, str]]:
    """
    Load data rekening dari Excel. Key = UPPERCASE principle name.
    Juga menyimpan normalized key untuk fuzzy matching.
    Nomor rekening disimpan sebagai string agar leading zeros tidak hilang.
    Data kosong (DF/VA tanpa nomor rekening) tetap disimpan apa adanya, tidak dikarang.
    """
    if not BANK_DATA_PATH or not os.path.exists(BANK_DATA_PATH):
        return {}
    # Try reading with header auto-detection: the Excel might have header at row 0 or row 2
    # First attempt: standard read
    try:
        df = pd.read_excel(BANK_DATA_PATH, dtype=str)
        cols = {c.strip().upper(): c for c in df.columns}
        required = ["PRINCIPLE", "NAMA BANK", "NOMOR REKENING", "NAMA PENERIMA"]
        missing = [c for c in required if c not in cols]
        if missing:
            # Try with header at row 2 (common format from rekprinciple.xlsx)
            df = pd.read_excel(BANK_DATA_PATH, header=2, dtype=str)
            cols = {c.strip().upper(): c for c in df.columns}
            missing = [c for c in required if c not in cols]
            if missing:
                return {}
    except Exception:
        return {}
    mp = {}
    for _, r in df.iterrows():
        raw_p = r.get(cols.get("PRINCIPLE", ""), None)
        p = s(raw_p) if raw_p is not None and not pd.isna(raw_p) else ""
        if not p:
            continue
        # Nomor rekening: pastikan string, jangan convert ke float
        raw_rek = r.get(cols.get("NOMOR REKENING", ""), None)
        if raw_rek is None or pd.isna(raw_rek):
            rek_str = ""
        else:
            rek_str = str(raw_rek).strip()
            # Hapus trailing .0 jika pandas membaca sebagai float
            if rek_str.endswith(".0"):
                rek_str = rek_str[:-2]

        raw_bank = r.get(cols.get("NAMA BANK", ""), None)
        bank_val = s(raw_bank) if raw_bank is not None and not pd.isna(raw_bank) else ""
        raw_pen = r.get(cols.get("NAMA PENERIMA", ""), None)
        penerima_val = s(raw_pen) if raw_pen is not None and not pd.isna(raw_pen) else ""

        # Jika bank = DF/VA dan rekening kosong, jangan mengarang
        # Simpan apa adanya
        mp[p.upper()] = {
            "principle": normalize_principle_name(p),
            "bank": bank_val,
            "rekening": rek_str,
            "penerima": penerima_val,
        }
    return mp


def load_bank_map_with_normalized_keys() -> Tuple[Dict[str, Dict[str, str]], Dict[str, str]]:
    """
    Load bank map + pre-build normalized keys untuk fuzzy matching.
    Returns: (bank_map, normalized_key_map)
    """
    bank_map = load_bank_map()
    norm_keys = build_normalized_key_map(bank_map)
    return bank_map, norm_keys

def write_invoice_excel(rows: List[Dict[str, Any]], out_path: str):
    cols = [
        "No",
        "Principle",
        "Tipe Pengajuan",
        "Nilai Invoice (Total)",
        "No. Invoice / Dokumen",
        "Potongan",
        "Nilai Pembayaran",
        "Jenis Pembayaran",
        "Keterangan",
    ]
    normalized_rows = []
    for row in rows or []:
        item = dict(row or {})
        if "No. Invoice / Dokumen" not in item:
            item["No. Invoice / Dokumen"] = s(item.get("No. Invoice", ""))
        if "Tipe Pengajuan" not in item:
            item["Tipe Pengajuan"] = s(item.get("Tipe", ""))
        normalized_rows.append(item)
    df = pd.DataFrame(normalized_rows, columns=cols)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="INVOICE")

def _normalize_yyyy_mm_dd(value: str) -> str:
    v = s(value)
    if not v:
        return ""
    try:
        dt = pd.to_datetime(v, errors="raise")
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""

def _style_excel_download_sheet(ws) -> None:
    money_tokens = ("NILAI", "AMOUNT", "TOTAL", "GAP", "POTONGAN", "PEMBAYARAN")
    date_tokens = ("TGL", "DATE", "J.T", "J TEMPO", "JATUH TEMPO")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1], start=1):
        header = s(cell.value).upper()
        width = max(12, min(34, len(header) + 3))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        if any(token in header for token in money_tokens):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = '#,##0.00'
                ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="right")
        elif any(token in header for token in date_tokens):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"

def _excel_download_response(rows: List[Dict[str, Any]], filename: str, sheet_name: str) -> Response:
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _style_excel_download_sheet(writer.book[sheet_name])
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

def safe_upload_filename(name: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(s(name)))
    return clean.strip("._") or "bukti-transfer"

def proof_public_url(file_name: str) -> str:
    return f"/payments/proofs/{os.path.basename(file_name)}"

def finance_mapping_key(principle: str) -> str:
    return s(principle).upper()

def get_finance_mapping(db: Dict[str, Any], principle: str) -> Dict[str, str]:
    mappings = db.setdefault("finance_mappings", {})
    item = mappings.get(finance_mapping_key(principle), {})
    if not isinstance(item, dict):
        return {}
    return {
        "principle": s(item.get("principle", principle)),
        "vendorNo": s(item.get("vendorNo", "")),
        "vendorName": s(item.get("vendorName", "")),
        "bankNo": s(item.get("bankNo", "")),
        "bankName": s(item.get("bankName", "")),
        "updated_at": s(item.get("updated_at", "")),
        "updated_by": s(item.get("updated_by", "")),
    }

def build_proof_metadata(proof_id: str, stored_name: str, original_name: str, content: bytes, uploaded_by: str) -> Dict[str, Any]:
    mime = mimetypes.guess_type(original_name)[0] or "application/octet-stream"
    return {
        "proof_id": proof_id,
        "original_filename": original_name,
        "stored_filename": stored_name,
        "mime": mime,
        "size": len(content),
        "sha256": hashlib.sha256(content).hexdigest(),
        "uploaded_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "uploaded_by": uploaded_by,
        "url": proof_public_url(stored_name),
    }

def xml_escape(text: str) -> str:
    return (
        s(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )

def render_docx_template(template_path: str, out_path: str, replacements: Dict[str, str], append_rows: str = "") -> None:
    if not os.path.exists(template_path):
        raise FileNotFoundError("Template SPPD tidak ditemukan.")
    with zipfile.ZipFile(template_path) as z:
        data = {name: z.read(name) for name in z.namelist()}
    if "word/document.xml" not in data:
        raise ValueError("Template SPPD tidak valid.")
    xml = data["word/document.xml"].decode("utf-8", errors="ignore")
    replaced = xml
    for k, v in replacements.items():
        replaced = replaced.replace("{{" + k + "}}", xml_escape(v))

    if "{{TRANSFER_ROWS}}" in replaced:
        replaced = replaced.replace("{{TRANSFER_ROWS}}", xml_escape(append_rows))
    elif append_rows:
        insert = (
            "<w:p><w:r><w:t>" + xml_escape(append_rows).replace("\n", "</w:t><w:br/><w:t>") + "</w:t></w:r></w:p>"
        )
        replaced = replaced.replace("</w:body>", insert + "</w:body>")

    data["word/document.xml"] = replaced.encode("utf-8")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for name, content in data.items():
            z.writestr(name, content)

def allocate_claim_by_mdstring(
    gross_amount: float,
    actual_discount: float,
    mdstring: str,
    rp_pabrik: float,
    pct_pabrik,
    pct_internal,
) -> Optional[Tuple[float, float, float]]:
    actual = float(actual_discount or 0.0)
    rp = float(rp_pabrik or 0.0)
    if actual <= 0:
        return (0.0, 0.0, 0.0)
    if actual < rp:
        return (actual, 0.0, 0.0)

    md_parts = normalize_pct_list(parse_mdstring_pct(mdstring))
    if not md_parts:
        return None

    p_parts = normalize_pct_list(pct_pabrik)
    i_parts = normalize_pct_list(pct_internal)

    base = max(float(gross_amount or 0.0) - rp, 0.0)
    rem = base
    p_idx = 0
    i_idx = 0
    p_pct_total = 0.0
    i_pct_total = 0.0
    matched_any = False

    for pct in md_parts:
        disc = rem * (pct / 100.0)
        p_next = p_parts[p_idx] if p_idx < len(p_parts) else None
        i_next = i_parts[i_idx] if i_idx < len(i_parts) else None
        p_match = (p_next is not None and abs(p_next - pct) < 1e-6)
        i_match = (i_next is not None and abs(i_next - pct) < 1e-6)

        if p_match and i_match:
            if (len(p_parts) - p_idx) >= (len(i_parts) - i_idx):
                p_pct_total += disc
                p_idx += 1
                matched_any = True
            else:
                i_pct_total += disc
                i_idx += 1
                matched_any = True
        elif p_match:
            p_pct_total += disc
            p_idx += 1
            matched_any = True
        elif i_match:
            i_pct_total += disc
            i_idx += 1
            matched_any = True
        else:
            return None

        rem -= disc

    if not matched_any:
        return None

    actual_pct_total = max(actual - rp, 0.0)
    pct_alloc = p_pct_total + i_pct_total
    if pct_alloc > 0:
        scale = min(actual_pct_total / pct_alloc, 1.0)
        p_pct_total *= scale
        i_pct_total *= scale

    p_total = rp + p_pct_total
    i_total = i_pct_total
    tanpat = max(actual - (p_total + i_total), 0.0)
    return (p_total, i_total, tanpat)

def sequential_pct_discount(base_amount: float, pct_parts: List[float]) -> float:
    remaining = float(base_amount or 0.0)
    total = 0.0
    for pct in pct_parts:
        disc = remaining * (pct / 100.0)
        total += disc
        remaining -= disc
    return total


def mdstring_is_bonus(md: str) -> bool:
    """Bonus line: MDSTRING represents 100% discount (free goods)."""
    parts = parse_mdstring_pct(md)
    if not parts:
        # sometimes users put plain "100" without %
        md_s = s(md).strip()
        return md_s == "100"
    mx = max(parts)
    if mx < 99.999:
        return False
    # allow 0% parts alongside 100%
    others = [p for p in parts if abs(p - mx) > 1e-6]
    return all(abs(p) < 1e-6 for p in others)

def pct_parts_full_discount(pct_parts: List[float]) -> bool:
    if not pct_parts:
        return False
    mx = max(pct_parts)
    if mx < 99.999:
        return False
    others = [p for p in pct_parts if abs(p - mx) > 1e-6]
    return all(abs(p) < 1e-6 for p in others)

def compute_expected_bonus_from_rules(
    sales_df: pd.DataFrame,
    rules_df: pd.DataFrame,
    channel_col: str,
    expected_col: str,
    debug_reason_prefix: str,
    program_name_col: Optional[str] = None,
):
    """
    Fill Expected bonus qty (PCS) per line.
    - Expected bonus is only written to rows that are 'bonus lines' (MDSTRING=100%).
    - For NON_GROUP: computed per line.
    - For grouped programs: computed per invoice+program key using invoice-level trigger metric,
      then allocated to bonus lines of the corresponding item(s).
    """
    if rules_df is None or rules_df.empty:
        sales_df.loc[:, expected_col] = 0.0
        return

    # normalize rules columns
    d = rules_df.copy()
    for c in ["KODE_BARANG","NAMA_BARANG","PROMO_LABEL","PROMO_GROUP_ID","PROMO_GROUP","TIER_NO","TRIGGER_QTY","TRIGGER_UNIT","BENEFIT_TYPE","BENEFIT_VALUE","BENEFIT_UNIT","PROMO_ACTIVE"]:
        if c not in d.columns:
            d.loc[:, c] = None
    d["KODE_BARANG"] = d["KODE_BARANG"].apply(s)
    d["NAMA_BARANG"] = d["NAMA_BARANG"].apply(s)
    d["_ALL_ITEM_RULE"] = (d["KODE_BARANG"] == "") & (d["NAMA_BARANG"] == "")

    # focus bonus programs only
    d = d[d["BENEFIT_TYPE"].astype(str).str.upper().isin(["BONUS_QTY","BONUSQTY","BONUS"])].copy()
    if d.empty:
        sales_df.loc[:, expected_col] = 0.0
        return

    # prep
    sales_df.loc[:, expected_col] = 0.0
    sales_df.loc[:, "_IS_BONUS_LINE"] = sales_df["MDSTRING"].apply(mdstring_is_bonus)

    # iterate invoices to keep it fast enough
    for inv, inv_df in sales_df.groupby("INVOICENO", sort=False):
        inv_idx = inv_df.index.tolist()

        # candidate rules for items present in this invoice
        items = set(inv_df["BRG"].astype(str))
        r0 = d[(d["KODE_BARANG"].astype(str).isin(items)) | (d["_ALL_ITEM_RULE"] == True)]
        if r0.empty:
            continue

        # group by program key
        for (pgid, plabel), pr in r0.groupby(["PROMO_GROUP_ID","PROMO_LABEL"], sort=False):
            pgid_s = norm(pgid)
            plabel_s = s(plabel)
            program_key = f"{pgid_s}||{plabel_s}"

            # channel filter (STD always allowed, otherwise must match effective channel)
            pr_ok = pr[pr.apply(
                lambda r: channel_ok(
                    r.get("PROMO_GROUP", ""),
                    inv_df.iloc[0].get(channel_col, ""),
                    r.get("PROMO_GROUP_ID", ""),
                    inv_df.iloc[0].get("SUB", "")
                ),
                axis=1
            )]
            if pr_ok.empty:
                continue

            # pick best tier (highest tier that passes trigger)
            # we decide trigger metric based on chosen trigger unit
            trigger_unit, _tu_source = choose_trigger_unit_satpam(pr_ok)
            if not trigger_unit:
                continue

            inv_non_bonus = inv_df[~inv_df["_IS_BONUS_LINE"]].copy()

            # Compute per-tier metrics to select highest satisfied tier
            best = None
            for tier_no in sorted(pr_ok["TIER_NO"].fillna(0).astype(int).unique(), reverse=True):
                tier_rows = pr_ok[pr_ok["TIER_NO"].fillna(0).astype(int) == int(tier_no)]
                if tier_rows.empty:
                    continue

                # trigger qty / benefit value (tier-level)
                tq = float(pd.to_numeric(tier_rows["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)
                if tq <= 0:
                    continue
                bq = float(pd.to_numeric(tier_rows["BENEFIT_VALUE"], errors="coerce").fillna(0).max() or 0.0)
                if bq <= 0:
                    continue

                if pgid_s == "NON_GROUP":
                    # non-group: per item using non-bonus lines for trigger
                    used_items: Set[str] = set()
                    items_tier = [x for x in tier_rows["KODE_BARANG"].astype(str).unique().tolist() if s(x)]
                    if bool(tier_rows["_ALL_ITEM_RULE"].any()):
                        items_tier = sorted(set(items_tier).union(set(inv_non_bonus["BRG"].astype(str).tolist())))
                    for item in items_tier:
                        if item in used_items:
                            continue
                        rows_item = tier_rows[tier_rows["KODE_BARANG"].astype(str) == str(item)]
                        if rows_item.empty:
                            rows_item = tier_rows[tier_rows["_ALL_ITEM_RULE"] == True]
                        if rows_item.empty:
                            continue

                        tq = float(pd.to_numeric(rows_item["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)
                        if tq <= 0:
                            continue
                        bq = float(pd.to_numeric(rows_item["BENEFIT_VALUE"], errors="coerce").fillna(0).max() or 0.0)
                        if bq <= 0:
                            continue

                        if trigger_unit == "GROSSAMOUNT":
                            metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str) == str(item)]["GROSSAMOUNT"].astype(float).fillna(0).sum())
                        else:
                            metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str) == str(item)].apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum())

                        if metric + 1e-9 < tq:
                            continue

                        total_exp = math.floor(metric / tq) * bq
                        if total_exp <= 0:
                            continue

                        bonus_idxs = [i for i in inv_idx if (str(sales_df.at[i, "BRG"]) == str(item)) and bool(sales_df.at[i, "_IS_BONUS_LINE"])]
                        if not bonus_idxs:
                            best = ("NON_GROUP_NO_BONUS_LINE", tier_no)
                            continue

                        if "Bonus_TriggerUnit" in sales_df.columns:
                            for i in bonus_idxs:
                                sales_df.at[i, "Bonus_TriggerUnit"] = trigger_unit
                                sales_df.at[i, "Bonus_TriggerQty"] = float(tq)
                                sales_df.at[i, "Bonus_TriggerMetric"] = float(metric)
                                sales_df.at[i, "Bonus_PickedTier"] = int(tier_no)

                        remaining = float(total_exp)
                        for i in bonus_idxs:
                            if remaining <= 0:
                                break
                            act_qty = float(qty_in_unit(sales_df.loc[i], "PCS"))
                            take = min(remaining, act_qty if act_qty > 0 else remaining)
                            sales_df.at[i, expected_col] = float(take)
                            remaining -= take
                            if program_name_col:
                                cur = s(sales_df.at[i, program_name_col])
                                sales_df.at[i, program_name_col] = add_list_item(cur, plabel_s)

                        used_items.add(item)
                        best = ("NON_GROUP", tier_no)
                    if used_items:
                        break

                # grouped program: invoice-level metric
                if bool(tier_rows["_ALL_ITEM_RULE"].any()):
                    elig_item_set = set(inv_non_bonus["BRG"].astype(str).tolist())
                else:
                    elig_item_set = set([x for x in tier_rows["KODE_BARANG"].astype(str).tolist() if s(x)])
                if not elig_item_set:
                    continue
                if trigger_unit == "GROSSAMOUNT":
                    metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str).isin(elig_item_set)]["GROSSAMOUNT"].astype(float).fillna(0).sum())
                else:
                    metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str).isin(elig_item_set)].apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum())

                if metric + 1e-9 < tq:
                    continue

                total_exp = math.floor(metric / tq) * bq
                if total_exp <= 0:
                    continue

                # allocate to bonus lines for eligible items
                bonus_idxs = [i for i in inv_idx if (str(sales_df.at[i,"BRG"]) in elig_item_set) and bool(sales_df.at[i,"_IS_BONUS_LINE"])]
                if not bonus_idxs:
                    # no bonus line present -> still mark debug but can't allocate
                    best = ("GROUPED_NO_BONUS_LINE", tier_no)
                    break

                if "Bonus_TriggerUnit" in sales_df.columns:
                    for i in bonus_idxs:
                        sales_df.at[i, "Bonus_TriggerUnit"] = trigger_unit
                        sales_df.at[i, "Bonus_TriggerQty"] = float(tq)
                        sales_df.at[i, "Bonus_TriggerMetric"] = float(metric)
                        sales_df.at[i, "Bonus_PickedTier"] = int(tier_no)

                remaining = float(total_exp)
                for i in bonus_idxs:
                    if remaining <= 0:
                        break
                    # allocate up to actual bonus qty (pcs) on that line
                    act_qty = float(qty_in_unit(sales_df.loc[i], "PCS"))
                    take = min(remaining, act_qty if act_qty > 0 else remaining)
                    sales_df.at[i, expected_col] = float(take)
                    remaining -= take
                    if program_name_col:
                        cur = s(sales_df.at[i, program_name_col])
                        sales_df.at[i, program_name_col] = add_list_item(cur, plabel_s)

                best = ("GROUPED", tier_no)
                break

            # Debug: record if needed
            if best:
                for i in inv_idx:
                    cur = s(sales_df.at[i, "Debug_SkipReason"]) if "Debug_SkipReason" in sales_df.columns else ""
                    # keep it lightweight: only for bonus lines
                    if bool(sales_df.at[i, "_IS_BONUS_LINE"]):
                        msg = f"{debug_reason_prefix}:{program_key}:{best[0]}:tier{best[1]}"
                        sales_df.at[i, "Debug_SkipReason"] = (cur + (" | " if cur else "") + msg)

    # cleanup helper col
    if "_IS_BONUS_LINE" in sales_df.columns:
        sales_df.drop(columns=["_IS_BONUS_LINE"], inplace=True, errors="ignore")
def dedupe_grouped_benefits(tier_rows: pd.DataFrame) -> pd.DataFrame:
    if tier_rows.empty:
        return tier_rows
    cols = ["BENEFIT_TYPE", "BENEFIT_VALUE", "BENEFIT_UNIT", "TIER_NO", "TRIGGER_QTY", "TRIGGER_UNIT"]
    for c in cols:
        if c not in tier_rows.columns:
            tier_rows.loc[:, c] = ""
    return tier_rows.drop_duplicates(subset=cols).copy()


# ---------------------------
# Internal program helpers
# ---------------------------
def internal_channel_match(rule_channel: str, effective_sales_channel: str) -> bool:
    rc = norm(rule_channel)
    sc = norm(effective_sales_channel)
    if rc in ("", "ALL", "STD", "GLOBAL"):
        return True
    if sc == "":
        return False
    return (rc == sc) or (rc in sc) or (sc in rc)

def load_internal_rules(internal_bytes: Optional[bytes]) -> Optional[pd.DataFrame]:
    if not internal_bytes:
        return None
    df = pd.read_excel(io.BytesIO(internal_bytes))
    for c in ["CHANNEL","SUB","KODE_BARANG","RULE_TYPE","MIN_GROSS","MAX_GROSS","DISC_PCT","ACTIVE"]:
        ensure_col(df, c, "")
    df["CHANNEL"] = df["CHANNEL"].apply(s)
    df["SUB"] = df["SUB"].apply(s)
    df["KODE_BARANG"] = df["KODE_BARANG"].apply(s)
    df["RULE_TYPE"] = df["RULE_TYPE"].apply(lambda x: norm(x))
    df["DISC_PCT"] = df["DISC_PCT"].apply(s)
    df["MIN_GROSS"] = df["MIN_GROSS"].apply(parse_number_id)
    df["MAX_GROSS"] = df["MAX_GROSS"].apply(parse_number_id)

    def is_active(v):
        if pd.isna(v):
            return True
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in ["true","1","yes","y"]
    df = df[df["ACTIVE"].apply(is_active)].copy()
    return df

def pick_internal_rule(rules: pd.DataFrame, invoice_gross: float) -> pd.DataFrame:
    if rules is None or rules.empty:
        return rules.iloc[0:0].copy()
    ig = float(invoice_gross or 0.0)
    d = rules.copy()
    d = d[d["MIN_GROSS"].astype(float) <= ig]
    d = d[(d["MAX_GROSS"].astype(float) <= 0) | (ig <= d["MAX_GROSS"].astype(float))]
    if d.empty:
        return d
    d = d.sort_values(["MIN_GROSS","MAX_GROSS"], ascending=[False, True])
    top_min = float(d["MIN_GROSS"].iloc[0])
    return d[d["MIN_GROSS"].astype(float) == top_min].copy()

# ---------------------------
# Channel selection
# ---------------------------
def channel_ok(promo_group: str, effective_sales_channel: str, promo_group_id: str, sales_sub: str) -> bool:
    """
    promo_group is dataset column PROMO_GROUP
    effective_sales_channel is computed from channel map (SUB->CHANNEL); fallback to sales field
    """
    pg = norm(promo_group)
    sc = norm(effective_sales_channel)

    # STD = global (all channels)
    if pg == "STD":
        return True

    # Outlet special: require SUB match PROMO_GROUP_ID
    if pg == "OUTLET":
        return norm(sales_sub) != "" and norm(promo_group_id) != "" and norm(sales_sub) == norm(promo_group_id)

    # Online vs non-online should not cross-match
    if ("ONLINE" in pg or "ONLINE" in sc) and (("ONLINE" in pg) != ("ONLINE" in sc)):
        return False

    # Otherwise: match channel
    if pg != "" and sc != "" and (pg == sc or pg in sc or sc in pg):
        return True
    return False

# ---------------------------
# Tier picking
# ---------------------------
def pick_best_tier(program_rows: pd.DataFrame, metric_value: float) -> Tuple[int, pd.DataFrame]:
    if program_rows.empty:
        return (0, program_rows)

    d = program_rows.copy()
    d.loc[:, "TIER_NO"] = pd.to_numeric(d["TIER_NO"], errors="coerce").fillna(0).astype(int)
    d = d.sort_values(["TIER_NO", "TRIGGER_QTY"], ascending=[False, False])

    mv = float(metric_value or 0.0)
    picked = None
    for _, r in d.iterrows():
        if mv >= float(r["TRIGGER_QTY"]):
            picked = int(r["TIER_NO"])
            break
    if picked is None:
        return (0, d.iloc[0:0].copy())

    return (picked, d[d["TIER_NO"] == picked].copy())

# ---------------------------
# SATPAM trigger-unit (v9)
# ---------------------------
VALID_UNITS = {"GROSSAMOUNT", "PCS", "CTN", "KRT"}

def choose_trigger_unit_satpam(program_rows: pd.DataFrame) -> Tuple[Optional[str], str]:
    if "TRIGGER_UNIT" in program_rows.columns:
        tu = [norm(x) for x in program_rows["TRIGGER_UNIT"].tolist() if s(x)]
        tu = [x for x in tu if x in VALID_UNITS]
        if tu:
            if "GROSSAMOUNT" in tu:
                return ("GROSSAMOUNT", "TRIGGER_UNIT")
            if any(x in ["CTN", "KRT"] for x in tu):
                return ("CTN", "TRIGGER_UNIT")
            if "PCS" in tu:
                return ("PCS", "TRIGGER_UNIT")
            return (tu[0], "TRIGGER_UNIT")

    if "BENEFIT_UNIT" in program_rows.columns:
        bu = [norm(x) for x in program_rows["BENEFIT_UNIT"].tolist() if s(x)]
        bu = [x for x in bu if x in VALID_UNITS]
        if bu:
            if "GROSSAMOUNT" in bu:
                return ("GROSSAMOUNT", "BENEFIT_UNIT")
            if any(x in ["CTN", "KRT"] for x in bu):
                return ("CTN", "BENEFIT_UNIT")
            if "PCS" in bu:
                return ("PCS", "BENEFIT_UNIT")
            return (bu[0], "BENEFIT_UNIT")

    return (None, "UNKNOWN")

def add_reason(existing: str, reason: str) -> str:
    existing = s(existing)
    if not existing:
        return reason
    if reason in existing:
        return existing
    return existing + "; " + reason

def add_list_item(existing: str, item: str, sep: str = "; ") -> str:
    existing = s(existing)
    item = s(item)
    if not item:
        return existing
    if not existing:
        return item
    parts = [p.strip() for p in existing.split(";") if p.strip()]
    if item in parts:
        return existing
    return existing + sep + item


# ---------------------------
# UI (v16 FIXED) — safe raw HTML string (no JS leaked into Python)
# ---------------------------

UI_HTML = r"""<!doctype html>
<html lang="id">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <link rel="icon" href="/favicon.ico">
  <link rel="icon" type="image/png" href="/favicon.png">
  <title>Discount Validator</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    .pulse-soft { animation: pulseSoft 1.4s ease-in-out infinite; }
    @keyframes pulseSoft { 0%,100%{opacity:.55} 50%{opacity:1} }
  </style>
</head>

<body class="min-h-screen bg-[#EFEFEF] text-[#1D1F1E]">
  <div class="max-w-[1400px] mx-auto px-4 py-7">
    <div class="grid xl:grid-cols-12 gap-5">

      <!-- LEFT SIDEBAR -->
      <aside class="xl:col-span-3 rounded-2xl overflow-hidden shadow-lg border border-[#8A4703] bg-[#EFEFEF]">
        <div class="p-5 bg-[#BD7401] text-[#EFEFEF]">
          <div class="flex items-start justify-between gap-3">
            <div>
              <div class="text-xs uppercase tracking-widest text-[#EFEFEF]">Validator</div>
              <div class="text-xl font-semibold leading-tight">Discount & Claim</div>
              <div class="mt-1 text-xs text-[#EFEFEF]">Patch: __PATCH_VERSION__</div>
            </div>
            <div class="flex items-center gap-2">
              <a href="/"
                 class="shrink-0 text-xs px-3 py-1.5 rounded-full bg-[#F2DA82] hover:bg-[#BD7401] hover:text-[#EFEFEF] border border-[#8A4703] transition">
                Dashboard
              </a>
              <a href="/logout"
                 class="shrink-0 text-xs px-3 py-1.5 rounded-full bg-[#F2DA82] hover:bg-[#BD7401] hover:text-[#EFEFEF] border border-[#8A4703] transition">
                Logout
              </a>
            </div>
          </div>

          <div class="mt-4 grid grid-cols-3 gap-2">
            <div class="rounded-xl p-3 bg-[#F2DA82] border border-[#8A4703]">
              <div class="text-[11px] text-[#1D1F1E]">C</div>
              <div id="cntC" class="text-2xl font-semibold">0</div>
            </div>
            <div class="rounded-xl p-3 bg-[#F2DA82] border border-[#8A4703]">
              <div class="text-[11px] text-[#1D1F1E]">B</div>
              <div id="cntB" class="text-2xl font-semibold">0</div>
            </div>
            <div class="rounded-xl p-3 bg-[#F2DA82] border border-[#8A4703]">
              <div class="text-[11px] text-[#1D1F1E]">A</div>
              <div id="cntA" class="text-2xl font-semibold">0</div>
            </div>
          </div>

          <div class="mt-4 rounded-xl bg-[#F2DA82] border border-[#8A4703] p-3">
            <div class="text-xs text-[#1D1F1E]">API Status</div>
            <div id="apiStatus" class="mt-1 text-sm font-medium pulse-soft">checking...</div>
          </div>
        </div>

        <div class="p-5">
          <div class="flex items-center justify-between">
            <div class="text-sm font-semibold">Checklist Upload</div>
            <div id="readyPill"
                 class="text-[11px] px-2 py-1 rounded-full border border-[#8A4703] bg-[#EFEFEF] text-[#1D1F1E]">
              belum siap
            </div>
          </div>

          <div class="mt-4 space-y-3">
            <div class="flex items-center gap-3">
              <div id="dotSales" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Data Penjualan</div>
                <div id="nameSales" class="text-xs text-[#1D1F1E] truncate">belum diupload</div>
              </div>
            </div>

            <div class="flex items-center gap-3">
              <div id="dotPromo" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Dataset Diskon Pabrik</div>
                <div id="namePromo" class="text-xs text-[#1D1F1E] truncate">belum diupload</div>
              </div>
            </div>

            <div class="flex items-center gap-3">
              <div id="dotChannel" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Data Channel by SUB</div>
                <div id="nameChannel" class="text-xs text-[#1D1F1E] truncate">belum diupload</div>
              </div>
            </div>

            <div class="flex items-center gap-3">
              <div id="dotInternal" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Dataset Diskon Internal <span class="text-xs text-[#1D1F1E]">(opsional)</span></div>
                <div id="nameInternal" class="text-xs text-[#1D1F1E] truncate">boleh kosong</div>
              </div>
            </div>
          </div>

          <div class="mt-5">
            <button id="downloadBtn" disabled
                    class="w-full rounded-xl px-4 py-3 text-sm font-semibold
                           bg-gradient-to-r from-[#BD7401] to-[#BD7401] text-[#EFEFEF]
                           hover:brightness-95 disabled:opacity-50 disabled:cursor-not-allowed transition">
              Download Result
            </button>
            <div class="mt-2 text-xs text-[#1D1F1E]">Setelah validate sukses, tombol ini aktif.</div>
          </div>

          <div class="mt-5 rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
            <div class="text-xs uppercase tracking-widest text-[#1D1F1E]">Tips</div>
            <div class="mt-2 text-sm text-[#1D1F1E]">
              Klik <b>Validate</b> kalau 3 file wajib sudah hijau.
              Kalau belum, nanti aku tegur manja dulu 😄
            </div>
          </div>
        </div>
      </aside>

      <!-- CENTER -->
      <main class="xl:col-span-6 rounded-2xl bg-[#EFEFEF] shadow-lg border border-[#8A4703] overflow-hidden">
        <div class="p-6 border-b border-[#8A4703]">
          <div class="flex items-start justify-between gap-4">
            <div>
              <h1 class="text-2xl md:text-3xl font-semibold tracking-tight">
                Validate Discount vs Program Dataset
              </h1>
              <p class="mt-2 text-[#1D1F1E]">__PATCH_TITLE__</p>
              <p class="mt-1 text-xs text-[#1D1F1E]">
                Required: Data Penjualan + Dataset Diskon Pabrik + Data Channel by SUB. Internal optional.
              </p>
            </div>
            <div class="hidden md:block rounded-2xl border border-[#8A4703] bg-[#EFEFEF] px-4 py-3">
              <div class="text-xs text-[#1D1F1E]">Mode</div>
              <div class="text-sm font-semibold text-[#1D1F1E]">Serius tapi ramah 😄</div>
            </div>
          </div>
        </div>

        <div class="p-6">
          <div id="toast" aria-live="polite" role="status" class="hidden mb-4 rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
            <div id="toastTitle" class="text-sm font-semibold text-[#1D1F1E]">Info</div>
            <div id="toastBody" class="mt-1 text-sm text-[#1D1F1E]"></div>
          </div>

          <form id="form" class="grid md:grid-cols-2 gap-4">
            <div class="md:col-span-2 rounded-xl border border-[#8A4703] bg-[#F8EFD1] p-3 text-xs text-[#1D1F1E]">
              <span class="font-semibold">Template Upload:</span>
              <a href="/validator/template/sales" class="underline ml-2">Data Penjualan</a>
              <span class="mx-1">|</span>
              <a href="/validator/template/promo" class="underline">Dataset Diskon Pabrik</a>
              <span class="mx-1">|</span>
              <a href="/validator/template/channel" class="underline">Data Channel by SUB</a>
            </div>
            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Data Penjualan <span class="text-xs text-[#1D1F1E]">(required)</span></div>
                <span id="badgeSales" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="sales" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Wajib. Tanpa ini, kita nggak bisa ngitung apa-apa.</div>
            </div>

            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Dataset Diskon Pabrik <span class="text-xs text-[#1D1F1E]">(required)</span></div>
                <span id="badgePromo" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="promo" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Program pabrik yang jadi patokan Expected_Pabrik.</div>
            </div>

            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Data Channel by SUB <span class="text-xs text-[#1D1F1E]">(required)</span></div>
                <span id="badgeChannel" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="channel" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Biar program by channel nggak salah pilih.</div>
            </div>

            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Dataset Diskon Internal <span class="text-xs text-[#1D1F1E]">(optional)</span></div>
                <span id="badgeInternal" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="internal" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Kalau diupload, Expected_Internal & FixClaim_Internal ikut dihitung.</div>
            </div>

            <div class="md:col-span-2">
              <button id="btn" type="submit" disabled
                      class="w-full rounded-2xl px-4 py-3 font-semibold text-[#EFEFEF]
                             bg-[#BD7401] hover:bg-[#8A4703] hover:text-[#EFEFEF] transition
                             disabled:opacity-50 disabled:cursor-not-allowed">
                Validate
              </button>

              <div class="mt-3 flex items-start gap-3">
                <div class="mt-0.5 h-2.5 w-2.5 rounded-full bg-[#EFEFEF]" id="readyDot"></div>
                <div class="text-sm text-[#1D1F1E]" id="msg">Upload 3 file wajib dulu, baru tombol Validate nyala.</div>
              </div>

              <div class="mt-3 grid md:grid-cols-3 gap-3">
                <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
                  <div class="text-xs text-[#1D1F1E]">Status A</div>
                  <div class="mt-1 text-sm font-semibold text-[#1D1F1E]">Tidak ada program cocok</div>
                  <div class="mt-1 text-xs text-[#1D1F1E]">Actual ada, expected 0</div>
                </div>
                <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
                  <div class="text-xs text-[#1D1F1E]">Status B</div>
                  <div class="mt-1 text-sm font-semibold text-[#1D1F1E]">Ada selisih</div>
                  <div class="mt-1 text-xs text-[#1D1F1E]">Selisih masuk TanpaTuan</div>
                </div>
                <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
                  <div class="text-xs text-[#1D1F1E]">Status C</div>
                  <div class="mt-1 text-sm font-semibold text-[#1D1F1E]">Sesuai</div>
                  <div class="mt-1 text-xs text-[#1D1F1E]">Dalam toleransi</div>
                </div>
              </div>
            </div>
          </form>
        </div>
      </main>

      <!-- RIGHT -->
      <section class="xl:col-span-3 space-y-5">
        __PATCH_NOTES_HTML__

        <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] shadow-lg overflow-hidden">
          <div class="p-5 border-b border-[#8A4703]">
            <div class="text-xs uppercase tracking-widest text-[#1D1F1E]">Insight</div>
            <div class="text-lg font-semibold text-[#1D1F1E]">Cara bacanya</div>
            <div class="mt-1 text-sm text-[#1D1F1E]">
              Goal: bikin kamu gampang ngecek “ini diskon beneran program atau nyasar”.
            </div>
          </div>

          <div class="p-5 space-y-3 text-sm text-[#1D1F1E]">
            <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
              <div class="font-semibold text-[#1D1F1E]">Urutan klaim (penting)</div>
              <div class="mt-1">Actual (LDISCAMOUNT) dibagi mengikuti urutan di <b>MDSTRING</b> untuk
                <b>Pabrik</b>/<b>Internal</b>, sisa masuk <b>TanpaTuan</b>.</div>
            </div>

            <details class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
              <summary class="cursor-pointer font-semibold text-[#1D1F1E]">Kenapa bisa “B”?</summary>
              <div class="mt-2 text-[#1D1F1E]">
                Karena expected (pabrik+internal) beda sama actual. Selisihnya otomatis ditaruh ke TanpaTuan,
                biar gampang investigasi.
              </div>
            </details>

            <details class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
              <summary class="cursor-pointer font-semibold text-[#1D1F1E]">Checklist sebelum validasi</summary>
              <ul class="mt-2 list-disc pl-5 space-y-1">
                <li>Pastikan 3 file wajib sudah hijau.</li>
                <li>Kalau internal belum ada, gap masuk TanpaTuan dulu.</li>
                <li>Kalau hasil “aneh”, cek kolom debug yang disediakan.</li>
              </ul>
            </details>
          </div>
        </div>
      </section>

    </div>
  </div>

<script>
const CSRF_TOKEN = "__CSRF_TOKEN__";
const CAN_VALIDATOR_EDIT = __CAN_VALIDATOR_EDIT__;
let downloadUrl = null;

function showToast(title, body){
  const t = document.getElementById('toast');
  document.getElementById('toastTitle').textContent = title;
  document.getElementById('toastBody').textContent = body;
  t.classList.remove('hidden');
  setTimeout(() => t.classList.add('hidden'), 5000);
}

async function checkHealth(){
  try{
    const r = await fetch('/health');
    const j = await r.json();
    document.getElementById('apiStatus').textContent = j.status || 'ok';
    document.getElementById('apiStatus').classList.remove('pulse-soft');
  }catch(e){
    document.getElementById('apiStatus').textContent = 'offline';
  }
}
checkHealth();

document.getElementById('downloadBtn').addEventListener('click', () => {
  if(downloadUrl) window.location.href = downloadUrl;
});

function setDot(id, ok, optional=false){
  const el = document.getElementById(id);
  el.className = "h-2.5 w-2.5 rounded-full " + (ok ? "bg-[#BD7401]" : "bg-[#EFEFEF] ring-1 ring-[#8A4703]");
}

function updateReadyState(){
  const sales = document.getElementById('sales').files[0];
  const promo = document.getElementById('promo').files[0];
  const channel = document.getElementById('channel').files[0];
  const internal = document.getElementById('internal').files[0];

  const ready = !!(sales && promo && channel);
  document.getElementById('btn').disabled = !ready;

  document.getElementById('readyDot').className = "mt-0.5 h-2.5 w-2.5 rounded-full " + (ready ? "bg-[#BD7401]" : "bg-[#EFEFEF] ring-1 ring-[#8A4703]");
  document.getElementById('readyPill').textContent = ready ? "siap" : "belum siap";
  document.getElementById('readyPill').className = "text-[11px] px-2 py-1 rounded-full border " + (ready ? "border-[#8A4703] bg-[#BD7401] text-[#EFEFEF]" : "border-[#8A4703] bg-[#EFEFEF] text-[#1D1F1E]");

  document.getElementById('msg').textContent = ready
    ? "Mantap. Kamu sudah bisa validate."
    : "Upload 3 file wajib dulu, baru tombol Validate nyala.";

  if(!CAN_VALIDATOR_EDIT){
    document.getElementById('btn').disabled = true;
    document.getElementById('msg').textContent = "Akses tidak diizinkan.";
  }

  document.getElementById('nameSales').textContent = sales ? sales.name : "belum diupload";
  document.getElementById('namePromo').textContent = promo ? promo.name : "belum diupload";
  document.getElementById('nameChannel').textContent = channel ? channel.name : "belum diupload";
  document.getElementById('nameInternal').textContent = internal ? internal.name : "boleh kosong";

  setDot('dotSales', !!sales);
  setDot('dotPromo', !!promo);
  setDot('dotChannel', !!channel);
  setDot('dotInternal', !!internal, true);

  document.getElementById('badgeSales').classList.toggle('hidden', !sales);
  document.getElementById('badgePromo').classList.toggle('hidden', !promo);
  document.getElementById('badgeChannel').classList.toggle('hidden', !channel);
  document.getElementById('badgeInternal').classList.toggle('hidden', !internal);
}

['sales','promo','channel','internal'].forEach(id => {
  document.getElementById(id).addEventListener('change', updateReadyState);
});
updateReadyState();

document.getElementById('form').addEventListener('submit', async (e) => {
  e.preventDefault();
  const btn = document.getElementById('btn');
  const dBtn = document.getElementById('downloadBtn');

  const sales = document.getElementById('sales').files[0];
  const promo = document.getElementById('promo').files[0];
  const channel = document.getElementById('channel').files[0];
  const internal = document.getElementById('internal').files[0];

  if(!(sales && promo && channel)){
    showToast("Eh belum bisa 😅", "__REQUIRED_MISSING_MSG__");
    return;
  }

  btn.disabled = true;
  btn.textContent = 'Validating... (lagi ngecek diskon satu-satu 😄)';
  dBtn.disabled = true;

  const fd = new FormData();
  fd.append('sales', sales);
  fd.append('promo', promo);
  fd.append('channel', channel);
  if(internal) fd.append('internal', internal);

  try{
    const res = await fetch('/validate_json', { method: 'POST', body: fd, headers: { 'X-CSRF-Token': CSRF_TOKEN } });
    const j = await res.json();
    if(!res.ok || !j.ok) throw new Error(j.error || ('HTTP ' + res.status));

    document.getElementById('cntA').textContent = j.counts.A;
    document.getElementById('cntB').textContent = j.counts.B;
    document.getElementById('cntC').textContent = j.counts.C;

    downloadUrl = j.download_url;
    dBtn.disabled = false;

    showToast("Selesai ✅", "Download hasilnya lewat tombol 'Download Result' di sidebar kiri.");
  }catch(err){
    showToast("Waduh error 😭", err.message || String(err));
  }finally{
    btn.disabled = false;
    btn.textContent = 'Validate';
  }
});
</script>
</body>
</html>"""












@app.get("/api/me")
def api_me(request: Request):
    user = get_current_user(request)
    token = get_or_create_csrf_token(request)
    samesite = _normalize_samesite(CSRF_COOKIE_SAMESITE)
    resp = JSONResponse({
        "ok": True, 
        "authenticated": bool(user), 
        "user": user if user else None,
        "csrf_token": token
    })
    resp.set_cookie(
        CSRF_COOKIE, token, httponly=False, max_age=CSRF_TTL_SECONDS,
        path="/", samesite=samesite, secure=AUTH_COOKIE_SECURE
    )
    return resp

@app.post("/api/logout")
def api_logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(AUTH_COOKIE)
    resp.delete_cookie(CSRF_COOKIE)
    return resp





@app.get("/favicon.ico")
@app.get("/favicon.png")
def favicon():
    path = find_favicon_path()
    if not path:
        return JSONResponse(status_code=404, content={"detail": "File not found"})
    return FileResponse(path, media_type=favicon_media_type(path))

@app.get("/payments/data")
def payments_data(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    db = load_payments_db()
    rows = []
    for key in sorted(db.get("lpb", {}).keys()):
        r = db["lpb"][key]
        row = dict(r)
        row["record_id"] = key
        row["tipe_pengajuan"] = normalize_pengajuan_type(row.get("tipe_pengajuan", "LPB"))
        row["jenis_dokumen"] = s(row.get("jenis_dokumen", ""))
        row["nomor_dokumen"] = s(row.get("nomor_dokumen", ""))
        if not s(row.get("no_lpb", "")) and row["tipe_pengajuan"] == "LPB":
            row["no_lpb"] = s(key)
        row["nilai_win_display"] = format_idr(float(row.get("nilai_win", 0) or 0.0))
        row["tgl_invoice"] = s(row.get("tgl_invoice", "")) or s(row.get("tgl_inv", ""))
        row["jt_invoice"] = s(row.get("jt_invoice", "")) or s(row.get("tgl_jtempo_pcp", ""))
        row["tgl_pembayaran"] = s(row.get("tgl_pembayaran", "")) or s(row.get("tgl_jtempo_pembayaran", ""))
        row["actual_date"] = s(row.get("actual_date", ""))
        nilai_invoice_raw = row.get("nilai_invoice", row.get("nilai_principle", ""))
        try:
            nilai_invoice_num = parse_number_id(nilai_invoice_raw)
        except Exception:
            nilai_invoice_num = 0.0
        raw_str = s(nilai_invoice_raw)
        empty_invoice = raw_str in ["", "0", "0.0"] or abs(nilai_invoice_num) < 1e-9
        row["nilai_invoice"] = "" if empty_invoice else format_idr(nilai_invoice_num)
        if empty_invoice:
            gap_val = 0.0
            row["gap_nilai"] = 0.0
            row["gap_nilai_display"] = ""
        else:
            try:
                gap_val = float(row.get("gap_nilai", 0) or 0.0)
            except Exception:
                gap_val = 0.0
            if abs(gap_val) < 1e-9:
                try:
                    gap_val = float(row.get("nilai_win", 0) or 0.0) - float(nilai_invoice_num or 0.0)
                except Exception:
                    gap_val = 0.0
            row["gap_nilai"] = gap_val
            row["gap_nilai_display"] = format_idr(gap_val)
        row["status_pembayaran"] = row.get("status_pembayaran", "")
        rows.append(row)
    return ORJSONResponse({"ok": True, "data": rows})

@app.get("/payments/export")
def payments_export(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})

    db = load_payments_db()
    rows: List[Dict[str, Any]] = []
    for key in sorted(db.get("lpb", {}).keys()):
        r = db["lpb"][key]
        nilai_sistem = parse_number_id(r.get("nilai_win", 0))
        nilai_invoice = parse_number_id(r.get("nilai_invoice", r.get("nilai_principle", 0)))
        potongan = parse_number_id(r.get("potongan", 0))
        if "nilai_pembayaran" in r:
            nilai_pembayaran = parse_number_id(r.get("nilai_pembayaran", 0))
        else:
            nilai_pembayaran = nilai_invoice
        if nilai_pembayaran < 0:
            nilai_pembayaran = 0.0
        if potongan < 0:
            potongan = 0.0
        gap_nilai = parse_number_id(r.get("gap_nilai", nilai_sistem - nilai_invoice))
        rows.append(
            {
                "Record ID": s(key),
                "Tipe Pengajuan": normalize_pengajuan_type(s(r.get("tipe_pengajuan", "LPB"))),
                "No LPB": s(r.get("no_lpb", key)),
                "Principle": s(r.get("principle", "")),
                "Tgl Setor": s(r.get("tgl_setor", "")),
                "Tgl Win": s(r.get("tgl_win", "")),
                "Tgl J.Tempo Win": s(r.get("tgl_jtempo_win", "")),
                "Nilai Sistem": nilai_sistem,
                "Tgl Terima Barang": s(r.get("tgl_terima_barang", "")),
                "Tgl Invoice": s(r.get("tgl_invoice", "")) or s(r.get("tgl_inv", "")),
                "No Invoice": s(r.get("invoice_no", "")) or s(r.get("nomor_dokumen", "")) or s(r.get("no_lpb", "")),
                "Nilai Invoice": nilai_invoice,
                "Potongan": potongan,
                "Nilai Pembayaran": nilai_pembayaran,
                "J.T Invoice": s(r.get("jt_invoice", "")) or s(r.get("tgl_jtempo_pcp", "")),
                "Gap Nilai": gap_nilai,
                "Actual Date": s(r.get("actual_date", "")),
                "Tgl Pembayaran": s(r.get("tgl_pembayaran", "")) or s(r.get("tgl_jtempo_pembayaran", "")),
                "Status Pembayaran": s(r.get("status_pembayaran", "")),
                "Metode Pembayaran": s(r.get("payment_method", "")),
                "Jenis Pembayaran": s(r.get("jenis_pembayaran", "")),
                "Tanggal Pengajuan Pembayaran": _normalize_yyyy_mm_dd(s(r.get("target_payment_date", ""))),
                "Jenis Dokumen": s(r.get("jenis_dokumen", "")),
                "Nomor Dokumen": s(r.get("nomor_dokumen", "")),
                "Keterangan": s(r.get("keterangan", "")),
                "Draft ID": s(r.get("draft_id", "")),
                "Submission ID": s(r.get("submission_id", "")),
                "SPPD No": s(r.get("sppd_no", "")),
                "Submitted At": s(r.get("submitted_at", "")),
                "Submitted By": s(r.get("submitted_by", "")),
                "Created At": s(r.get("created_at", "")),
                "Created By": s(r.get("created_by", "")),
            }
        )

    ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    filename = f"backup_payments_{ts}.xlsx"
    return _excel_download_response(rows, filename, "PAYMENTS")


@app.get("/payments/template")
def payments_template_download(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    if not user_has_permission(user, "payments", "view"):
        return HTMLResponse("Forbidden", status_code=403)
    rows = lpb_upload_template_rows()
    return _excel_download_response(rows, "template_upload_lpb.xlsx", "LPB_TEMPLATE")

@app.post("/payments/upload")
async def payments_upload(request: Request, file: UploadFile = File(None)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    if file is None:
        return JSONResponse(status_code=400, content={"ok": False, "error": "File belum diupload."})
    try:
        content = await read_upload_file_limited(
            file,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="File LPB",
        )
        preview_df = pd.read_excel(io.BytesIO(content), nrows=1)
        preview_cols = {str(c).strip().upper(): c for c in preview_df.columns}
        if looks_like_payments_backup(preview_cols):
            restore_rows = parse_payments_backup_upload(content)
            if not restore_rows:
                return JSONResponse(status_code=400, content={"ok": False, "error": "Data backup PAYMENTS kosong."})
            async with _PAYMENTS_DB_LOCK:
                db = load_payments_db()
                conflicts = validate_backup_restore_conflicts(db, restore_rows)
                if conflicts:
                    return JSONResponse(status_code=400, content={"ok": False, "error": "Restore backup dibatalkan: " + "; ".join(conflicts[:5])})
                for key, rec in restore_rows:
                    db["lpb"][key] = rec
                rebuild_payment_submissions(db)
                max_seq = max_sppd_sequence_from_records([rec for _, rec in restore_rows])
                if max_seq:
                    db["sppd_seq"] = max(int(db.get("sppd_seq", 0) or 0), max_seq)
                save_payments_db(db)
            append_audit_log(user, "payments_restore_backup", "lpb", {"added": len(restore_rows), "max_sppd_seq": max_seq})
            return JSONResponse({"ok": True, "added": len(restore_rows), "mode": "restore_backup", "message": f"Restore backup berhasil: {len(restore_rows)} record."})

        rows = parse_lpb_upload(content)
        if not rows:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Data LPB kosong."})
        now = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
        async with _PAYMENTS_DB_LOCK:
            db = load_payments_db()
            dups = []
            for r in rows:
                no_lpb = s(r.get("no_lpb", ""))
                if find_lpb_duplicate_key(db, no_lpb):
                    dups.append(no_lpb)
            if dups:
                return JSONResponse(status_code=400, content={"ok": False, "error": f"No. LPB {dups[0]} sudah ada di sistem, gagal upload"})
            for r in rows:
                key = normalize_lpb_no(r["no_lpb"])
                nilai_invoice = parse_number_id(r.get("nilai_invoice", 0))
                gap_nilai = parse_number_id(r.get("gap_nilai", 0))
                db["lpb"][key] = {
                    **r,
                    "record_id": key,
                    "tipe_pengajuan": "LPB",
                    "tgl_invoice": s(r.get("tgl_invoice", "")),
                    "jt_invoice": s(r.get("jt_invoice", "")),
                    "tgl_pembayaran": s(r.get("tgl_pembayaran", "")),
                    "actual_date": s(r.get("actual_date", "")),
                    "nilai_invoice": nilai_invoice if nilai_invoice > 0 else "",
                    "gap_nilai": gap_nilai,
                    "invoice_no": s(r.get("invoice_no", "")),
                    "status_pembayaran": "",
                    "payment_method": "",
                    "submitted_at": "",
                    "submitted_by": "",
                    "submission_id": "",
                    "draft_id": "",
                    "potongan": 0.0,
                    "nilai_pembayaran": 0.0,
                    "target_payment_date": "",
                    "jenis_dokumen": s(r.get("jenis_dokumen", "")),
                    "nomor_dokumen": s(r.get("nomor_dokumen", "")),
                    "keterangan": s(r.get("keterangan", "")),
                    "created_at": now,
                    "created_by": user,
                }
            save_payments_db(db)
        append_audit_log(user, "payments_upload", "lpb", {"added": len(rows)})
        return JSONResponse({"ok": True, "added": len(rows)})
    except ValueError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        append_error_log("payments_upload", e, {"user": user})
        return JSONResponse(status_code=500, content={"ok": False, "error": "Gagal memproses upload. Silakan coba lagi."})


@app.post("/payments/manual/add")
async def payments_manual_add(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}

    tipe = normalize_pengajuan_type(payload.get("tipe_pengajuan", "CBD"))
    no_lpb = s(payload.get("no_lpb", ""))
    principle = s(payload.get("principle", ""))
    invoice_no = s(payload.get("invoice_no", ""))
    nilai_invoice = parse_number_id(payload.get("nilai_invoice", 0))
    jenis_dokumen = s(payload.get("jenis_dokumen", ""))
    nomor_dokumen = s(payload.get("nomor_dokumen", ""))

    if not principle:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Principle wajib diisi."})
    if nilai_invoice <= 0:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Nilai Invoice wajib > 0."})
    if tipe == "LPB" and not no_lpb:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Tipe LPB wajib isi No. LPB."})
    if tipe == "NON_LPB" and (not jenis_dokumen or not nomor_dokumen):
        return JSONResponse(status_code=400, content={"ok": False, "error": "NON_LPB wajib isi Jenis Dokumen dan Nomor Dokumen."})

    now = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    async with _PAYMENTS_DB_LOCK:
      db = load_payments_db()
      if no_lpb and find_lpb_duplicate_key(db, no_lpb):
          return JSONResponse(status_code=400, content={"ok": False, "error": f"No. LPB {no_lpb} sudah ada di sistem."})
      if no_lpb:
          key = normalize_lpb_no(no_lpb)
          if key in db.get("lpb", {}):
              key = make_payment_record_id(tipe)
      else:
          key = make_payment_record_id(tipe)
      while key in db.get("lpb", {}):
          key = make_payment_record_id(tipe)
      db["lpb"][key] = {
        "record_id": key,
        "tipe_pengajuan": tipe,
        "no_lpb": no_lpb,
        "tgl_setor": "",
        "tgl_win": "",
        "tgl_jtempo_win": "",
        "principle": principle,
        "nilai_win": nilai_invoice,
        "tgl_terima_barang": "",
        "tgl_invoice": "",
        "jt_invoice": "",
        "tgl_pembayaran": "",
        "actual_date": "",
        "nilai_invoice": nilai_invoice,
        "gap_nilai": 0.0,
        "invoice_no": invoice_no,
        "status_pembayaran": "",
        "payment_method": "",
        "submitted_at": "",
        "submitted_by": "",
        "submission_id": "",
        "draft_id": "",
        "potongan": 0.0,
        "nilai_pembayaran": 0.0,
        "target_payment_date": "",
          "jenis_dokumen": jenis_dokumen,
          "nomor_dokumen": nomor_dokumen,
          "created_at": now,
          "created_by": user,
      }
      save_payments_db(db)
    append_audit_log(user, "payments_manual_add", "lpb", {"record_id": key, "tipe_pengajuan": tipe})
    return JSONResponse({"ok": True, "record_id": key})

@app.post("/payments/update")
async def payments_update(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "update"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    items = payload.get("items", [])
    if not isinstance(items, list):
        return JSONResponse(status_code=400, content={"ok": False, "error": "Format data tidak valid."})
    updated = []
    skipped = []
    async with _PAYMENTS_DB_LOCK:
        db = load_payments_db()
        for item in items:
            if not isinstance(item, dict):
                skipped.append("")
                continue
            row_id = s(item.get("record_id", "")) or s(item.get("id", "")) or s(item.get("no_lpb", ""))
            key = resolve_payment_record_key(db, row_id)
            if not key or key not in db.get("lpb", {}):
                skipped.append(row_id)
                continue
            rec = db["lpb"][key]
            tipe = normalize_pengajuan_type(item.get("tipe_pengajuan", rec.get("tipe_pengajuan", "LPB")))
            no_lpb = s(item.get("no_lpb", rec.get("no_lpb", "")))
            jenis_dokumen = s(item.get("jenis_dokumen", rec.get("jenis_dokumen", "")))
            nomor_dokumen = s(item.get("nomor_dokumen", rec.get("nomor_dokumen", "")))

            if tipe == "LPB" and not no_lpb:
                return JSONResponse(status_code=400, content={"ok": False, "error": "Tipe LPB wajib isi No. LPB."})
            if tipe == "NON_LPB" and (not jenis_dokumen or not nomor_dokumen):
                return JSONResponse(status_code=400, content={"ok": False, "error": "NON_LPB wajib isi Jenis Dokumen dan Nomor Dokumen."})
            if no_lpb:
                dup_key = find_lpb_duplicate_key(db, no_lpb, exclude_key=key)
                if dup_key:
                    return JSONResponse(status_code=400, content={"ok": False, "error": f"No. LPB {no_lpb} sudah dipakai record lain."})

            changed = False
            rec["record_id"] = key
            if "tipe_pengajuan" in item:
                rec["tipe_pengajuan"] = tipe
                changed = True
            if "no_lpb" in item:
                rec["no_lpb"] = no_lpb
                changed = True
            if "jenis_dokumen" in item:
                rec["jenis_dokumen"] = jenis_dokumen
                changed = True
            if "nomor_dokumen" in item:
                rec["nomor_dokumen"] = nomor_dokumen
                changed = True
            if "tgl_invoice" in item:
                rec["tgl_invoice"] = s(item.get("tgl_invoice", ""))
                changed = True
            if "jt_invoice" in item:
                rec["jt_invoice"] = s(item.get("jt_invoice", ""))
                changed = True
            if "tgl_pembayaran" in item:
                rec["tgl_pembayaran"] = s(item.get("tgl_pembayaran", ""))
                changed = True
            if "actual_date" in item:
                rec["actual_date"] = s(item.get("actual_date", ""))
                changed = True
            if "nilai_invoice" in item:
                rec["nilai_invoice"] = parse_number_id(item.get("nilai_invoice", 0))
                changed = True
            if "invoice_no" in item:
                rec["invoice_no"] = s(item.get("invoice_no", ""))
                changed = True
            if "principle" in item:
                rec["principle"] = s(item.get("principle", rec.get("principle", "")))
                changed = True
            if "ajukan" in item:
                rec["ajukan"] = bool(item.get("ajukan"))
                changed = True
            try:
                rec["gap_nilai"] = float(rec.get("nilai_win", 0) or 0.0) - float(rec.get("nilai_invoice", 0) or 0.0)
            except Exception:
                rec["gap_nilai"] = 0.0
            if changed:
                updated.append(key)
        save_payments_db(db)
    append_audit_log(user, "payments_update", "lpb", {"count": len(updated), "samples": updated[:10], "skipped": skipped[:10]})
    return JSONResponse({"ok": True, "updated": len(updated), "updated_ids": updated, "skipped": len(skipped)})

SPPD_EXCEL_FORBIDDEN_FIELDS = {
    "ajukan",
    "gap_nilai",
    "gap_nilai_display",
    "status_pembayaran",
    "status_tracking",
    "tracking_status",
    "submitted_at",
    "submitted_by",
    "submission_id",
    "draft_id",
    "sppd_no",
    "finance_status",
    "payment_proof",
    "proof_metadata",
    "accurate_posting",
}

SPPD_EXCEL_FIELD_ALIASES = {
    "recordid": "record_id",
    "record_id": "record_id",
    "id": "record_id",
    "tippepengajuan": "tipe_pengajuan",
    "tipepengajuan": "tipe_pengajuan",
    "tipe_pengajuan": "tipe_pengajuan",
    "nolpb": "no_lpb",
    "no_lpb": "no_lpb",
    "lpb": "no_lpb",
    "principle": "principle",
    "supplier": "principle",
    "vendor": "principle",
    "tglsetor": "tgl_setor",
    "tgl_setor": "tgl_setor",
    "tglwin": "tgl_win",
    "tgl_win": "tgl_win",
    "tgljtempowin": "tgl_jtempo_win",
    "tgl_jtempo_win": "tgl_jtempo_win",
    "tglterimabarang": "tgl_terima_barang",
    "tgl_terima_barang": "tgl_terima_barang",
    "tglinvoice": "tgl_invoice",
    "tgl_invoice": "tgl_invoice",
    "noinvoice": "invoice_no",
    "invoice": "invoice_no",
    "invoice_no": "invoice_no",
    "nomorinvoice": "invoice_no",
    "jenisdokumen": "jenis_dokumen",
    "jenis_dokumen": "jenis_dokumen",
    "nomordokumen": "nomor_dokumen",
    "nomor_dokumen": "nomor_dokumen",
    "nilaigiro": "nilai_invoice",
    "nilaiinvoice": "nilai_invoice",
    "nilai_invoice": "nilai_invoice",
    "nilaisistem": "nilai_win",
    "nilaiwin": "nilai_win",
    "nilai_win": "nilai_win",
    "potongan": "potongan",
    "nilaipembayaran": "nilai_pembayaran",
    "nilai_pembayaran": "nilai_pembayaran",
    "jtinvoice": "jt_invoice",
    "jt_invoice": "jt_invoice",
    "jatuhtempoinvoice": "jt_invoice",
    "actualdate": "actual_date",
    "actual_date": "actual_date",
    "tglpembayaran": "tgl_pembayaran",
    "tgl_pembayaran": "tgl_pembayaran",
    "tanggalpembayaran": "tgl_pembayaran",
    "tanggalpengajuanpembayaran": "target_payment_date",
    "targetpaymentdate": "target_payment_date",
    "target_payment_date": "target_payment_date",
    "metodepembayaran": "payment_method",
    "paymentmethod": "payment_method",
    "payment_method": "payment_method",
    "jenispembayaran": "jenis_pembayaran",
    "jenis_pembayaran": "jenis_pembayaran",
    "keterangan": "keterangan",
    "ajukan": "ajukan",
    "gap": "gap_nilai",
    "gapnilai": "gap_nilai",
    "gap_nilai": "gap_nilai",
    "status": "status_pembayaran",
    "statuspembayaran": "status_pembayaran",
    "status_pembayaran": "status_pembayaran",
    "statustracking": "status_tracking",
    "status_tracking": "status_tracking",
    "trackingstatus": "tracking_status",
    "tracking_status": "tracking_status",
    "submittedat": "submitted_at",
    "submitted_at": "submitted_at",
    "submittedby": "submitted_by",
    "submitted_by": "submitted_by",
    "submissionid": "submission_id",
    "submission_id": "submission_id",
    "draftid": "draft_id",
    "draft_id": "draft_id",
    "sppdno": "sppd_no",
    "sppd_no": "sppd_no",
}

SPPD_EXCEL_NUMERIC_FIELDS = {"nilai_invoice", "nilai_win", "potongan", "nilai_pembayaran"}
SPPD_EXCEL_DATE_FIELDS = {
    "tgl_setor",
    "tgl_win",
    "tgl_jtempo_win",
    "tgl_terima_barang",
    "tgl_invoice",
    "jt_invoice",
    "actual_date",
    "tgl_pembayaran",
    "target_payment_date",
}

def normalize_excel_field_name(value: Any) -> str:
    text = s(value).lower()
    text = text.replace(".", "")
    text = re.sub(r"[^a-z0-9_]+", "", text)
    return text

def normalize_sppd_excel_value(field: str, value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if field in SPPD_EXCEL_NUMERIC_FIELDS:
        return parse_number_id(value)
    if field in SPPD_EXCEL_DATE_FIELDS:
        return parse_sppd_date_ddmmyyyy(value)
    if field == "tipe_pengajuan":
        return normalize_pengajuan_type(value)
    return s(value)

def parse_sppd_excel_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    df = pd.read_excel(io.BytesIO(content))
    if df.empty:
        return [], [], []
    columns: Dict[str, str] = {}
    ignored_columns: List[str] = []
    blocked_columns: List[str] = []
    for col in df.columns:
        normalized = normalize_excel_field_name(col)
        field = SPPD_EXCEL_FIELD_ALIASES.get(normalized)
        if field in SPPD_EXCEL_FORBIDDEN_FIELDS:
            blocked_columns.append(s(col))
            continue
        if field:
            columns[s(col)] = field
        elif normalized:
            ignored_columns.append(s(col))
    rows: List[Dict[str, Any]] = []
    date_errors: List[str] = []
    for idx, (_, row) in enumerate(df.iterrows()):
        excel_row = idx + 2  # header di baris 1
        item: Dict[str, Any] = {}
        for col, field in columns.items():
            try:
                value = normalize_sppd_excel_value(field, row.get(col))
            except ValueError as de:
                date_errors.append(f"kolom '{s(col)}' baris {excel_row} = {str(de)}")
                continue
            if value is None or s(value) == "":
                continue
            item[field] = value
        if item:
            rows.append(item)
    if date_errors:
        shown = "; ".join(date_errors[:5])
        extra = len(date_errors) - 5
        suffix = f"; dan {extra} lainnya" if extra > 0 else ""
        raise ValueError(
            f"Tanggal tidak valid (harus DD/MM/YYYY): {shown}{suffix}. Upload dibatalkan."
        )
    return rows, ignored_columns, blocked_columns

@app.post("/payments/sppd/upload")
async def payments_sppd_upload(request: Request, file: UploadFile = File(None)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "sppd", "upload_excel"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    if file is None:
        return JSONResponse(status_code=400, content={"ok": False, "error": "File Excel belum diupload."})
    try:
        content = await read_upload_file_limited(
            file,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="File SPPD",
        )
        rows, ignored_columns, blocked_columns = parse_sppd_excel_rows(content)
        if not rows:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Tidak ada baris valid untuk diupdate."})
        db = load_payments_db()
        updated: List[str] = []
        not_found: List[str] = []
        changed_fields: Dict[str, int] = {}
        for item in rows:
            row_id = s(item.get("record_id", "")) or s(item.get("no_lpb", ""))
            key = resolve_payment_record_key(db, row_id)
            if not key or key not in db.get("lpb", {}):
                not_found.append(row_id or "-")
                continue
            rec = db["lpb"][key]
            next_no_lpb = s(item.get("no_lpb", rec.get("no_lpb", "")))
            if next_no_lpb:
                dup_key = find_lpb_duplicate_key(db, next_no_lpb, exclude_key=key)
                if dup_key:
                    return JSONResponse(status_code=400, content={"ok": False, "error": f"No. LPB {next_no_lpb} sudah dipakai record lain."})
            for field, value in item.items():
                if field == "record_id" or field in SPPD_EXCEL_FORBIDDEN_FIELDS:
                    continue
                rec[field] = value
                changed_fields[field] = changed_fields.get(field, 0) + 1
            if "nilai_invoice" in item or "nilai_win" in item:
                try:
                    rec["gap_nilai"] = float(parse_number_id(rec.get("nilai_win", 0))) - float(parse_number_id(rec.get("nilai_invoice", 0)))
                except Exception:
                    rec["gap_nilai"] = 0.0
            updated.append(key)
        if not updated:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Tidak ada record yang cocok untuk diupdate.", "not_found": not_found[:20]})
        save_payments_db(db)
        append_audit_log(user, "payments_sppd_excel_upload", "lpb", {
            "updated": len(updated),
            "not_found": len(not_found),
            "changed_fields": changed_fields,
            "blocked_columns": blocked_columns,
        })
        return JSONResponse({
            "ok": True,
            "updated": len(updated),
            "not_found": not_found[:20],
            "ignored_columns": ignored_columns[:30],
            "blocked_columns": blocked_columns[:30],
            "changed_fields": changed_fields,
        })
    except ValueError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        append_error_log("payments_sppd_upload", e, {"user": user})
        return JSONResponse(status_code=500, content={"ok": False, "error": "Gagal memproses upload Excel SPPD."})

@app.post("/payments/delete")
async def payments_delete(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "delete"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    record_ids = payload.get("record_ids", payload.get("no_lpbs", []))
    if not isinstance(record_ids, list) or not record_ids:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Data belum dipilih."})

    db = load_payments_db()
    deleted = 0
    for row_id in record_ids:
        key = resolve_payment_record_key(db, s(row_id))
        if key in db.get("lpb", {}):
            del db["lpb"][key]
            deleted += 1
    save_payments_db(db)
    append_audit_log(user, "payments_delete", "lpb", {"count": deleted, "samples": [s(n) for n in record_ids][:10]})
    return JSONResponse({"ok": True, "deleted": deleted})

@app.post("/payments/clear")
async def payments_clear(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not is_admin_user(user):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Hanya admin yang bisa clear seluruh data payments."})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    if s(payload.get("confirm", "")) != "CLEAR PAYMENTS":
        return JSONResponse(status_code=400, content={"ok": False, "error": "Konfirmasi wajib tepat: CLEAR PAYMENTS"})
    if not PAYMENTS_DB_PATH:
        return JSONResponse(status_code=500, content={"ok": False, "error": "PAYMENTS_DB_PATH belum dikonfigurasi."})

    backup_name = ""
    try:
        async with _PAYMENTS_DB_LOCK:
            db = load_payments_db()
            before_counts = {
                "lpb": len(db.get("lpb", {}) or {}),
                "submissions": len(db.get("submissions", {}) or {}),
                "drafts": len(db.get("drafts", {}) or {}),
                "proofs": len(db.get("proofs", {}) or {}),
            }
            os.makedirs(os.path.dirname(PAYMENTS_DB_PATH), exist_ok=True)
            if os.path.exists(PAYMENTS_DB_PATH):
                ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                backup_path = f"{PAYMENTS_DB_PATH}.backup_before_clear_{ts}"
                import shutil
                shutil.copy2(PAYMENTS_DB_PATH, backup_path)
                backup_name = os.path.basename(backup_path)
            save_payments_db(empty_payments_db_preserving_config(db))
        append_audit_log(user, "payments_clear_all", "payments", {"backup": backup_name, "before": before_counts})
        return JSONResponse({
            "ok": True,
            "cleared": before_counts,
            "backup_file": backup_name,
            "preserved": ["finance_mappings", "sppd_settings", "sppd_seq"],
        })
    except Exception as e:
        append_error_log("payments_clear", e, {"user": user, "backup": backup_name})
        return JSONResponse(status_code=500, content={"ok": False, "error": "Gagal clear data payments."})

def _can_access_draft(user: str, draft: Dict[str, Any]) -> bool:
    if not user or not draft:
        return False
    if is_admin_user(user):
        return True
    return s(draft.get("created_by", "")) == s(user)

@app.post("/payments/cart/create")
async def payments_cart_create(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    method = s(payload.get("method", ""))
    record_ids = payload.get("record_ids", payload.get("no_lpbs", []))
    target_payment_date = _normalize_yyyy_mm_dd(s(payload.get("target_payment_date", "")))
    if not target_payment_date:
        target_payment_date = (pd.Timestamp.now() + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
    if method not in ["NON_PANIN", "BANK_PANIN"]:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Metode pembayaran tidak valid."})
    if not isinstance(record_ids, list) or not record_ids:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Data belum dipilih."})

    async with _PAYMENTS_DB_LOCK:
      db = load_payments_db()
      selected = []
      for row_id in record_ids:
          key = resolve_payment_record_key(db, s(row_id))
          rec = db.get("lpb", {}).get(key)
          if rec:
              selected.append({**rec, "record_id": key})
      if not selected:
          return JSONResponse(status_code=400, content={"ok": False, "error": "Data pengajuan tidak ditemukan."})

    for rec in selected:
        tipe = normalize_pengajuan_type(rec.get("tipe_pengajuan", "LPB"))
        no_lpb = s(rec.get("no_lpb", ""))
        principle = s(rec.get("principle", ""))
        tgl_invoice = s(rec.get("tgl_invoice", "")) or s(rec.get("tgl_inv", ""))
        jt_invoice = s(rec.get("jt_invoice", "")) or s(rec.get("tgl_jtempo_pcp", ""))
        invoice_no = s(rec.get("invoice_no", ""))
        jenis_dokumen = s(rec.get("jenis_dokumen", ""))
        nomor_dokumen = s(rec.get("nomor_dokumen", ""))
        nilai_invoice = parse_number_id(rec.get("nilai_invoice", rec.get("nilai_principle", 0)))
        if not principle:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Principle wajib diisi sebelum diajukan."})
        if tipe == "LPB" and not no_lpb:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Tipe LPB wajib isi No. LPB (record {rec.get('record_id','')})."})
        if tipe == "LPB" and (not tgl_invoice or not jt_invoice):
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Lengkapi tanggal invoice & jatuh tempo invoice untuk LPB {no_lpb or rec.get('record_id','')}."})
        if tipe == "LPB" and not invoice_no:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Nomor Invoice kosong untuk LPB {no_lpb or rec.get('record_id','')}."})
        if tipe == "CBD" and (not invoice_no and not no_lpb):
            return JSONResponse(status_code=400, content={"ok": False, "error": f"CBD wajib isi No. Invoice atau No. LPB manual untuk principle {principle}."})
        if tipe == "NON_LPB" and (not jenis_dokumen or not nomor_dokumen):
            return JSONResponse(status_code=400, content={"ok": False, "error": f"NON_LPB wajib isi Jenis Dokumen dan Nomor Dokumen untuk principle {principle}."})
        if float(nilai_invoice or 0) <= 0:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Nilai Invoice kosong untuk principle {principle}."})
        if tipe == "LPB" and has_submitted_duplicate_payment(db, s(rec.get("record_id", "")), rec):
            return JSONResponse(status_code=400, content={"ok": False, "error": f"LPB untuk principle {principle} terindikasi sudah pernah diajukan (kemungkinan case CBD). Cek data finance terlebih dulu."})
        if s(rec.get("submission_id", "")) and s(rec.get("status_pembayaran", "")).lower() != "ajukan ulang":
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Record {no_lpb or rec.get('record_id','')} sudah pernah diajukan ke finance."})

    order_keys: List[str] = []
    groups: Dict[str, List[Dict[str, Any]]] = {}
    group_meta: Dict[str, Dict[str, str]] = {}
    for rec in selected:
        pr = s(rec.get("principle", ""))
        tipe = normalize_pengajuan_type(rec.get("tipe_pengajuan", "LPB"))
        gk = f"{pr}||{tipe}"
        if gk not in groups:
            groups[gk] = []
            group_meta[gk] = {"principle": pr, "tipe_pengajuan": tipe}
            order_keys.append(gk)
        groups[gk].append(rec)

    items = []
    idx = 1
    for gk in order_keys:
        recs = groups[gk]
        meta = group_meta[gk]
        inv_list = [s(r.get("invoice_no", "")) for r in recs if s(r.get("invoice_no", ""))]
        ref_docs = [s(r.get("nomor_dokumen", "")) for r in recs if s(r.get("nomor_dokumen", ""))]
        lpb_refs = [s(r.get("no_lpb", "")) for r in recs if s(r.get("no_lpb", ""))]
        total = sum(parse_number_id(r.get("nilai_invoice", r.get("nilai_principle", 0))) for r in recs)
        refs = inv_list if inv_list else (ref_docs if ref_docs else lpb_refs)
        ref_concat = ", ".join(refs)
        items.append({
            "no": idx,
            "group_key": gk,
            "principle": meta["principle"],
            "tipe_pengajuan": meta["tipe_pengajuan"],
            "total": total,
            "invoice_list": inv_list,
            "invoice_concat": ref_concat,
            "potongan": 0.0,
            "nilai_pembayaran": total,
            "jenis_pembayaran": "",
            "keterangan": "",
        })
        idx += 1

    draft_id = str(uuid.uuid4())[:8]
    now = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    async with _PAYMENTS_DB_LOCK:
        db2 = load_payments_db()
        db2["drafts"][draft_id] = {
            "id": draft_id,
            "created_at": now,
            "created_by": user,
            "method": method,
            "target_payment_date": target_payment_date,
            "record_ids": [s(r.get("record_id", "")) for r in selected],
            "items": items,
        }
        save_payments_db(db2)
    append_audit_log(user, "payments_cart_create", "draft", {"draft_id": draft_id, "count": len(selected), "types": sorted({normalize_pengajuan_type(x.get("tipe_pengajuan", "")) for x in selected})})
    return JSONResponse({"ok": True, "draft_id": draft_id})


@app.get("/payments/cart-info")
def payments_cart_data(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    draft_id = s(request.query_params.get("draft", ""))
    db = load_payments_db()
    drafts = db.get("drafts", {})
    draft = drafts.get(draft_id) or drafts.get(draft_id.lower()) or drafts.get(draft_id.upper())
    if not draft or not _can_access_draft(user, draft):
        if is_admin_user(user):
            keys = list(drafts.keys())
            preview = ", ".join(keys[:8])
            msg = f"Draft tidak ditemukan. PATH={PAYMENTS_DB_PATH}. Drafts: {preview}"
            return JSONResponse(status_code=404, content={"ok": False, "error": msg})
        return JSONResponse(status_code=404, content={"ok": False, "error": "Draft tidak ditemukan."})
    items = []
    for it in draft.get("items", []):
        total = float(it.get("total", 0) or 0.0)
        potongan = parse_number_id(it.get("potongan", 0))
        if potongan < 0:
            potongan = 0.0
        if potongan > total:
            potongan = total
        nilai_pembayaran = parse_number_id(it.get("nilai_pembayaran", total - potongan))
        expected_pay = max(total - potongan, 0.0)
        if abs(nilai_pembayaran - expected_pay) > 0.5:
            nilai_pembayaran = expected_pay
        items.append({
            "no": it.get("no", ""),
            "group_key": s(it.get("group_key", "")),
            "principle": it.get("principle", ""),
            "tipe_pengajuan": normalize_pengajuan_type(it.get("tipe_pengajuan", "LPB")),
            "total": total,
            "total_display": format_idr(total),
            "invoice_concat": it.get("invoice_concat", ""),
            "potongan": potongan,
            "potongan_display": format_idr(potongan),
            "nilai_pembayaran": nilai_pembayaran,
            "nilai_pembayaran_display": format_idr(nilai_pembayaran),
            "jenis_pembayaran": it.get("jenis_pembayaran", ""),
            "keterangan": it.get("keterangan", ""),
        })
    method = s(draft.get("method", ""))
    method_label = "Bank Panin" if method == "BANK_PANIN" else ("Non Panin" if method == "NON_PANIN" else "")
    target_payment_date = _normalize_yyyy_mm_dd(s(draft.get("target_payment_date", "")))
    if not target_payment_date:
        target_payment_date = (pd.Timestamp.now() + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
    return JSONResponse({
        "ok": True,
        "items": items,
        "method": method,
        "method_label": method_label,
        "target_payment_date": target_payment_date,
    })

@app.post("/payments/cart/submit")
async def payments_cart_submit(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    draft_id = s(payload.get("draft_id", ""))
    items = payload.get("items", [])
    target_payment_date = _normalize_yyyy_mm_dd(s(payload.get("target_payment_date", "")))
    if not draft_id:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Draft tidak valid."})
    if not isinstance(items, list):
        return JSONResponse(status_code=400, content={"ok": False, "error": "Format data tidak valid."})

    # Validasi item_input_map bisa dilakukan sebelum lock (pure CPU, tidak baca DB)
    item_input_map: Dict[str, Dict[str, Any]] = {}
    for it in items:
        group_key = s(it.get("group_key", ""))
        pr = s(it.get("principle", ""))
        jenis = s(it.get("jenis_pembayaran", "")).upper()
        ket = s(it.get("keterangan", ""))
        potongan = parse_number_id(it.get("potongan", 0))
        if not group_key and pr:
            tipe = normalize_pengajuan_type(it.get("tipe_pengajuan", "LPB"))
            group_key = f"{pr}||{tipe}"
        if not group_key:
            continue
        if jenis not in ["TRF", "DF", "VA"]:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Jenis pembayaran wajib diisi untuk {pr or group_key}."})
        if potongan < 0:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Potongan tidak boleh minus untuk {pr or group_key}."})
        item_input_map[group_key] = {"jenis_pembayaran": jenis, "keterangan": ket, "potongan": potongan}

    # Lock mulai dari sini — semua read-modify-write dilindungi satu lock
    async with _PAYMENTS_DB_LOCK:
        db = load_payments_db()
        draft = db.get("drafts", {}).get(draft_id)
        if not draft or not _can_access_draft(user, draft):
            return JSONResponse(status_code=404, content={"ok": False, "error": "Draft tidak ditemukan."})
        method = s(draft.get("method", ""))
        if method not in ["NON_PANIN", "BANK_PANIN"]:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Metode pembayaran tidak valid."})
        if not target_payment_date:
            target_payment_date = _normalize_yyyy_mm_dd(s(draft.get("target_payment_date", "")))
        if not target_payment_date:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Tanggal pengajuan pembayaran wajib diisi."})

        selected = []
        selected_ids = draft.get("record_ids", draft.get("lpb", []))
        for row_id in selected_ids:
            key = resolve_payment_record_key(db, s(row_id))
            rec = db.get("lpb", {}).get(key)
            if rec:
                selected.append({**rec, "record_id": key})
        if not selected:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Data pengajuan tidak ditemukan."})

    submission_id = str(uuid.uuid4())[:8]
    submit_dt = pd.Timestamp.now()
    now = submit_dt.strftime("%Y-%m-%d %H:%M:%S")
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for rec in selected:
        pr = s(rec.get("principle", ""))
        tipe = normalize_pengajuan_type(rec.get("tipe_pengajuan", "LPB"))
        gk = f"{pr}||{tipe}"
        groups.setdefault(gk, []).append(rec)

    item_map: Dict[str, Dict[str, Any]] = {}
    for group_key, recs in groups.items():
        principle = s(recs[0].get("principle", ""))
        tipe = normalize_pengajuan_type(recs[0].get("tipe_pengajuan", "LPB"))
        if group_key not in item_input_map:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Data cart untuk {principle} ({tipe}) tidak lengkap."})
        total_invoice = sum(parse_number_id(r.get("nilai_invoice", r.get("nilai_principle", 0))) for r in recs)
        info = dict(item_input_map[group_key])
        potongan = float(info.get("potongan", 0.0) or 0.0)
        if potongan > total_invoice:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Potongan melebihi Nilai Invoice untuk {principle} ({tipe})."})
        info["group_key"] = group_key
        info["principle"] = principle
        info["tipe_pengajuan"] = tipe
        info["total_invoice"] = total_invoice
        info["nilai_pembayaran"] = max(total_invoice - potongan, 0.0)
        item_map[group_key] = info

    files = []
    for group_key, recs in groups.items():
        principle = s(recs[0].get("principle", ""))
        tipe = normalize_pengajuan_type(recs[0].get("tipe_pengajuan", "LPB"))
        inv_list = [s(r.get("invoice_no", "")) for r in recs if s(r.get("invoice_no", ""))]
        doc_list = [s(r.get("nomor_dokumen", "")) for r in recs if s(r.get("nomor_dokumen", ""))]
        lpb_refs = [s(r.get("no_lpb", "")) for r in recs if s(r.get("no_lpb", ""))]
        refs = inv_list if inv_list else (doc_list if doc_list else lpb_refs)
        cart_info = item_map.get(group_key, {})
        total = float(cart_info.get("total_invoice", 0.0) or 0.0)
        potongan = float(cart_info.get("potongan", 0.0) or 0.0)
        nilai_pembayaran = float(cart_info.get("nilai_pembayaran", total) or 0.0)
        rows = [{
            "No": 1,
            "Tipe Pengajuan": tipe,
            "Principle": principle,
            "Nilai Invoice (Total)": total,
            "No. Invoice / Dokumen": ", ".join(refs),
            "Potongan": potongan,
            "Nilai Pembayaran": nilai_pembayaran,
            "Jenis Pembayaran": cart_info.get("jenis_pembayaran", ""),
            "Keterangan": cart_info.get("keterangan", ""),
        }]
        fname = f"invoice_{submission_id}_{slugify(principle)}_{slugify(tipe)}.xlsx"
        out_path = os.path.join(PAYMENTS_FILES_DIR, fname)
        write_invoice_excel(rows, out_path)
        files.append({"label": f"Invoice {principle} ({tipe})", "url": f"/payments/files/{fname}"})

    sppd_file = None
    sppd_no = None
    if method == "BANK_PANIN":
        bank_map, norm_keys = load_bank_map_with_normalized_keys()
        transfer_items = []
        total_all = 0.0
        for group_key, recs in groups.items():
            principle = s(recs[0].get("principle", ""))
            info, match_status = find_best_match(principle, bank_map, norm_keys)
            if not info or match_status == "ambiguous":
                return JSONResponse(status_code=400, content={"ok": False, "error": f"Data rekening untuk principle '{principle}' tidak ditemukan. Status: {match_status}"})
            cart_info = item_map.get(group_key, {})
            amount = float(cart_info.get("nilai_pembayaran", 0.0) or 0.0)
            total_all += amount
            transfer_items.append(
                {
                    "principle": info["principle"],
                    "bank": info["bank"],
                    "rekening": info["rekening"],
                    "penerima": info["penerima"],
                    "amount": amount,
                }
            )
        _, sppd_no, sppd_settings = next_sppd_number(db, submit_dt)
        sppd_name = f"sppd_{submission_id}.docx"
        sppd_path = os.path.join(PAYMENTS_FILES_DIR, sppd_name)
        render_sppd_docx(SPPD_TEMPLATE_PATH, sppd_path, submit_dt, sppd_no, transfer_items, sppd_settings)
        files.append({"label": "SPPD Bank Panin", "url": f"/payments/files/{sppd_name}"})
        sppd_file = sppd_name

    payment_alloc_by_lpb: Dict[str, float] = {}
    potongan_alloc_by_lpb: Dict[str, float] = {}
    for group_key, recs in groups.items():
        cart_info = item_map.get(group_key, {})
        total_invoice = float(cart_info.get("total_invoice", 0.0) or 0.0)
        total_pembayaran = float(cart_info.get("nilai_pembayaran", 0.0) or 0.0)
        remain = total_pembayaran
        for idx, rec in enumerate(recs):
            key = s(rec.get("record_id", ""))
            nilai_invoice = parse_number_id(rec.get("nilai_invoice", rec.get("nilai_principle", 0)))
            if idx == len(recs) - 1:
                nilai_bayar = remain
            elif total_invoice <= 0:
                nilai_bayar = 0.0
            else:
                nilai_bayar = total_pembayaran * (nilai_invoice / total_invoice)
                remain -= nilai_bayar
            nilai_bayar = max(0.0, min(float(nilai_invoice or 0.0), float(nilai_bayar or 0.0)))
            payment_alloc_by_lpb[key] = nilai_bayar
            potongan_alloc_by_lpb[key] = max(float(nilai_invoice or 0.0) - nilai_bayar, 0.0)

    # Tulis semua perubahan ke DB di dalam lock
    async with _PAYMENTS_DB_LOCK:
        db = load_payments_db()
        for rec in selected:
            key = s(rec.get("record_id", ""))
            if key in db.get("lpb", {}):
                principle = s(db["lpb"][key].get("principle", ""))
                tipe = normalize_pengajuan_type(db["lpb"][key].get("tipe_pengajuan", "LPB"))
                group_key = f"{principle}||{tipe}"
                cart_info = item_map.get(group_key, {})
                db["lpb"][key]["payment_method"] = "Bank Panin" if method == "BANK_PANIN" else "Non Panin"
                db["lpb"][key]["status_pembayaran"] = "Belum Transfer"
                db["lpb"][key]["submitted_at"] = now
                db["lpb"][key]["submitted_by"] = user
                db["lpb"][key]["submission_id"] = submission_id
                db["lpb"][key]["draft_id"] = draft_id
                db["lpb"][key]["target_payment_date"] = target_payment_date
                db["lpb"][key]["jenis_pembayaran"] = cart_info.get("jenis_pembayaran", "")
                db["lpb"][key]["keterangan"] = cart_info.get("keterangan", "")
                db["lpb"][key]["potongan"] = float(potongan_alloc_by_lpb.get(key, 0.0) or 0.0)
                db["lpb"][key]["nilai_pembayaran"] = float(payment_alloc_by_lpb.get(key, 0.0) or 0.0)
                if sppd_no:
                    db["lpb"][key]["sppd_no"] = sppd_no

        db["submissions"][submission_id] = {
            "id": submission_id,
            "created_at": now,
            "created_by": user,
            "draft_id": draft_id,
            "method": method,
            "target_payment_date": target_payment_date,
            "record_ids": [s(r.get("record_id", "")) for r in selected],
            "files": files,
            "sppd_file": sppd_file,
            "sppd_no": sppd_no,
            "cart_items": item_map,
        }
        if draft_id in db.get("drafts", {}):
            del db["drafts"][draft_id]
        save_payments_db(db)
    append_audit_log(
        user,
        "payments_cart_submit",
        "submission",
        {"submission_id": submission_id, "method": method, "count": len(selected)},
    )
    return JSONResponse({"ok": True, "submission_id": submission_id, "files": files})

@app.post("/payments/submit")
async def payments_submit(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    return JSONResponse(status_code=400, content={"ok": False, "error": "Gunakan keranjang dulu sebelum diajukan ke finance."})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    method = s(payload.get("method", ""))
    no_lpbs = payload.get("no_lpbs", [])
    if method not in ["NON_PANIN", "BANK_PANIN"]:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Metode pembayaran tidak valid."})
    if not isinstance(no_lpbs, list) or not no_lpbs:
        return JSONResponse(status_code=400, content={"ok": False, "error": "LPB belum dipilih."})

    db = load_payments_db()
    selected = []
    for no in no_lpbs:
        key = normalize_lpb_no(s(no))
        rec = db.get("lpb", {}).get(key)
        if rec:
            selected.append(rec)
    if not selected:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Data LPB tidak ditemukan."})

    for rec in selected:
        tgl_invoice = s(rec.get("tgl_invoice", "")) or s(rec.get("tgl_inv", ""))
        jt_invoice = s(rec.get("jt_invoice", "")) or s(rec.get("tgl_jtempo_pcp", ""))
        invoice_no = s(rec.get("invoice_no", ""))
        nilai_invoice = parse_number_id(rec.get("nilai_invoice", rec.get("nilai_principle", 0)))
        if not tgl_invoice or not jt_invoice:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Lengkapi tanggal invoice & jatuh tempo invoice untuk LPB {rec.get('no_lpb','')}"})
        if not invoice_no:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Nomor Invoice kosong untuk LPB {rec.get('no_lpb','')}"})
        if float(nilai_invoice or 0) <= 0:
            return JSONResponse(status_code=400, content={"ok": False, "error": f"Nilai Invoice kosong untuk LPB {rec.get('no_lpb','')}"})

    submission_id = str(uuid.uuid4())[:8]
    submit_dt = pd.Timestamp.now()
    now = submit_dt.strftime("%Y-%m-%d %H:%M:%S")
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for rec in selected:
        key = s(rec.get("principle", ""))
        groups.setdefault(key, []).append(rec)

    files = []
    for principle, recs in groups.items():
        inv_list = [s(r.get("invoice_no", "")) for r in recs if s(r.get("invoice_no", ""))]
        total = sum(parse_number_id(r.get("nilai_invoice", r.get("nilai_principle", 0))) for r in recs)
        rows = [{
            "No": 1,
            "Principle": principle,
            "Nilai Invoice (Total)": total,
            "No. Invoice": ", ".join(inv_list),
            "Potongan": 0.0,
            "Nilai Pembayaran": total,
            "Jenis Pembayaran": "NON PANIN" if method == "NON_PANIN" else "BANK PANIN",
            "Keterangan": "",
        }]
        fname = f"invoice_{submission_id}_{slugify(principle)}.xlsx"
        out_path = os.path.join(PAYMENTS_FILES_DIR, fname)
        write_invoice_excel(rows, out_path)
        files.append({"label": f"Invoice {principle}", "url": f"/payments/files/{fname}"})

    sppd_file = None
    sppd_no = None
    if method == "BANK_PANIN":
        bank_map, norm_keys = load_bank_map_with_normalized_keys()
        transfer_items = []
        total_all = 0.0
        for principle, recs in groups.items():
            info, match_status = find_best_match(principle, bank_map, norm_keys)
            if not info or match_status == "ambiguous":
                return JSONResponse(status_code=400, content={"ok": False, "error": f"Data rekening untuk principle '{principle}' tidak ditemukan. Status: {match_status}"})
            amount = sum(parse_number_id(r.get("nilai_invoice", r.get("nilai_principle", 0))) for r in recs)
            total_all += amount
            transfer_items.append(
                {
                    "principle": info["principle"],
                    "bank": info["bank"],
                    "rekening": info["rekening"],
                    "penerima": info["penerima"],
                    "amount": amount,
                }
            )
        _, sppd_no, sppd_settings = next_sppd_number(db, submit_dt)
        sppd_name = f"sppd_{submission_id}.docx"
        sppd_path = os.path.join(PAYMENTS_FILES_DIR, sppd_name)
        render_sppd_docx(SPPD_TEMPLATE_PATH, sppd_path, submit_dt, sppd_no, transfer_items, sppd_settings)
        files.append({"label": "SPPD Bank Panin", "url": f"/payments/files/{sppd_name}"})
        sppd_file = sppd_name

    for rec in selected:
        key = normalize_lpb_no(rec.get("no_lpb", ""))
        if key in db.get("lpb", {}):
            db["lpb"][key]["payment_method"] = "Bank Panin" if method == "BANK_PANIN" else "Non Panin"
            db["lpb"][key]["status_pembayaran"] = "Belum Transfer"
            db["lpb"][key]["submitted_at"] = now
            db["lpb"][key]["submitted_by"] = user
            db["lpb"][key]["submission_id"] = submission_id
            db["lpb"][key]["potongan"] = 0.0
            db["lpb"][key]["nilai_pembayaran"] = parse_number_id(db["lpb"][key].get("nilai_invoice", db["lpb"][key].get("nilai_principle", 0)))
            if sppd_no:
                db["lpb"][key]["sppd_no"] = sppd_no

    db["submissions"][submission_id] = {
        "id": submission_id,
        "created_at": now,
        "created_by": user,
        "method": method,
        "lpb": [normalize_lpb_no(r.get("no_lpb", "")) for r in selected],
        "files": files,
        "sppd_file": sppd_file,
        "sppd_no": sppd_no,
    }
    save_payments_db(db)
    append_audit_log(
        user,
        "payments_submit",
        "submission",
        {"submission_id": submission_id, "method": method, "count": len(selected)},
    )
    return JSONResponse({"ok": True, "submission_id": submission_id, "files": files})

@app.get("/payments/files/{file_name}")
def payments_files(request: Request, file_name: str):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    safe_name = os.path.basename(file_name)
    path = os.path.join(PAYMENTS_FILES_DIR, safe_name)
    if not os.path.exists(path):
        return JSONResponse(status_code=404, content={"detail": "File not found"})
    return FileResponse(path, filename=safe_name)

@app.get("/payments/sppd/settings")
def payments_sppd_settings_get(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "sppd", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    db = load_payments_db()
    settings = get_sppd_settings(db)
    preview_date = _normalize_yyyy_mm_dd(s(request.query_params.get("date", ""))) or pd.Timestamp.today().strftime("%Y-%m-%d")
    preview_dt = pd.to_datetime(preview_date)
    next_seq = int(settings.get("last_sequence", 0)) + 1
    return JSONResponse({
        "ok": True,
        "settings": settings,
        "next_sequence": next_seq,
        "preview_number": format_sppd_number_with_template(next_seq, preview_dt, s(settings.get("number_template", ""))),
        "preview_date": preview_date,
        "template_path": SPPD_TEMPLATE_PATH,
    })

@app.post("/payments/sppd/settings")
async def payments_sppd_settings_save(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "sppd", "edit_settings"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    db = load_payments_db()
    current = get_sppd_settings(db)
    settings = normalize_sppd_settings({**current, **(payload if isinstance(payload, dict) else {})}, db)
    settings["updated_at"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    settings["updated_by"] = user
    db["sppd_settings"] = settings
    db["sppd_seq"] = int(settings.get("last_sequence", 0))
    save_payments_db(db)
    append_audit_log(user, "payments_sppd_settings_save", "sppd_settings", {
        "last_sequence": settings.get("last_sequence"),
        "fixed_jaminan_date": settings.get("fixed_jaminan_date"),
        "maturity_months": settings.get("maturity_months"),
    })
    next_seq = int(settings.get("last_sequence", 0)) + 1
    preview_dt = pd.Timestamp.today()
    return JSONResponse({
        "ok": True,
        "settings": settings,
        "next_sequence": next_seq,
        "preview_number": format_sppd_number_with_template(next_seq, preview_dt, s(settings.get("number_template", ""))),
    })


# ---------------------------
# Bank Data / Rekening Principle Endpoints
# ---------------------------

@app.get("/api/bank-data")
def get_bank_data(request: Request):
    """Get all bank/rekening data from Excel master."""
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    bank_map = load_bank_map()
    items = []
    for key, info in bank_map.items():
        items.append({
            "principle": info["principle"],
            "bank": info["bank"],
            "rekening": info["rekening"],
            "penerima": info["penerima"],
            "has_rekening": bool(info["rekening"].strip()),
        })
    return {"ok": True, "items": items, "total": len(items), "source": BANK_DATA_PATH}


@app.get("/api/bank-data/match-report")
def get_bank_data_match_report(request: Request):
    """
    Generate report matching antara nama principle di payments.json vs data rekening Excel.
    Berguna untuk verifikasi sebelum submit SPPD.
    """
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    bank_map = load_bank_map()
    # Kumpulkan semua principle names dari payments DB
    db = load_payments_db()
    web_principles = set()
    for rec_key, r in db.get("lpb", {}).items():
        p = s(r.get("principle", ""))
        if p:
            web_principles.add(p)
    report = generate_import_report(bank_map, list(web_principles))
    return {
        "ok": True,
        "report": report,
        "total_excel_principles": len(bank_map),
        "total_web_principles": len(web_principles),
    }


@app.post("/api/bank-data/upload")
async def upload_bank_data(request: Request, file: UploadFile = File(None)):
    """
    Upload file Excel rekening principle baru. Menggantikan file existing.
    Validasi: harus punya kolom PRINCIPLE, NAMA BANK, NOMOR REKENING, NAMA PENERIMA.
    """
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden: butuh permission payments.view"})
    if file is None:
        return JSONResponse(status_code=400, content={"ok": False, "error": "File Excel belum diupload."})
    try:
        content = await read_upload_file_limited(
            file,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="File Rekening",
        )
        # Validasi kolom
        df = pd.read_excel(io.BytesIO(content), dtype=str)
        cols = {c.strip().upper(): c for c in df.columns}
        required = ["PRINCIPLE", "NAMA BANK", "NOMOR REKENING", "NAMA PENERIMA"]
        missing = [c for c in required if c not in cols]
        if missing:
            # Try header at row 2
            df = pd.read_excel(io.BytesIO(content), header=2, dtype=str)
            cols = {c.strip().upper(): c for c in df.columns}
            missing = [c for c in required if c not in cols]
        if missing:
            return JSONResponse(status_code=400, content={
                "ok": False,
                "error": f"Kolom wajib tidak ditemukan: {', '.join(missing)}. Kolom yang ada: {', '.join(df.columns.tolist())}"
            })
        # Hitung jumlah baris valid
        valid_count = 0
        for _, r in df.iterrows():
            p = s(r[cols["PRINCIPLE"]])
            if p:
                valid_count += 1
        if valid_count == 0:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Tidak ada baris dengan data Principle valid."})
        # Simpan file (overwrite)
        with open(BANK_DATA_PATH, "wb") as f:
            f.write(content)
        # Generate report
        new_bank_map = load_bank_map()
        db = load_payments_db()
        web_principles = set()
        for rec_key, r in db.get("lpb", {}).items():
            p = s(r.get("principle", ""))
            if p:
                web_principles.add(p)
        report = generate_import_report(new_bank_map, list(web_principles))
        append_audit_log(user, "bank_data_upload", "bank_data", {
            "filename": file.filename,
            "total_principles": valid_count,
            "matched": len(report["matched"]),
            "unmatched": len(report["unmatched"]),
            "ambiguous": len(report["ambiguous"]),
        })
        return {
            "ok": True,
            "message": f"File rekening berhasil diupload. {valid_count} principle ditemukan.",
            "total_principles": valid_count,
            "report": report,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": f"Gagal memproses file: {str(e)}"})


@app.get("/api/bank-data/lookup")
def lookup_bank_data(request: Request):
    """
    Lookup rekening untuk satu principle name tertentu (fuzzy match).
    Query param: ?principle=NAMA
    """
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    principle_name = s(request.query_params.get("principle", ""))
    if not principle_name:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Parameter 'principle' wajib diisi."})
    bank_map, norm_keys = load_bank_map_with_normalized_keys()
    info, status = find_best_match(principle_name, bank_map, norm_keys)
    if info:
        return {
            "ok": True,
            "status": status,
            "data": {
                "principle": info["principle"],
                "bank": info["bank"],
                "rekening": info["rekening"],
                "penerima": info["penerima"],
                "has_rekening": bool(info["rekening"].strip()),
            },
        }
    return {"ok": True, "status": status, "data": None, "message": f"Tidak ditemukan rekening untuk '{principle_name}'."}


@app.post("/api/bank-data/replace-principle-name")
async def replace_principle_name(request: Request):
    """
    Replace All: Ganti semua occurrence nama principle lama dengan nama baru di payments.json.
    Body JSON: { "old_name": "...", "new_name": "..." }
    Berguna ketika nama principle di web tidak sesuai data rekening.
    """
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden: butuh permission payments.view"})
    try:
        payload = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Body JSON tidak valid."})
    old_name = s(payload.get("old_name", ""))
    new_name = s(payload.get("new_name", ""))
    if not old_name:
        return JSONResponse(status_code=400, content={"ok": False, "error": "old_name wajib diisi."})
    if not new_name:
        return JSONResponse(status_code=400, content={"ok": False, "error": "new_name wajib diisi."})
    if old_name == new_name:
        return JSONResponse(status_code=400, content={"ok": False, "error": "old_name dan new_name tidak boleh sama."})

    db = load_payments_db()
    replaced_count = 0
    replaced_keys: List[str] = []
    for rec_key, rec in db.get("lpb", {}).items():
        current = s(rec.get("principle", ""))
        # Case-insensitive comparison for matching
        if current.upper() == old_name.upper():
            rec["principle"] = new_name
            replaced_count += 1
            replaced_keys.append(rec_key)
    if replaced_count > 0:
        save_payments_db(db)
        append_audit_log(user, "replace_principle_name", "lpb", {
            "old_name": old_name,
            "new_name": new_name,
            "count": replaced_count,
            "samples": replaced_keys[:20],
        })
    return {
        "ok": True,
        "replaced": replaced_count,
        "old_name": old_name,
        "new_name": new_name,
        "message": f"Berhasil mengganti {replaced_count} record dari '{old_name}' menjadi '{new_name}'." if replaced_count > 0 else f"Tidak ada record dengan principle '{old_name}'.",
    }


@app.post("/api/bank-data/auto-fix-names")
async def auto_fix_principle_names(request: Request):
    """
    Auto-fix: Untuk semua principle di payments.json yang bisa di-match (non-ambiguous)
    ke data rekening Excel, ganti namanya agar konsisten dengan format Excel.
    Dry-run by default (preview only). Kirim { "confirm": true } untuk eksekusi.
    """
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "payments", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden: butuh permission payments.view"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    confirm = bool(payload.get("confirm", False))

    bank_map, norm_keys = load_bank_map_with_normalized_keys()
    db = load_payments_db()

    # Collect unique principle names from payments
    name_counts: Dict[str, int] = {}
    for rec_key, rec in db.get("lpb", {}).items():
        p = s(rec.get("principle", ""))
        if p:
            name_counts[p] = name_counts.get(p, 0) + 1

    changes: List[Dict[str, Any]] = []
    skipped: List[Dict[str, str]] = []
    already_correct: List[str] = []

    for web_name, count in name_counts.items():
        info, status = find_best_match(web_name, bank_map, norm_keys)
        if status == "matched" and info:
            excel_name = info["principle"]
            if web_name != excel_name:
                changes.append({"old": web_name, "new": excel_name, "count": count})
            else:
                already_correct.append(web_name)
        elif status == "ambiguous":
            skipped.append({"name": web_name, "reason": "ambiguous", "count": str(count)})
        else:
            skipped.append({"name": web_name, "reason": "unmatched", "count": str(count)})

    total_records_affected = sum(c["count"] for c in changes)

    if confirm and changes:
        # Execute the renames
        for change in changes:
            old = change["old"]
            new = change["new"]
            for rec_key, rec in db.get("lpb", {}).items():
                if s(rec.get("principle", "")) == old:
                    rec["principle"] = new
        save_payments_db(db)
        append_audit_log(user, "auto_fix_principle_names", "lpb", {
            "changes": changes,
            "total_records": total_records_affected,
        })

    return {
        "ok": True,
        "executed": confirm and bool(changes),
        "changes": changes,
        "skipped": skipped,
        "already_correct": already_correct,
        "total_records_affected": total_records_affected,
        "message": (
            f"Berhasil mengupdate {total_records_affected} record ({len(changes)} nama principle)."
            if confirm and changes
            else f"Preview: {len(changes)} nama akan diubah ({total_records_affected} record). Kirim confirm=true untuk eksekusi."
        ),
    }


@app.get("/payments/finance/data")
def payments_finance_data(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "finance", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    date_filter = s(request.query_params.get("date", ""))
    if not date_filter:
        date_filter = pd.Timestamp.today().strftime("%Y-%m-%d")
    db = load_payments_db()
    groups: Dict[str, Dict[str, Any]] = {}
    total_all = 0.0
    for rec_key, r in db.get("lpb", {}).items():
        submitted_at = s(r.get("submitted_at", ""))
        if not submitted_at:
            continue
        target_payment_date = _normalize_yyyy_mm_dd(s(r.get("target_payment_date", "")))
        if not target_payment_date:
            target_payment_date = _normalize_yyyy_mm_dd(submitted_at.split(" ")[0])
        if date_filter and target_payment_date != date_filter:
            continue
        principle = s(r.get("principle", "")) or "-"
        tipe_pengajuan = normalize_pengajuan_type(r.get("tipe_pengajuan", "LPB"))
        submission_id = s(r.get("submission_id", ""))
        draft_id = s(r.get("draft_id", ""))
        group_ref = draft_id or submission_id or submitted_at or "-"
        group_key = f"{principle}||{tipe_pengajuan}||{group_ref}"
        nilai_invoice = parse_number_id(r.get("nilai_invoice", r.get("nilai_principle", 0)))
        if "nilai_pembayaran" in r:
            amount = parse_number_id(r.get("nilai_pembayaran", 0))
        else:
            amount = nilai_invoice
        if amount < 0:
            amount = 0.0
        if "potongan" in r:
            potongan = parse_number_id(r.get("potongan", 0))
        else:
            potongan = max(nilai_invoice - amount, 0.0)
        if potongan < 0:
            potongan = 0.0
        total_all += amount
        g = groups.setdefault(group_key, {
            "principle": principle,
            "tipe_pengajuan": tipe_pengajuan,
            "submission_id": submission_id,
            "draft_id": draft_id,
            "total_nilai": 0.0,
            "total_invoice": 0.0,
            "total_potongan": 0.0,
            "methods": set(),
            "statuses": [],
            "invoice_list": [],
            "detail_invoices": [],
            "jenis_pembayaran": "",
            "keterangan": "",
            "sppd_no": "",
            "transfer_date": "",
            "transfer_proof": {},
            "accurate_post_status": "",
            "accurate_post_error": "",
            "accurate_purchase_payment_number": "",
            "accurate_purchase_payment_id": "",
            "submitted_date": target_payment_date,
        })
        g["total_nilai"] += amount
        g["total_invoice"] += nilai_invoice
        g["total_potongan"] += potongan
        pm = s(r.get("payment_method", ""))
        if pm:
            g["methods"].add(pm)
        st = s(r.get("status_pembayaran", ""))
        if st:
            g["statuses"].append(st)
        inv_no = s(r.get("invoice_no", ""))
        if not inv_no:
            inv_no = s(r.get("nomor_dokumen", ""))
        if not inv_no:
            inv_no = s(r.get("no_lpb", ""))
        if inv_no:
            g["invoice_list"].append(inv_no)
        g["detail_invoices"].append({
            "record_id": s(rec_key),
            "invoiceNo": inv_no,
            "paymentAmount": amount,
            "paymentAmountDisplay": format_idr(amount),
        })
        if not g["jenis_pembayaran"]:
            g["jenis_pembayaran"] = s(r.get("jenis_pembayaran", ""))
        if not g["keterangan"]:
            g["keterangan"] = s(r.get("keterangan", ""))
        if not g["sppd_no"]:
            g["sppd_no"] = s(r.get("sppd_no", ""))
        if not g["transfer_date"]:
            g["transfer_date"] = s(r.get("transfer_date", ""))
        if not g["transfer_proof"] and isinstance(r.get("transfer_proof"), dict):
            g["transfer_proof"] = r.get("transfer_proof", {})
        if not g["accurate_post_status"]:
            g["accurate_post_status"] = s(r.get("accurate_post_status", ""))
        if not g["accurate_post_error"]:
            g["accurate_post_error"] = s(r.get("accurate_post_error", ""))
        if not g["accurate_purchase_payment_number"]:
            g["accurate_purchase_payment_number"] = s(r.get("accurate_purchase_payment_number", ""))
        if not g["accurate_purchase_payment_id"]:
            g["accurate_purchase_payment_id"] = s(r.get("accurate_purchase_payment_id", ""))

    def pick_status(statuses: List[str]) -> str:
        lower = [s(x).lower() for x in statuses if s(x)]
        if any("ajukan ulang" in x for x in lower):
            return "Ajukan Ulang"
        if any("belum" in x for x in lower):
            return "Belum Transfer"
        if any("sudah" in x for x in lower):
            return "Sudah Transfer"
        return "Belum Transfer"

    rows = []
    ordered_groups = sorted(
        groups.values(),
        key=lambda x: (
            s(x.get("submitted_date", "")),
            s(x.get("draft_id", "")),
            s(x.get("submission_id", "")),
            s(x.get("tipe_pengajuan", "")),
            s(x.get("principle", "")),
        ),
    )
    for g in ordered_groups:
        pr = s(g.get("principle", "")) or "-"
        tipe_pengajuan = normalize_pengajuan_type(g.get("tipe_pengajuan", "LPB"))
        mapping = get_finance_mapping(db, pr)
        methods = list(g["methods"])
        method_val = methods[0] if len(methods) == 1 else ("Mixed" if methods else "")
        status_val = pick_status(g["statuses"])
        draft_label = s(g.get("draft_id", "")) or (f"SUB-{s(g.get('submission_id', ''))}" if s(g.get("submission_id", "")) else "-")
        rows.append({
            "principle": pr,
            "tipe_pengajuan": tipe_pengajuan,
            "submission_id": s(g.get("submission_id", "")),
            "draft_id": s(g.get("draft_id", "")),
            "draft_label": draft_label,
            "total_nilai": g["total_nilai"],
            "total_nilai_display": format_idr(g["total_nilai"]),
            "total_invoice": g["total_invoice"],
            "total_invoice_display": format_idr(g["total_invoice"]),
            "total_potongan": g["total_potongan"],
            "total_potongan_display": format_idr(g["total_potongan"]),
            "invoice_concat": ", ".join(g.get("invoice_list", [])),
            "detail_invoices": g.get("detail_invoices", []),
            "jenis_pembayaran": g.get("jenis_pembayaran", ""),
            "keterangan": g.get("keterangan", ""),
            "sppd_no": g.get("sppd_no", ""),
            "transfer_date": g.get("transfer_date", ""),
            "transfer_proof": g.get("transfer_proof", {}),
            "accurate_post_status": g.get("accurate_post_status", ""),
            "accurate_post_error": g.get("accurate_post_error", ""),
            "accurate_purchase_payment_number": g.get("accurate_purchase_payment_number", ""),
            "accurate_purchase_payment_id": g.get("accurate_purchase_payment_id", ""),
            "mapping": mapping,
            "payment_method": method_val,
            "status_pembayaran": status_val,
            "submitted_date": g["submitted_date"],
        })

    return JSONResponse({"ok": True, "data": rows, "total_all": total_all, "total_all_display": format_idr(total_all), "date": date_filter})

@app.get("/payments/finance/export")
def payments_finance_export(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "finance", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})

    from_q = s(request.query_params.get("from", ""))
    to_q = s(request.query_params.get("to", ""))
    from_date = _normalize_yyyy_mm_dd(from_q) if from_q else ""
    to_date = _normalize_yyyy_mm_dd(to_q) if to_q else ""
    if from_q and not from_date:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Format tanggal 'from' tidak valid (YYYY-MM-DD)."})
    if to_q and not to_date:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Format tanggal 'to' tidak valid (YYYY-MM-DD)."})
    if from_date and to_date and from_date > to_date:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Tanggal 'from' tidak boleh lebih besar dari 'to'."})

    db = load_payments_db()
    rows: List[Dict[str, Any]] = []
    for key in sorted(db.get("lpb", {}).keys()):
        r = db["lpb"][key]
        submitted_at = s(r.get("submitted_at", ""))
        if not submitted_at:
            continue
        target_payment_date = _normalize_yyyy_mm_dd(s(r.get("target_payment_date", "")))
        if not target_payment_date:
            target_payment_date = _normalize_yyyy_mm_dd(submitted_at.split(" ")[0])
        if from_date and (not target_payment_date or target_payment_date < from_date):
            continue
        if to_date and (not target_payment_date or target_payment_date > to_date):
            continue

        nilai_invoice = parse_number_id(r.get("nilai_invoice", r.get("nilai_principle", 0)))
        if "nilai_pembayaran" in r:
            nilai_pembayaran = parse_number_id(r.get("nilai_pembayaran", 0))
        else:
            nilai_pembayaran = nilai_invoice
        if nilai_pembayaran < 0:
            nilai_pembayaran = 0.0
        if "potongan" in r:
            potongan = parse_number_id(r.get("potongan", 0))
        else:
            potongan = max(nilai_invoice - nilai_pembayaran, 0.0)
        if potongan < 0:
            potongan = 0.0

        rows.append(
            {
                "Tanggal Pengajuan Pembayaran": target_payment_date,
                "Waktu Submit Sistem": submitted_at,
                "Draft ID": s(r.get("draft_id", "")),
                "Submission ID": s(r.get("submission_id", "")),
                "Tipe Pengajuan": normalize_pengajuan_type(r.get("tipe_pengajuan", "LPB")),
                "No LPB": s(r.get("no_lpb", key)),
                "Principle": s(r.get("principle", "")),
                "No Invoice": s(r.get("invoice_no", "")) or s(r.get("nomor_dokumen", "")) or s(r.get("no_lpb", "")),
                "Jenis Dokumen": s(r.get("jenis_dokumen", "")),
                "Nomor Dokumen": s(r.get("nomor_dokumen", "")),
                "Nilai Invoice": nilai_invoice,
                "Potongan": potongan,
                "Nilai Pembayaran": nilai_pembayaran,
                "Status Pembayaran": s(r.get("status_pembayaran", "")),
                "Metode Pembayaran": s(r.get("payment_method", "")),
                "Jenis Pembayaran": s(r.get("jenis_pembayaran", "")),
                "Keterangan": s(r.get("keterangan", "")),
                "SPPD No": s(r.get("sppd_no", "")),
                "Tgl Setor": s(r.get("tgl_setor", "")),
                "Tgl Win": s(r.get("tgl_win", "")),
                "Tgl J.Tempo Win": s(r.get("tgl_jtempo_win", "")),
                "Tgl Terima Barang": s(r.get("tgl_terima_barang", "")),
                "Tgl Invoice": s(r.get("tgl_invoice", "")) or s(r.get("tgl_inv", "")),
                "J.T Invoice": s(r.get("jt_invoice", "")) or s(r.get("tgl_jtempo_pcp", "")),
                "Actual Date": s(r.get("actual_date", "")),
                "Tgl Pembayaran": s(r.get("tgl_pembayaran", "")) or s(r.get("tgl_jtempo_pembayaran", "")),
                "Submitted By": s(r.get("submitted_by", "")),
            }
        )

    rows.sort(
        key=lambda x: (
            s(x.get("Tanggal Pengajuan Pembayaran", "")),
            s(x.get("Draft ID", "")),
            s(x.get("Submission ID", "")),
            s(x.get("Tipe Pengajuan", "")),
            s(x.get("Principle", "")),
            s(x.get("No LPB", "")),
        )
    )
    ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    suffix_from = from_date or "ALL"
    suffix_to = to_date or "ALL"
    filename = f"backup_finance_{suffix_from}_to_{suffix_to}_{ts}.xlsx"
    return _excel_download_response(rows, filename, "FINANCE")

@app.get("/payments/proofs/{file_name}")
def payments_proof_file(request: Request, file_name: str):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not (user_has_permission(user, "finance", "view") or user_has_permission(user, "payments", "view")):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    safe_name = os.path.basename(file_name)
    path = os.path.join(PAYMENTS_PROOFS_DIR, safe_name)
    if not os.path.exists(path):
        return JSONResponse(status_code=404, content={"ok": False, "error": "Bukti transfer tidak ditemukan."})
    return FileResponse(path, filename=safe_name)

@app.get("/payments/finance/mappings")
def payments_finance_mappings(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "finance", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    db = load_payments_db()
    mappings = db.get("finance_mappings", {})
    return JSONResponse({"ok": True, "data": list(mappings.values()) if isinstance(mappings, dict) else []})

@app.post("/payments/finance/mapping")
async def payments_finance_mapping_save(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "finance", "update"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    principle = s(payload.get("principle", ""))
    vendor_no = s(payload.get("vendorNo", ""))
    bank_no = s(payload.get("bankNo", ""))
    if not principle:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Principle wajib diisi."})
    if not vendor_no or not bank_no:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Vendor No dan Bank No Accurate wajib diisi."})
    db = load_payments_db()
    mapping = {
        "principle": principle,
        "vendorNo": vendor_no,
        "vendorName": s(payload.get("vendorName", "")),
        "bankNo": bank_no,
        "bankName": s(payload.get("bankName", "")),
        "updated_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_by": user,
    }
    db.setdefault("finance_mappings", {})[finance_mapping_key(principle)] = mapping
    save_payments_db(db)
    append_audit_log(user, "payments_finance_mapping_save", "finance_mapping", {"principle": principle, "vendorNo": vendor_no, "bankNo": bank_no})
    return JSONResponse({"ok": True, "mapping": mapping})

@app.post("/payments/finance/proof")
async def payments_finance_proof_upload(request: Request, file: UploadFile = File(None)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "finance", "update"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    if file is None:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Bukti transfer wajib diupload."})
    try:
        content = await read_upload_file_limited(
            file,
            max_bytes=MAX_PROOF_UPLOAD_BYTES,
            allowed_exts=(".pdf", ".jpg", ".jpeg", ".png"),
            label="Bukti Transfer",
        )
        original_name = safe_upload_filename(file.filename or "bukti-transfer")
        ext = os.path.splitext(original_name)[1].lower()
        proof_id = str(uuid.uuid4())[:12]
        stored_name = f"proof_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}_{proof_id}{ext}"
        os.makedirs(PAYMENTS_PROOFS_DIR, exist_ok=True)
        out_path = os.path.join(PAYMENTS_PROOFS_DIR, stored_name)
        with open(out_path, "wb") as f:
            f.write(content)
        meta = build_proof_metadata(proof_id, stored_name, original_name, content, user)
        db = load_payments_db()
        db.setdefault("proofs", {})[proof_id] = meta
        save_payments_db(db)
        append_audit_log(user, "payments_finance_proof_upload", "proof", {"proof_id": proof_id, "stored_filename": stored_name, "size": len(content)})
        return JSONResponse({"ok": True, "proof": meta})
    except ValueError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        append_error_log("payments_finance_proof_upload", e, {"user": user})
        return JSONResponse(status_code=500, content={"ok": False, "error": "Gagal menyimpan bukti transfer."})

@app.post("/payments/finance/update")
async def payments_finance_update(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "finance", "update"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    items = payload.get("items", [])
    if not isinstance(items, list):
        return JSONResponse(status_code=400, content={"ok": False, "error": "Format data tidak valid."})
    if not items:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Data finance yang akan diupdate tidak boleh kosong."})
    db = load_payments_db()
    updated_count = 0
    for item in items:
        no = normalize_lpb_no(s(item.get("no_lpb", "")))
        status = s(item.get("status_pembayaran", ""))
        principle = s(item.get("principle", "")).upper()
        tipe_pengajuan = normalize_pengajuan_type(item.get("tipe_pengajuan", ""))
        date_filter = s(item.get("date", ""))
        submission_id = s(item.get("submission_id", ""))
        draft_id = s(item.get("draft_id", ""))
        if status not in ["Belum Transfer", "Sudah Transfer", "Ajukan Ulang"]:
            continue
        transfer_date = _normalize_yyyy_mm_dd(s(item.get("transfer_date", "")))
        proof_id = s(item.get("proof_id", ""))
        proof_meta = db.get("proofs", {}).get(proof_id, {}) if proof_id else {}
        accurate_post_status = s(item.get("accurate_post_status", ""))
        if accurate_post_status not in ["", "posted", "failed", "skipped"]:
            accurate_post_status = "failed"
        if status == "Sudah Transfer":
            if not transfer_date:
                return JSONResponse(status_code=400, content={"ok": False, "error": "Tanggal transfer wajib diisi untuk status Sudah Transfer."})
            if not proof_id or not isinstance(proof_meta, dict) or not proof_meta:
                return JSONResponse(status_code=400, content={"ok": False, "error": "Bukti transfer wajib diupload sebelum status Sudah Transfer."})

        def apply_finance_update(rec: Dict[str, Any]) -> None:
            nonlocal updated_count
            rec["status_pembayaran"] = status
            if status == "Sudah Transfer":
                rec["transfer_date"] = transfer_date
                rec["proof_id"] = proof_id
                rec["transfer_proof"] = proof_meta
                rec["accurate_post_status"] = accurate_post_status or "skipped"
                rec["accurate_post_error"] = s(item.get("accurate_post_error", ""))
                rec["accurate_purchase_payment_number"] = s(item.get("accurate_purchase_payment_number", ""))
                rec["accurate_purchase_payment_id"] = s(item.get("accurate_purchase_payment_id", ""))
                rec["accurate_post_response"] = item.get("accurate_post_response", {})
                rec["accurate_payload_digest"] = s(item.get("accurate_payload_digest", ""))
                rec["accurate_posted_at"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S") if rec["accurate_post_status"] == "posted" else s(rec.get("accurate_posted_at", ""))
                rec["accurate_posted_by"] = user if rec["accurate_post_status"] == "posted" else s(rec.get("accurate_posted_by", ""))
            updated_count += 1

        if no and no in db.get("lpb", {}):
            apply_finance_update(db["lpb"][no])
            continue
        if principle:
            for k, r in db.get("lpb", {}).items():
                submitted_at = s(r.get("submitted_at", ""))
                submitted_date = _normalize_yyyy_mm_dd(submitted_at.split(" ")[0]) if submitted_at else ""
                target_date = _normalize_yyyy_mm_dd(s(r.get("target_payment_date", ""))) or submitted_date
                if date_filter and target_date != date_filter:
                    continue
                if submission_id and s(r.get("submission_id", "")) != submission_id:
                    continue
                if draft_id and s(r.get("draft_id", "")) != draft_id:
                    continue
                if tipe_pengajuan and normalize_pengajuan_type(r.get("tipe_pengajuan", "LPB")) != tipe_pengajuan:
                    continue
                if s(r.get("principle", "")).upper() == principle:
                    apply_finance_update(db["lpb"][k])
    if updated_count <= 0:
        return JSONResponse(status_code=404, content={"ok": False, "error": "Tidak ada data finance yang cocok untuk diupdate."})
    save_payments_db(db)
    append_audit_log(user, "payments_finance_update", "lpb", {"count": updated_count, "items": len(items)})
    return JSONResponse({"ok": True, "updated": updated_count})

@app.get("/health")
def health():
    return {"status": "ok", "patch": PATCH_VERSION}

# ---------------------------
# Finalize
# ---------------------------
def finalize(sales_df: pd.DataFrame, sales_cols_fixed: List[str]) -> pd.DataFrame:
    # Ensure output columns exist before we reference them
    sales_df = sales_df.copy()
    for c, dv in [
        ("Expected_Pabrik", 0.0),
        ("Expected_Internal", 0.0),
        ("Expected_Total", 0.0),
        ("Actual_BonusQty", 0.0),
        ("Expected_Bonus_Pabrik", 0.0),
        ("Expected_Bonus_Internal", 0.0),
        ("Expected_Bonus_Total", 0.0),
        ("FixClaimBonus_Pabrik", 0.0),
        ("FixClaimBonus_Internal", 0.0),
        ("FixClaimBonus_TanpaTuan", 0.0),
        ("Bonus_Selisih", 0.0),
        ("FixClaim_Pabrik", 0.0),
        ("FixClaim_Internal", 0.0),
        ("FixClaim_TanpaTuan", 0.0),
        ("Fix Claim", 0.0),
        ("Selisih", 0.0),
        ("StatusValidasi", ""),
        ("PesanValidasi", ""),
        ("HAS_ANY_PROGRAM", False),
        ("HAS_ELIGIBLE_PROGRAM", False),
        ("Program_Pabrik_Eligible", ""),
        ("Program_Pabrik_Applied", ""),
        ("Program_Bonus_Applied", ""),
        ("Debug_TriggerQty", 0.0),
        ("Bonus_TriggerUnit", ""),
        ("Bonus_TriggerQty", 0.0),
        ("Bonus_TriggerMetric", 0.0),
        ("Bonus_PickedTier", 0),
        ("Penjelasan_Claim", ""),
        ("Penjelasan_Bonus", ""),
    ]:
        if c not in sales_df.columns:
            sales_df.loc[:, c] = dv

    actual = sales_df["LDISCAMOUNT"].astype(float).fillna(0.0)
    exp_p = sales_df["Expected_Pabrik"].astype(float).fillna(0.0)
    exp_i = sales_df["Expected_Internal"].astype(float).fillna(0.0)

    sales_df.loc[:, "Expected_Total"] = exp_p + exp_i

    use_mdstring_claim = all(
        c in sales_df.columns
        for c in ["_PCT_PARTS_PABRIK", "_PCT_PARTS_INTERNAL", "_RP_PABRIK"]
    )

    if use_mdstring_claim:
        def alloc_row(r):
            md_parts = normalize_pct_list(parse_mdstring_pct(r.get("MDSTRING", "")))
            if not md_parts:
                act = float(r.get("LDISCAMOUNT", 0.0) or 0.0)
                exp_p_row = float(r.get("Expected_Pabrik", 0.0) or 0.0)
                exp_i_row = float(r.get("Expected_Internal", 0.0) or 0.0)
                fix_p_row = min(act, exp_p_row)
                rem_row = max(act - fix_p_row, 0.0)
                fix_i_row = min(rem_row, exp_i_row)
                fix_t_row = max(rem_row - fix_i_row, 0.0)
                return pd.Series([fix_p_row, fix_i_row, fix_t_row])

            res = allocate_claim_by_mdstring(
                gross_amount=r.get("GROSSAMOUNT", 0.0),
                actual_discount=r.get("LDISCAMOUNT", 0.0),
                mdstring=r.get("MDSTRING", ""),
                rp_pabrik=r.get("_RP_PABRIK", 0.0),
                pct_pabrik=r.get("_PCT_PARTS_PABRIK", []),
                pct_internal=r.get("_PCT_PARTS_INTERNAL", []),
            )
            if res is None:
                act = float(r.get("LDISCAMOUNT", 0.0) or 0.0)
                exp_p_row = float(r.get("Expected_Pabrik", 0.0) or 0.0)
                exp_i_row = float(r.get("Expected_Internal", 0.0) or 0.0)
                fix_p_row = min(act, exp_p_row)
                rem_row = max(act - fix_p_row, 0.0)
                fix_i_row = min(rem_row, exp_i_row)
                fix_t_row = max(rem_row - fix_i_row, 0.0)
                return pd.Series([fix_p_row, fix_i_row, fix_t_row])

            fix_p_row, fix_i_row, fix_t_row = res
            return pd.Series([fix_p_row, fix_i_row, fix_t_row])

        alloc = sales_df.apply(alloc_row, axis=1)
        sales_df.loc[:, "FixClaim_Pabrik"] = alloc[0].values
        sales_df.loc[:, "FixClaim_Internal"] = alloc[1].values
        sales_df.loc[:, "FixClaim_TanpaTuan"] = alloc[2].values
        sales_df.loc[:, "Fix Claim"] = sales_df["FixClaim_Pabrik"] + sales_df["FixClaim_Internal"]
    else:
        # Allocate Actual -> Pabrik -> Internal -> TanpaTuan (legacy)
        fix_p = pd.concat([actual, exp_p], axis=1).min(axis=1)
        rem = (actual - fix_p).clip(lower=0.0)
        fix_i = pd.concat([rem, exp_i], axis=1).min(axis=1)
        rem2 = (rem - fix_i).clip(lower=0.0)

        sales_df.loc[:, "FixClaim_Pabrik"] = fix_p
        sales_df.loc[:, "FixClaim_Internal"] = fix_i
        sales_df.loc[:, "FixClaim_TanpaTuan"] = rem2
        sales_df.loc[:, "Fix Claim"] = fix_p + fix_i

    # ---------------------------
    # Bonus allocation (PCS)
    # Allocate Actual_BonusQty -> Expected_Bonus_Pabrik -> Expected_Bonus_Internal -> TanpaTuan
    # ---------------------------
    b_actual = sales_df["Actual_BonusQty"].astype(float).fillna(0.0)
    b_exp_p = sales_df["Expected_Bonus_Pabrik"].astype(float).fillna(0.0)
    b_exp_i = sales_df["Expected_Bonus_Internal"].astype(float).fillna(0.0)
    sales_df.loc[:, "Expected_Bonus_Total"] = b_exp_p + b_exp_i

    b_fix_p = pd.concat([b_actual, b_exp_p], axis=1).min(axis=1)
    b_rem = (b_actual - b_fix_p).clip(lower=0.0)
    b_fix_i = pd.concat([b_rem, b_exp_i], axis=1).min(axis=1)
    b_rem2 = (b_rem - b_fix_i).clip(lower=0.0)

    sales_df.loc[:, "FixClaimBonus_Pabrik"] = b_fix_p
    sales_df.loc[:, "FixClaimBonus_Internal"] = b_fix_i
    sales_df.loc[:, "FixClaimBonus_TanpaTuan"] = b_rem2
    sales_df.loc[:, "Bonus_Selisih"] = sales_df["Expected_Bonus_Total"].astype(float).fillna(0.0) - b_actual

    sales_df.loc[:, "Selisih"] = actual - (exp_p + exp_i)

    TOL = 1.0

    def status_row(r):
        act = float(r["LDISCAMOUNT"] or 0.0)
        exp = float(r["Expected_Total"] or 0.0)
        anyp = bool(r.get("HAS_ANY_PROGRAM", False))
        elig = bool(r.get("HAS_ELIGIBLE_PROGRAM", False))

        if act <= 0 and exp <= 0:
            return "C"
        if (not anyp) and act > 0:
            return "A"
        if anyp and (not elig) and act > 0 and exp <= 0:
            return "B"
        if abs(act - exp) <= TOL:
            return "C"
        return "B"

    sales_df.loc[:, "StatusValidasi"] = sales_df.apply(status_row, axis=1)

    def pesan_row(r):
        st = r["StatusValidasi"]
        act = float(r["LDISCAMOUNT"] or 0.0)
        exp = float(r["Expected_Total"] or 0.0)
        if st == "A":
            return "Tidak ada program yang cocok untuk item/outlet/channel ini (masuk TanpaTuan)"
        if st == "B":
            if bool(r.get("HAS_ANY_PROGRAM", False)) and (not bool(r.get("HAS_ELIGIBLE_PROGRAM", False))) and exp <= 0 and act > 0:
                return "Ada program di dataset, tapi faktur tidak memenuhi trigger/tier (expected 0). Masuk TanpaTuan."
            return f"Expected total {exp:.2f}, aktual {act:.2f}, selisih {(act-exp):.2f}. Selisih masuk TanpaTuan."
        return "Sesuai (toleransi)"

    sales_df.loc[:, "PesanValidasi"] = sales_df.apply(pesan_row, axis=1)

    def explain_claim_row(r):
        parts = []
        fix_p = float(r.get("FixClaim_Pabrik", 0.0) or 0.0)
        fix_i = float(r.get("FixClaim_Internal", 0.0) or 0.0)
        fix_t = float(r.get("FixClaim_TanpaTuan", 0.0) or 0.0)

        tu = s(r.get("Debug_TriggerUnit", ""))
        tq = float(r.get("Debug_TriggerQty", 0.0) or 0.0)
        metric = float(r.get("Debug_TriggerMetric", 0.0) or 0.0)
        tier = int(r.get("Debug_PickedTier", 0) or 0)

        trigger = ""
        if tu and tq > 0:
            trigger = f"trigger min {fmt_metric(tq, tu)} {tu}, tercapai {fmt_metric(metric, tu)} {tu}"
            if tier > 0:
                trigger += f", tier {tier}"

        md_parts = normalize_pct_list(parse_mdstring_pct(r.get("MDSTRING", "")))
        elig_parts = []
        elig_parts.extend(normalize_pct_list(r.get("_PCT_PARTS_PABRIK", [])))
        elig_parts.extend(normalize_pct_list(r.get("_PCT_PARTS_INTERNAL", [])))
        md_sum = sum(md_parts) if md_parts else 0.0
        elig_sum = sum(elig_parts) if elig_parts else 0.0
        leftover = max(md_sum - elig_sum, 0.0) if md_parts else 0.0

        if fix_p > 0:
            msg = "Pabrik: memenuhi trigger"
            if trigger:
                msg = f"Pabrik: hanya mencukupi {trigger}"
            if elig_sum > 0:
                msg += f" (disc {fmt_pct(elig_sum)}%)"
            parts.append(msg)

        if fix_i > 0:
            parts.append("Internal: rule match")

        if fix_t > 0:
            msg = "TanpaTuan: selisih diskon masuk TanpaTuan"
            if md_parts:
                md_s = pct_list_str(md_parts)
                if elig_sum > 0 and leftover > 0:
                    msg = f"TanpaTuan: sisa {fmt_pct(leftover)}% dari MDSTRING {fmt_pct(md_sum)}% masuk TanpaTuan"
                elif elig_sum <= 0:
                    msg = f"TanpaTuan: MDSTRING {md_s}% tidak ada program eligible; semua masuk TanpaTuan"
            parts.append(msg)

        return " | ".join([p for p in parts if p])

    def explain_bonus_row(r):
        actual = float(r.get("Actual_BonusQty", 0.0) or 0.0)
        expected = float(r.get("Expected_Bonus_Total", 0.0) or 0.0)
        if actual <= 0 and expected <= 0:
            return ""
        tanp = float(r.get("FixClaimBonus_TanpaTuan", 0.0) or 0.0)
        prog = s(r.get("Program_Bonus_Applied", "")) or s(r.get("Program_Pabrik_Applied", ""))
        tu = s(r.get("Bonus_TriggerUnit", ""))
        tq = float(r.get("Bonus_TriggerQty", 0.0) or 0.0)
        metric = float(r.get("Bonus_TriggerMetric", 0.0) or 0.0)
        tier = int(r.get("Bonus_PickedTier", 0) or 0)

        trigger = ""
        if tu and tq > 0:
            trigger = f"trigger min {fmt_metric(tq, tu)} {tu}, tercapai {fmt_metric(metric, tu)} {tu}"
            if tier > 0:
                trigger += f", tier {tier}"

        if tanp > 0:
            msg = f"Bonus eligible {fmt_metric(expected, 'PCS')} pcs, actual {fmt_metric(actual, 'PCS')} pcs; {fmt_metric(tanp, 'PCS')} pcs masuk TanpaTuan"
        elif expected > actual + 1e-9:
            diff = expected - actual
            msg = f"Bonus eligible {fmt_metric(expected, 'PCS')} pcs, actual {fmt_metric(actual, 'PCS')} pcs (kurang {fmt_metric(diff, 'PCS')} pcs)"
        else:
            msg = f"Bonus sesuai {fmt_metric(actual, 'PCS')} pcs"

        if trigger:
            msg += f" ({trigger})"
        return msg

    sales_df.loc[:, "Penjelasan_Claim"] = sales_df.apply(explain_claim_row, axis=1)
    sales_df.loc[:, "Penjelasan_Bonus"] = sales_df.apply(explain_bonus_row, axis=1)

    def program_name_row(r):
        name = s(r.get("Program_Pabrik_Applied", "")) or s(r.get("Program_Pabrik_Eligible", ""))
        if not name and bool(r.get("_IS_BONUS_LINE_MAIN", False)):
            name = s(r.get("Program_Bonus_Applied", ""))
        return name

    sales_df.loc[:, "Nama_Program_Pabrik"] = sales_df.apply(program_name_row, axis=1)

    out_cols = sales_cols_fixed + [
        "EffectiveChannel",
        "Nama_Program_Pabrik",
        "FixClaim_Pabrik", "FixClaim_Internal", "FixClaim_TanpaTuan",
        "StatusValidasi",
        "Penjelasan_Claim", "Penjelasan_Bonus",
    ]
    return sales_df[out_cols].copy()



def write_excel(out_df: pd.DataFrame, out_path: str):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="RESULT")

def accel_or_file_response(out_path: str, download_name: str):
    if not os.path.exists(out_path):
        return JSONResponse(status_code=404, content={"detail": "File not found"})

    if USE_X_ACCEL:
        prefix = X_ACCEL_PREFIX.rstrip("/") or "/protected-downloads"
        accel_path = f"{prefix}/{os.path.basename(out_path)}"
        headers = {
            "X-Accel-Redirect": accel_path,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f'attachment; filename="{download_name}"',
        }
        return Response(status_code=200, headers=headers)

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name
    )

# ---------------------------
# Load channel mapping
# ---------------------------
def load_channel_map(channel_bytes: bytes) -> Dict[str, str]:
    df = pd.read_excel(io.BytesIO(channel_bytes))
    # accept either column names:
    # - "KODE OUTLET" + "CHANNEL" (your current template)
    # - "SUB" + "CHANNEL_CODE"
    cols = {c.strip().upper(): c for c in df.columns}
    sub_col = cols.get("KODE OUTLET") or cols.get("SUB")
    ch_col = cols.get("CHANNEL") or cols.get("CHANNEL_CODE") or cols.get("CHANNELNAME") or cols.get("CHANNEL_NAME")
    if not sub_col or not ch_col:
        raise ValueError("Data Channel by SUB harus punya kolom 'KODE OUTLET'/'SUB' dan 'CHANNEL' (contoh: KODE OUTLET, CHANNEL).")
    df[sub_col] = df[sub_col].apply(s)
    df[ch_col] = df[ch_col].apply(s)
    mp = {}
    for _, r in df.iterrows():
        if s(r[sub_col]) and s(r[ch_col]):
            mp[s(r[sub_col])] = s(r[ch_col])
    return mp

# ---------------------------
# Engine
# ---------------------------
def run_engine(sales_bytes: bytes, promo_bytes: bytes, channel_bytes: bytes, internal_bytes: Optional[bytes]) -> pd.DataFrame:
    sales_df = pd.read_excel(io.BytesIO(sales_bytes))
    promo_df = pd.read_excel(io.BytesIO(promo_bytes))

    # Load channel map (required)
    channel_map = load_channel_map(channel_bytes)

    # Load internal (optional; used in v15)
    internal_rules = load_internal_rules(internal_bytes)
    if internal_bytes:
        try:
            internal_df = pd.read_excel(io.BytesIO(internal_bytes))
        except:
            internal_df = None

    sales_cols_fixed = [
        "SUBNAME", "SUB", "SUBADDRESS", "SUBKOTA", "JPSNAMA_CHANNEL",
        "BRG", "BRGNAME", "UNIT", "PACKAGING", "INVOICEDATE", "INVOICENO",
        "SALESORDER", "SALESMAN", "QTYSTR", "QTYPCS", "QTY", "QTYBONUSSTR",
        "PRICE", "GROSSAMOUNT", "MDSTRING", "LDISCAMOUNT", "DPP", "TAXVAL", "NETAMOUNT"
    ]
    for c in sales_cols_fixed:
        if c in ["QTYPCS", "QTY", "QTYBONUSSTR", "PRICE", "GROSSAMOUNT", "LDISCAMOUNT", "DPP", "TAXVAL", "NETAMOUNT"]:
            ensure_col(sales_df, c, 0.0)
        else:
            ensure_col(sales_df, c, "")

    for c in ["SUBNAME", "SUB", "SUBADDRESS", "SUBKOTA", "JPSNAMA_CHANNEL", "BRG", "BRGNAME",
              "UNIT", "PACKAGING", "INVOICENO", "SALESORDER", "SALESMAN", "QTYSTR", "MDSTRING"]:
        sales_df[c] = sales_df[c].apply(s)

    sales_df = to_float_series(
        sales_df,
        ["QTYPCS", "QTY", "QTYBONUSSTR", "PRICE", "GROSSAMOUNT", "LDISCAMOUNT", "DPP", "TAXVAL", "NETAMOUNT"]
    )
    sales_df = sales_df.copy()
    sales_df["_ROWID"] = range(len(sales_df))
    sales_df["_IS_BONUS_LINE_MAIN"] = sales_df["MDSTRING"].apply(mdstring_is_bonus)

    # Effective channel (from SUB map)
    def eff_channel(row):
        sub = s(row.get("SUB", ""))
        if sub in channel_map and channel_map[sub]:
            return channel_map[sub]
        return s(row.get("JPSNAMA_CHANNEL", ""))

    sales_df["EffectiveChannel"] = sales_df.apply(eff_channel, axis=1)

    promo_needed = [
        "KODE_BARANG", "NAMA_BARANG", "PROMO_LABEL", "PROMO_GROUP_ID", "PROMO_GROUP", "PROMO_ACTIVE",
        "TIER_NO", "TRIGGER_QTY", "TRIGGER_UNIT", "BENEFIT_TYPE", "BENEFIT_VALUE", "BENEFIT_UNIT"
    ]
    for c in promo_needed:
        ensure_col(promo_df, c, "")

    promo_df["KODE_BARANG"] = promo_df["KODE_BARANG"].apply(s)
    promo_df["NAMA_BARANG"] = promo_df["NAMA_BARANG"].apply(s)
    promo_df["PROMO_LABEL"] = promo_df["PROMO_LABEL"].apply(s)
    promo_df["PROMO_GROUP_ID"] = promo_df["PROMO_GROUP_ID"].apply(s)
    promo_df["PROMO_GROUP"] = promo_df["PROMO_GROUP"].apply(s)
    promo_df["TRIGGER_UNIT"] = promo_df["TRIGGER_UNIT"].apply(s)

    promo_df["TRIGGER_QTY"] = promo_df["TRIGGER_QTY"].apply(parse_number_id)
    promo_df["TIER_NO"] = pd.to_numeric(promo_df["TIER_NO"], errors="coerce").fillna(0).astype(int)

    promo_df["BENEFIT_UNIT"] = promo_df["BENEFIT_UNIT"].apply(s)
    promo_df["BENEFIT_TYPE"] = promo_df["BENEFIT_TYPE"].apply(s).str.upper()
    promo_df["BENEFIT_VALUE"] = promo_df["BENEFIT_VALUE"].apply(s)
    promo_df["_ALL_ITEM_RULE"] = (promo_df["KODE_BARANG"].apply(s) == "") & (promo_df["NAMA_BARANG"].apply(s) == "")

    if "PROMO_ACTIVE" in promo_df.columns:
        def is_active(v):
            if pd.isna(v):
                return True
            if isinstance(v, bool):
                return v
            return str(v).strip().lower() in ["true", "1", "yes", "y"]
        promo_df = promo_df[promo_df["PROMO_ACTIVE"].apply(is_active)].copy()

    # outputs + debug
    sales_df.loc[:, "HAS_ANY_PROGRAM"] = False
    sales_df.loc[:, "HAS_ELIGIBLE_PROGRAM"] = False
    sales_df.loc[:, "Expected_Pabrik"] = 0.0

    sales_df.loc[:, "Debug_ProgramKeysEligible"] = ""
    sales_df.loc[:, "Debug_ProgramKeysApplied"] = ""
    sales_df.loc[:, "Debug_TriggerUnit"] = ""
    sales_df.loc[:, "Debug_TriggerUnitSource"] = ""
    sales_df.loc[:, "Debug_TriggerMetric"] = 0.0
    sales_df.loc[:, "Debug_TriggerQty"] = 0.0
    sales_df.loc[:, "Debug_PickedTier"] = 0
    sales_df.loc[:, "Debug_SkipReason"] = ""
    sales_df.loc[:, "Debug_ProgramLock_ClaimedPct"] = ""
    sales_df.loc[:, "Debug_ProgramLock_SelectedPctPrograms"] = ""
    sales_df.loc[:, "Program_Pabrik_Eligible"] = ""
    sales_df.loc[:, "Program_Pabrik_Applied"] = ""
    sales_df.loc[:, "Program_Bonus_Applied"] = ""
    sales_df.loc[:, "_INTERNAL_BONUS_MATCH"] = False
    sales_df.loc[:, "Bonus_TriggerUnit"] = ""
    sales_df.loc[:, "Bonus_TriggerQty"] = 0.0
    sales_df.loc[:, "Bonus_TriggerMetric"] = 0.0
    sales_df.loc[:, "Bonus_PickedTier"] = 0
    sales_df.loc[:, "_PCT_PARTS_PABRIK"] = [list() for _ in range(len(sales_df))]
    sales_df.loc[:, "_PCT_PARTS_INTERNAL"] = [list() for _ in range(len(sales_df))]
    sales_df.loc[:, "_RP_PABRIK"] = 0.0

    def scope_program_rows(df: pd.DataFrame, pgid: str, plabel: str, brgs: List[str]) -> pd.DataFrame:
        r = df[
            (df["PROMO_GROUP_ID"].apply(s) == pgid) &
            (df["PROMO_LABEL"].apply(s) == plabel)
        ].copy()
        if r.empty:
            return r
        return r[(r["KODE_BARANG"].apply(s).isin(brgs)) | (r["_ALL_ITEM_RULE"] == True)].copy()

    def pick_rows_for_item(rows: pd.DataFrame, brg: str) -> pd.DataFrame:
        exact = rows[rows["KODE_BARANG"].apply(s) == brg].copy()
        if not exact.empty:
            return exact
        return rows[rows["_ALL_ITEM_RULE"] == True].copy()

    # Build candidate rows by channel (use EffectiveChannel!)
    promo_exact = promo_df[promo_df["_ALL_ITEM_RULE"] == False].copy()
    merged_exact = sales_df.merge(promo_exact, left_on="BRG", right_on="KODE_BARANG", how="left", suffixes=("", "_P"))
    merged_all = pd.DataFrame(columns=merged_exact.columns)
    promo_all = promo_df[promo_df["_ALL_ITEM_RULE"] == True].copy()
    if not promo_all.empty:
        merged_all = (
            sales_df.assign(_J=1)
            .merge(promo_all.assign(_J=1), on="_J", how="inner", suffixes=("", "_P"))
            .drop(columns=["_J"])
        )
    merged = pd.concat([merged_exact, merged_all], ignore_index=True, sort=False)
    merged["CHANNEL_OK"] = merged.apply(
        lambda r: channel_ok(r.get("PROMO_GROUP", ""), r.get("EffectiveChannel", ""), r.get("PROMO_GROUP_ID", ""), r.get("SUB", "")),
        axis=1
    )
    cand = merged[(merged["KODE_BARANG"].notna()) & (merged["CHANNEL_OK"] == True)].copy()
    if cand.empty:
        return finalize(sales_df, sales_cols_fixed)

    any_rowids = cand["_ROWID"].dropna().astype(int).unique().tolist()
    sales_df.loc[sales_df["_ROWID"].isin(any_rowids), "HAS_ANY_PROGRAM"] = True

    cand["PGID"] = cand["PROMO_GROUP_ID"].apply(s)
    cand["PLABEL"] = cand["PROMO_LABEL"].apply(s)
    cand["PROGRAM_KEY"] = cand["PGID"] + "||" + cand["PLABEL"]
    cand["IS_NON_GROUP"] = cand["PGID"].str.upper().eq("NON_GROUP")

    # Program scope per invoice+program_key
    program_scopes: Dict[Tuple[str, str], List[int]] = {}
    program_defs: Dict[Tuple[str, str], Dict] = {}
    for (inv, program_key), g in cand.groupby(["INVOICENO", "PROGRAM_KEY"], sort=True):
        inv = s(inv); program_key = s(program_key)
        rowids = g["_ROWID"].dropna().astype(int).unique().tolist()
        idxs = sales_df.index[sales_df["_ROWID"].isin(rowids)].tolist()
        program_scopes[(inv, program_key)] = idxs
        program_defs[(inv, program_key)] = {
            "pgid": s(g["PGID"].iloc[0]),
            "plabel": s(g["PLABEL"].iloc[0]),
            "is_non_group": bool(g["IS_NON_GROUP"].iloc[0])
        }
        for idx in idxs:
            cur = sales_df.at[idx, "Debug_ProgramKeysEligible"]
            sales_df.at[idx, "Debug_ProgramKeysEligible"] = (cur + (";" if cur else "") + program_key)
            sales_df.at[idx, "Program_Pabrik_Eligible"] = add_list_item(
                sales_df.at[idx, "Program_Pabrik_Eligible"], s(g["PLABEL"].iloc[0])
            )

    # PROGRAM LOCK (DISC_PCT grouped only) — uses mdstring per invoice
    inv_claimed_pct: Dict[str, Set[float]] = {}
    for inv, g in sales_df.groupby("INVOICENO"):
        claimed = []
        for md in g["MDSTRING"].tolist():
            claimed.extend(parse_mdstring_pct(md))
        claimed = [round(x, 4) for x in claimed if abs(x) > 1e-9]
        inv_claimed_pct[s(inv)] = set(claimed)

    inv_selected_pct_programs: Dict[str, Set[str]] = {}
    for inv, g in cand.groupby("INVOICENO"):
        inv = s(inv)
        claimed = inv_claimed_pct.get(inv, set())
        if not claimed:
            inv_selected_pct_programs[inv] = set()
            continue

        scores: Dict[str, int] = {}
        for program_key in g["PROGRAM_KEY"].unique().tolist():
            program_key = s(program_key)
            if program_defs.get((inv, program_key), {}).get("is_non_group", False):
                continue
            idxs = program_scopes.get((inv, program_key), [])
            if not idxs:
                continue
            scope = sales_df.loc[idxs]
            sub = s(scope["SUB"].iloc[0])
            eff_channel = s(scope["EffectiveChannel"].iloc[0])
            pgid = program_defs[(inv, program_key)]["pgid"]
            plabel = program_defs[(inv, program_key)]["plabel"]

            brgs = scope["BRG"].apply(s).unique().tolist()
            prog_rows = scope_program_rows(promo_df, pgid, plabel, brgs)
            if prog_rows.empty:
                continue

            prog_rows = prog_rows[prog_rows.apply(lambda r: channel_ok(r.get("PROMO_GROUP", ""), eff_channel, r.get("PROMO_GROUP_ID", ""), sub), axis=1)].copy()
            if prog_rows.empty:
                continue

            pct_rows = prog_rows[prog_rows["BENEFIT_TYPE"].str.upper() == "DISC_PCT"]
            if pct_rows.empty:
                continue

            program_pcts: Set[float] = set()
            for v in pct_rows["BENEFIT_VALUE"].tolist():
                for p in parse_pct_chain(s(v)):
                    if abs(p) > 1e-9:
                        program_pcts.add(round(p, 4))
            if not program_pcts:
                continue

            score = len(program_pcts.intersection(claimed))
            scores[program_key] = score

        if not scores:
            inv_selected_pct_programs[inv] = set()
            continue

        max_score = max(scores.values())
        if max_score <= 0:
            inv_selected_pct_programs[inv] = set()
        else:
            inv_selected_pct_programs[inv] = {k for k, sc in scores.items() if sc == max_score}

    for idx, r in sales_df.iterrows():
        inv = s(r["INVOICENO"])
        claimed = sorted(inv_claimed_pct.get(inv, set()))
        selected = sorted(inv_selected_pct_programs.get(inv, set()))
        sales_df.at[idx, "Debug_ProgramLock_ClaimedPct"] = ",".join([str(x).rstrip("0").rstrip(".") for x in claimed])
        sales_df.at[idx, "Debug_ProgramLock_SelectedPctPrograms"] = ";".join(selected)

    # Compute expected
    rp_line = pd.Series(0.0, index=sales_df.index)
    pct_tasks: List[Dict] = []
    eligible_rowids = set()

    # PASS 1 - RP
    for (inv, program_key) in sorted(program_scopes.keys()):
        idxs = program_scopes[(inv, program_key)]
        if not idxs:
            continue

        scope = sales_df.loc[idxs].copy()
        sub = s(scope["SUB"].iloc[0])
        eff_channel = s(scope["EffectiveChannel"].iloc[0])
        pgid = program_defs[(inv, program_key)]["pgid"]
        plabel = program_defs[(inv, program_key)]["plabel"]
        is_non_group = program_defs[(inv, program_key)]["is_non_group"]

        brgs = scope["BRG"].apply(s).unique().tolist()
        prog_rows = scope_program_rows(promo_df, pgid, plabel, brgs)
        if prog_rows.empty:
            continue

        prog_rows = prog_rows[prog_rows.apply(lambda r: channel_ok(r.get("PROMO_GROUP", ""), eff_channel, r.get("PROMO_GROUP_ID", ""), sub), axis=1)].copy()
        if prog_rows.empty:
            continue

        trigger_unit, tu_source = choose_trigger_unit_satpam(prog_rows)
        if trigger_unit is None:
            for idx in idxs:
                sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (rp)")
                sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
            continue

        if not is_non_group:
            metric = float(scope["GROSSAMOUNT"].sum() or 0.0) if trigger_unit == "GROSSAMOUNT" else float(scope.apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum() or 0.0)
            picked_tier, tier_rows = pick_best_tier(prog_rows, metric)
            if tier_rows.empty:
                continue

            tier_rows = dedupe_grouped_benefits(tier_rows)
            tq_val = float(pd.to_numeric(tier_rows["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)

            for idx in idxs:
                eligible_rowids.add(int(sales_df.at[idx, "_ROWID"]))
                sales_df.at[idx, "Debug_TriggerUnit"] = trigger_unit
                sales_df.at[idx, "Debug_TriggerUnitSource"] = tu_source
                sales_df.at[idx, "Debug_TriggerMetric"] = metric
                sales_df.at[idx, "Debug_TriggerQty"] = tq_val
                sales_df.at[idx, "Debug_PickedTier"] = picked_tier

            disc_rp_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_RP"].copy()
            if disc_rp_rows.empty:
                continue

            line_gross = scope["GROSSAMOUNT"].astype(float); sum_gross = float(line_gross.sum() or 0.0)
            line_pcs = scope.apply(lambda r: qty_in_unit(r, "PCS"), axis=1).astype(float); sum_pcs = float(line_pcs.sum() or 0.0)
            line_ctn = scope.apply(lambda r: qty_in_unit(r, "CTN"), axis=1).astype(float); sum_ctn = float(line_ctn.sum() or 0.0)

            for _, tr in disc_rp_rows.iterrows():
                bunit = norm(tr.get("BENEFIT_UNIT", "PCS"))
                rp = parse_number_id(tr.get("BENEFIT_VALUE", "0"))
                rp_dpp = rp / 1.11

                if bunit == "GROSSAMOUNT":
                    total_rp = rp_dpp
                    add = (total_rp * (line_gross / sum_gross)) if sum_gross > 0 else 0.0
                elif bunit == "PCS":
                    total_rp = rp_dpp * sum_pcs
                    add = (total_rp * (line_pcs / sum_pcs)) if sum_pcs > 0 else 0.0
                elif bunit in ["CTN", "KRT"]:
                    total_rp = rp_dpp * sum_ctn
                    add = (total_rp * (line_ctn / sum_ctn)) if sum_ctn > 0 else 0.0
                else:
                    total_rp = rp_dpp * sum_pcs
                    add = (total_rp * (line_pcs / sum_pcs)) if sum_pcs > 0 else 0.0

                if hasattr(add, "index"):
                    for idx in idxs:
                        rp_line.loc[idx] += float(add.loc[idx])
                else:
                    for idx in idxs:
                        rp_line.loc[idx] += float(add)

                for idx in idxs:
                    cur = sales_df.at[idx, "Debug_ProgramKeysApplied"]
                    sales_df.at[idx, "Debug_ProgramKeysApplied"] = (cur + (";" if cur else "") + program_key)
                    sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                        sales_df.at[idx, "Program_Pabrik_Applied"], program_defs[(inv, program_key)]["plabel"]
                    )

        else:
            for idx in idxs:
                line = sales_df.loc[idx]
                brg = s(line["BRG"])
                rows_item = pick_rows_for_item(prog_rows, brg)
                if rows_item.empty:
                    continue

                trigger_unit_line, tu_source_line = choose_trigger_unit_satpam(rows_item)
                if trigger_unit_line is None:
                    sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (rp/non_group)")
                    sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
                    continue

                metric = float(line["GROSSAMOUNT"]) if trigger_unit_line == "GROSSAMOUNT" else qty_in_unit(line, trigger_unit_line)
                picked_tier, tier_rows = pick_best_tier(rows_item, metric)
                if tier_rows.empty:
                    continue
                tq_val = float(pd.to_numeric(tier_rows["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)

                eligible_rowids.add(int(line["_ROWID"]))
                sales_df.at[idx, "Debug_TriggerUnit"] = trigger_unit_line
                sales_df.at[idx, "Debug_TriggerUnitSource"] = tu_source_line
                sales_df.at[idx, "Debug_TriggerMetric"] = metric
                sales_df.at[idx, "Debug_TriggerQty"] = tq_val
                sales_df.at[idx, "Debug_PickedTier"] = picked_tier

                disc_rp_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_RP"].copy()
                if disc_rp_rows.empty:
                    continue

                rp_add = 0.0
                for _, tr in disc_rp_rows.iterrows():
                    bunit = norm(tr.get("BENEFIT_UNIT", "PCS"))
                    rp = parse_number_id(tr.get("BENEFIT_VALUE", "0"))
                    rp_dpp = rp / 1.11
                    rp_add += rp_dpp * qty_in_unit(line, bunit)

                if rp_add:
                    rp_line.loc[idx] += rp_add
                    cur = sales_df.at[idx, "Debug_ProgramKeysApplied"]
                    sales_df.at[idx, "Debug_ProgramKeysApplied"] = (cur + (";" if cur else "") + program_key)
                    sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                        sales_df.at[idx, "Program_Pabrik_Applied"], program_defs[(inv, program_key)]["plabel"]
                    )

    sales_df.loc[:, "_RP_PABRIK"] = rp_line.values

    # PASS 2 - PCT collect (with lock)
    for (inv, program_key) in sorted(program_scopes.keys()):
        idxs = program_scopes[(inv, program_key)]
        if not idxs:
            continue
        scope = sales_df.loc[idxs].copy()
        sub = s(scope["SUB"].iloc[0])
        eff_channel = s(scope["EffectiveChannel"].iloc[0])
        pgid = program_defs[(inv, program_key)]["pgid"]
        plabel = program_defs[(inv, program_key)]["plabel"]
        is_non_group = program_defs[(inv, program_key)]["is_non_group"]

        if not is_non_group:
            selected = inv_selected_pct_programs.get(s(inv), set())
            if selected and (program_key not in selected):
                for idx in idxs:
                    sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], f"LOCK_SKIP: pct program {program_key} not matching MDSTRING")
                continue

        brgs = scope["BRG"].apply(s).unique().tolist()
        prog_rows = scope_program_rows(promo_df, pgid, plabel, brgs)
        if prog_rows.empty:
            continue
        prog_rows = prog_rows[prog_rows.apply(lambda r: channel_ok(r.get("PROMO_GROUP", ""), eff_channel, r.get("PROMO_GROUP_ID", ""), sub), axis=1)].copy()
        if prog_rows.empty:
            continue

        trigger_unit, tu_source = choose_trigger_unit_satpam(prog_rows)
        if trigger_unit is None:
            for idx in idxs:
                sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (pct)")
                sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
            continue

        if is_non_group:
            for idx in idxs:
                line = sales_df.loc[idx]
                brg = s(line["BRG"])
                rows_item = pick_rows_for_item(prog_rows, brg)
                if rows_item.empty:
                    continue

                trigger_unit_line, tu_source_line = choose_trigger_unit_satpam(rows_item)
                if trigger_unit_line is None:
                    sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (pct/non_group)")
                    sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
                    continue

                metric = float(line["GROSSAMOUNT"]) if trigger_unit_line == "GROSSAMOUNT" else qty_in_unit(line, trigger_unit_line)
                picked_tier, tier_rows = pick_best_tier(rows_item, metric)
                if tier_rows.empty:
                    continue
                disc_pct_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_PCT"].copy()
                if disc_pct_rows.empty:
                    continue
                pct_parts = []
                for _, tr in disc_pct_rows.iterrows():
                    pct_parts.extend(parse_pct_chain(s(tr.get("BENEFIT_VALUE", ""))))
                if pct_parts:
                    pct_tasks.append({"indices": [idx], "pct_parts": pct_parts, "program_key": program_key})
                    cur = sales_df.at[idx, "_PCT_PARTS_PABRIK"]
                    if not isinstance(cur, list):
                        cur = []
                    cur.extend(pct_parts)
                    sales_df.at[idx, "_PCT_PARTS_PABRIK"] = cur
        else:
            metric = float(scope["GROSSAMOUNT"].sum() or 0.0) if trigger_unit == "GROSSAMOUNT" else float(scope.apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum() or 0.0)
            picked_tier, tier_rows = pick_best_tier(prog_rows, metric)
            if tier_rows.empty:
                continue

            tier_rows = dedupe_grouped_benefits(tier_rows)

            disc_pct_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_PCT"].copy()
            if disc_pct_rows.empty:
                continue

            pct_parts_scope = []
            for _, tr in disc_pct_rows.iterrows():
                pct_parts_scope.extend(parse_pct_chain(s(tr.get("BENEFIT_VALUE", ""))))
            if pct_parts_scope:
                pct_tasks.append({"indices": idxs, "pct_parts": pct_parts_scope, "program_key": program_key})
                for idx in idxs:
                    cur = sales_df.at[idx, "_PCT_PARTS_PABRIK"]
                    if not isinstance(cur, list):
                        cur = []
                    cur.extend(pct_parts_scope)
                    sales_df.at[idx, "_PCT_PARTS_PABRIK"] = cur

    # Apply % after rp_line known
    expected = rp_line.copy()
    for task in pct_tasks:
        idxs = task["indices"]; pct_parts = task["pct_parts"]; program_key = task.get("program_key", "")
        if not idxs or not pct_parts:
            continue
        scope = sales_df.loc[idxs].copy()
        rem = (scope["GROSSAMOUNT"].astype(float) - rp_line.loc[idxs]).clip(lower=0.0)
        base = float(rem.sum() or 0.0)
        if base <= 0:
            continue
        pct_total = sequential_pct_discount(base, pct_parts)
        if pct_total <= 0:
            continue
        srem = float(rem.sum() or 0.0)
        shares = (rem / srem) if srem > 0 else 0.0
        if hasattr(shares, "iloc"):
            for i, idx in enumerate(scope.index):
                expected.loc[idx] += float(pct_total * shares.iloc[i])

        for idx in idxs:
            cur = sales_df.at[idx, "Debug_ProgramKeysApplied"]
            if program_key:
                sales_df.at[idx, "Debug_ProgramKeysApplied"] = (cur + (";" if cur else "") + program_key)
                label = program_key.split("||", 1)[1] if "||" in program_key else program_key
                sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                    sales_df.at[idx, "Program_Pabrik_Applied"], label
                )

    sales_df.loc[:, "Expected_Pabrik"] = expected.values
    sales_df.loc[sales_df["_ROWID"].isin(list(eligible_rowids)), "HAS_ELIGIBLE_PROGRAM"] = True

    # ---------------------------
    # INTERNAL expected (v15)
    # ---------------------------
    sales_df.loc[:, "Expected_Internal"] = 0.0
    if internal_rules is not None and (not internal_rules.empty):
        for inv, g in sales_df.groupby("INVOICENO", sort=False):
            inv = s(inv)
            idxs = g.index.tolist()
            if not idxs:
                continue

            invoice_gross = float(g["GROSSAMOUNT"].sum() or 0.0)
            eff_channel = s(g["EffectiveChannel"].iloc[0])
            sub = s(g["SUB"].iloc[0])

            r0 = internal_rules[internal_rules["CHANNEL"].apply(lambda x: internal_channel_match(x, eff_channel))].copy()
            r0 = r0[(r0["SUB"].apply(lambda x: (s(x) == "") or (norm(x) == norm(sub))))].copy()
            if r0.empty:
                continue

            rem_line = (g["GROSSAMOUNT"].astype(float) - g["Expected_Pabrik"].astype(float)).clip(lower=0.0)
            rem_total = float(rem_line.sum() or 0.0)
            if rem_total <= 0:
                continue

            bonus_mask = g["_IS_BONUS_LINE_MAIN"] if "_IS_BONUS_LINE_MAIN" in g.columns else g["MDSTRING"].apply(mdstring_is_bonus)

            def apply_internal_indices(idxs_apply: List[int], pct_parts: List[float], mark_bonus: bool) -> bool:
                if not idxs_apply:
                    return False
                rem = rem_line.loc[idxs_apply]
                base = float(rem.sum() or 0.0)
                if base <= 0:
                    return False
                disc = sequential_pct_discount(base, pct_parts)
                if disc <= 0:
                    return False
                for idx in idxs_apply:
                    cur = sales_df.at[idx, "_PCT_PARTS_INTERNAL"]
                    if not isinstance(cur, list):
                        cur = []
                    cur.extend(pct_parts)
                    sales_df.at[idx, "_PCT_PARTS_INTERNAL"] = cur
                    if mark_bonus:
                        sales_df.at[idx, "_INTERNAL_BONUS_MATCH"] = True
                shares = rem / float(rem.sum() or 1.0)
                for idx in idxs_apply:
                    sales_df.at[idx, "Expected_Internal"] += float(disc * shares.loc[idx])
                return True

            used_any_discount = False

            # Item-specific rules first (if any)
            for brg, gi in g.groupby("BRG", sort=False):
                brg = s(brg)
                rr_item = r0[(r0["KODE_BARANG"].apply(lambda x: norm(x) == norm(brg)))].copy()
                if rr_item.empty:
                    continue
                picked = pick_internal_rule(rr_item, invoice_gross)
                if picked.empty:
                    continue
                pct_parts = []
                for v in picked["DISC_PCT"].tolist():
                    pct_parts.extend(parse_pct_chain(s(v)))
                if not pct_parts:
                    continue

                gi_idxs = gi.index.tolist()
                gi_bonus = [i for i in gi_idxs if bool(bonus_mask.loc[i])]
                gi_non_bonus = [i for i in gi_idxs if not bool(bonus_mask.loc[i])]

                applied_any = False
                if gi_non_bonus:
                    applied_any |= apply_internal_indices(gi_non_bonus, pct_parts, False)
                if pct_parts_full_discount(pct_parts) and gi_bonus:
                    applied_any |= apply_internal_indices(gi_bonus, pct_parts, True)

                if applied_any:
                    used_any_discount = True

            # General rule if no item-specific matched
            if not used_any_discount:
                rr_gen = r0[r0["KODE_BARANG"].apply(lambda x: s(x) == "")].copy()
                picked = pick_internal_rule(rr_gen, invoice_gross)
                if picked.empty:
                    continue
                pct_parts = []
                for v in picked["DISC_PCT"].tolist():
                    pct_parts.extend(parse_pct_chain(s(v)))
                if not pct_parts:
                    continue
                gen_bonus = [i for i in idxs if bool(bonus_mask.loc[i])]
                gen_non_bonus = [i for i in idxs if not bool(bonus_mask.loc[i])]
                if gen_non_bonus:
                    apply_internal_indices(gen_non_bonus, pct_parts, False)
                if pct_parts_full_discount(pct_parts) and gen_bonus:
                    apply_internal_indices(gen_bonus, pct_parts, True)

    
# ---------------------------
    # BONUS satpam (v16+bonus)
    # Bonus definition:
    # - MDSTRING contains 100% (free goods) -> treated as BONUS line
    # - Expected bonus is computed from Dataset Diskon (pabrik) where BENEFIT_TYPE = BONUS_QTY/BONUS
    #   using the same Trigger/Tier rules as discount engine (NON_GROUP per-line; grouped per-invoice).
    # - Expected bonus is allocated to bonus lines (MDSTRING=100%) for the eligible item(s).
    # ---------------------------
    sales_df.loc[:, "Actual_BonusQty"] = sales_df.apply(
        lambda r: float(qty_in_unit(r, "PCS")) if mdstring_is_bonus(r.get("MDSTRING", "")) else 0.0,
        axis=1
    )
    compute_expected_bonus_from_rules(
        sales_df=sales_df,
        rules_df=promo_df,
        channel_col="EffectiveChannel",
        expected_col="Expected_Bonus_Pabrik",
        debug_reason_prefix="BONUS_PABRIK",
        program_name_col="Program_Bonus_Applied"
    )
    # Internal bonus: template internal currently focuses on discount pct; keep 0 unless future internal rules add bonus columns.
    sales_df.loc[:, "Expected_Bonus_Internal"] = 0.0

    # Bonus discount allocation:
    # - Internal only if internal rule is 100% (marked by _INTERNAL_BONUS_MATCH)
    # - Otherwise, if bonus program matched (Expected_Bonus_Pabrik > 0) -> discount goes to Pabrik
    # - Else -> TanpaTuan
    bonus_mask = sales_df["_IS_BONUS_LINE_MAIN"] if "_IS_BONUS_LINE_MAIN" in sales_df.columns else sales_df["MDSTRING"].apply(mdstring_is_bonus)
    if bonus_mask.any():
        sales_df.loc[bonus_mask, "Expected_Pabrik"] = 0.0
        sales_df.loc[bonus_mask, "Expected_Internal"] = 0.0
        unit_disc = pd.Series(0.0, index=sales_df.index)
        bonus_qty = sales_df.loc[bonus_mask, "Actual_BonusQty"].astype(float).fillna(0.0)
        denom = bonus_qty.replace(0, float("nan"))
        unit_disc.loc[bonus_mask] = sales_df.loc[bonus_mask, "LDISCAMOUNT"].astype(float).fillna(0.0).div(denom).fillna(0.0)

        bonus_internal = bonus_mask & (sales_df["_INTERNAL_BONUS_MATCH"] == True)
        if bonus_internal.any():
            sales_df.loc[bonus_internal, "Expected_Internal"] = unit_disc.loc[bonus_internal] * sales_df.loc[bonus_internal, "Actual_BonusQty"].astype(float).fillna(0.0)

        bonus_pabrik = bonus_mask & (~bonus_internal) & (sales_df["Expected_Bonus_Pabrik"].astype(float) > 0)
        if bonus_pabrik.any():
            sales_df.loc[bonus_pabrik, "Expected_Pabrik"] = unit_disc.loc[bonus_pabrik] * sales_df.loc[bonus_pabrik, "Expected_Bonus_Pabrik"].astype(float).fillna(0.0)
            for idx in sales_df.index[bonus_pabrik]:
                sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                    sales_df.at[idx, "Program_Pabrik_Applied"],
                    sales_df.at[idx, "Program_Bonus_Applied"]
                )
    return finalize(sales_df, sales_cols_fixed)

# ---------------------------
# Endpoints
# ---------------------------
@app.post("/validate_json")
async def validate_json(
    request: Request,
    sales: UploadFile = File(None),
    promo: UploadFile = File(None),
    channel: UploadFile = File(None),
    internal: UploadFile = File(None),
):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "validator", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    # Server-side guard (in case someone calls API directly)
    if sales is None or promo is None or channel is None:
        return JSONResponse(status_code=400, content={"ok": False, "error": REQUIRED_MISSING_MSG})

    try:
        sales_bytes = await read_upload_file_limited(
            sales,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="Data Penjualan",
        )
        promo_bytes = await read_upload_file_limited(
            promo,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="Dataset Diskon Pabrik",
        )
        channel_bytes = await read_upload_file_limited(
            channel,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="Data Channel by SUB",
        )
        internal_bytes = None
        if internal is not None:
            try:
                internal_bytes = await read_upload_file_limited(
                    internal,
                    max_bytes=MAX_EXCEL_UPLOAD_BYTES,
                    allowed_exts=(".xlsx", ".xls"),
                    label="Dataset Diskon Internal",
                )
            except:
                internal_bytes = None

        out_df = run_engine(sales_bytes, promo_bytes, channel_bytes, internal_bytes)

        file_id = str(uuid.uuid4())[:8]
        base_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base_dir, "output")
        out_path = os.path.join(out_dir, f"hasil_validasi_{file_id}.xlsx")
        write_excel(out_df, out_path)

        counts = out_df["StatusValidasi"].value_counts(dropna=False).to_dict()
        return JSONResponse({
            "ok": True,
            "file_id": file_id,
            "download_url": f"/download/{file_id}",
            "counts": {"A": int(counts.get("A", 0)), "B": int(counts.get("B", 0)), "C": int(counts.get("C", 0))},
        })
    except ValueError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        append_error_log("validate_json", e, {"user": user})
        payload = {"ok": False, "error": "Terjadi kesalahan saat memproses validasi."}
        if APP_DEBUG and is_admin_user(user):
            payload["detail"] = str(e)
        return JSONResponse(status_code=500, content=payload)

@app.post("/validate")
async def validate(
    request: Request,
    sales: UploadFile = File(None),
    promo: UploadFile = File(None),
    channel: UploadFile = File(None),
    internal: UploadFile = File(None),
):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "validator", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    if sales is None or promo is None or channel is None:
        return JSONResponse(status_code=400, content={"ok": False, "error": REQUIRED_MISSING_MSG})

    try:
        sales_bytes = await read_upload_file_limited(
            sales,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="Data Penjualan",
        )
        promo_bytes = await read_upload_file_limited(
            promo,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="Dataset Diskon Pabrik",
        )
        channel_bytes = await read_upload_file_limited(
            channel,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="Data Channel by SUB",
        )
        internal_bytes = None
        if internal is not None:
            try:
                internal_bytes = await read_upload_file_limited(
                    internal,
                    max_bytes=MAX_EXCEL_UPLOAD_BYTES,
                    allowed_exts=(".xlsx", ".xls"),
                    label="Dataset Diskon Internal",
                )
            except:
                internal_bytes = None

        out_df = run_engine(sales_bytes, promo_bytes, channel_bytes, internal_bytes)

        base_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base_dir, "output")
        out_path = os.path.join(out_dir, "hasil_validasi.xlsx")
        write_excel(out_df, out_path)

        return accel_or_file_response(out_path, "hasil_validasi.xlsx")
    except ValueError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        append_error_log("validate", e, {"user": user})
        payload = {"ok": False, "error": "Terjadi kesalahan saat memproses validasi."}
        if APP_DEBUG and is_admin_user(user):
            payload["detail"] = str(e)
        return JSONResponse(status_code=500, content=payload)

@app.get("/download/{file_id}")
def download(request: Request, file_id: str):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"detail": "Unauthorized"})
    if not user_has_permission(user, "validator", "view"):
        return JSONResponse(status_code=403, content={"detail": "Forbidden"})
    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(base_dir, "output", f"hasil_validasi_{file_id}.xlsx")
    return accel_or_file_response(out_path, "hasil_validasi.xlsx")


@app.get("/validator/template/sales")
def validator_template_sales(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    if not user_has_permission(user, "validator", "view"):
        return HTMLResponse("Forbidden", status_code=403)
    return _excel_download_response(validator_sales_template_rows(), "template_data_penjualan.xlsx", "SALES_TEMPLATE")


@app.get("/validator/template/promo")
def validator_template_promo(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    if not user_has_permission(user, "validator", "view"):
        return HTMLResponse("Forbidden", status_code=403)
    return _excel_download_response(validator_promo_template_rows(), "template_dataset_diskon_pabrik.xlsx", "PROMO_TEMPLATE")


@app.get("/validator/template/channel")
def validator_template_channel(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login")
    if not user_has_permission(user, "validator", "view"):
        return HTMLResponse("Forbidden", status_code=403)
    return _excel_download_response(validator_channel_template_rows(), "template_data_channel_by_sub.xlsx", "CHANNEL_TEMPLATE")

@app.post("/summary/manual")
async def summary_manual_auto_generate(
    request: Request,
    file: UploadFile = File(None),
    list_mode: str = Form("TANPA LIST"),
    template: str = Form("GUMINDO"),
    engine: str = Form("ai"),  # "ai" or "manual"
    model: str = Form("kimi-k2-250905"),
):
    """
    Auto summary generator (formerly /summary_generate).
    - engine="ai": use SumoPod model (Kimi/DeepSeek) to extract structured rows
    - engine="manual": use template parser (build_summary_rows)
    """
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "summary", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    if file is None:
        return JSONResponse(status_code=400, content={"ok": False, "error": "File belum dipilih"})

    try:
        raw = await read_upload_file_limited(
            file,
            max_bytes=MAX_PDF_UPLOAD_BYTES,
            allowed_exts=(".pdf", ".xlsx", ".xls"),
            label="File Summary",
        )
        if file.filename and file.filename.lower().endswith((".xlsx", ".xls")):
            return JSONResponse(status_code=400, content={"ok": False, "error": "Excel belum didukung, upload PDF dulu."})

        text = extract_pdf_text(raw)
        if len(s(text)) < 50:
            return JSONResponse(status_code=400, content={"ok": False, "error": "OCR/PDF text kosong. Pastikan OCR tersedia di server."})

        engine_l = s(engine).lower()
        rows: List[Dict[str, str]] = []
        if engine_l in ("manual", "rule", "template"):
            rows = build_summary_rows(text, list_mode, s(template).upper())
        else:
            # AI first, fallback to template parser if AI fails
            try:
                rows = ai_extract_summary_rows(text, list_mode, s(template).upper(), model=model)
            except Exception as ai_err:
                rows = build_summary_rows(text, list_mode, s(template).upper())
                if not rows:
                    raise ai_err

        if not rows:
            return JSONResponse(status_code=400, content={"ok": False, "error": "Template belum dikenali / hasil kosong."})

        file_id = str(uuid.uuid4())[:8]
        base_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base_dir, "output")
        out_path = os.path.join(out_dir, f"summary_{file_id}.xlsx")
        write_summary_excel(rows, out_path)

        return JSONResponse({"ok": True, "file_id": file_id, "download_url": f"/summary_download/{file_id}"})
    except ValueError as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        append_error_log("summary_manual_auto_generate", e, {"user": user})
        payload = {"ok": False, "error": "Gagal memproses summary otomatis."}
        if APP_DEBUG and is_admin_user(user):
            payload["detail"] = str(e)
        return JSONResponse(status_code=500, content=payload)

@app.get("/summary_download/{file_id}")
def summary_download(request: Request, file_id: str):
    # Auth check removed to support direct downloads from Next.js cross-origin links
    # The UUID acts as the access token.
    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(base_dir, "output", f"summary_{file_id}.xlsx")
    return accel_or_file_response(out_path, "summary.xlsx")



# ======================================================================================
# Summary Program Manual Web Input (Tom Select checkbox_options)
# - Variant & Gramasi: multi-select checkbox
# - "ALL VARIANT" / "ALL GRAMASI" exclusive (cannot be selected with other options)
# - User can create new values not in master (create:true)
# ======================================================================================

import openpyxl  # added for master parsing + excel output

MANUAL_MASTER_CACHE: dict = {}   # token -> {"kelompok": [...], "variant_map": {...}, "gramasi_map": {...}}
MANUAL_OUTPUTS: dict = {}        # file_id -> {"form": path, "dataset": path}

def _norm_col(x: object) -> str:
    return " ".join(str(x or "").strip().split()).upper()

def _parse_master_barang_xlsx(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {}

    for c, h in enumerate(headers, start=1):
        hn = _norm_col(h)

        # Kode / Nama barang
        if hn == "KODE BARANG":
            idx["kode"] = c
        if hn == "NAMA BARANG":
            idx["nama"] = c
            
        # Principle Name Target (e.g. Camellia, Bellagio)
        if hn == "NAMA BARANG PRINCIPLE" or hn == "PRINCIPLE":
            idx["principle"] = c

        # Kelompok fields
        if ("NAMA" in hn and "KLP" in hn and "SUB" not in hn) or hn == "NAMA KLP":
            idx["klp"] = c
        if ("NAMA" in hn and "SUB" in hn and "KLP 2" not in hn and "KLP2" not in hn) or hn == "NAMA SUB KLP":
            idx["sub1"] = c
        if ("NAMA" in hn and "SUB" in hn and ("KLP 2" in hn or "KLP2" in hn)) or hn == "NAMA SUB KLP 2":
            idx["sub2"] = c

        # Variant = Aroma/Rasa
        if ("AROMA" in hn) or ("RASA" in hn):
            idx["variant"] = c

        # Gramasi
        if ("GRAMASI" in hn) or (("PACK" in hn or "PCS" in hn) and ("CTN" in hn or "KARTON" in hn or "CARTON" in hn)):
            idx["gramasi"] = c

    missing = [k for k in ["kode", "nama", "klp", "sub1", "sub2", "variant", "gramasi"] if k not in idx]
    if missing:
        raise ValueError("Kolom master barang tidak ketemu: " + ", ".join(missing))

    variant_map = {}
    gramasi_map = {}
    items = []

    for r in range(2, ws.max_row + 1):
        kode_barang = str(ws.cell(r, idx["kode"]).value or "").strip()
        nama_barang = str(ws.cell(r, idx["nama"]).value or "").strip()
        principle = str(ws.cell(r, idx["principle"]).value or "").strip() if "principle" in idx else ""
        
        klp = str(ws.cell(r, idx["klp"]).value or "").strip()
        sub1 = str(ws.cell(r, idx["sub1"]).value or "").strip()
        sub2 = str(ws.cell(r, idx["sub2"]).value or "").strip()
        kelompok = " - ".join([x for x in [klp, sub1, sub2] if x]).strip()
        if not kelompok:
            continue

        v = str(ws.cell(r, idx["variant"]).value or "").strip()
        g = str(ws.cell(r, idx["gramasi"]).value or "").strip()

        if v:
            variant_map.setdefault(kelompok, set()).add(v)
        if g:
            gramasi_map.setdefault(kelompok, set()).add(g)

        if kode_barang or nama_barang:
            items.append({
                "kode_barang": kode_barang,
                "nama_barang": nama_barang,
                "principle": principle,
                "kelompok": kelompok,
                "variant": v,
                "gramasi": g,
            })

    kelompok_list = sorted(set(list(variant_map.keys()) + list(gramasi_map.keys())))
    variant_map = {k: sorted(list(v)) for k, v in variant_map.items()}
    gramasi_map = {k: sorted(list(v)) for k, v in gramasi_map.items()}
    return kelompok_list, variant_map, gramasi_map, items

def _parse_master_customer_xlsx(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {}
    for c, h in enumerate(headers, start=1):
        hn = str(h or "").strip().upper()
        if hn == "KODE CUSTOMER": idx["kode"] = c
        elif hn == "NAMA CUSTOMER": idx["nama"] = c
            
    if "kode" not in idx or "nama" not in idx:
        raise ValueError("Kolom Master Customer tidak valid. Pastikan ada KODE CUSTOMER dan NAMA CUSTOMER.")
        
    customers = []
    for r in range(2, ws.max_row + 1):
        kode = str(ws.cell(r, idx["kode"]).value or "").strip()
        nama = str(ws.cell(r, idx["nama"]).value or "").strip()
        if kode and nama:
            customers.append({"kode_customer": kode, "nama_customer": nama})
            
    return customers

def _ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)


@app.get("/dev/dump_context")
def dev_dump_context(token: str, principle_name: str = "Priskila (Default)"):
    if token not in MANUAL_MASTER_CACHE: return {"error": "no token"}
    cache = MANUAL_MASTER_CACHE[token]
    raw_items = cache.get("items", [])
    
    if principle_name and principle_name.strip():
        items = [it for it in raw_items if principle_name.upper() in str(it.get("Nama Barang Principle", "")).upper()]
        if not items:
            items = raw_items 
    else:
        items = raw_items
        
    item_names_cache = set()
    kode_barang_map = {}
    for item in items:
        name = str(item.get("Nama Barang", "")).strip().upper()
        code = str(item.get("Kode Barang", "")).strip()
        if name: item_names_cache.add(name)
        if name not in kode_barang_map: kode_barang_map[name] = []
        if code and code not in kode_barang_map[name]: kode_barang_map[name].append(code)
            
    master_names_context = ""
    for n, kodes in kode_barang_map.items():
        s_kodes = ",".join(kodes)
        for master_item in items:
            nama_barang = str(master_item.get("Nama Barang", "")).strip().upper()
            nama_principle = str(master_item.get("Nama Barang Principle", "")).strip().upper()
            nama_aroma = ""
            for k, v in master_item.items():
                if "aroma" in str(k).lower() or "rasa" in str(k).lower() or "variant" in str(k).lower():
                    nama_aroma = str(v).strip()
                    break
            kelompok_asli = str(master_item.get("kelompok", "")).strip()
            if nama_barang == n:
                master_names_context += f"REF: {nama_principle} - {nama_barang} -> OUTPUT_KELOMPOK: {kelompok_asli} | OUTPUT_VARIANT: {nama_aroma} | OUTPUT_KODE: {s_kodes}\n"
                break
                
    return {"ok": True, "count": len(items), "context": master_names_context}

@app.post("/summary/manual/master/upload")
async def summary_manual_master_upload(
    request: Request, 
    master: UploadFile = File(...),
    master_customer: UploadFile = File(None)
):

    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "summary", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        file_bytes = await read_upload_file_limited(
            master,
            max_bytes=MAX_EXCEL_UPLOAD_BYTES,
            allowed_exts=(".xlsx", ".xls"),
            label="Master Barang",
        )
        kelompok_list, variant_map, gramasi_map, items = _parse_master_barang_xlsx(file_bytes)
        
        customers = []
        if master_customer and master_customer.filename:
            cust_bytes = await read_upload_file_limited(
                master_customer,
                max_bytes=MAX_EXCEL_UPLOAD_BYTES,
                allowed_exts=(".xlsx", ".xls"),
                label="Master Customer",
            )
            customers = _parse_master_customer_xlsx(cust_bytes)
            
        token = str(uuid.uuid4())
        MANUAL_MASTER_CACHE[token] = {
            "kelompok": kelompok_list,
            "variant_map": variant_map,
            "gramasi_map": gramasi_map,
            "items": items,
            "customers": customers
        }
        return {"ok": True, "token": token, "kelompok_list": kelompok_list}
    except ValueError as e:
        return {"ok": False, "error": str(e)}
    except Exception as e:
        append_error_log("summary_manual_master_upload", e, {"user": user})
        if APP_DEBUG and is_admin_user(user):
            return {"ok": False, "error": str(e)}
        return {"ok": False, "error": "Gagal membaca master barang."}

@app.get("/summary/manual/master/options")
def summary_manual_master_options(request: Request, token: str, group: str):

    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "summary", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    try:
        if token not in MANUAL_MASTER_CACHE:
            return {"ok": False, "error": "Token master tidak ditemukan / expired"}

        cache = MANUAL_MASTER_CACHE[token]
        vlist = cache["variant_map"].get(group, [])
        glist = cache["gramasi_map"].get(group, [])

        variants = [{"value": "ALL VARIANT", "text": "ALL VARIANT", "disabled": False}] +                    [{"value": v, "text": v, "disabled": False} for v in vlist]
        gramasies = [{"value": "ALL GRAMASI", "text": "ALL GRAMASI", "disabled": False}] +                     [{"value": g, "text": g, "disabled": False} for g in glist]

        return {"ok": True, "variants": variants, "gramasis": gramasies}
    except Exception as e:
        append_error_log("summary_manual_master_options", e, {"user": user, "group": group, "token": token})
        payload = {"ok": False, "error": "Gagal memuat opsi master."}
        if APP_DEBUG and is_admin_user(user):
            payload["detail"] = str(e)
        return payload

@app.post("/summary/manual/generate")
def summary_manual_generate(request: Request, token: str = Form(...), rows_json: str = Form(...)):

    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "summary", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        if token not in MANUAL_MASTER_CACHE:
            return {"ok": False, "error": "Token master tidak ditemukan / expired"}
        rows = json.loads(rows_json)
        # FASE 5: identitas input DIUKUR DI SINI (rows msh murni, sblm diproses/dimutasi
        # di bawah -- _matched_items_cache dll ditempel belakangan). Order-sensitive: run
        # dokumen yg sama menghasilkan urutan baris yg sama (ekstraksi kini deterministik).
        _golden_input_key = canonical_signature(rows)

        try:
            with open("d:/disc_web/debug_payload.json", "w") as f:
                json.dump(rows, f, indent=4)
        except: pass
        
        cache = MANUAL_MASTER_CACHE[token]
        items = cache.get("items", [])

        def norm(x: object) -> str:
            return " ".join(str(x or "").strip().split()).upper()

        def split_list(val: str) -> List[str]:
            return [x.strip() for x in str(val or "").split(",") if x.strip()]

        def unit_from_text(text: str) -> str:
            t = norm(text)
            if "CTN" in t or "KRT" in t:
                return "CTN"
            return "PCS"

        def has_number(text: str) -> bool:
            return bool(re.search(r"\d", str(text or "")))

        # ponytail: nilai kelompok/variant/gramasi digabung pakai " & " saat konsolidasi, BUKAN koma.
        # split_list (koma) tak pernah memecahnya -> dulu tak ada compression/dedup. Pisah pakai " & ".
        # Pakai spasi wajib di sekitar & supaya varian spt "H.SHIN&STR HLD" tidak ikut terpecah.
        def split_amp(val: str) -> List[str]:
            return [x.strip() for x in re.split(r'\s+&\s+', str(val or "")) if x.strip()]

        def join_human(items: List[str]) -> str:
            if not items: return ""
            if len(items) == 1: return items[0]
            if len(items) == 2: return f"{items[0]} & {items[1]}"
            return ", ".join(items[:-1]) + f" & {items[-1]}"

        def format_array_human_readable(raw_str: str) -> str:
            seen = set()
            unique_arr = []
            for g in split_amp(raw_str):
                if g not in seen:
                    unique_arr.append(g)
                    seen.add(g)
            return join_human(unique_arr)

        def format_kelompoks_human_readable(raw_str: str) -> str:
            seen = set()
            unique_k = []
            for k in split_amp(raw_str):
                if k not in seen:
                    unique_k.append(k)
                    seen.add(k)

            from collections import defaultdict
            groups = defaultdict(list)
            order = []
            for k in unique_k:
                if " - " in k:
                    prefix, suffix = k.split(" - ", 1)
                    # "EDP - PRESTIGE" -> "EDP PRESTIGE" (rapikan dash internal sub-kelompok)
                    suffix = suffix.replace(" - ", " ").strip()
                else:
                    prefix, suffix = k, ""
                if prefix not in groups:
                    order.append(prefix)
                if suffix and suffix not in groups[prefix]:
                    groups[prefix].append(suffix)

            result_parts = []
            for prefix in order:
                clean = groups[prefix]
                if not clean:
                    result_parts.append(prefix)
                else:
                    result_parts.append(f"{prefix} - {join_human(clean)}")
            return join_human(result_parts)

        # Row Consolidation Algorithm: Merge rows that share exactly the same Base Prefix, Ketentuan, Benefit, and Channel.
        consolidated_rows_dict = {}
        idx_counter = 1
        
        for r in rows:
            # Safely extract prefix (e.g., 'BLAGIO HM' from 'BLAGIO HM - EDT')
            raw_k = str(r.get("kelompok", "")).strip()
            prefix = raw_k.split(" - ")[0] if " - " in raw_k else raw_k
            
            # The composite key dictates what gets merged together
            merge_key = (
                r.get("surat_program", ""),
                r.get("nama_program", ""),
                r.get("channel_gtmt", ""),
                r.get("periode", ""),
                prefix,
                norm(r.get("ketentuan", "")),
                norm(r.get("benefit", "")),
                norm(r.get("benefit_type", ""))
            )
            
            if merge_key not in consolidated_rows_dict:
                # First time seeing this combination, clone the row
                r_copy = dict(r)
                # Keep _matched_items_cache as a list of dictionaries if it exists
                cache = r_copy.get("_matched_items_cache", [])
                r_copy["_matched_items_cache"] = list(cache) if isinstance(cache, list) else []
                consolidated_rows_dict[merge_key] = r_copy
            else:
                # Merge into existing row
                target = consolidated_rows_dict[merge_key]
                
                # Append string fields using '&'
                for field in ["kelompok", "variant", "gramasi"]:
                    val1 = target.get(field, "")
                    val2 = r.get(field, "")
                    if val2:
                        target[field] = f"{val1} & {val2}" if val1 else val2
                        
                # Append comma separated fields
                for field in ["kode_barangs"]:
                    val1 = target.get(field, "")
                    val2 = r.get(field, "")
                    if val2:
                        target[field] = f"{val1},{val2}" if val1 else val2
                        
                # Merge caches
                incoming_cache = r.get("_matched_items_cache", [])
                if isinstance(incoming_cache, list):
                    target["_matched_items_cache"].extend(incoming_cache)

        # Re-assign the consolidated rows back to the main list
        rows = list(consolidated_rows_dict.values())
        
        # Apply Prefix Compressor and Formatter globally to all rows before PDF & Excel generation
        for i, r in enumerate(rows):
            r["no"] = str(i + 1)
            r["kelompok"] = format_kelompoks_human_readable(r.get("kelompok", ""))
            r["variant"] = format_array_human_readable(r.get("variant", ""))
            r["gramasi"] = format_array_human_readable(r.get("gramasi", ""))
            
            # Rebuild clean Kode Barangs without duplicates
            k_list = split_list(r.get("kode_barangs", ""))
            r["kode_barangs"] = ",".join(list(dict.fromkeys(k_list)))

        out_dir = os.path.join(BASE_DIR, "output", "summary_manual")
        _ensure_dir(out_dir)
        file_id = str(uuid.uuid4())

        form_path = os.path.join(out_dir, f"{file_id}_Form_Summary_Program.pdf")
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.units import inch
        from datetime import datetime
        
        # Get current date formatted for Indonesia
        now = datetime.now()
        months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
        print_date = f"{now.month}/{now.day}/{now.year}"
        dibuat_date = f"{now.day:02d} {months[now.month - 1]} {now.year}"

        def my_canvas(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setFont('Helvetica-Bold', 7)
            # Top Left
            canvas_obj.drawString(doc_obj.leftMargin, landscape(A4)[1] - 0.5*cm, f"Di Print Tgl : {print_date}")
            # Top Right (Page Number)
            page_str = f"Hal {doc_obj.page}"
            canvas_obj.drawRightString(landscape(A4)[0] - doc_obj.rightMargin, landscape(A4)[1] - 0.5*cm, page_str)
            
            # Bottom of Header (Before Table Starts)
            canvas_obj.setFont('Helvetica', 8)
            canvas_obj.drawString(doc_obj.leftMargin, landscape(A4)[1] - 2*cm, f"Dibuat Tanggal : {dibuat_date}")
            canvas_obj.setFont('Helvetica-Bold', 8)
            canvas_obj.drawRightString(landscape(A4)[0] - doc_obj.rightMargin, landscape(A4)[1] - 2*cm, "(ON PRINCIPLE COKLAT)")
            
            canvas_obj.restoreState()

        # FASE 6: PDF reproducible (CreationDate/ModDate/doc-id tetap) -> byte-identik antar-run
        enable_pdf_determinism()
        # Update margins to give space for the custom canvas headers
        doc = SimpleDocTemplate(form_path, pagesize=landscape(A4), rightMargin=0.5*cm, leftMargin=0.5*cm, topMargin=2.2*cm, bottomMargin=0.5*cm)
        elements = []
        styles = getSampleStyleSheet()

        # Add Title and Subtitle properly centered
        title_style = styles["Heading3"].clone("TitleStyle")
        title_style.alignment = TA_CENTER
        title_style.fontSize = 12
        elements.append(Paragraph("<b>SUMMARY PROGRAM ON FAKTUR BEBAN PRINCIPLE (COKLAT)</b>", title_style))
        
        sub_style = styles["Normal"].clone("SubStyle")
        sub_style.alignment = TA_CENTER
        sub_style.fontSize = 11
        sub_style.fontName = 'Helvetica'
        
        # Determine Period globally from the first row if available
        global_period = rows[0].get("periode", "") if rows else ""
        period_text = f" PERIODE {global_period.upper()}" if global_period else ""
        elements.append(Paragraph(f"CV. SURYA PERKASA {period_text}", sub_style))
        elements.append(Spacer(1, 15))

        headers_str = ["No", "Surat Program", "Nama Program", "Channel", "Periode", 
                   "Kelompok Barang", "Variant", "Gramasi", "Ketentuan", "Benefit", "Syarat Claim", "Keterangan"]
        
        header_style = styles["Normal"].clone("HeaderStyle")
        header_style.fontSize = 7
        header_style.leading = 8
        header_style.fontName = 'Helvetica-Bold'
        header_style.alignment = TA_CENTER
        header_style.textColor = colors.whitesmoke
        
        table_data = [[Paragraph(h, header_style) for h in headers_str]]

        cell_style = styles["Normal"].clone("CellStyle")
        cell_style.fontSize = 6
        cell_style.leading = 7
        cell_style.alignment = TA_CENTER

        def benefit_display(r) -> str:
            # ponytail: cut price (DISC_RP) tampilkan per-satuan (default PCS) -> "4700/PCS".
            # BONUS_QTY sudah "1 PCS", DISC_PCT sudah "5%" -> biarkan apa adanya.
            b = str(r.get("benefit", "") or "").strip()
            bt = norm(r.get("benefit_type", ""))
            if bt == "DISC_RP" and has_number(b) and "/" not in b:
                unit = unit_from_text(str(r.get("ketentuan", "")) + " " + b)
                return f"{b}/{unit}"
            return b

        # ============================================================
        # SINGLE SOURCE OF TRUTH: cocokkan tiap baris surat ke item per-SKU SEKALI di sini.
        # PDF (loop di bawah) & Excel (setelah PDF, lihat blok Excel) SAMA-SAMA dibangun dari
        # hasil ini -- tidak ada lagi 2 pipeline independen yang bisa saling berbeda (akar
        # divergensi PDF vs Excel yg dilaporkan user).
        # ============================================================
        channel_seq: Dict[str, int] = {}
        pdf_meta: Dict[int, dict] = {}
        pdf_items: Dict[int, list] = {}
        excel_rows: List[Dict[str, Any]] = []
        flagged_mismatches: List[Dict[str, Any]] = []
        flagged_conflicts: List[Dict[str, Any]] = []

        for i, r in enumerate(rows):
            kelompok = str(r.get("kelompok","") or "").strip()
            vlist = split_list(r.get("variant",""))
            glist_raw = str(r.get("gramasi","") or "")
            glist = split_list(glist_raw)
            v_all = (not vlist) or any(norm(x) == "ALL VARIANT" for x in vlist)
            g_all = (not glist) or any(norm(x) == "ALL GRAMASI" for x in glist)

            ket = str(r.get("ketentuan","") or "").strip()
            promo_label = str(r.get("nama_program","") or "").strip()
            promo_group_id = str(r.get("promo_group_id","") or "").strip()
            promo_group = str(r.get("channel_gtmt","") or "").strip()
            periode = str(r.get("periode","") or "").strip()
            if not promo_group_id or promo_group_id.upper() == "NON_GROUP":
                promo_group_id = promo_group

            pdf_meta[i] = {
                "no": i + 1, "surat_program": r.get("surat_program",""), "nama_program": promo_label,
                "channel_gtmt": promo_group, "periode": periode, "ketentuan": ket,
                "benefit_type": r.get("benefit_type",""), "benefit": r.get("benefit",""),
                "syarat_claim": r.get("syarat_claim",""), "keterangan": r.get("keterangan",""),
                "variant_display": r.get("variant",""), "kelompok_fallback": r.get("kelompok",""),
            }
            pdf_items[i] = []

            matched_items = []
            klist = [k.strip() for k in str(r.get("kode_barangs", "")).split(",") if k.strip()]

            if klist:
                for it in items:
                    if str(it.get("kode_barang", "")).strip() in klist:
                        matched_items.append(it)
                # ponytail: guard V3b -- kode dari AI divalidasi silang ke gramasi yg BENERAN
                # diklaim baris surat ini (terbukti live: "Pomade 80gr" salah nyantol ke master
                # "PMD KIDZ" gramasi 40gr krn AI cuma percaya kode_barangs GPT tanpa cross-check).
                # Kalau gramasi master tidak muncul sama sekali di klaim surat -> buang & catat,
                # JANGAN diam-diam dipakai (akurasi finansial wajib, bukan tebak-tebakan).
                if not g_all and glist:
                    _kept = []
                    for it in matched_items:
                        it_gram = norm(it.get("gramasi"))
                        if any(it_gram and (norm(g) == it_gram or it_gram in norm(g) or norm(g) in it_gram) for g in glist):
                            _kept.append(it)
                        else:
                            flagged_mismatches.append({
                                "kode_barang": it.get("kode_barang"), "nama_barang": it.get("nama_barang"),
                                "gramasi_master": it.get("gramasi"), "gramasi_klaim_surat": glist_raw,
                                "channel": promo_group, "ketentuan": ket,
                            })
                    matched_items = _kept

            if not matched_items:
                fb_kelompok = kelompok
                if fb_kelompok and any(skip in fb_kelompok.lower() for skip in ["- kelompok -", "bisa meleset"]):
                    fb_kelompok = ""
                pool = [it for it in items if norm(it.get("kelompok")) == norm(fb_kelompok)] if fb_kelompok else items
                if not pool: pool = items
                for it in pool:
                    it_variant = norm(it.get("variant")); it_nama = norm(it.get("nama_barang"))
                    variant_match = v_all
                    if not variant_match:
                        for v in [norm(x) for x in vlist]:
                            if "- variant -" in v.lower() or "all variant" in v.lower() or "bisa meleset" in v.lower():
                                variant_match = True; break
                            if v == it_nama or v == it_variant or (len(v) > 5 and v in it_nama):
                                variant_match = True; break
                    if not variant_match: continue
                    it_gramasi = norm(it.get("gramasi"))
                    gramasi_match = g_all
                    if not gramasi_match:
                        for g in [norm(x) for x in glist]:
                            if "- gramasi -" in g.lower() or "all gramasi" in g.lower() or "bisa meleset" in g.lower():
                                gramasi_match = True; break
                            if g == it_gramasi or (len(g) > 2 and g in it_nama):
                                gramasi_match = True; break
                    if not gramasi_match: continue
                    matched_items.append(it)

            _seen_kb, _dedup = set(), []
            for _it in matched_items:
                _kb = str(_it.get("kode_barang","")).strip()
                if _kb and _kb in _seen_kb: continue
                _seen_kb.add(_kb); _dedup.append(_it)
            matched_items = _dedup

            trig_has_num = has_number(ket)
            trig_qty = parse_number_id(ket) if trig_has_num else ""
            trig_unit = unit_from_text(ket) if trig_has_num else ""
            benefit_text = str(r.get("benefit","") or "").strip()
            benefit_type = str(r.get("benefit_type","") or "").strip()
            benefit_unit = unit_from_text(benefit_text) if benefit_text else ""

            if not matched_items:
                excel_rows.append({"pdf_key": i, "channel": promo_group, "kode_barang": "", "nama_barang": "",
                                    "promo_label": promo_label, "pg_id": promo_group_id, "periode": periode,
                                    "trig_qty": trig_qty, "trig_unit": trig_unit, "benefit_type": benefit_type,
                                    "benefit_text": benefit_text, "benefit_unit": benefit_unit})
                continue

            pdf_items[i].extend(matched_items)

            from collections import defaultdict
            grouped_by_master_kel = defaultdict(list)
            for it in matched_items:
                grouped_by_master_kel[str(it.get("kelompok","")).strip()].append(it)

            for master_kel, items_in_kel in grouped_by_master_kel.items():
                current_pg_id = promo_group_id
                if not current_pg_id or current_pg_id.upper() == "NON_GROUP" or current_pg_id == promo_group:
                    prefix = "".join(e for e in promo_group if e.isalnum())
                    if not prefix: prefix = "Retail"
                    channel_seq[prefix] = channel_seq.get(prefix, 0) + 1
                    current_pg_id = f"{prefix}_{channel_seq[prefix]}"
                for it in items_in_kel:
                    excel_rows.append({"pdf_key": i, "channel": promo_group, "kode_barang": str(it.get("kode_barang","")),
                                        "nama_barang": str(it.get("nama_barang","")), "promo_label": promo_label,
                                        "pg_id": current_pg_id, "periode": periode, "trig_qty": trig_qty,
                                        "trig_unit": trig_unit, "benefit_type": benefit_type, "benefit_text": benefit_text,
                                        "benefit_unit": benefit_unit})

        # ponytail: guard V4 -- 1 kode fisik tidak boleh nyantol di >1 baris-surat (tier beda) dlm
        # channel yg sama (terbukti live: "Pmd Wtr Bas" muncul di baris 4+1 DAN 7+1 sekaligus).
        # Akurasi finansial wajib -> JANGAN menebak salah satu benar, buang dari SEMUA sisi (Excel
        # & PDF) dan wajib direview manusia lewat tombol Laporkan Salah.
        kode_channel_to_pdfkeys: Dict[Tuple[str, str], set] = {}
        for er in excel_rows:
            if er["kode_barang"]:
                kode_channel_to_pdfkeys.setdefault((er["channel"], er["kode_barang"]), set()).add(er["pdf_key"])
        conflicted = {k for k, v in kode_channel_to_pdfkeys.items() if len(v) > 1}
        if conflicted:
            kept_excel_rows = []
            for er in excel_rows:
                ck = (er["channel"], er["kode_barang"])
                if er["kode_barang"] and ck in conflicted:
                    flagged_conflicts.append({"kode_barang": er["kode_barang"], "nama_barang": er["nama_barang"],
                                               "channel": er["channel"], "pg_id": er["pg_id"]})
                else:
                    kept_excel_rows.append(er)
            excel_rows = kept_excel_rows
            conflicted_kodes = {kb for (_, kb) in conflicted}
            for i in pdf_items:
                pdf_items[i] = [it for it in pdf_items[i] if str(it.get("kode_barang","")).strip() not in conflicted_kodes]

        for i in range(len(rows)):
            meta = pdf_meta[i]
            items_in_row = pdf_items.get(i, [])

            seen_kel, kel_order = set(), []
            for it in items_in_row:
                k = str(it.get("kelompok","")).strip()
                if k and k not in seen_kel:
                    seen_kel.add(k); kel_order.append(k)
            # ponytail: JANGAN jatuh ke teks kelompok mentah klaim AI kalau semua item-nya sudah
            # dibuang guard (terbukti live: "BLAGIO HM - PMD KIDZ" halusinasi tetap bocor ke PDF
            # walau Excel-nya sudah bersih -- itu justru meniadakan tujuan guard V3b/V4). Kalau
            # tidak ada item valid sama sekali, tandai jelas utk direview, jangan tampilkan nama
            # yang belum tentu benar.
            if kel_order:
                kelompok_display = format_kelompoks_human_readable(" & ".join(kel_order))
            elif meta.get("kelompok_fallback",""):
                kelompok_display = "(TIDAK ADA ITEM COCOK DI MASTER -- PERLU REVIEW MANUAL)"
            else:
                kelompok_display = ""

            if not items_in_row or norm(meta.get("variant_display","")).replace(" ","") == "ALLVARIANT":
                variant_display = meta.get("variant_display","")
            else:
                seen_v, v_order = set(), []
                for it in items_in_row:
                    v = str(it.get("variant","")).strip()
                    if v and v not in seen_v:
                        seen_v.add(v); v_order.append(v)
                variant_display = format_array_human_readable(" & ".join(v_order)) if v_order else meta.get("variant_display","")

            # guard/fix E: SATU ENTRI GRAMASI PER KELOMPOK BARANG, urutan sesuai kelompok,
            # JANGAN di-dedupe lintas kelompok walau angkanya kebetulan sama (mis. 2 kelompok
            # sama-sama 100ml tetap harus tampil 2x sesuai posisi kelompoknya).
            gramasi_parts = []
            for k in kel_order:
                gram_for_k, seen_g = [], set()
                for it in items_in_row:
                    if str(it.get("kelompok","")).strip() == k:
                        g = str(it.get("gramasi","")).strip()
                        if g and g not in seen_g:
                            seen_g.add(g); gram_for_k.append(g)
                if gram_for_k:
                    gramasi_parts.append(",".join(gram_for_k))
            gramasi_display = join_human(gramasi_parts)

            table_data.append([
                Paragraph(str(meta["no"]), cell_style),
                Paragraph(str(meta.get("surat_program","")), cell_style),
                Paragraph(str(meta.get("nama_program","")), cell_style),
                Paragraph(str(meta.get("channel_gtmt","")), cell_style),
                Paragraph(str(meta.get("periode","")), cell_style),
                Paragraph(kelompok_display, cell_style),
                Paragraph(variant_display, cell_style),
                Paragraph(gramasi_display, cell_style),
                Paragraph(str(meta.get("ketentuan","")), cell_style),
                Paragraph(benefit_display(meta), cell_style),
                Paragraph(str(meta.get("syarat_claim","")), cell_style),
                Paragraph(str(meta.get("keterangan","")), cell_style),
            ])
            
        # Total A4 landscape width is ~842. Margins are 0.5cm each (approx 14 points each, total 28 pts margin)
        # Usable width = 842 - 28 = 814 points
        usable = landscape(A4)[0] - (1 * cm)
        cw = [
            usable * 0.03, # No
            usable * 0.12, # Surat Program
            usable * 0.12, # Nama Program
            usable * 0.05, # Channel
            usable * 0.08, # Periode
            usable * 0.09, # Kelompok
            usable * 0.14, # Variant
            usable * 0.08, # Gramasi
            usable * 0.08, # Ketentuan
            usable * 0.08, # Benefit
            usable * 0.07, # Syarat Claim
            usable * 0.06  # Keterangan
        ]
            
        t = Table(table_data, repeatRows=1, colWidths=cw)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9E7C85')), # Match the brownish pink header color
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        elements.append(t)
        
        # Add the Footer Signatures
        elements.append(Spacer(1, 25))
        
        footer_style_left = styles["Normal"].clone("FooterLeft")
        footer_style_left.fontSize = 8
        footer_style_left.fontName = 'Helvetica-Bold'
        
        footer_style_right = styles["Normal"].clone("FooterRight")
        footer_style_right.fontSize = 8
        footer_style_right.fontName = 'Helvetica-Bold'
        footer_style_right.alignment = TA_RIGHT

        sig_data = [
            [Paragraph(f"Makassar , {dibuat_date}", footer_style_left), ""],
            [Paragraph("Diajukan Oleh,", footer_style_left), Paragraph("Disetujui Oleh,", footer_style_right)],
            [Spacer(1, 40), Spacer(1, 40)], # Space for signature
            [Paragraph("SM<br/>(.................................................)", footer_style_left), 
             Paragraph("OPERATIONAL MANAGER<br/>(.................................................)", footer_style_right)]
        ]
        
        # Table takes up full usable width so left is left, right is right
        sig_table = Table(sig_data, colWidths=[usable/2.0, usable/2.0])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(sig_table)
        
        doc.build(elements, onFirstPage=my_canvas, onLaterPages=my_canvas)

        dataset_path = os.path.join(out_dir, f"{file_id}_Dataset_Diskon_With_Channel.xlsx")
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Dataset"
        headers2 = [
            "KODE_BARANG", "NAMA_BARANG", "PROMO_LABEL", "PROMO_GROUP_ID", "PROMO_GROUP",
            "PERIODE", "PROMO_ACTIVE", "TIER_NO", "TRIGGER_QTY", "TRIGGER_UNIT",
            "BENEFIT_TYPE", "BENEFIT_VALUE", "BENEFIT_UNIT", "BENEFIT_BEBAN"
        ]
        ws2.append(headers2)

        # FASE 4b: override koreksi manusia (stable key kode_barang+channel+no_surat) --
        # menang atas HASIL APA PUN (parser posisional, variant_resolver, atau LLM lama).
        # Key TIDAK PERNAH pakai posisi/index baris -> aman walau urutan OCR run berikutnya beda.
        for _er in excel_rows:
            _er["no_surat"] = pdf_meta.get(_er.get("pdf_key"), {}).get("surat_program", "")
        excel_rows, _correction_log = apply_stable_corrections(excel_rows, load_stable_corrections())

        # excel_rows sudah dihitung SEKALI di atas (dipakai jg utk PDF) -- di sini cuma tulis ke sheet.
        excel_tier_counter: Dict[Tuple[str, str, str, str], int] = {}
        for er in excel_rows:
            tkey = (er["promo_label"], er["pg_id"], er["channel"], er["kode_barang"])
            excel_tier_counter[tkey] = excel_tier_counter.get(tkey, 0) + 1
            ws2.append([
                er["kode_barang"], er["nama_barang"], er["promo_label"], er["pg_id"], er["channel"],
                er["periode"], True, excel_tier_counter[tkey], er["trig_qty"], er["trig_unit"],
                er["benefit_type"], er["benefit_text"], er["benefit_unit"], "PABRIK",
            ])
        wb2.save(dataset_path)
        # FASE 6: paku timestamp entry-zip + core.xml modified -> xlsx byte-identik antar-run
        finalize_xlsx(dataset_path)

        if flagged_mismatches or flagged_conflicts:
            append_error_log("summary_manual_generate_flags", Exception("data quality flags"), {
                "mismatches": flagged_mismatches, "conflicts": flagged_conflicts, "user": user,
            })

        MANUAL_OUTPUTS[file_id] = {"form": form_path, "dataset": dataset_path}

        # FASE 5: golden snapshot. Input baris identik (dok+approval sama) HARUS -> output identik.
        # input_key: urutan baris diabaikan (identitas dok). output_sig: urutan DIPERTAHANKAN
        # (drift urutan pun terdeteksi). status "drift" = regresi non-determinisme -> dilaporkan
        # ke UI, golden TIDAK ditimpa diam2 (butuh approve manual). Gagal apa pun di sini tak
        # boleh menggagalkan generate -> dibungkus try (fitur audit, bukan jalur kritikal output).
        _determinism = None
        try:
            _output_sig = canonical_signature(excel_rows)
            _g = golden_check_and_freeze(_golden_input_key, _output_sig, {"user": user, "file_id": file_id})
            _determinism = _g["status"]
            if _determinism == "drift":
                append_error_log("summary_golden_drift", Exception("output berbeda utk input identik"), {
                    "input_key": _golden_input_key, "golden_sig": _g.get("golden_sig"),
                    "current_sig": _g.get("current_sig"), "user": user,
                })
        except Exception as _ge:
            append_error_log("summary_golden_error", _ge, {"user": user})

        return {"ok": True, "file_id": file_id, "flagged_mismatches": flagged_mismatches,
                "flagged_conflicts": flagged_conflicts, "determinism": _determinism}
    except Exception as e:
        append_error_log("summary_manual_generate", e, {"user": user, "token": token})
        payload = {"ok": False, "error": "Gagal membuat output summary manual."}
        if APP_DEBUG and is_admin_user(user):
            payload["detail"] = str(e)
        return payload

@app.get("/summary/manual/download/{file_id}/{kind}")
@app.get("/summary/manual/download/{file_id}/{kind}/{dummy:path}")
def summary_manual_download(request: Request, file_id: str, kind: str, dummy: str = None):
    # Auth check removed to support direct downloads from Next.js cross-origin links
    # The UUID acts as the access token.
    if file_id not in MANUAL_OUTPUTS:
        return JSONResponse({"ok": False, "error": "File ID tidak ditemukan"}, status_code=404)
    if kind not in ["form","dataset"]:
        return JSONResponse({"ok": False, "error": "Kind harus form/dataset"}, status_code=400)
    path = MANUAL_OUTPUTS[file_id].get(kind)
    if not path or not os.path.exists(path):
        return JSONResponse({"ok": False, "error": "File tidak ditemukan di server"}, status_code=404)
        
    filename = os.path.basename(path)
    
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if kind == "form" and filename.endswith(".pdf"):
        content_type = "application/pdf"
        
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Access-Control-Expose-Headers": "Content-Disposition",
        "Content-Type": content_type
    }
    return FileResponse(path, filename=filename, headers=headers)

@app.post("/api/principles/add")
async def add_principle(request: Request, name: str = Form(...), file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user: return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    pid = str(uuid.uuid4())
    safe_name = "".join(c for c in file.filename if c.isalnum() or c in " ._-")
    filename = f"{pid}_{safe_name}"
    filepath = os.path.join(MASTERS_DIR, filename)
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)
    
    ps = _load_principles()
    ps[pid] = {"name": name, "filename": filename, "uploaded_by": user, "created_at": datetime.date.today().isoformat()}
    _save_principles(ps)
    return {"ok": True, "pid": pid}

@app.post("/api/principles/{pid}/delete")
def delete_principle(request: Request, pid: str):
    user = get_current_user(request)
    if not user: return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    ps = _load_principles()
    if pid in ps:
        filepath = os.path.join(MASTERS_DIR, ps[pid]["filename"])
        if os.path.exists(filepath):
            os.remove(filepath)
        del ps[pid]
        _save_principles(ps)
    return {"ok": True}

@app.post("/api/summary/manual/master/load_principle/{pid}")
def load_principle_master(request: Request, pid: str):
    user = get_current_user(request)
    if not user: return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "summary", "view"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    
    ps = _load_principles()
    if pid not in ps:
        return {"ok": False, "error": "Principle tidak ditemukan."}
        
    filepath = os.path.join(MASTERS_DIR, ps[pid]["filename"])
    if not os.path.exists(filepath):
        return {"ok": False, "error": "File Excel Principle hilang."}
        
    try:
        with open(filepath, "rb") as f:
            file_bytes = f.read()
            
        kelompok_list, variant_map, gramasi_map, items = _parse_master_barang_xlsx(file_bytes)
        k_list = sorted(list(set(str(x.get("kelompok", "")).strip() for x in items if x.get("kelompok"))))
        token = str(uuid.uuid4())
        
        MANUAL_MASTER_CACHE[token] = {
            "expires": time.time() + 7200,
            "items": items,
            "kelompok_list": k_list,
            "variant_map": variant_map,
            "gramasi_map": gramasi_map,
        }
        return {"ok": True, "token": token, "kelompok_list": k_list}
    except Exception as e:
        return {"ok": False, "error": f"Gagal membaca Excel: {str(e)}"}

# --- RESTORED MISSING ROUTES ---

@app.get("/api/job_status/{job_id}")
async def get_job_status(job_id: str, request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    
    job = BACKGROUND_JOBS.get(job_id)
    if not job:
        return JSONResponse(status_code=404, content={"ok": False, "error": "Job not found"})
        
    return JSONResponse(content={"ok": True, "status": job["status"], "result": job.get("result"), "error": job.get("error")})

@app.post("/summary/manual/parse_pdf_regex")
async def summary_manual_parse_pdf_regex(request: Request, token: str = Form(...), pdf: UploadFile = File(...)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token): return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    
    try:
        if token not in MANUAL_MASTER_CACHE: return {"ok": False, "error": "Token master tidak ditemukan / expired"}
        file_bytes = await read_upload_file_limited(pdf, max_bytes=MAX_PDF_UPLOAD_BYTES, allowed_exts=(".pdf",))
        
        import fitz
        import re
        import uuid
        
        pdf_text = ""
        with fitz.open(stream=file_bytes, filetype="pdf") as doc:
            for page in doc: pdf_text += page.get_text() + "\n"

        # ponytail: PDF scan (image-only) -> get_text() kosong -> regex mustahil. Jangan balas 0 baris diam-diam.
        if not pdf_text.strip():
            return {"ok": False, "error": "PDF ini hasil scan (tanpa teks). Regex tidak bisa membaca gambar — gunakan tombol 'Ekstrak Cerdas' (AI/OCR)."}

        rows = []
        # Basic Regex implementation (Fragile, structure-dependent)
        # Looks for lines starting with "PROID-..." and captures nearby context naively
        matches = re.finditer(r"(PROID-[A-Z0-9/\-]+)", pdf_text)
        idx = 1
        for match in matches:
            surat_program = match.group(1)
            rows.append({
                "id": str(uuid.uuid4()),
                "no": str(idx),
                "principle": "Auto (Regex)",
                "surat_program": surat_program,
                "nama_program": "Hasil RegEx Terbatas",
                "channel_gtmt": "MT",
                "kelompok": "Bisa Meleset",
                "variant": "...",
                "gramasi": "...",
                "ketentuan": "Beli XX",
                "benefit_type": "DISC_PCT",
                "benefit": "5%",
                "syarat_claim": "Faktur",
                "keterangan": "Automated OCR/Regex"
            })
            idx += 1
            
        rows = _apply_native_kelompok(rows, MANUAL_MASTER_CACHE[token].get("items", []))
        return {"ok": True, "rows": rows}
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "error": f"Regex Parser Error: {str(e)}"}
def _apply_native_kelompok(rows_to_check, master_items):
    final_rows_out = []
    def norm(x: object) -> str:
        return " ".join(str(x or "").strip().split()).upper()

    def split_list(val: str) -> list:
        import re
        return [x.strip() for x in re.split(r'[,&]', str(val or "")) if x.strip()]

    for r in rows_to_check:
        matched_items = []
        # FASE 3b: resolusi varian data-driven (variant_mapping.json) DULU, sebelum LLM/
        # kode_barangs manapun -- kasus terbukti: "Spray Cologne Series" harus jadi White SR
        # + Black SR (bukan cuma GLASS/salah satu), "EDT Sport" harus 4 varian tertentu
        # (Azzuro/Bianco/Nero/Rosso), bukan tebakan LLM yg terbukti salah/tak konsisten.
        _variant_hit = resolve_variant(str(r.get("kelompok", "")), master_items, _VARIANT_MAPPING)
        if _variant_hit is not None:
            matched_items = _variant_hit
            klist = []  # skip jalur matching lama sepenuhnya utk baris ini
        else:
            klist = [k.strip() for k in str(r.get("kode_barangs", "")).split(",") if k.strip()]
        if klist:
            for it in master_items:
                if str(it.get("kode_barang", "")).strip() in klist:
                    matched_items.append(it)
                    
        # Fallback Match: String matching if AI failed to provide API keys or "ALL VARIANT"
        if not matched_items:            
            kelompok = str(r.get("kelompok","") or "").strip()
            vlist = split_list(r.get("variant",""))
            glist = split_list(r.get("gramasi",""))
            v_all = (not vlist) or any(norm(x) == "ALL VARIANT" for x in vlist)
            g_all = (not glist) or any(norm(x) == "ALL GRAMASI" for x in glist)

            # Treat generic placeholders as empty
            if kelompok and any(skip in kelompok.lower() for skip in ["- kelompok -", "bisa meleset"]):
                kelompok = ""
                
            # Filter pool initially based on Kelompok if provided
            pool = []
            if kelompok:
                pool = [it for it in master_items if norm(it.get("kelompok")) == norm(kelompok)]
            else:
                pool = master_items

            # If pool is empty (hallucinated kelompok), fallback to all items
            if not pool: pool = master_items

            # Now filter the pool based on variants and gramasi
            for it in pool:
                it_variant = norm(it.get("variant"))
                it_nama = norm(it.get("nama_barang"))
                variant_match = False
                if v_all:
                    variant_match = True
                else:
                    for v in [norm(x) for x in vlist]:
                        if "- variant -" in v.lower() or "all variant" in v.lower() or "bisa meleset" in v.lower():
                            variant_match = True; break
                            
                        # Fuzzy Token Matching: Resolve LLM Hallucinations like 'BELLAGIO EAU DE TOILETTE'
                        if v == it_nama or v == it_variant or (len(v) > 5 and v in it_nama):
                            variant_match = True; break
                            
                        # Token overlap with Brand/Variant Aliasing
                        ALIASES = {"BELLAGIO": "BLAGIO", "CASABLANCA": "CSBNCA", "TOILETTE": "EDT", "PARFUME": "EDP", "PARFUM": "EDP"}
                        v_tokens = []
                        for t in v.split():
                            if len(t) >= 4:
                                t_mapped = ALIASES.get(t, t)
                                v_tokens.append(t_mapped)
                        v_tokens = set(v_tokens)
                        
                        db_tokens = set([t for t in it_nama.split() if len(t) >= 4])
                        if v_tokens and len(v_tokens.intersection(db_tokens)) >= 1:
                            variant_match = True; break
                            
                if not variant_match: continue
                    
                it_gramasi = norm(it.get("gramasi"))
                gramasi_match = False
                if g_all:
                    gramasi_match = True
                else:
                    for g in [norm(x) for x in glist]:
                        if "- gramasi -" in g.lower() or "all gramasi" in g.lower() or "bisa meleset" in g.lower():
                            gramasi_match = True; break
                        if g == it_gramasi or (len(g) > 2 and g in it_nama):
                            gramasi_match = True; break
                            
                if not gramasi_match: continue
                    
                matched_items.append(it)
                    
        # Deterministic Kelompok String Builder strictly from Master DB
        # EXPLODE MAGIC: If a single AI row contains items from DIFFERENT Brand Prefixes (Nama KLPs),
        # we must split it into separate rows so the frontend and generator handle them cleanly!
        from collections import defaultdict
        groups_by_prefix = defaultdict(list)
        
        for it in matched_items:
            real_kelompok = str(it.get("kelompok", "")).strip()
            if real_kelompok and real_kelompok.lower() not in ("nan", "null", "none"):
                # Extract prefix to group strictly by Nama KLP
                prefix = real_kelompok.split(" - ")[0] if " - " in real_kelompok else real_kelompok
                groups_by_prefix[prefix].append(it)
                
        if not groups_by_prefix:
            # Fallback for completely unmatched row
            final_rows_out.append(r)
        else:
            # Explode into multiple rows based on the prefix
            import copy
            is_first = True
            for prefix, its in groups_by_prefix.items():
                new_row = r if is_first else copy.deepcopy(r)
                is_first = False
                
                # Assign this prefix's valid kelompoks & matched items
                k_list = []
                k_codes = []
                for it in its:
                    k_name = str(it.get("kelompok", "")).strip()
                    k_str = str(it.get("kode_barang", "")).strip()
                    if k_name and k_name not in k_list: k_list.append(k_name)
                    if k_str and k_str not in k_codes: k_codes.append(k_str)
                    
                new_row["kelompok"] = " & ".join(k_list)
                new_row["kode_barangs"] = ",".join(k_codes)
                new_row["_matched_items_cache"] = its
                
                # Let's ensure no ID collisions if exploded
                if not is_first:
                     import uuid
                     new_row["id"] = str(uuid.uuid4())
                     
                final_rows_out.append(new_row)
                
    return final_rows_out

@app.post("/summary/manual/parse_pdf_ai")
async def summary_manual_parse_pdf_ai(request: Request, token: str = Form(...), pdf: UploadFile = File(...), n8n_webhook: str = Form(default=""), principle_name: str = Form(default=""), ai_mode: str = Form(default="split")):
    
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "summary", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    
    try:
        if token not in MANUAL_MASTER_CACHE:
            return {"ok": False, "error": "Token master tidak ditemukan / expired"}
            
        file_bytes = await read_upload_file_limited(pdf, max_bytes=MAX_PDF_UPLOAD_BYTES, allowed_exts=(".pdf",), label="PDF Program")

        # FASE 1b: hasil parse dibekukan per (dokumen, principle). Hit -> lewati OCR+LLM total
        # (0 biaya API) & rows IDENTIK dgn run pertama -> pipeline dok->rows deterministik penuh.
        _parse_key = parse_cache_key(file_bytes, principle_name)
        _cached_rows = parse_cache_get(_parse_key)
        if _cached_rows is not None:
            return {"ok": True, "rows": _cached_rows}

        # Direct Python native parsing via OpenAI SDK (Sumopod)
        api_key = os.getenv("SUMOPOD_API_KEY") or os.getenv("OPENAI_API_KEY")
        if not api_key:
             return {"ok": False, "error": "API Key belum dikonfigurasi. Pastikan SUMOPOD_API_KEY atau OPENAI_API_KEY ada di file .env."}
             
        import fitz
        from openai import AsyncOpenAI
        import json
        import uuid
        import re
        
        # Fetch Official DB Product Names to ground the AI's deductions
        cache = MANUAL_MASTER_CACHE[token]
        raw_items = cache.get("items", [])
        
        # SANGAT PENTING: Filter context supaya API Proxy tidak memuntah/terpotong (Tokens Limit)
        # Hanya gunakan barang yang sesuai dengan Principle yang sedang diproses!
        if principle_name and principle_name.strip():
            # ponytail: dulu dicocokkan ke kolom 'principle' (nama produk) -> "Priskila" tak pernah match -> 0.
            # Cocokkan tiap keyword principal (>=4 huruf) ke principle ATAU nama_barang; kosong -> pakai semua
            # (master per-principal itu normal; kolom 'Nama Pcpl' memang blank).
            _kw = [w for w in re.sub(r"\(.*?\)", "", principle_name).upper().split() if len(w) >= 4]
            def _match_principal(it):
                blob = (str(it.get("principle", "")) + " " + str(it.get("nama_barang", ""))).upper()
                return any(w in blob for w in _kw)
            items = [it for it in raw_items if _match_principal(it)] if _kw else raw_items
            if not items:
                items = raw_items # Fallback: master ini memang katalog 1 principal
        else:
            items = raw_items
        
        # Bikin mapping text yang panjang tapi detail
        item_names_cache = set()
        kode_barang_map = {}
        for item in items:
            name = str(item.get("nama_barang", "")).strip().upper()
            code = str(item.get("kode_barang", "")).strip()
            
            if name: item_names_cache.add(name)
            
            # Kita map nama ke array kode karena 1 nama bisa banyak gramasi
            if name not in kode_barang_map: kode_barang_map[name] = []
            
            if code and code not in kode_barang_map[name]: kode_barang_map[name].append(code)
            
        master_names_context = ""
        for n, kodes in kode_barang_map.items():
            s_kodes = ",".join(kodes)
            
            # Cari baris yang bener-bener punya nama barang ini
            for master_item in items:
                nama_barang = str(master_item.get("nama_barang", "")).strip().upper()
                nama_principle = str(master_item.get("principle", "")).strip().upper()
                
                # Dynamic matching for Aroma / Variant
                nama_aroma = ""
                for k, v in master_item.items():
                    if "aroma" in str(k).lower() or "rasa" in str(k).lower() or "variant" in str(k).lower():
                        nama_aroma = str(v).strip()
                        break
                        # Fetch pre-combined Kelompok string from Master cache
                kelompok_asli = str(master_item.get("kelompok", "")).strip()
                
                # Format: REF: [Principle] - [Nama Barang] -> OUTPUT_KELOMPOK: [Kelompok Asli] | OUTPUT_VARIANT: [Variant] | OUTPUT_KODE: [Kode]
                if nama_barang == n:
                    master_names_context += f"REF: {nama_principle} - {nama_barang} -> OUTPUT_KELOMPOK: {kelompok_asli} | OUTPUT_VARIANT: {nama_aroma} | OUTPUT_KODE: {s_kodes}\n"
                    break
        
        # Fetch Master Customers if available
        db_customers = cache.get("customers", [])
        customer_names = sorted([f"{c.get('kode_customer','')} | {c.get('nama_customer','')}" for c in db_customers])
        master_customers_context = "\n".join(customer_names) if customer_names else "TIDAK ADA DATA CUSTOMER"
        
        try:
            with open("d:/disc_web/debug_ai_context.txt", "w", encoding="utf-8") as _f:
                _f.write(master_names_context)
        except: pass
        
        try:
            import httpx
            import base64
            # 1. Extract pure text and highly compressed images from PDF
            pdf_text = ""
            base64_images = []
            
            with fitz.open(stream=file_bytes, filetype="pdf") as doc:
                # ponytail: dulu doc[:10] -> halaman 11+ hilang diam-diam. Proses semua sampai cap longgar.
                MAX_OCR_PAGES = int(os.getenv("SUMMARY_MAX_OCR_PAGES", "40"))
                pages_total = doc.page_count
                for page in doc[:MAX_OCR_PAGES]:
                    txt = page.get_text()
                    pdf_text += txt + "\n"
                    # Compress the image so the Base64 string doesn't eat the token budget
                    pix = page.get_pixmap(matrix=fitz.Matrix(1.0, 1.0))
                    img_bytes = pix.tobytes("jpeg", 80)
                    b64 = base64.b64encode(img_bytes).decode("utf-8")
                    base64_images.append(b64)
            pages_truncated = pages_total > len(base64_images)
                    
            if not pdf_text.strip() and not base64_images:
                return {"ok": False, "error": "PDF kosong atau tidak memiliki halaman valid."}

            # 2. Call Sumopod proxy directly via HTTP POST
            prompt = f"""
SANGAT PENTING: Dokumen promosi ini berkaitan dengan Brand / Keluarga Produk: {principle_name.upper()}.
Meskipun teks di PDF mungkin buram atau terpotong, JIKA ada kemiripan, Anda WAJIB memprioritaskan penyocokan kode barang dengan nama-nama resmi yang mengandung kata kunci brand ini!

Tugas Anda: EKSTRAK SEMUA TABEL PROMO/DISKON dari dokumen ini ke dalam ARRAY JSON.
KEMBALIKAN HASILNYA SAJA DALAM FORMAT JSON VALID! (JANGAN ada teks pembuka/penutup).

ATURAN REASONING & EKSTRAKSI (WAJIB DIIKUTI 100%):
1. COST RATIO (CR): Kolom 'CR' / 'Cost Ratio' HANYA angka referensi internal (BUKAN benefit/diskon) -- ABAIKAN KOLOM ITU SAJA. JANGAN abaikan kolom LAIN di tabel yang sama hanya karena tabel tersebut JUGA punya kolom CR! Kolom 'CUT PRICE' / 'HET' / 'PAKET' di tabel yang sama TETAP WAJIB diekstrak sebagai benefit (lihat aturan 2 & 3) -- CR cuma 1 kolom yang diabaikan, bukan alasan mengosongkan seluruh baris/tabel.
2. BONUS QTY (Beli X Gratis Y): JIKA mekanismenya memberikan gratis barang (misal: "Beli 2 gratis 1"), isi 'benefit_type' dengan "BONUS_QTY". Nilai 'benefit' adalah jumlah barang gratisnya (1).
   ATURAN FORMAT "X+Y" (SANGAT PENTING - kolom PAKET biasa tertulis "7+1", "4+1", "10+2", "65+7"):
   Angka ini BUKAN penjumlahan! "7+1" artinya BELI 7 GRATIS 1. Maka: 'ketentuan' = "Beli 7" (angka KIRI saja, JANGAN dijumlah jadi "Beli 8"!), 'benefit_type' = "BONUS_QTY", 'benefit' = "1 PCS" (angka KANAN + satuan). Contoh: "10+2" -> ketentuan "Beli 10", benefit "2 PCS". "65+7" -> ketentuan "Beli 65", benefit "7 PCS". DILARANG KERAS menjumlahkan kiri+kanan.
3. POTONGAN HARGA (Cut Price): Isi 'benefit_type' dengan "DISC_RP" dan 'benefit' angkanya SAJA (TANPA huruf "Cut Price" atau "Potongan"). JANGAN SEKALI-KALI MENGGANTI NILAI `ketentuan` (Trigger Beli) DENGAN TEKS POTONGAN HARGA INI! `ketentuan` WAJIB TETAP BERISI "Beli 1", "Beli 2", dll.
4. KETENTUAN TRIGGER QTY: Jika di surat tertulis "Setiap pembelian", "Setiap pengambilan", ATAU "TIDAK ADA ANGKA MINIMAL", WAJIB ubah teks 'ketentuan' menjadi "Beli 1".
4b. TABEL FORMAT "CUT PRICE" TANPA TEKS TRIGGER SAMA SEKALI (channel MTI/Modern Trade biasanya
    begini -- WAJIB tetap diekstrak, JANGAN dikosongkan/dilewati hanya karena tidak ada frasa "Beli X"):
    Kalau tabel HANYA berisi kolom seperti "CUT PRICE | HET | CR" tanpa kolom "PAKET" dan tanpa kalimat
    trigger apa pun, maka SETIAP baris tabel = 1 promo terpisah dengan 'ketentuan'="Beli 1",
    'benefit_type'="DISC_RP", 'benefit'=angka di kolom CUT PRICE SAJA (tanpa titik/koma ribuan diubah
    jadi angka polos). ABAIKAN kolom HET dan CR (bukan benefit). WAJIB ekstrak SEMUA baris/brand
    sampai baris TERAKHIR tabel, walau tabelnya panjang dan tidak ada kalimat pemicu di setiap barisnya.
    CONTOH KONKRET (WAJIB DIIKUTI POLA INI):
      Input tabel: "| BELLAGIO | Bellagio Eau de Toilette 100ml | 4,700 | 31,628 | 13% |"
      Output JSON: {{"kelompok": "Bellagio Eau de Toilette 100ml", "ketentuan": "Beli 1",
                    "benefit_type": "DISC_RP", "benefit": "4700"}}
      Input tabel: "|          | Bellagio Pomade Kidz 40gr        | 1,400 | 12,600 | 10% |"
      Output JSON: {{"kelompok": "Bellagio Pomade Kidz 40gr", "ketentuan": "Beli 1",
                    "benefit_type": "DISC_RP", "benefit": "1400"}}
      (Catatan: "Pomade Kidz" BEDA dari "Pomade" biasa -- keduanya bisa SAMA-SAMA ada di channel yang
      sama dengan harga cut price berbeda, JANGAN dianggap duplikat/salah satu dibuang.)
5. CHANNEL PROMO: Isi 'channel_gtmt' dengan NAMA ASLI channel sesuai di surat.
6. ATURAN PEMISAHAN & PENGGABUNGAN MEREK (MUTLAK - PROMPT EXPLODER):
   - HANYA BOLEH GABUNGKAN item-item promo ke dalam 1 baris JSON APABILA mereka memiliki MEREK UTAMA (Brand Keluarga) yang sama 100%. (Misal: Sesama Bellagio Homme boleh digabung).
     * JIKA Anda menggabungkan beberapa produk/varian/gramasi ke dalam 1 baris (karena mereknya sama), MAKA string `ketentuan` WAJIB ditambah " Boleh Mix Kelompok dan Gramasi Barang Sama" di akhir teks! (Contoh: "Beli 7 Boleh Mix Kelompok dan Gramasi Barang Sama").
   - JIKA dalam satu tabel/promo dokumen PDF mencakup beberapa MEREK UTAMA yang berbeda (Misal: "Bellagio" dan "Camellia" mendapat promo diskon yang sama), ANDA DILARANG KERAS menggabungkannya ke dalam 1 object array JSON!
   - ANDA WAJIB MENDUPLIKASI / MEMECAH (EXPLODE) promo tersebut menjadi beberapa baris object JSON yang terpisah secara independen!
     * JSON Object 1: KHUSUS berisi kelompok "Bellagio" dengan `kode_barangs` yang HANYA milik Bellagio.
     * JSON Object 2: KHUSUS berisi kelompok "Camellia" dengan `kode_barangs` yang HANYA milik Camellia.
     * (Keduanya memiliki isi ketentuan, benefit_type, dan benefit yang sama dari hasil duplikasi. JANGAN LUPA tambahkan " Boleh Mix Kelompok dan Gramasi Barang Sama" pada masing-masing baris jika di dalamnya masih merupakan gabungan varian dari merek tersebut).
   - Ingat: 1 Object JSON = MAKSIMAL 1 MEREK UTAMA (KELOMPOK)! Jangan pernah ada penggabungan silang brand di kolom `kelompok` atau `kode_barangs`!
7. TIERING PROMO: Beda 'Ketentuan' (trigger qty) = baris JSON harus dipisah! (e.g., Beli 1 diskon 5%, Beli 10 diskon 10% -> 2 baris json).
8. PARTISI ITEM PER PAKET (MUTLAK - PENYEBAB UTAMA KESALAHAN!):
   Di dalam SATU merek, TIAP baris GROUP ITEM punya nilai PAKET/CUT PRICE-nya SENDIRI. Kamu WAJIB mengelompokkan item berdasarkan nilai paket yang PERSIS SAMA, lalu buat 1 baris JSON per nilai paket.
   - SATU item (kode_barang) hanya boleh masuk ke SATU baris JSON -- yaitu baris dengan paket yang sesuai barisnya di surat. DILARANG KERAS memasukkan item yang sama ke lebih dari satu baris ketentuan!
   - CONTOH BENAR (Bellagio, channel Retail):
       * "Bellagio Eau de Toilette 100ml"=7+1 dan "Bellagio EDP Prestige 50ml"=7+1 -> 1 baris: ketentuan "Beli 7", kode HANYA kedua item itu.
       * "Bellagio Roll On 50ml"=4+1, "Bellagio EDP 50ml"=4+1, "Bellagio Pomade 80gr"=4+1, "Bellagio Clay 90gr"=4+1, "Bellagio Body Spray 80ml"=4+1 -> 1 baris TERPISAH: ketentuan "Beli 4", kode HANYA kelima item itu.
   - CONTOH SALAH (JANGAN LAKUKAN): menaruh SEMUA item Bellagio ke baris "Beli 7" DAN juga ke baris "Beli 4". Item EDT 100ml TIDAK boleh muncul di baris Beli 4, dan Roll On TIDAK boleh muncul di baris Beli 7.
   - Jadi jumlah `kode_barangs` gabungan dari semua baris 1 merek = TEPAT sama dengan jumlah item merek itu di surat (tidak ada item dobel lintas baris).

=== DAFTAR REFERENSI BARANG ===
{master_names_context}
=== AKHIR DAFTAR REFERENSI BARANG ===

=== DAFTAR DATA CUSTOMER (KODE | NAMA) ===
{master_customers_context}
=== AKHIR DAFTAR DATA CUSTOMER ===

TUGAS PENCOCOKAN KEYWORD DAN KODE (ATURAN MUTLAK!):
Ubah logika pencarianmu dari Exact Match menjadi Keyword Mapping cerdas!
Saat dokumen PDF menyebutkan nama/varian barang (misal: 'Bellagio Eau de Toilete'), silakan cari baris 'REF:' yang paling relevan di "DAFTAR REFERENSI BARANG" di atas berdasarkan MEREK dan JENISNYA. 
Catat KESELURUHAN angka `Kode Barang` (-kode angka) dari referensi yang cocok tersebut dan gabungkan dengan koma di `kode_barangs`.

ATURAN PENGISIAN PROPERTI JSON (HURUF KECIL):
- "principle": (String) Nama Perusahaan
- "surat_program": (String) Nomor surat program
- "nama_program": (String) Nama Promo / Program
- "promo_group_id": (String) Isi NON_GROUP jika channel umum. Isi KODE CUSTOMER (C-XXX) jika ini adalah program khusus OUTLET/Toko tertentu.
- "channel_gtmt": (String) Nama Spesifik Channel (Misal: Retail, MTI, Star Outlet).
- "periode": (String) Ekstrak periode dari surat (misal "Februari 2024").
- "kelompok": (String) Jika nama kelompok tidak spesifik, JIBLAK EXACT dari `OUTPUT_KELOMPOK` referensi. Jika tidak ada referensi, isi string kosong "".
- "variant": (String) ATURAN MUTLAK: Jika surat program menyebut semua tipe/wangi, WAJIB isi dengan 'All Variant'.
- "gramasi": (String) Gramasi/volume LENGKAP DENGAN SATUAN persis seperti di surat (mis. "22ml", "100gr", "50ml", "80gr"). JANGAN buang satuannya. Kalau ada beberapa, pisah koma.
- "kode_barangs": (String) Angka Kode Barang dari `OUTPUT_KODE`. Pisahkan koma jika > 1.
- "ketentuan": (String) Syarat Beli (Misal "Beli 7"). JIKA PROMO BERLAKU UNTUK GABUNGAN VARIAN/GRAMASI, WAJIB tambahkan kalimat " Boleh Mix Kelompok dan Gramasi Barang Sama" di akhir teks! (Contoh: "Beli 7 Boleh Mix Kelompok dan Gramasi Barang Sama").
- "benefit_type": (String) DISC_RP, DISC_PCT, atau BONUS_QTY
- "benefit": (String) KHUSUS BONUS QTY (Brg fisik), WAJIB TULIS SATUAN (Misal "1 PCS" / "1 Grt"). Jika DISC_RP/PCT biarkan angkanya saja.
- "syarat_claim": (String) KOSONGKAN SAJA
- "keterangan": (String) KOSONGKAN SAJA

SANGAT PENTING: JANGAN BERIKAN TEKS APAPUN SELAIN JSON ARRAY VALID! PASTIKAN JSON DITUTUP SEMPURNA DENGAN `]` PADA AKHIRNYA!
"""
            # ponytail: "AI learning" dari koreksi manual user (tombol Laporkan Salah) -- bukan fine-tune
            # model, tapi few-shot: inject before->after koreksi lama ke prompt supaya kesalahan yg sama
            # tidak terulang utk principal yg sama.
            prompt += _format_corrections_for_prompt(_load_corrections(principle_name))

            # Prepare multimodal payload for Gemini 2.5 Flash
            import httpx
            import json
            import uuid
            
            all_rows = []
            
            async with httpx.AsyncClient(timeout=300.0) as client_http:
                # ==========================
                # SPLIT MODE LOGIC
                # ==========================
                if True: # Split Mode (Gemini OCR per-halaman + deepseek JSON parse)
                    # --- Phase 1: OCR PER-HALAMAN ---
                    # ponytail: dulu 1 request untuk SEMUA gambar -> mentok max_tokens (finish_reason=length)
                    # dan buang ~12% teks di dok 9 hlm (terbukti live). Per-halaman bikin tiap call finish=stop.
                    # Model OCR: gemini/gemini-2.5-flash (kualitas OCR dokumen unggul, dipilih user).
                    # Parse JSON pakai gpt-4.1-mini (non-reasoning, murah). mimo-v2.5/deepseek = reasoning
                    # (risiko token kebakar di halaman padat), gpt-4.1-mini juga vision kalau perlu fallback.
                    ocr_prompt = "Tugas Anda adalah melakukan OCR (Optical Character Recognition). Baca gambar halaman dokumen ini. Ekstrak SELURUH teks di dalamnya persis seperti aslinya, baris demi baris, tabel demi tabel. JANGAN diringkas, JANGAN ada kata atau angka yang terlewat sekecil apapun!"
                    headers = {
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json"
                    }
                    ocr_model = os.getenv("SUMOPOD_OCR_MODEL", "gemini/gemini-2.5-flash")
                    # FASE 1: OCR cache by content hash -- surat byte-identik TIDAK di-OCR ulang
                    # (fondasi determinisme: run ke-2 ambil teks beku, Gemini 0 panggilan).
                    from ocr_cache import ocr_cache_key, ocr_cache_get, ocr_cache_put
                    _doc_hash = ocr_cache_key(file_bytes)
                    _cached_ocr = ocr_cache_get(_doc_hash)
                    ocr_chunks = []
                    for _pg_idx, b64 in enumerate([] if _cached_ocr is not None else base64_images):
                        ocr_payload = {
                            "model": ocr_model,
                            "messages": [
                                {"role": "system", "content": "You are an expert Data Entry and OCR assistant."},
                                {"role": "user", "content": [
                                    {"type": "text", "text": ocr_prompt},
                                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                                ]}
                            ],
                            "temperature": 0.1,
                            "max_tokens": 8192
                        }
                        # ponytail: Gemini kadang balas 503 "high demand, usually temporary" -> retry
                        # singkat dgn backoff sebelum nyerah ke placeholder (terbukti live: 4/9 hlm gagal
                        # tanpa retry -> combined_text penuh placeholder -> parser downstream balik kosong).
                        _pg_text = None
                        _last_ocr_err = None
                        for _attempt in range(3):
                            try:
                                ocr_resp = await client_http.post("https://ai.sumopod.com/v1/chat/completions", json=ocr_payload, headers=headers)
                                if ocr_resp.status_code in (503, 429) and _attempt < 2:
                                    append_error_log("gemini_ocr_retry", Exception(f"HTTP {ocr_resp.status_code} hlm {_pg_idx+1} attempt {_attempt+1}"), {"text": ocr_resp.text[:300]})
                                    await asyncio.sleep(3 * (_attempt + 1))
                                    continue
                                if ocr_resp.status_code != 200:
                                    append_error_log("gemini_ocr_error", Exception(f"HTTP {ocr_resp.status_code} hlm {_pg_idx+1}"), {"text": ocr_resp.text})
                                    ocr_resp.raise_for_status()
                                _pg_text = ocr_resp.json().get("choices", [{}])[0].get("message", {}).get("content", "").strip()
                                # ponytail: halaman minim-konten (mis. tanda tangan) kadang bikin model OCR
                                # nyasar ke loop repetisi karakter tanpa henti (terbukti live: 1 halaman jadi
                                # >120rb karakter underscore). Deteksi & retry spt 503, biar tak membengkakkan
                                # chunk lain yang menempel di akhir teks.
                                _degenerate = bool(re.search(r'(.)\1{199,}', _pg_text))
                                if _degenerate and _attempt < 2:
                                    append_error_log("gemini_ocr_degenerate", Exception(f"repetition loop hlm {_pg_idx+1} attempt {_attempt+1}"), {"len": len(_pg_text)})
                                    _pg_text = None
                                    await asyncio.sleep(3 * (_attempt + 1))
                                    continue
                                if _degenerate:
                                    # percobaan terakhir masih degenerate -> potong sebelum repetisi mulai,
                                    # simpan konten valid yg sempat terbaca drpd buang seluruh halaman
                                    _pg_text = re.split(r'(.)\1{199,}', _pg_text, maxsplit=1)[0].strip() or None
                                break
                            except Exception as _ocr_e:
                                _last_ocr_err = _ocr_e
                                if _attempt < 2:
                                    await asyncio.sleep(3 * (_attempt + 1))
                                    continue
                        if _pg_text is not None:
                            ocr_chunks.append(f"--- HALAMAN {_pg_idx+1} ---\n{_pg_text}")
                        else:
                            append_error_log("gemini_ocr_page_fail", _last_ocr_err or Exception("unknown"), {"page": _pg_idx + 1})
                            ocr_chunks.append(f"--- HALAMAN {_pg_idx+1} (OCR GAGAL setelah 3x percobaan) ---")
                    if _cached_ocr is not None:
                        ocr_text = _cached_ocr.get("ocr_text", "")  # cache hit: teks OCR beku, Gemini tak dipanggil
                    else:
                        ocr_text = "\n\n".join(ocr_chunks)
                        ocr_cache_put(_doc_hash, getattr(pdf, "filename", "") or "", ocr_text, pages_total, len(base64_images))
                    combined_text = pdf_text + "\n\n=== HASIL OCR DARI GAMBAR ===\n\n" + ocr_text
                    if pages_truncated:
                        combined_text += f"\n\n[PERINGATAN: surat {pages_total} halaman, hanya {len(base64_images)} halaman pertama diproses. Naikkan SUMMARY_MAX_OCR_PAGES.]"

                    # --- Phase 2: Parsing JSON PER-CHANNEL ---
                    # ponytail: 1 panggilan utk SELURUH dokumen kehabisan max_tokens di tengah channel
                    # pertama (Retail) krn format JSON per-SKU sangat verbose (puluhan kode per baris) ->
                    # channel MTI/Grosir/Star Outlet tak pernah ke-generate meski JSON yg dihasilkan valid
                    # (terbukti live: qwen3.6-flash & gpt-4.1-mini keduanya berhenti persis stlh Retail).
                    # Fix: pecah teks per section "N. ... CHANNEL ..." lalu parse tiap bagian terpisah
                    # (pola sama dgn fix OCR per-halaman) -> tiap panggilan dapat budget token penuh.
                    import re
                    _hdr_re = re.compile(r'^\s*\d+\.\s*.{0,80}?channel', re.IGNORECASE | re.MULTILINE)
                    _hdrs = list(_hdr_re.finditer(combined_text))
                    if len(_hdrs) >= 2:
                        _preamble = combined_text[:_hdrs[0].start()]
                        channel_chunks = [
                            combined_text[_hdrs[i].start(): _hdrs[i + 1].start() if i + 1 < len(_hdrs) else len(combined_text)]
                            for i in range(len(_hdrs))
                        ]
                    else:
                        _preamble = ""
                        channel_chunks = [combined_text]

                    # ponytail: guard V1/D -- deteksi TERBUKTI SALAH kalau pakai regex baris-tabel
                    # markdown (rapuh, gagal kalau OCR format tabelnya beda per run -- chunk MTI live
                    # sempat balik expected_brands KOSONG shg guard lolos trivial padahal brand hilang
                    # total). Ganti: alias brand (ejaan surat vs abrev master, spt "BELLAGIO"->"BLAGIO"
                    # yg sudah dipakai di _apply_native_kelompok) + substring BEBAS di teks mentah,
                    # BUKAN bergantung struktur tabel -- tahan terhadap variasi format OCR apa pun.
                    _BRAND_ALIASES = {
                        "BLAGIO": ["BLAGIO", "BELLAGIO"],
                        "CAMELLIA": ["CAMELLIA"],
                        "CSBNCA": ["CSBNCA", "CASABLANCA"],
                        "EXCELO": ["EXCELO", "EXCELLO"],
                        "MARIE JOSE": ["MARIE JOSE", "MARIE-JOSE", "MARIEJOSE"],
                        "REGAZZA": ["REGAZZA", "REGZZA", "REGGAZZA"],
                    }
                    def _extract_expected_brands(_chunk_text_for_ai):
                        _upper = _chunk_text_for_ai.upper()
                        return {p for p, aliases in _BRAND_ALIASES.items() if any(a in _upper for a in aliases)}

                    async def _parse_json_chunk(_chunk_text_for_ai, _label):
                        _parsing_prompt = f"{prompt}\n\n====================\nBERIKUT ADALAH TEKS DOKUMEN PROMO ({_label}):\n{_chunk_text_for_ai}\n\n====================\nPENTING: Ekstrak tabel dari teks di atas dan KEMBALIKAN ARRAY JSON SEKARANG JUGA DIAWALI DENGAN SIMBOL '['. JANGAN TULIS HAL LAIN."
                        _payload = {
                            "model": os.getenv("SUMOPOD_MODEL", "gpt-4.1-mini"),
                            "messages": [
                                {"role": "system", "content": "You are a STRICT data extraction AI. You ONLY speak in valid JSON array format starting with '['. You NEVER output regular text, markdown, or greetings. You MUST obey the schema."},
                                {"role": "user", "content": _parsing_prompt}
                            ],
                            "temperature": 0.1,
                            # ponytail: HANYA "max_tokens" -- OpenAI/litellm menolak 400 kalau max_tokens &
                            # max_completion_tokens dikirim BERSAMAAN (terbukti live dgn gpt-4.1-mini).
                            "max_tokens": int(os.getenv("SUMMARY_PARSE_MAX_TOKENS", "16000"))
                        }
                        _full_raw = ""
                        for _loop_idx in range(5):
                            _resp = await client_http.post("https://ai.sumopod.com/v1/chat/completions", json=_payload, headers=headers)
                            if _resp.status_code != 200:
                                append_error_log("claude_400_debug", Exception(f"HTTP {_resp.status_code} chunk={_label}"), {"text": _resp.text})
                            _resp.raise_for_status()
                            _chunk_msg = _resp.json().get("choices", [{}])[0].get("message", {}).get("content", "").strip()
                            _clean = re.sub(r"```json", "", _chunk_msg, flags=re.IGNORECASE)
                            _clean = re.sub(r"```", "", _clean)
                            _overlap = 0
                            if _full_raw:
                                for _i in range(min(100, len(_full_raw), len(_clean)), 0, -1):
                                    if _full_raw[-_i:] == _clean[:_i]:
                                        _overlap = _i
                                        break
                            _full_raw += _clean[_overlap:]
                            if _full_raw.strip().endswith("]"):
                                break
                            if not _chunk_msg:
                                break
                            _payload["messages"].append({"role": "assistant", "content": _chunk_msg})
                            _payload["messages"].append({"role": "user", "content": "Teks JSON terpotong karena batas token! WAJIB lanjutkan string JSON di atas TEPAT mulai dari huruf/simbol yang terputus tanpa basa-basi pengantar, tanpa markdown ```json. Langsung sambung karakternya!"})
                        return _full_raw.strip()

                    all_rows = []
                    _debug_dump = []
                    for _ci, _chunk in enumerate(channel_chunks):
                        # ponytail: preamble berisi No surat (mis. "002/PPM/NSPM/III/2026") -> WAJIB ikut
                        # ke SEMUA chunk termasuk chunk pertama (dulu _ci>0 bikin Retail kosong surat_program).
                        _chunk_full = (_preamble + "\n\n" + _chunk) if _preamble else _chunk
                        _label = f"bagian {_ci + 1}/{len(channel_chunks)}"
                        _chunk_rows = []
                        # ponytail: 2 kegagalan berbeda yg pernah kejadian live: (a) model balas [] valid
                        # padahal chunk jelas ada datanya (non-deterministik), (b) model balas SEBAGIAN
                        # brand saja lalu berhenti (MTI: cuma 2 dari 6 brand tertangkap, JSON tetap valid
                        # jadi tidak "kosong"). Guard emptiness SAJA tidak menangkap (b) -- tambah cek
                        # kelengkapan brand (dari kolom BRAND tabel OCR, generik, tidak spesifik principal),
                        # retry sampai lengkap, simpan percobaan TERBAIK (paling sedikit brand hilang).
                        _expected_brands = _extract_expected_brands(_chunk_full)
                        _best_rows, _best_missing = [], None
                        for _retry in range(6):  # ponytail: akurasi wajib > biaya -- retry lebih banyak sampai brand lengkap
                            _raw = await _parse_json_chunk(_chunk_full, _label)
                            _debug_dump.append(f"=== CHUNK {_label} (percobaan {_retry + 1}) ===\n{_raw}\n")
                            if not _raw:
                                append_error_log("chunk_empty_response", Exception("No text returned"), {"chunk": _label, "attempt": _retry + 1, "user": user})
                                continue
                            _match = re.search(r"\[.*\]", _raw, re.DOTALL)
                            _clean_text = _match.group(0).strip() if _match else _raw.strip()
                            if _clean_text.startswith("[") and not _clean_text.endswith("]"):
                                _last_brace = _clean_text.rfind("}")
                                if _last_brace != -1:
                                    _clean_text = _clean_text[:_last_brace + 1] + "\n]"
                            try:
                                _chunk_data = json.loads(_clean_text, strict=False)
                            except json.JSONDecodeError:
                                append_error_log("chunk_invalid_json", Exception("Non-JSON payload"), {"chunk": _label, "attempt": _retry + 1, "raw": _raw[:500]})
                                continue
                            if isinstance(_chunk_data, list):
                                _chunk_rows = _chunk_data
                            elif isinstance(_chunk_data, dict) and "rows" in _chunk_data:
                                _chunk_rows = _chunk_data["rows"]
                            if not _chunk_rows:
                                continue
                            # ponytail: cek kelengkapan brand HARUS pakai kelompok HASIL MATCHING ke
                            # master (yg sungguhan masuk Excel/PDF), BUKAN teks tebakan AI mentah --
                            # terbukti live keduanya bisa berbeda (raw text bilang "hilang" padahal
                            # setelah di-match ke master brand-nya sebenarnya ketemu, atau sebaliknya).
                            import copy as _copy
                            try:
                                _resolved_probe = _apply_native_kelompok(_copy.deepcopy(_chunk_rows), items)
                            except Exception:
                                _resolved_probe = _chunk_rows
                            _got_text = " ".join(str(_r.get("kelompok","")) + " " + str(_r.get("principle","")) for _r in _resolved_probe).upper()
                            _missing = [b for b in _expected_brands if b not in _got_text]
                            if _best_missing is None or len(_missing) < len(_best_missing):
                                _best_rows, _best_missing = _chunk_rows, _missing
                            if not _missing:
                                break
                        if _best_missing:
                            append_error_log("chunk_incomplete_brands", Exception("Brand tidak lengkap setelah retry"), {"chunk": _label, "missing_brands": _best_missing, "user": user})
                        all_rows.extend(_best_rows)

                    try:
                        with open(os.path.join(BASE_DIR, "data", "debug_ai.txt"), "w", encoding="utf-8") as f:
                            f.write(f"=== OCR PHASE 1 ===\n{ocr_text if 'ocr_text' in locals() else 'N/A'}\n\n=== JSON PHASE 2 (per-channel) ===\n" + "\n".join(_debug_dump))
                    except Exception:
                        pass

                    if not all_rows:
                        return {"ok": False, "error": "AI tidak menemukan tabel promo valid di dalam dokumen, atau gagal mengekstrak."}

                    for idx, row in enumerate(all_rows):
                        if "id" not in row:
                            row["id"] = str(uuid.uuid4())
                        if "no" not in row:
                            row["no"] = str(idx + 1)

                    # TAHAP 2: Native Master DB Mapping (Injects Kelompok perfectly)
                    all_rows = _apply_native_kelompok(all_rows, items)

                    # FASE 2b: regroup baris berdasarkan tier OTORITATIF dari tabel OCR (bukan LLM) --
                    # kode_barang yg terbukti (keyakinan tinggi) py trigger/benefit sama digabung jadi
                    # 1 baris (kasus nyata: Bellagio EDT & EDP Prestige ke-split LLM padahal 7+1 sama).
                    # Kode yg tak ter-bridge dgn keyakinan tinggi TIDAK disentuh (aman, no silent guess).
                    all_rows, _tier_regroup_log = regroup_rows_by_tier(all_rows, items, ocr_text)

                    # FASE 1b: bekukan rows hasil parse (freeze-on-first-write) -> run berikut
                    # dok+principle sama pakai ini, tanpa OCR/LLM lagi (deterministik + hemat).
                    parse_cache_put(_parse_key, all_rows, getattr(pdf, "filename", "") or "", principle_name)

                    # TAHAP 3: Return raw rows directly to frontend so the user can see/edit individual variants natively.
                    # (The actual grouping and Prefix Compression runs natively during summary_manual_generate)
                    return {"ok": True, "rows": all_rows}

                # ==========================
                # FULL MODE LOGIC (dead code, ai_mode selalu "split" di UI -- dibiarkan sbg fallback lama)
                # ==========================
                else: 
                    # Original logic using only Gemini
                    user_content = [{"type": "text", "text": prompt}]
                    # GLM-5 and Claude sonnet on some proxy setups reject image_url.
                    current_model = os.getenv("SUMOPOD_MODEL", "glm/glm-5").lower()
                    if "glm-5" not in current_model and "kimi" not in current_model and "deepseek" not in current_model:
                        for b64 in base64_images:
                            user_content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}})
                        
                    messages = [
                        {"role": "system", "content": "You are a helpful AI assistant that extracts precise JSON tables from text and scanned images."},
                        {"role": "user", "content": user_content}
                    ]
                    
                    payload = {
                        "model": os.getenv("SUMOPOD_MODEL", "glm/glm-5"),
                        "messages": messages,
                        "temperature": 0.1,
                        "max_completion_tokens": 8192,
                        "max_tokens": 8192
                    }
                    
                    headers = {
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json"
                    }
    
                    full_raw_text = ""
                    for loop_idx in range(5):
                        resp = await client_http.post(
                            "https://ai.sumopod.com/v1/chat/completions",
                            json=payload,
                            headers=headers
                        )
                        
                        if resp.status_code != 200:
                            append_error_log("gemini_400_debug", Exception(f"HTTP {resp.status_code}"), {"text": resp.text})
                        resp.raise_for_status()
                        
                        response_json = resp.json()
                        chunk_text = response_json.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
                        
                        # Bersihkan markdown formatting di tengah-tengah jika ini adalah sambungan
                        import re
                        clean_chunk = re.sub(r"```json", "", chunk_text, flags=re.IGNORECASE)
                        clean_chunk = re.sub(r"```", "", clean_chunk)
                        
                        # Gemini sometimes repeats the last few characters when asked to continue.
                        # We must find the overlap and slice it out before appending.
                        overlap_len = 0
                        if full_raw_text:
                            # Check overlapping strings up to 100 characters max
                            for i in range(min(100, len(full_raw_text), len(clean_chunk)), 0, -1):
                                if full_raw_text[-i:] == clean_chunk[:i]:
                                    overlap_len = i
                                    break
                                    
                        full_raw_text += clean_chunk[overlap_len:]
                        
                        # Cek apakah JSON sudah tertutup seutuhnya
                        if full_raw_text.strip().endswith("]"):
                            break # Yey selesai!
                        
                        # Jika belum selesai tapi chunk kosong, AI nyerah
                        if not chunk_text:
                            break
                            
                        # Minta AI melanjutkan TEPAT dari karakter terakhir yang terpotong
                        payload["messages"].append({"role": "assistant", "content": chunk_text})
                        payload["messages"].append({"role": "user", "content": "Teks JSON terpotong karena batas token! WAJIB lanjutkan string JSON di atas TEPAT mulai dari huruf/simbol yang terputus tanpa basa-basi pengantar, tanpa markdown ```json. Langsung sambung karakternya!"})


                raw_text = full_raw_text.strip()
                try:
                    with open(os.path.join(BASE_DIR, "data", "debug_ai.txt"), "w", encoding="utf-8") as f:
                        f.write(f"=== OCR PHASE 1 ===\n{ocr_text if 'ocr_text' in locals() else 'N/A'}\n\n=== JSON PHASE 2 ===\n{raw_text}\n")
                except Exception:
                    pass
                if not raw_text:
                    append_error_log("gemini_empty_response", Exception("No text returned"), {"user": user})
                    return {"ok": False, "error": "AI mengembalikan respons kosong."}
                    
                # Extract JSON block just in case
                match = re.search(r"\[.*\]", raw_text, re.DOTALL)
                if match:
                    clean_text = match.group(0).strip()
                else:
                    clean_text = raw_text.strip()
                    
                # Auto-heal truncated JSON jika proxy benar-benar mati
                if clean_text.startswith("[") and not clean_text.endswith("]"):
                    last_brace = clean_text.rfind("}")
                    if last_brace != -1:
                        clean_text = clean_text[:last_brace+1] + "\n]"
                        
                try:
                    batch_data = json.loads(clean_text, strict=False)
                except json.JSONDecodeError:
                    append_error_log("gemini_invalid_json", Exception("Non-JSON Payload String Loop"), {"raw": raw_text[:500]})
                    return {"ok": False, "error": f"AI gagal mengirim struktur JSON yang benar.\n\nContoh respons:\n{raw_text[:200]}"}
                    
                if isinstance(batch_data, list):
                    all_rows = batch_data
                elif isinstance(batch_data, dict) and "rows" in batch_data:
                    all_rows = batch_data["rows"]
                
            if not all_rows:
                return {"ok": False, "error": "AI tidak menemukan tabel promo valid di dalam dokumen, atau gagal mengekstrak."}
                
            try:
                import json
                with open("/tmp/ai_dump.json", "w") as f:
                    json.dump(all_rows, f, indent=2)
            except: pass
            
            for idx, row in enumerate(all_rows):
                if "id" not in row:
                    row["id"] = str(uuid.uuid4())
                if "no" not in row:
                    row["no"] = str(idx + 1)
                    
            # TAHAP 2: Native Master DB Mapping (Injects Kelompok perfectly)
            all_rows = _apply_native_kelompok(all_rows, items)
            
            # TAHAP 3: Return raw rows directly to frontend so the user can see/edit individual variants natively.
            # (The actual grouping and Prefix Compression runs natively during summary_manual_generate)
            return {"ok": True, "rows": all_rows}
                
        except Exception as api_err:
             append_error_log("gemini_api_error", api_err, {"user": user})
             return {"ok": False, "error": f"Gagal menghubungi Google Gemini AI: {str(api_err)}"}

        
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        print("====== FATAL N8N PARSE ERROR ======")
        print(err_msg)
        return {"ok": False, "error": f"Internal Server Error: {str(e)}"}
        print("===================================")
        append_error_log("summary_manual_parse_pdf_n8n", e, {"user": user, "token": token})
        payload = {"ok": False, "error": "Kegagalan sistem internal saat memproses PDF."}
        if APP_DEBUG and is_admin_user(user):
            payload["detail"] = str(e)
        return payload

# ======================================================================================
# "AI LEARNING" DARI KOREKSI MANUAL (tombol Laporkan Salah di grid)
# Bukan fine-tuning model -- few-shot memory: tiap koreksi user (before->after) disimpan,
# lalu di-inject balik ke prompt parse_pdf_ai supaya kesalahan yg sama tidak berulang.
# ======================================================================================
CORRECTIONS_PATH = os.path.join(BASE_DIR, "data", "parse_corrections.jsonl")
_CORRECTION_IGNORE_KEYS = {"id", "_matched_items_cache", "no"}

def _load_corrections(principle_name: str, limit: int = 15) -> List[Dict[str, Any]]:
    if not os.path.exists(CORRECTIONS_PATH):
        return []
    items = []
    try:
        with open(CORRECTIONS_PATH, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    items.append(json.loads(line))
                except Exception:
                    continue
    except Exception:
        return []
    pn = s(principle_name).upper()
    matched = [it for it in items if pn and pn in s(it.get("principle_name", "")).upper()]
    pool = matched if matched else items
    return pool[-limit:]

def _format_corrections_for_prompt(corrections: List[Dict[str, Any]]) -> str:
    blocks = []
    for c in corrections:
        before = c.get("before") or {}
        after = c.get("after") or {}
        diffs = []
        for k in set(before.keys()) | set(after.keys()):
            if k in _CORRECTION_IGNORE_KEYS:
                continue
            bv, av = before.get(k, ""), after.get(k, "")
            if s(bv) != s(av):
                diffs.append(f'    {k}: SALAH="{bv}" -> BENAR="{av}"')
        if diffs:
            note = s(c.get("note", ""))
            header = f"KOREKSI ({note}):" if note else "KOREKSI:"
            blocks.append(header + "\n" + "\n".join(diffs))
    if not blocks:
        return ""
    return ("\n\n=== KOREKSI DARI USER SEBELUMNYA (WAJIB DIPATUHI, JANGAN ULANGI KESALAHAN INI) ===\n"
            + "\n".join(blocks) + "\n=== AKHIR KOREKSI ===")

@app.post("/summary/manual/report_correction")
async def summary_manual_report_correction(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    if not user_has_permission(user, "summary", "edit"):
        return JSONResponse(status_code=403, content={"ok": False, "error": "Forbidden"})
    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})
    try:
        body = await request.json()
        before = body.get("before") if isinstance(body.get("before"), dict) else {}
        after = body.get("after") if isinstance(body.get("after"), dict) else {}
        if not before and not after:
            return {"ok": False, "error": "Data before/after kosong."}
        from datetime import datetime
        entry = {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "user": user,
            "principle_name": s(body.get("principle_name", "")),
            "before": {k: v for k, v in before.items() if k not in _CORRECTION_IGNORE_KEYS},
            "after": {k: v for k, v in after.items() if k not in _CORRECTION_IGNORE_KEYS},
            "note": s(body.get("note", "")),
        }
        _ensure_dir(os.path.join(BASE_DIR, "data"))
        with open(CORRECTIONS_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
        return {"ok": True}
    except Exception as e:
        append_error_log("summary_manual_report_correction", e, {"user": user})
        return {"ok": False, "error": "Gagal menyimpan koreksi."}

def process_summary_generation_job(job_id: str, token: str, rows: List[Dict[str, Any]], user: str):
    try:
        BACKGROUND_JOBS[job_id]["status"] = "processing"
        cache = MANUAL_MASTER_CACHE[token]
        items = cache.get("items", [])

        def norm(x: object) -> str:
            return " ".join(str(x or "").strip().split()).upper()

        def smart_normalize(text: str) -> str:
            t = norm(text)
            aliases = {
                "SHM": "SHAMPOO",
                "F.WASH": "FACIAL WASH",
                "COND.": "CONDITIONER",
                "COND ": "CONDITIONER ",
                "ALOVERA": "ALOE",
                "ALOEVERA": "ALOE",
                "ALOE VERA": "ALOE",
                "ALOE VERE": "ALOE",
                "HABBATS": "HABBATUSSAUDA",
                "ANTI DANDRUFF": "TEA TREE",
                "ANTIDANDRUFF": "TEA TREE",
                "VIT ": "VITAMIN ",
                "165ML": "160ML",
                "BRIGHT&DC": "BRIGHTENING",
                "CR&OC": "ACNE",
                "EXTRACT": "",
                "&": " ",
                "OIL": ""
            }
            for k, v in aliases.items():
                t = t.replace(k, v)
            t = re.sub(r'[^A-Z0-9\s]', ' ', t)
            return " ".join(t.split())

        def split_list(val: str) -> List[str]:
            return [x.strip() for x in str(val or "").split(",") if x.strip()]

        def unit_from_text(text: str) -> str:
            t = norm(text)
            if "CTN" in t or "KRT" in t:
                return "CTN"
            return "PCS"

        def has_number(text: str) -> bool:
            return bool(re.search(r"\d", str(text or "")))

        out_dir = os.path.join(BASE_DIR, "output", "summary_manual")
        _ensure_dir(out_dir)
        file_id = str(uuid.uuid4())

        form_path = os.path.join(out_dir, f"{file_id}_Form_Summary_Program.pdf")
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.units import inch
        from datetime import datetime
        
        # Get current date formatted for Indonesia
        now = datetime.now()
        months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
        print_date = f"{now.month}/{now.day}/{now.year}"
        dibuat_date = f"{now.day:02d} {months[now.month - 1]} {now.year}"

        # TAHAP 2: Native Master DB Mapping (Injects Kelompok perfectly)
        _apply_native_kelompok(rows, items)
            
        def my_canvas(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setFont('Helvetica-Bold', 7)
            # Top Left
            canvas_obj.drawString(doc_obj.leftMargin, landscape(A4)[1] - 0.5*cm, f"Di Print Tgl : {print_date}")
            # Top Right (Page Number)
            page_str = f"Hal {doc_obj.page}"
            canvas_obj.drawRightString(landscape(A4)[0] - doc_obj.rightMargin, landscape(A4)[1] - 0.5*cm, page_str)
            
            # Bottom of Header (Before Table Starts)
            canvas_obj.setFont('Helvetica', 8)
            canvas_obj.drawString(doc_obj.leftMargin, landscape(A4)[1] - 2*cm, f"Dibuat Tanggal : {dibuat_date}")
            canvas_obj.setFont('Helvetica-Bold', 8)
            canvas_obj.drawRightString(landscape(A4)[0] - doc_obj.rightMargin, landscape(A4)[1] - 2*cm, "(ON PRINCIPLE COKLAT)")
            
            canvas_obj.restoreState()

        # FASE 6: PDF reproducible (CreationDate/ModDate/doc-id tetap) -> byte-identik antar-run
        enable_pdf_determinism()
        # Update margins to give space for the custom canvas headers
        doc = SimpleDocTemplate(form_path, pagesize=landscape(A4), rightMargin=0.5*cm, leftMargin=0.5*cm, topMargin=2.2*cm, bottomMargin=0.5*cm)
        elements = []
        styles = getSampleStyleSheet()

        # Add Title and Subtitle properly centered
        title_style = styles["Heading3"].clone("TitleStyle")
        title_style.alignment = TA_CENTER
        title_style.fontSize = 12
        elements.append(Paragraph("<b>SUMMARY PROGRAM ON FAKTUR BEBAN PRINCIPLE (COKLAT)</b>", title_style))
        
        sub_style = styles["Normal"].clone("SubStyle")
        sub_style.alignment = TA_CENTER
        sub_style.fontSize = 11
        sub_style.fontName = 'Helvetica'
        
        # Determine Period globally from the first row if available
        global_period = rows[0].get("periode", "") if rows else ""
        period_text = f" PERIODE {global_period.upper()}" if global_period else ""
        elements.append(Paragraph(f"CV. SURYA PERKASA {period_text}", sub_style))
        elements.append(Spacer(1, 15))

        headers_str = ["No", "Surat Program", "Nama Program", "Channel", "Periode", 
                   "Kelompok Barang", "Variant", "Gramasi", "Ketentuan", "Benefit", "Syarat Claim", "Keterangan"]
        
        header_style = styles["Normal"].clone("HeaderStyle")
        header_style.fontSize = 7
        header_style.leading = 8
        header_style.fontName = 'Helvetica-Bold'
        header_style.alignment = TA_CENTER
        header_style.textColor = colors.whitesmoke
        
        table_data = [[Paragraph(h, header_style) for h in headers_str]]

        cell_style = styles["Normal"].clone("CellStyle")
        cell_style.fontSize = 6
        cell_style.leading = 7
        cell_style.alignment = TA_CENTER

        for r in rows:
            benefit_pdf_text = str(r.get("benefit",""))
            if str(r.get("benefit_type","")).upper() == "DISC_RP":
                if benefit_pdf_text and not benefit_pdf_text.lower().startswith("cut price"):
                    benefit_pdf_text = f"cut price {benefit_pdf_text}"
                    
            table_data.append([
                Paragraph(str(r.get("no","")), cell_style),
                Paragraph(str(r.get("surat_program","")), cell_style),
                Paragraph(str(r.get("nama_program","")), cell_style),
                Paragraph(str(r.get("channel_gtmt","")), cell_style),
                Paragraph(str(r.get("periode","")), cell_style),
                Paragraph(str(r.get("kelompok","")), cell_style),
                Paragraph(str(r.get("variant","")), cell_style),
                Paragraph(str(r.get("gramasi","")), cell_style),
                Paragraph(str(r.get("ketentuan","")), cell_style),
                Paragraph(benefit_pdf_text, cell_style),
                Paragraph(str(r.get("syarat_claim","")), cell_style),
                Paragraph(str(r.get("keterangan","")), cell_style),
            ])
            
        # Total A4 landscape width is ~842. Margins are 0.5cm each (approx 14 points each, total 28 pts margin)
        # Usable width = 842 - 28 = 814 points
        usable = landscape(A4)[0] - (1 * cm)
        cw = [
            usable * 0.03, # No
            usable * 0.12, # Surat Program
            usable * 0.12, # Nama Program
            usable * 0.05, # Channel
            usable * 0.08, # Periode
            usable * 0.09, # Kelompok
            usable * 0.14, # Variant
            usable * 0.08, # Gramasi
            usable * 0.08, # Ketentuan
            usable * 0.08, # Benefit
            usable * 0.07, # Syarat Claim
            usable * 0.06  # Keterangan
        ]
            
        t = Table(table_data, repeatRows=1, colWidths=cw)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9E7C85')), # Match the brownish pink header color
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        elements.append(t)
        
        # Add the Footer Signatures
        elements.append(Spacer(1, 25))
        
        footer_style_left = styles["Normal"].clone("FooterLeft")
        footer_style_left.fontSize = 8
        footer_style_left.fontName = 'Helvetica-Bold'
        
        footer_style_right = styles["Normal"].clone("FooterRight")
        footer_style_right.fontSize = 8
        footer_style_right.fontName = 'Helvetica-Bold'
        footer_style_right.alignment = TA_RIGHT

        sig_data = [
            [Paragraph(f"Makassar , {dibuat_date}", footer_style_left), ""],
            [Paragraph("Diajukan Oleh,", footer_style_left), Paragraph("Disetujui Oleh,", footer_style_right)],
            [Spacer(1, 40), Spacer(1, 40)], # Space for signature
            [Paragraph("SM<br/>(.................................................)", footer_style_left), 
             Paragraph("OPERATIONAL MANAGER<br/>(.................................................)", footer_style_right)]
        ]
        
        # Table takes up full usable width so left is left, right is right
        sig_table = Table(sig_data, colWidths=[usable/2.0, usable/2.0])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(sig_table)
        
        doc.build(elements, onFirstPage=my_canvas, onLaterPages=my_canvas)

        dataset_path = os.path.join(out_dir, f"{file_id}_Dataset_Diskon_With_Channel.xlsx")
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Dataset"
        headers2 = [
            "KODE_BARANG", "NAMA_BARANG", "PROMO_LABEL", "PROMO_GROUP_ID", "PROMO_GROUP",
            "PERIODE", "PROMO_ACTIVE", "TIER_NO", "TRIGGER_QTY", "TRIGGER_UNIT",
            "BENEFIT_TYPE", "BENEFIT_VALUE", "BENEFIT_UNIT", "BENEFIT_BEBAN"
        ]
        ws2.append(headers2)

        tier_counter: Dict[Tuple[str, str, str, str], int] = {}
        channel_group_counters: Dict[str, int] = {}
        kelompok_group_map: Dict[Tuple[str, str, str], str] = {}

        for r in rows:
            kelompok = str(r.get("kelompok","") or "").strip()
            vlist = split_list(r.get("variant",""))
            glist = split_list(r.get("gramasi",""))
            v_all = (not vlist) or any(norm(x) == "ALL VARIANT" for x in vlist)
            g_all = (not glist) or any(norm(x) == "ALL GRAMASI" for x in glist)

            ket = str(r.get("ketentuan","") or "").strip()
            trig_has_num = has_number(ket)
            trig_qty = parse_number_id(ket) if trig_has_num else ""
            trig_unit = unit_from_text(ket) if trig_has_num else ""
            benefit_text_raw = str(r.get("benefit","") or "").strip()
            benefit_type = str(r.get("benefit_type","") or "").strip()
            
            import re
            # Extract only the numbers, avoiding capturing standalone periods like in "Pot."
            nums = re.findall(r'\d[\d\.,]*', benefit_text_raw)
            benefit_text = nums[0].replace(',', '.') if nums else benefit_text_raw
                
            benefit_unit = unit_from_text(benefit_text_raw) if benefit_text_raw else ""
            benefit_beban = "PABRIK"

            promo_label = str(r.get("nama_program","") or "").strip()
            promo_group_id = str(r.get("promo_group_id","") or "").strip()
            promo_group = str(r.get("channel_gtmt","") or "").strip()
            periode = str(r.get("periode","") or "").strip()
            
            if not promo_group_id or promo_group_id.upper() == "NON_GROUP":
                promo_group_id = promo_group

            matched_items = r.get("_matched_items_cache", [])
            
            if not matched_items:
                key_unmatched = (promo_label, promo_group_id, promo_group, "UNMATCHED")
                tier_counter[key_unmatched] = tier_counter.get(key_unmatched, 0) + 1
                tier_no = tier_counter[key_unmatched]
                
                ws2.append([
                    "",
                    "",
                    promo_label,
                    promo_group_id,
                    promo_group,
                    periode,
                    True,
                    tier_no,
                    trig_qty,
                    trig_unit,
                    benefit_type,
                    benefit_text,
                    benefit_unit,
                    benefit_beban,
                ])
                continue

            for it in matched_items:
                kb = str(it.get("kode_barang", ""))
                real_kelompok = str(it.get("kelompok", "")).strip()
                
                # Determine native promo group id progressively based on original valid kelompok
                base_pg_id = promo_group_id
                map_key = (promo_label, base_pg_id, real_kelompok)
                
                if map_key not in kelompok_group_map:
                    channel_group_counters[base_pg_id] = channel_group_counters.get(base_pg_id, 0) + 1
                    kelompok_group_map[map_key] = f"{base_pg_id}_{channel_group_counters[base_pg_id]}"
                
                final_promo_group_id = kelompok_group_map[map_key]
                
                key_item = (promo_label, final_promo_group_id, promo_group, kb)
                tier_counter[key_item] = tier_counter.get(key_item, 0) + 1
                tier_no = tier_counter[key_item]
                
                ws2.append([
                    kb,
                    str(it.get("nama_barang", "")),
                    promo_label,
                    final_promo_group_id,
                    promo_group,
                    periode,
                    True,
                    tier_no,
                    trig_qty,
                    trig_unit,
                    benefit_type,
                    benefit_text,
                    benefit_unit,
                    benefit_beban,
                ])
        wb2.save(dataset_path)

        MANUAL_OUTPUTS[file_id] = {"form": form_path, "dataset": dataset_path}
        
        BACKGROUND_JOBS[job_id]["status"] = "done"
        BACKGROUND_JOBS[job_id]["result"] = {
            "file_id": file_id
        }
        
    except Exception as e:
        import traceback
        with open(os.path.join(BASE_DIR, "debug_traceback.txt"), "w") as f:
            traceback.print_exc(file=f)
        BACKGROUND_JOBS[job_id]["status"] = "error"
        BACKGROUND_JOBS[job_id]["error"] = str(e)
        if hasattr(e, '__traceback__'):
            append_error_log("background_summary_manual", e, {"user": user, "job_id": job_id})

@app.post("/summary/manual/email")
async def summary_manual_email(
    request: Request,
    background_tasks: BackgroundTasks,
    email: str = Form(...),
    file_id: str = Form(...)
):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    
    if file_id not in MANUAL_OUTPUTS:
        return JSONResponse(status_code=404, content={"ok": False, "error": "Generated files not found."})

    csrf_token = request.headers.get("X-CSRF-Token", "")
    if not validate_csrf_request(request, csrf_token):
        return JSONResponse(status_code=403, content={"ok": False, "error": "CSRF token invalid"})

    if not EMAIL_USER or not EMAIL_PASSWORD:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Fitur email belum dikonfigurasi di server (EMAIL_USER / EMAIL_PASSWORD kosong)."})

    email = email.strip()
    if not email or "@" not in email:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Format email tidak valid."})
        
    try:
        background_tasks.add_task(send_email_background, user, email, file_id)
        return JSONResponse({"ok": True})
    except Exception as e:
         return JSONResponse(status_code=500, content={"ok": False, "error": "Gagal memulai tugas pengiriman email."})

PRINCIPLES_JSON_PATH = os.path.join(BASE_DIR, "data", "principles.json")
MASTERS_DIR = os.path.join(BASE_DIR, "data", "masters")

def _load_principles() -> dict:
    try:
        import sqlite3
        conn = sqlite3.connect("database.sqlite")
        c = conn.cursor()
        c.execute("SELECT id, name, filename, uploaded_by, created_at FROM principles")
        res = {}
        for row in c.fetchall():
            res[row[0]] = {
                "name": row[1],
                "filename": row[2],
                "uploaded_by": row[3],
                "created_at": row[4]
            }
        conn.close()
        return res
    except:
        return {}

def _save_principles(data: dict):
    try:
        import sqlite3
        conn = sqlite3.connect("database.sqlite")
        c = conn.cursor()
        c.execute("DELETE FROM principles")
        for uid, info in data.items():
            name = info.get("name", "")
            filename = info.get("filename", "")
            upb = info.get("uploaded_by", "")
            ca = info.get("created_at", "")
            c.execute("INSERT INTO principles (id, name, filename, uploaded_by, created_at) VALUES (?, ?, ?, ?, ?)", (uid, name, filename, upb, ca))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error saving principles to SQLite: {e}")

@app.get("/api/principles")
def get_principles(request: Request):
    user = get_current_user(request)
    if not user: return JSONResponse(status_code=401, content={"ok": False, "error": "Unauthorized"})
    ps = _load_principles()
    return {"ok": True, "principles": ps}
# Tujuan: FastAPI backend untuk validator diskon, summary, payments, finance, dan principle.
# Caller: Next.js dashboard, halaman HTML legacy FastAPI, dan workflow internal operasional.
# Dependensi: pandas/openpyxl, auth helpers, validator_engine, payments, SumoPod/OpenAI-compatible API.
# Main Functions: `app`, endpoint validator/summary/payments/finance/principles.
# Side Effects: DB/file read-write runtime, HTTP call AI/SMTP, generate Excel/download artifacts.
#
# Tujuan: Backend FastAPI utama untuk validator, summary, payments/SPPD, finance, dan principle.
# Caller: Next.js dashboard, halaman legacy FastAPI, dan API browser internal AccAPI.
# Dependensi: pandas/openpyxl, payments.py, validator_engine.py, payments.json, file output, Accurate session dari UI.
# Main Functions: app routes, load/save payments DB, upload/restore payments, render SPPD, finance proof/mapping/update APIs.
# Side Effects: HTTP response, JSON/file I/O, DOCX/XLSX generation, audit/error log writes.


# =======================================================================================================
# Laporan Harian per SPV/SM — proses FIX LAP PENJ (Paste Acc only, retur sudah minus) -> agregat & split.
# Dipanggil Next.js app/api/laporan-harian/upload. Tidak kirim email (email = gate terpisah di Next.js).
# =======================================================================================================
LH_RUNTIME_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "runtime", "laporan-harian")

@app.post("/laporan-harian/process")
async def laporan_harian_process(
    penjualan: Optional[UploadFile] = File(None),
    retur: Optional[UploadFile] = File(None),
    fix: Optional[UploadFile] = File(None),
    stock: Optional[UploadFile] = File(None),
    run_id: Optional[str] = Form(None),
    report_date: Optional[str] = Form(None),
    write_files: Optional[str] = Form(None),
):
    import tempfile, os as _os
    import laporan_harian as LH

    tmpdir = tempfile.mkdtemp(prefix="lh_")

    async def _save(uf, name):
        if uf is None:
            return None
        p = _os.path.join(tmpdir, name)
        with open(p, "wb") as f:
            f.write(await uf.read())
        return p

    penj_path = await _save(penjualan, "penjualan.xlsx")
    ret_path = await _save(retur, "retur.xlsx")
    fix_path = await _save(fix, "fix.xlsx")
    stock_path = await _save(stock, "stock.xlsx")

    try:
        lk = LH.load_lookups_json()   # master GOLONGAN/JENIS PRODUK/Mapping (untuk NAMA SM juga)
        if penj_path:
            fix_df = LH.build_fix_from_accurate(penj_path, ret_path, lk)   # web bangun 2.ToFormat sendiri
        elif fix_path:
            fix_df = LH.load_fix(fix_path)                                  # backward-compat: upload FIX jadi
        else:
            return ORJSONResponse({"ok": False, "error": "Wajib upload 'penjualan' (+retur) atau 'fix'."}, status_code=400)
        sb = LH.build_salesbase(fix_df, lk)
        progress = LH.aggregate_progress(sb)
        summary = (sb.groupby("GOLONGAN", dropna=True)
                     .agg(rows=("NO_NOTA", "size"), dpp=("DPP", "sum"),
                          ao=("AO", "sum"), ec=("EC", "sum"), ia=("Item Aktif", "sum"))
                     .reset_index().sort_values("GOLONGAN"))
        # periode dominan
        pm = int(progress["periodMonth"].dropna().mode().iloc[0]) if len(progress) else None
        py = int(progress["periodYear"].dropna().mode().iloc[0]) if len(progress) else None
        prog_records = []
        for r in progress.to_dict("records"):
            prog_records.append({
                "salesCode": None if r["salesCode"] is None else str(r["salesCode"]),
                "principle": None if r["principle"] is None else str(r["principle"]),
                "branch": None if r["branch"] is None else str(r["branch"]),
                "date": None if r["date"] is None else str(r["date"]),
                "periodMonth": None if r["periodMonth"] is None else int(r["periodMonth"]),
                "periodYear": None if r["periodYear"] is None else int(r["periodYear"]),
                "achievedValueDpp": float(r["achievedValueDpp"] or 0),
                "achievedEc": int(r["achievedEc"] or 0),
                "achievedAo": int(r["achievedAo"] or 0),
                "achievedIa": int(r["achievedIa"] or 0),
            })
        stock_spv = []
        if stock_path:
            try:
                st = LH.build_stock(stock_path, sb, lk)
                stock_spv = [k for k in st.keys() if k != "__error__"]
            except Exception:
                stock_spv = []
        files_written = []
        if write_files and run_id:
            import re as _re, datetime as _dt
            safe_run = _re.sub(r"[^A-Za-z0-9_-]", "", str(run_id))[:64]
            rdate = report_date or _dt.date.today().strftime("%Y-%m-%d")
            out_dir = _os.path.join(LH_RUNTIME_DIR, safe_run)
            files_written = LH.write_per_spv_files(sb, out_dir, rdate)

        return ORJSONResponse({
            "ok": True,
            "files": files_written,
            "sales_rows": int(len(sb)),
            "net_dpp": float(sb["DPP"].sum()),
            "period": {"month": pm, "year": py},
            "spv_list": [str(x) for x in summary["GOLONGAN"].tolist()],
            "summary": [
                {"spv": str(r["GOLONGAN"]), "rows": int(r["rows"]), "dpp": float(r["dpp"]),
                 "ao": int(r["ao"]), "ec": int(r["ec"]), "ia": int(r["ia"])}
                for r in summary.to_dict("records")
            ],
            "progress": prog_records,
            "stock_spv": stock_spv,
        })
    except Exception as e:
        import traceback
        return ORJSONResponse({"ok": False, "error": str(e), "trace": traceback.format_exc()[-1500:]}, status_code=500)
    finally:
        try:
            import shutil as _sh; _sh.rmtree(tmpdir, ignore_errors=True)
        except Exception:
            pass



@app.get("/laporan-harian/file")
async def laporan_harian_file(run: str, name: str):
    """Stream 1 file laporan per-SPV (run-scoped). Guard path traversal."""
    import re as _re
    safe_run = _re.sub(r"[^A-Za-z0-9_-]", "", str(run))[:64]
    if "/" in name or "\\" in name or ".." in name:
        return ORJSONResponse({"error": "nama file tidak valid"}, status_code=400)
    path = os.path.join(LH_RUNTIME_DIR, safe_run, name)
    if not os.path.isfile(path):
        return ORJSONResponse({"error": "file tidak ditemukan"}, status_code=404)
    return FileResponse(path, filename=name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
