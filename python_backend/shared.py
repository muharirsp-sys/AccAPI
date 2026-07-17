# Tujuan: Menyediakan konfigurasi, state, dan helper bersama untuk runtime FastAPI.
# Caller: main.py dan routers/*.py; modul ini tidak mengimpor main/routers agar tidak circular.
# Dependensi: FastAPI, pandas/openpyxl, konfigurasi environment, dan modul domain backend.
# Main Functions: Helper pemrosesan dokumen, pembayaran, serta laporan harian per SPV/SM.
# Side Effects: Membaca/menulis data runtime, file hasil, cache, dan melakukan integrasi eksternal.

from fastapi import FastAPI, UploadFile, File, Request, Form, Response, Cookie, BackgroundTasks
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse, ORJSONResponse
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import asyncio
import io, os, re, uuid, json, math, base64, hashlib, hmac, time, zipfile, copy, mimetypes
from dotenv import load_dotenv
load_dotenv(override=False)
import traceback
import xml.etree.ElementTree as ET
# Modul determinisme Summary Program (FASE 1-4): cache OCR, parser tier posisional,
# resolusi varian tabel, correction store stable-key. Lihat SYSTEM_MAP.md.
from variant_resolver import load_variant_mapping, resolve_variant
from correction_store import load_corrections as load_stable_corrections, apply_corrections as apply_stable_corrections, save_correction, correction_key
from tier_parser import regroup_rows_by_tier
from golden_store import canonical_signature, golden_check_and_freeze
from deterministic_output import enable_pdf_determinism, finalize_xlsx
from parse_cache import parse_cache_key, parse_cache_get, parse_cache_put
# PASS 3 self-correction ala Reducto: editor QA LLM memverifikasi rows vs teks sumber (patch-based).
from self_correction import verify_and_correct_rows
_VARIANT_MAPPING = load_variant_mapping()
# Kelompok yg SELALU di-EXCLUDE dari promo (dari exclude_kelompok tiap rule variant_mapping,
# mis. Casablanca Spray Cologne GLASS/GLAS -- aturan surat, dikonfirmasi user 2026-07-15).
# LLM kadang mencantumkan kode GLASS langsung di kode_barangs -> lolos klist-match; drop-list
# global ini menutup SEMUA jalur (klist, ekspansi, fallback), simetris dgn exclude banded (BND).
_EXCLUDED_KELOMPOKS = {
    " ".join(str(_k).strip().split()).upper()
    for _rule in _VARIANT_MAPPING.values()
    for _k in _rule.get("exclude_kelompok", [])
}
from typing import List, Dict, Tuple, Optional, Set, Any
from urllib.parse import urlparse, unquote
from auth import (
    LoginRateLimiter,
    build_security_headers,
    enforce_non_default_auth_secret,
    is_production_env,
    login_rate_key,
    parse_bool_env,
)
from validator_engine import extract_pdf_text_safe, read_upload_file_limited
from payments import (
    lpb_upload_template_rows,
    validator_channel_template_rows,
    validator_promo_template_rows,
    validator_sales_template_rows,
)
from principle_matcher import (
    normalize_principle_name,
    get_principle_match_key,
    find_best_match,
    build_normalized_key_map,
    generate_import_report,
)
try:
    import bcrypt as _bcrypt  # type: ignore
except Exception:
    _bcrypt = None

try:
    from argon2 import PasswordHasher as _Argon2Hasher  # type: ignore
    _ARGON2 = _Argon2Hasher()
except Exception:
    _ARGON2 = None

PATCH_VERSION = "v15"
PATCH_TITLE = "Channel lookup required + internal discount engine (expected internal + allocation)"

REQUIRED_MISSING_MSG = "Data Penjualan masih kosong tuh, upload dulu dong adiks-adiks/kakaks-kakaks"

# Nginx X-Accel-Redirect (optional)
USE_X_ACCEL = str(os.getenv("USE_X_ACCEL", "1")).strip().lower() in ["1", "true", "yes", "y"]
X_ACCEL_PREFIX = str(os.getenv("X_ACCEL_PREFIX", "/protected-downloads")).strip()

# Auth (no storage) - set via env
AUTH_USERS = str(os.getenv("AUTH_USERS", "admin:admin")).strip()
AUTH_SECRET = str(os.getenv("AUTH_SECRET", "dev-secret")).strip()
AUTH_COOKIE = str(os.getenv("AUTH_COOKIE", "dv_auth")).strip()
AUTH_TTL_SECONDS = int(os.getenv("AUTH_TTL_SECONDS", "43200"))  # 12 hours
AUTH_COOKIE_SECURE = str(os.getenv("AUTH_COOKIE_SECURE", "1")).strip().lower() in ["1", "true", "yes", "y"]
AUTH_COOKIE_SAMESITE = str(os.getenv("AUTH_COOKIE_SAMESITE", "strict")).strip().lower()
AUTH_PASSWORD_SCHEME = str(os.getenv("AUTH_PASSWORD_SCHEME", "auto")).strip().lower()
AUTH_PBKDF2_ITERATIONS = int(os.getenv("AUTH_PBKDF2_ITERATIONS", "260000"))
AUTH_BCRYPT_ROUNDS = int(os.getenv("AUTH_BCRYPT_ROUNDS", "12"))
CSRF_COOKIE = str(os.getenv("CSRF_COOKIE", "dv_csrf")).strip()
CSRF_COOKIE_SAMESITE = str(os.getenv("CSRF_COOKIE_SAMESITE", "lax")).strip().lower()
CSRF_TTL_SECONDS = int(os.getenv("CSRF_TTL_SECONDS", "7200"))
APP_ENV = str(os.getenv("APP_ENV", "development")).strip().lower()
APP_DEBUG = parse_bool_env(os.getenv("APP_DEBUG"), default=False)
# Audit F11: tarif PPN satu sumber (selaras lib/claim-workflow/calculations.ts yang parametrized).
PPN_RATE = float(os.getenv("PPN_RATE", "0.11"))
APP_IS_PRODUCTION = is_production_env(APP_ENV)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BETTER_AUTH_DB_PATH = str(os.getenv("BETTER_AUTH_DB_PATH", os.path.join(BASE_DIR, "..", "sqlite.db"))).strip()
LOGIN_BG_PATH = os.path.join(BASE_DIR, "login page", "bg.png")
AUTH_USERS_JSON = str(os.getenv("AUTH_USERS_JSON", os.path.join(BASE_DIR, "users.json"))).strip()
AUTH_ADMINS = set([u.strip() for u in str(os.getenv("AUTH_ADMINS", "admin")).split(",") if u.strip()])
AUTH_FINANCE = set([u.strip() for u in str(os.getenv("AUTH_FINANCE", "")).split(",") if u.strip()])
FAVICON_PATH = str(os.getenv("FAVICON_PATH", os.path.join(os.path.dirname(os.path.abspath(__file__)), "favicon.png"))).strip()
AUDIT_LOG_PATH = str(os.getenv("AUDIT_LOG_PATH", os.path.join(BASE_DIR, "data", "audit_log.jsonl"))).strip()
ERROR_LOG_PATH = str(os.getenv("ERROR_LOG_PATH", os.path.join(BASE_DIR, "data", "error_log.jsonl"))).strip()
PAYMENTS_DB_PATH = str(os.getenv("PAYMENTS_DB_PATH", os.path.join(BASE_DIR, "data", "payments.json"))).strip()
PAYMENTS_FILES_DIR = str(os.getenv("PAYMENTS_FILES_DIR", os.path.join(BASE_DIR, "output", "payments"))).strip()
PAYMENTS_PROOFS_DIR = str(os.getenv("PAYMENTS_PROOFS_DIR", os.path.join(PAYMENTS_FILES_DIR, "proofs"))).strip()
BANK_DATA_PATH = str(os.getenv("BANK_DATA_PATH", os.path.join(BASE_DIR, "Data No Rekening Principle.xlsx"))).strip()
SPPD_TEMPLATE_PATH = str(os.getenv("SPPD_TEMPLATE_PATH", os.path.join(BASE_DIR, "SPPD TGL 24 FEBRUARI 2026.docx"))).strip()
MAX_EXCEL_UPLOAD_BYTES = int(os.getenv("MAX_EXCEL_UPLOAD_BYTES", str(15 * 1024 * 1024)))
MAX_PROOF_UPLOAD_BYTES = int(os.getenv("MAX_PROOF_UPLOAD_BYTES", str(8 * 1024 * 1024)))
MAX_PDF_UPLOAD_BYTES = int(os.getenv("MAX_PDF_UPLOAD_BYTES", str(12 * 1024 * 1024)))
MAX_PDF_OCR_PAGES = int(os.getenv("MAX_PDF_OCR_PAGES", "8"))
MAX_PDF_OCR_TIMEOUT_SECONDS = int(os.getenv("MAX_PDF_OCR_TIMEOUT_SECONDS", "20"))
MAX_PDF_OCR_PAGE_TIMEOUT_SECONDS = int(os.getenv("MAX_PDF_OCR_PAGE_TIMEOUT_SECONDS", "4"))
LOGIN_MAX_FAILED_ATTEMPTS = int(os.getenv("LOGIN_MAX_FAILED_ATTEMPTS", "5"))
LOGIN_FAILED_WINDOW_SECONDS = int(os.getenv("LOGIN_FAILED_WINDOW_SECONDS", "300"))
LOGIN_LOCKOUT_SECONDS = int(os.getenv("LOGIN_LOCKOUT_SECONDS", "300"))

PERMISSION_MODULES = [
    "dashboard",
    "api_wrapper",
    "payments",
    "sppd",
    "finance",
    "principles",
    "summary",
    "validator",
    "users",
]
PERMISSION_ACTIONS = [
    "view",
    "create",
    "edit",
    "update",
    "delete",
    "upload",
    "export",
    "submit",
    "execute",
    "edit_settings",
    "upload_excel",
    "generate",
    "download",
    "approve",
    "transfer",
    "upload_proof",
    "post_accurate",
    "retry_post",
    "run",
    "email",
    "sync",
    "manage",
    "create_user",
    "edit_user",
    "delete_user",
    "set_role",
    "set_permission",
]

if APP_IS_PRODUCTION:
    enforce_non_default_auth_secret(AUTH_SECRET)

LOGIN_LIMITER = LoginRateLimiter(
    max_attempts=LOGIN_MAX_FAILED_ATTEMPTS,
    window_seconds=LOGIN_FAILED_WINDOW_SECONDS,
    cooldown_seconds=LOGIN_LOCKOUT_SECONDS,
)

PATCH_NOTES_HTML = r"""
<div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] shadow-lg overflow-hidden">
  <div class="p-5 border-b border-[#8A4703]">
    <div class="flex items-start justify-between gap-3">
      <div>
        <div class="text-xs uppercase tracking-widest text-[#1D1F1E]">Patch Notes</div>
        <div class="text-lg font-semibold text-[#1D1F1E]">PATCH v15</div>
        <div class="mt-1 text-sm text-[#1D1F1E]">Channel lookup + hitung diskon Internal (Expected_Internal + FixClaim_Internal).</div>
      </div>
      <div class="shrink-0 rounded-xl px-3 py-2 bg-[#BD7401] text-[#EFEFEF] text-xs font-semibold">
        Engine + UI
      </div>
    </div>
  </div>

  <div class="p-5 space-y-4">
    <div class="rounded-xl bg-[#EFEFEF] border border-[#8A4703] p-4">
      <div class="text-sm font-semibold text-[#1D1F1E]">Yang berubah di v15</div>
      <ul class="mt-2 text-sm text-[#1D1F1E] list-disc pl-5 space-y-1">
        <li><b>Internal aktif</b>: jika dataset diskon internal diupload, engine akan hitung <code>Expected_Internal</code>.</li>
        <li><b>Alokasi claim</b>: Actual discount mengikuti urutan diskon di <code>MDSTRING</code> untuk
          menentukan porsi <code>FixClaim_Pabrik</code> / <code>FixClaim_Internal</code>; sisanya <code>TanpaTuan</code>.
        </li>
        <li><b>Rule internal</b>: match by <code>CHANNEL</code> (hasil lookup SUB), optional by <code>SUB</code> dan <code>KODE_BARANG</code>, tier by <code>MIN_GROSS/MAX_GROSS</code>.</li>
        <li>Dataset internal tetap optional (kalau kosong, Expected_Internal=0).</li>
      </ul>
    </div>
  </div>
</div>
"""

# ----- Background Jobs Storage -----
# In a real enterprise app, use Redis/Celery. For now, in-memory dict mapped by job_id.
BACKGROUND_JOBS: Dict[str, Dict[str, Any]] = {}

# CORS allowlist — hanya origin frontend yang sah yang boleh memanggil API dengan
# credentials. Mengganti pola lama `allow_origin_regex="^https?://.*$"` yang
# mengizinkan SEMUA origin (celah pencurian data lintas-situs saat digabung
# allow_credentials=True).
# Sumber origin (prioritas): env CORS_ORIGINS (pisahkan koma) > NEXT_PUBLIC_APP_URL
# > localhost dev. Origin lokal dikunci ke http://localhost:3000 agar cookie
# auth tidak pecah karena campur localhost/127.0.0.1 atau port dev lain.
_cors_env = str(os.getenv("CORS_ORIGINS", "")).strip()
if _cors_env:
    CORS_ALLOWED_ORIGINS = [o.strip() for o in _cors_env.split(",") if o.strip()]
else:
    CORS_ALLOWED_ORIGINS = [
        o for o in [
            str(os.getenv("NEXT_PUBLIC_APP_URL", "")).strip(),
            "http://localhost:3000",
        ] if o
    ]

def is_allowed_cors_origin(origin: str) -> bool:
    try:
        parsed = urlparse(origin)
    except Exception:
        return False
    host = str(parsed.hostname or "").strip().lower()
    is_local = host in {"localhost", "127.0.0.1", "0.0.0.0"}
    return (not is_local) or origin == "http://localhost:3000"

CORS_ALLOWED_ORIGINS = [o for o in CORS_ALLOWED_ORIGINS if is_allowed_cors_origin(o)]
# Dedup sambil menjaga urutan.
CORS_ALLOWED_ORIGINS = list(dict.fromkeys(CORS_ALLOWED_ORIGINS))

# ---------------------------
# Helpers
# ---------------------------
def s(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

def norm(x) -> str:
    return s(x).upper()

def ensure_col(df: pd.DataFrame, col: str, default):
    if col not in df.columns:
        df[col] = default
    return df

def to_float_series(df: pd.DataFrame, cols: List[str]):
    for c in cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    return df

def parse_number_id(x) -> float:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return 0.0
    if isinstance(x, (int, float)) and not pd.isna(x):
        return float(x)

    t = str(x).strip()
    if t == "":
        return 0.0

    t = re.sub(r"[^0-9\.\,\-]", "", t)

    # Both '.' and ',' => EU style "1.234,56"
    if "." in t and "," in t:
        t2 = t.replace(".", "").replace(",", ".")
        try:
            return float(t2)
        except:
            return 0.0

    # Only comma
    if "," in t and "." not in t:
        parts = t.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            t2 = t.replace(",", ".")
        else:
            t2 = t.replace(",", "")
        try:
            return float(t2)
        except:
            return 0.0

    # Only dot
    if "." in t and "," not in t:
        groups = t.split(".")
        if len(groups) >= 2 and all(g.isdigit() for g in groups if g != ""):
            if len(groups[-1]) == 3:
                t2 = "".join(groups)
                try:
                    return float(t2)
                except:
                    return 0.0
        try:
            return float(t)
        except:
            return 0.0

    try:
        return float(t)
    except:
        return 0.0

def packaging_to_int(packaging: str) -> int:
    packaging = s(packaging).upper()
    m = re.search(r"(\d+)", packaging)
    return int(m.group(1)) if m else 1

def qty_in_unit(row: pd.Series, unit: str) -> float:
    unit = norm(unit)
    qty = float(row.get("QTY", 0) or 0.0)
    qtypcs = float(row.get("QTYPCS", 0) or 0.0)
    u = norm(row.get("UNIT", ""))
    pack = packaging_to_int(row.get("PACKAGING", "")) or 1

    if unit == "PCS":
        if qtypcs > 0:
            return qtypcs
        if u in ["CTN", "KRT"]:
            return qty * float(pack)
        return qty

    if unit in ["CTN", "KRT"]:
        if u in ["CTN", "KRT"]:
            return qty
        pcs = qtypcs if qtypcs > 0 else qty
        return float(pcs) / float(pack)

    return qty

def parse_pct_chain(value: str) -> List[float]:
    value = s(value)
    if not value:
        return []
    parts = re.split(r"\s*\+\s*", value.replace("%", ""))
    out: List[float] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        try:
            out.append(float(p))
        except:
            pass
    return out

def parse_mdstring_pct(md: str) -> List[float]:
    md = s(md)
    if not md:
        return []
    nums = re.findall(r"(-?\d+(?:\.\d+)?)\s*%", md)
    if not nums:
        return parse_pct_chain(md)
    out = []
    for n in nums:
        try:
            out.append(float(n))
        except:
            pass
    return out

def normalize_pct_list(pcts) -> List[float]:
    if pcts is None:
        return []
    if isinstance(pcts, str):
        raw = parse_pct_chain(pcts)
    elif isinstance(pcts, (list, tuple, set)):
        raw = list(pcts)
    else:
        raw = [pcts]
    out: List[float] = []
    for p in raw:
        try:
            f = float(p)
        except:
            continue
        if abs(f) < 1e-9:
            continue
        out.append(round(f, 4))
    return out

def fmt_pct(p: float) -> str:
    if abs(p - round(p)) < 1e-6:
        return str(int(round(p)))
    return f"{p:.2f}".rstrip("0").rstrip(".")

def pct_list_str(pcts: List[float]) -> str:
    if not pcts:
        return ""
    return "+".join([fmt_pct(p) for p in pcts])

def pct_diff(md_parts: List[float], elig_parts: List[float]) -> Tuple[List[float], List[float]]:
    remaining = list(elig_parts)
    matched: List[float] = []
    missing: List[float] = []
    for p in md_parts:
        found = False
        for i, ep in enumerate(remaining):
            if abs(ep - p) < 1e-6:
                found = True
                matched.append(p)
                remaining.pop(i)
                break
        if not found:
            missing.append(p)
    return matched, missing

def fmt_metric(value: float, unit: str = "") -> str:
    v = float(value or 0.0)
    u = norm(unit)
    if u in ["PCS", "CTN", "KRT"]:
        return str(int(round(v)))
    if u == "GROSSAMOUNT":
        return f"{v:,.0f}"
    return f"{v:.2f}".rstrip("0").rstrip(".")

def find_favicon_path() -> Optional[str]:
    path = s(FAVICON_PATH)
    if not path:
        return None
    if os.path.isdir(path):
        for name in ["favicon.ico", "favicon.png", "favicon.jpg", "favicon.jpeg"]:
            p = os.path.join(path, name)
            if os.path.exists(p):
                return p
        return None
    if os.path.exists(path):
        return path
    for ext in [".ico", ".png", ".jpg", ".jpeg"]:
        p = path + ext
        if os.path.exists(p):
            return p
    return None

def favicon_media_type(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".ico":
        return "image/x-icon"
    if ext == ".png":
        return "image/png"
    if ext in [".jpg", ".jpeg"]:
        return "image/jpeg"
    return "application/octet-stream"

def extract_pdf_text(pdf_bytes: bytes) -> str:
    return extract_pdf_text_safe(
        pdf_bytes=pdf_bytes,
        max_pdf_bytes=MAX_PDF_UPLOAD_BYTES,
        max_pages=MAX_PDF_OCR_PAGES,
        total_timeout_seconds=MAX_PDF_OCR_TIMEOUT_SECONDS,
        per_page_ocr_timeout_seconds=MAX_PDF_OCR_PAGE_TIMEOUT_SECONDS,
    )

def normalize_ocr_text(text: str) -> str:
    # basic cleanup for OCR noise
    t = text.replace("\u2013", "-").replace("\u2014", "-")
    t = re.sub(r"[ \t]+", " ", t)
    return t

def find_line_with_pattern(lines: List[str], pattern: str) -> Optional[str]:
    for line in lines:
        if re.search(pattern, line, flags=re.IGNORECASE):
            return line
    return None

def parse_gumindo_program(text: str, list_mode: str) -> Dict[str, str]:
    raw = normalize_ocr_text(text)
    lines = [s(l) for l in raw.splitlines() if s(l)]

    surat_program = ""
    nama_program = ""
    for line in lines:
        m = re.search(r"\b\d{3,4}/[A-Z0-9]+/[A-Z0-9]+/[A-Z0-9]+/\d{2}\b", line, flags=re.IGNORECASE)
        if m:
            surat_program = m.group(0)
            tail = s(line.replace(surat_program, ""))
            if tail:
                nama_program = tail
            break

    periode = ""
    bulan = "(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)"
    for line in lines:
        if re.search(bulan, line, flags=re.IGNORECASE):
            periode = line
            break

    channel = ""
    for line in lines:
        if re.search(r"\b(GT|MT)\b", line, flags=re.IGNORECASE):
            channel = line
            break

    mekanisme_line = find_line_with_pattern(lines, r"Gramasi")
    if not mekanisme_line:
        mekanisme_line = find_line_with_pattern(lines, r"\bgr\b")

    mekanisme_text = ""
    mech_idx = None
    for i, line in enumerate(lines):
        if re.search(r"\bGramasi\b", line, flags=re.IGNORECASE):
            mech_idx = i
            break
    if mech_idx is None:
        for i, line in enumerate(lines):
            if re.search(r"\bgr\b", line, flags=re.IGNORECASE):
                mech_idx = i
                break

    if mech_idx is not None:
        block_lines = []
        for j in range(mech_idx, min(mech_idx + 6, len(lines))):
            line = lines[j]
            if j > mech_idx:
                if re.search(r"klaim.*mencantum|mencantum.*klaim", line, flags=re.IGNORECASE):
                    break
                if re.match(r"^\d+\s+", line):
                    if re.search(r"Surat|Cover|Klaim|Invoice|Harga|Tembusan|Rekap|Proses|Mengirim|Return", line, flags=re.IGNORECASE):
                        break
                if re.search(r"Demikian|Hormat|Tangerang|Terima kasih", line, flags=re.IGNORECASE):
                    break
            block_lines.append(line)
        mekanisme_text = " ".join(block_lines).strip()
    elif mekanisme_line:
        mekanisme_text = mekanisme_line.strip()

    gramasi = ""
    kelompok_barang = ""
    variant = ""
    ketentuan = ""
    benefit = ""

    def extract_ketentuan_benefit_lists(src: str) -> Tuple[List[str], List[str]]:
        if not src:
            return ([], [])
        t = normalize_ocr_text(src)
        t = re.sub(r"[=<>]", " ", t)
        t = re.sub(r"(ctn|krt|pcs)[\.\,]", r"\1", t, flags=re.IGNORECASE)
        t = re.sub(r"(ctn|krt|pcs)\s*u\s*p", r"\1 up", t, flags=re.IGNORECASE)
        t = re.sub(r"(ctn|krt|pcs)\s*up", r"\1 up", t, flags=re.IGNORECASE)

        unit_pat = r"(ctn|krt|pcs)"
        ketentuan_list = []
        ket_re = re.compile(
            r"(?:>=|=>|=|>|<)?\s*(\d+)\s*-\s*(\d+)\s*" + unit_pat + r"\b|(?:>=|=>|=|>|<)?\s*(\d+)\s*" + unit_pat + r"\s*(?:up|\+|ke\s*atas)\b",
            flags=re.IGNORECASE,
        )
        for m in ket_re.finditer(t):
            if m.group(1) and m.group(2) and m.group(3):
                ketentuan_list.append(f"{m.group(1)}-{m.group(2)} {m.group(3).lower()}")
            elif m.group(4) and m.group(5):
                ketentuan_list.append(f"{m.group(4)} {m.group(5).lower()} up")

        rp_vals = []
        for m in re.finditer(r"R\s*P\.?\s*([0-9][0-9\.\,]*)", t, flags=re.IGNORECASE):
            val = s(m.group(1))
            if val:
                rp_vals.append("Rp " + val)

        return (ketentuan_list, rp_vals)

    if mekanisme_text:
        mg = re.search(r"(\d+)\s*gr", mekanisme_text, flags=re.IGNORECASE)
        if mg:
            gramasi = mg.group(1) + " gr"
        after_gr = ""
        if mg:
            after_gr = mekanisme_text[mg.end():].strip()
        mword = re.search(r"\b([A-Za-z][A-Za-z0-9\-]+)\b", after_gr)
        if mword:
            kelompok_barang = mword.group(1)

        if "all variant" in s(nama_program).lower() or "all variant" in s(mekanisme_text).lower():
            variant = "All Variant"

        ket_list, ben_list = extract_ketentuan_benefit_lists(mekanisme_text)
        ketentuan = ", ".join(ket_list) if ket_list else ""
        benefit = ", ".join(ben_list) if ben_list else ""

        if not ketentuan or not benefit:
            rp_parts = re.split(r"\bRp\b", mekanisme_text, flags=re.IGNORECASE)
            if len(rp_parts) > 1:
                if not ketentuan:
                    ketentuan = s(rp_parts[0].replace("Gramasi", ""))
                if not benefit:
                    rp_vals = []
                    for part in rp_parts[1:]:
                        val = s(re.split(r"\s", part, maxsplit=1)[0])
                        if val:
                            rp_vals.append("Rp " + val)
                    if rp_vals:
                        benefit = " / ".join(rp_vals)

    mech_block = ""
    if not ketentuan or not benefit:
        m_start = re.search(r"mekan", raw, flags=re.IGNORECASE)
        if m_start:
            tail = raw[m_start.start():]
            m_end = re.search(r"klaim", tail, flags=re.IGNORECASE)
            mech_block = tail[:m_end.start()] if m_end else tail
        k2, b2 = extract_ketentuan_benefit_lists(mech_block)
        if not ketentuan:
            ketentuan = ", ".join(k2) if k2 else ""
        if not benefit:
            benefit = ", ".join(b2) if b2 else ""

    if not ketentuan or not benefit:
        k3, b3 = extract_ketentuan_benefit_lists(raw)
        if not ketentuan:
            ketentuan = ", ".join(k3) if k3 else ""
        if not benefit:
            benefit = ", ".join(b3) if b3 else ""

    ket_list_final, ben_list_final = extract_ketentuan_benefit_lists(mekanisme_text)
    if not ket_list_final and not ben_list_final:
        ket_list_final, ben_list_final = extract_ketentuan_benefit_lists(mech_block)
    if not ket_list_final and not ben_list_final:
        ket_list_final, ben_list_final = extract_ketentuan_benefit_lists(raw)

    syarat_claim = ""
    idx = None
    for i, line in enumerate(lines):
        if re.search(r"klaim.*mencantum", line, flags=re.IGNORECASE):
            idx = i
            break
    if idx is not None:
        syarat_lines = []
        for j in range(idx + 1, len(lines)):
            if re.search(r"Demikian|Hormat|Tangerang|Terima kasih", lines[j], flags=re.IGNORECASE):
                break
            if len(lines[j]) < 2:
                continue
            syarat_lines.append(lines[j])
        syarat_claim = "; ".join(syarat_lines)

    result = {
        "Surat Program": surat_program,
        "Nama Program": nama_program,
        "GT / MT": channel,
        "ADA LIST / TANPA LIST": list_mode,
        "Periode": periode,
        "Kelompok Barang": kelompok_barang,
        "Variant": variant,
        "Gramasi": gramasi,
        "Ketentuan Pengambilan": ketentuan,
        "Benefit": benefit,
        "Syarat Claim": syarat_claim,
        "Update": pd.Timestamp.today().strftime("%d-%m-%Y"),
        "Keterangan": "",
        "_KET_LIST": ket_list_final,
        "_BEN_LIST": ben_list_final,
    }
    return result

def build_summary_rows(text: str, list_mode: str, template: str) -> List[Dict[str, str]]:
    if template == "GUMINDO":
        base = parse_gumindo_program(text, list_mode)
        ket_list = base.pop("_KET_LIST", []) or []
        ben_list = base.pop("_BEN_LIST", []) or []
        if ket_list or ben_list:
            max_len = max(len(ket_list), len(ben_list))
            rows = []
            for i in range(max_len):
                row = dict(base)
                row["No."] = i + 1
                row["Ketentuan Pengambilan"] = ket_list[i] if i < len(ket_list) else ""
                row["Benefit"] = ben_list[i] if i < len(ben_list) else ""
                rows.append(row)
            return rows
        base["No."] = 1
        return [base]
    return []

def write_summary_excel(rows: List[Dict[str, str]], out_path: str):
    cols = [
        "No.",
        "Surat Program",
        "Nama Program",
        "GT / MT",
        "ADA LIST / TANPA LIST",
        "Periode",
        "Kelompok Barang",
        "Variant",
        "Gramasi",
        "Ketentuan Pengambilan",
        "Benefit",
        "Syarat Claim",
        "Update",
        "Keterangan",
    ]
    df = pd.DataFrame(rows, columns=cols)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SUMMARY")


# ---------------------------
# AI Summary (SumoPod OpenAI-compatible)
# ---------------------------
SUMOPOD_API_KEY = str(os.getenv("SUMOPOD_API_KEY", "")).strip()
SUMOPOD_BASE_URL = str(os.getenv("SUMOPOD_BASE_URL", "https://ai.sumopod.com/v1")).strip()
SUMOPOD_TIMEOUT = int(os.getenv("SUMOPOD_TIMEOUT", "120"))
DEFAULT_SUMMARY_AI_MODEL = str(os.getenv("SUMOPOD_MODEL", "kimi-k2-250905")).strip()

def _sumopod_url(path: str) -> str:
    base = SUMOPOD_BASE_URL.rstrip("/")
    return f"{base}/{path.lstrip('/')}"

def _strip_code_fences(t: str) -> str:
    t = (t or "").strip()
    # remove ```json ... ``` fences
    t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*```$", "", t)
    return t.strip()

def sumopod_chat_completion(
    model: str,
    messages: List[Dict[str, str]],
    temperature: float = 0.0,
    max_tokens: int = 1200,
) -> str:
    """
    Minimal OpenAI-compatible call to SumoPod using requests.
    """
    if not SUMOPOD_API_KEY:
        raise RuntimeError("SUMOPOD_API_KEY belum diset (export SUMOPOD_API_KEY='sk-...').")

    import requests  # requests is already used widely and usually available on VPS

    payload = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    r = requests.post(
        _sumopod_url("/chat/completions"),
        headers={
            "Authorization": f"Bearer {SUMOPOD_API_KEY}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=SUMOPOD_TIMEOUT,
    )
    if not r.ok:
        txt = (r.text or "")[:800]
        raise RuntimeError(f"SumoPod API error {r.status_code}: {txt}")

    data = r.json()
    try:
        return data["choices"][0]["message"]["content"]
    except Exception:
        raise RuntimeError("Format response SumoPod tidak sesuai (choices/message/content).")

def ai_extract_summary_rows(
    text: str,
    list_mode: str,
    template: str,
    model: Optional[str] = None,
) -> List[Dict[str, str]]:
    """
    Extract rows for write_summary_excel() from program letter text.
    Returns list of dicts with keys matching write_summary_excel columns.
    """
    model = s(model) or DEFAULT_SUMMARY_AI_MODEL
    template = s(template).upper()
    list_mode = s(list_mode).upper()

    cols = [
        "No.",
        "Surat Program",
        "Nama Program",
        "GT / MT",
        "ADA LIST / TANPA LIST",
        "Periode",
        "Kelompok Barang",
        "Variant",
        "Gramasi",
        "Ketentuan Pengambilan",
        "Benefit",
        "Syarat Claim",
        "Update",
        "Keterangan",
    ]

    sys_prompt = (
        "Kamu adalah asisten data yang mengekstrak isi surat program promo menjadi tabel ringkas. "
        "Keluaran HARUS JSON valid saja (tanpa penjelasan, tanpa markdown). "
        "Jika ragu, isi string kosong, tapi jangan mengarang angka. "
        "Jika ada beberapa ketentuan/benefit, buat beberapa baris (rows) dengan No. berurutan mulai 1."
    )

    user_prompt = f"""
Template: {template}
Mode List: {list_mode}

Ambil informasi dari teks surat program berikut dan hasilkan JSON dengan format:

{{
  "rows": [
    {{
      "No.": 1,
      "Surat Program": "...",
      "Nama Program": "...",
      "GT / MT": "GT" atau "MT" atau "",
      "ADA LIST / TANPA LIST": "{list_mode}",
      "Periode": "...",
      "Kelompok Barang": "...",
      "Variant": "...",
      "Gramasi": "...",
      "Ketentuan Pengambilan": "...",
      "Benefit": "...",
      "Syarat Claim": "...",
      "Update": "...",
      "Keterangan": "..."
    }}
  ],
  "confidence": 0-100
}}

Aturan:
- Jangan buat field tambahan selain "rows" dan "confidence".
- Setiap baris wajib punya semua kolom di atas (boleh kosong).
- "No." harus integer berurutan mulai 1.
- Jika ada bullet/daftar ketentuan/benefit, pecah menjadi beberapa baris.
- Jangan menebak tanggal/nominal kalau tidak tertulis.

Teks surat:
{text}
""".strip()

    content = sumopod_chat_completion(
        model=model,
        messages=[
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=0.0,
        max_tokens=1600,
    )

    raw = _strip_code_fences(content)
    # try to locate JSON object
    m = re.search(r"\{.*\}", raw, flags=re.S)
    if m:
        raw = m.group(0)

    data = json.loads(raw)
    rows = data.get("rows") if isinstance(data, dict) else None
    if not isinstance(rows, list):
        raise RuntimeError("AI output tidak berisi 'rows' list.")

    out: List[Dict[str, str]] = []
    for i, r in enumerate(rows, start=1):
        if not isinstance(r, dict):
            continue
        row: Dict[str, str] = {}
        for c in cols:
            row[c] = s(r.get(c, ""))
        # enforce No.
        try:
            row["No."] = int(r.get("No.", i))
        except Exception:
            row["No."] = i
        if not row.get("ADA LIST / TANPA LIST"):
            row["ADA LIST / TANPA LIST"] = list_mode
        out.append(row)

    return out


def normalize_permissions(raw) -> Dict[str, Set[str]]:
    perms: Dict[str, Set[str]] = {}
    if isinstance(raw, str):
        raw_text = s(raw)
        if not raw_text:
            return {}
        try:
            raw = json.loads(raw_text)
        except Exception:
            raw = [item.strip() for item in raw_text.replace(";", ",").split(",") if item.strip()]
    if isinstance(raw, dict) and "__custom" in raw:
        raw = raw.get("permissions", {})
    if isinstance(raw, dict):
        for mod, actions in raw.items():
            mod = s(mod).lower()
            if mod not in PERMISSION_MODULES:
                continue
            acts: List[str] = []
            if isinstance(actions, str):
                acts = [a.strip() for a in actions.replace(",", " ").split()]
            elif isinstance(actions, (list, tuple, set)):
                acts = [s(a) for a in actions]
            for a in acts:
                a = s(a).lower()
                if a in PERMISSION_ACTIONS:
                    perms.setdefault(mod, set()).add(a)
        return perms
    if isinstance(raw, (list, tuple, set)):
        for item in raw:
            item = s(item)
            if not item:
                continue
            if ":" in item:
                mod, act = item.split(":", 1)
            elif "." in item:
                mod, act = item.split(".", 1)
            else:
                continue
            mod = s(mod).lower()
            act = s(act).lower()
            if mod in PERMISSION_MODULES and act in PERMISSION_ACTIONS:
                perms.setdefault(mod, set()).add(act)
    return perms

def parse_permission_profile(raw) -> Tuple[Dict[str, Set[str]], bool]:
    if raw is None:
        return ({}, False)
    parsed = raw
    if isinstance(raw, str):
        raw_text = s(raw)
        if not raw_text or raw_text == "{}":
            return ({}, False)
        try:
            parsed = json.loads(raw_text)
        except Exception:
            return (normalize_permissions(raw_text), True)
    if isinstance(parsed, dict) and parsed.get("__custom") is True:
        return (normalize_permissions(parsed.get("permissions", {})), True)
    if isinstance(parsed, dict) and parsed.get("__custom") is False:
        return ({}, False)
    perms = normalize_permissions(parsed)
    return (perms, bool(perms))



def get_auth_user_records() -> Dict[str, Dict[str, str]]:
    # ponytail: store kredensial paralel (users.json + env AUTH_USERS) dimatikan (#7).
    # Identitas & RBAC kini hanya dari Better Auth (sqlite.db user/session). Fungsi
    # dipertahankan agar pemanggil legacy aman; selalu kosong.
    return {}

def get_auth_user_map() -> Dict[str, str]:
    records = get_auth_user_records()
    return {u: s(info.get("password", "")) for u, info in records.items() if s(info.get("password", ""))}

def get_user_role(username: str) -> str:
    username = s(username)
    if not username:
        return "user"
    if username.startswith("betterauth|"):
        parts = username.split("|", 2)
        role = s(parts[1] if len(parts) > 1 else "").lower()
        if role in {"admin", "manager", "finance", "staff", "viewer"}:
            return role
        return "viewer"
    rec = get_auth_user_records().get(username)
    if rec:
        role = s(rec.get("role", ""))
        if role:
            return role
    if username in AUTH_ADMINS:
        return "admin"
    if username in AUTH_FINANCE:
        return "finance"
    return "user"

def _b64url_encode(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).rstrip(b"=").decode("ascii")

def _b64url_decode(sv: str) -> bytes:
    pad = "=" * (-len(sv) % 4)
    return base64.urlsafe_b64decode(sv + pad)

def _normalize_samesite(value: str) -> str:
    v = s(value).lower()
    if v in ["lax", "strict", "none"]:
        return v
    return "strict"


def _pbkdf2_hash(password: str, iterations: Optional[int] = None, salt: Optional[bytes] = None) -> str:
    iters = int(iterations or AUTH_PBKDF2_ITERATIONS)
    salt_bytes = salt or os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt_bytes, iters)
    return f"pbkdf2${iters}${_b64url_encode(salt_bytes)}${_b64url_encode(dk)}"

def _pbkdf2_verify(password: str, stored: str) -> bool:
    parts = s(stored).split("$")
    if len(parts) != 4:
        return False
    try:
        iters = int(parts[1])
        salt = _b64url_decode(parts[2])
        expected = _b64url_decode(parts[3])
    except Exception:
        return False
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iters)
    return hmac.compare_digest(dk, expected)



def make_csrf_token() -> str:
    exp = int(time.time()) + int(CSRF_TTL_SECONDS)
    nonce = _b64url_encode(os.urandom(16))
    payload = f"{nonce}|{exp}"
    sig = hmac.new(AUTH_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{_b64url_encode(payload.encode('utf-8'))}.{sig}"

def validate_csrf_token(token: str) -> bool:
    token = s(token)
    if not token or "." not in token:
        return False
    payload_b64, sig = token.split(".", 1)
    try:
        payload = _b64url_decode(payload_b64).decode("utf-8")
    except Exception:
        return False
    expected_sig = hmac.new(AUTH_SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sig, expected_sig):
        return False
    if "|" not in payload:
        return False
    _nonce, exp_s = payload.rsplit("|", 1)
    try:
        exp = int(exp_s)
    except Exception:
        return False
    return exp >= int(time.time())

def get_or_create_csrf_token(request: Request) -> str:
    token = s(request.cookies.get(CSRF_COOKIE, ""))
    if token and validate_csrf_token(token):
        return token
    return make_csrf_token()

def validate_csrf_request(request: Request, token: str) -> bool:
    token = s(token)
    cookie_token = s(request.cookies.get(CSRF_COOKIE, ""))
    if not token:
        return is_same_origin_request(request)
    if not validate_csrf_token(token):
        return False
    if cookie_token and hmac.compare_digest(token, cookie_token):
        return True
    # Fallback: allow signed token for same-origin requests to reduce false negatives
    # when browser cookies are stale/not yet synchronized.
    return is_same_origin_request(request)


def is_same_origin_request(request: Request) -> bool:
    host = s(request.headers.get("host", ""))
    if not host:
        return False
    host = host.lower()
    for hdr in ["origin", "referer"]:
        raw = s(request.headers.get(hdr, ""))
        if not raw:
            continue
        try:
            p = urlparse(raw)
        except Exception:
            return False
        src_host = s(p.netloc).lower()
        if not src_host:
            return False
        
        # Allow exact match
        if src_host == host:
            continue
            
        # Allow locked local frontend (localhost:3000) to call the local backend.
        src_hostname = src_host.split(":")[0] if ":" in src_host else src_host
        dst_hostname = host.split(":")[0] if ":" in host else host
        if src_host == "localhost:3000" and dst_hostname == "localhost":
            continue

        return False
    return True


# Audit F9/D4: verifikasi sesi via HTTP ke Next (/api/auth/verify) bila AUTH_VERIFY_URL
# di-set — wajib saat DB pindah dari file sqlite lokal (Postgres). Cache TTL 60s per token.
# AUTH_VERIFY_URL kosong = jalur lama (baca sqlite langsung), zero perubahan perilaku.
AUTH_VERIFY_URL = str(os.getenv("AUTH_VERIFY_URL", "")).strip()
_AUTH_VERIFY_CACHE: Dict[str, Tuple[float, Optional[str]]] = {}
# D4: permissions per email hasil verify — pengganti baca kolom user.permissions dari sqlite.
_AUTH_VERIFY_PERMS: Dict[str, Tuple[float, Any]] = {}
_AUTH_VERIFY_TTL = 60.0

def _verify_session_via_next(ba_token: str, raw_cookie: str) -> Optional[str]:
    import time
    cached = _AUTH_VERIFY_CACHE.get(ba_token)
    if cached and cached[0] > time.time():
        return cached[1]
    result: Optional[str] = None
    try:
        import requests
        r = requests.get(AUTH_VERIFY_URL, headers={"cookie": raw_cookie}, timeout=5)
        if r.status_code == 200:
            data = r.json()
            if data.get("ok"):
                role = s(data.get("role")).lower() or "viewer"
                if role not in {"admin", "manager", "finance", "staff", "viewer"}:
                    role = "viewer"
                identity = s(data.get("email")).lower() or s(data.get("name"))
                result = f"betterauth|{role}|{identity}"
                if s(data.get("email")):
                    _AUTH_VERIFY_PERMS[s(data.get("email")).lower()] = (
                        time.time() + _AUTH_VERIFY_TTL, data.get("permissions"))
    except Exception as e:
        print(f"[AUTH VERIFY] gagal panggil {AUTH_VERIFY_URL}: {e}")
        return None  # jangan cache kegagalan network — fallback sqlite di caller
    _AUTH_VERIFY_CACHE[ba_token] = (time.time() + _AUTH_VERIFY_TTL, result)
    return result

def get_current_user(request: Request) -> Optional[str]:
    # ponytail: auth Python paralel dihapus (#7) — satu-satunya sumber identitas
    # adalah sesi Better Auth (cookie better-auth.session_token) divalidasi ke sqlite.db.

    # 2. Try Better-Auth SQLite Session
    ba_token = None
    raw_cookie = request.headers.get("cookie", "")
    if raw_cookie:
        for chunk in raw_cookie.split(";"):
            chunk = chunk.strip()
            if chunk.startswith("better-auth.session_token="):
                ba_token = unquote(chunk.split("=", 1)[1]).split(".")[0]
                break
            elif chunk.startswith("__Secure-better-auth.session_token="):
                ba_token = unquote(chunk.split("=", 1)[1]).split(".")[0]
                break

    # print(f"[DEBUG AUTH] Manual ba_token: {ba_token}")

    if ba_token and AUTH_VERIFY_URL:
        verified = _verify_session_via_next(ba_token, raw_cookie)
        if verified is not None:
            return verified
        # None = network error ATAU sesi invalid; utk sesi invalid cache menyimpan None
        # dan kita tetap coba fallback sqlite di bawah (aman: sqlite juga akan menolak).

    # D4: bila AUTH_VERIFY_URL aktif (DB utama = Postgres), file sqlite lokal stale —
    # JANGAN fallback ke sana (risiko sesi kadaluarsa/di-revoke tetap lolos). Deny saja.
    if ba_token and not AUTH_VERIFY_URL:
        try:
            import sqlite3
            db_path = BETTER_AUTH_DB_PATH
            if os.path.exists(db_path):
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute('''
                    SELECT user.email, user.name, COALESCE(user.role, 'viewer'), session.expiresAt
                    FROM session 
                    JOIN user ON session.userId = user.id 
                    WHERE session.token = ?
                ''', (ba_token,))
                row = c.fetchone()
                conn.close()
                if row:
                    email, name, role, expiresAt = row
                    import time
                    now = time.time()
                    exp = expiresAt / 1000.0 if expiresAt > 20000000000 else expiresAt
                    if exp > now:
                        role = s(role).lower() or "viewer"
                        if role not in {"admin", "manager", "finance", "staff", "viewer"}:
                            role = "viewer"
                        return f"betterauth|{role}|{s(email).lower() or s(name)}"
        except Exception as e:
            print(f"Error checking Better-Auth session: {e}")

    return None

def is_admin_user(username: Optional[str]) -> bool:
    if not username:
        return False
    return get_user_role(username) == "admin"

def is_finance_user(username: Optional[str]) -> bool:
    if not username:
        return False
    return get_user_role(username) == "finance"

def get_user_permissions_info(username: str) -> Tuple[Dict[str, Set[str]], bool]:
    username = s(username)
    if not username:
        return ({}, False)
    if username.startswith("betterauth|"):
        parts = username.split("|", 2)
        email = s(parts[2] if len(parts) > 2 else "").lower()
        if not email:
            return ({}, False)
        # D4: jalur verify — permissions ikut respons /api/auth/verify (cache 60s,
        # terisi oleh _verify_session_via_next pada request yang sama). Sqlite stale dilewati.
        if AUTH_VERIFY_URL:
            import time
            cached = _AUTH_VERIFY_PERMS.get(email)
            if cached and cached[0] > time.time():
                return parse_permission_profile(cached[1])
            return ({}, False)
        try:
            import sqlite3
            if not os.path.exists(BETTER_AUTH_DB_PATH):
                return ({}, False)
            conn = sqlite3.connect(BETTER_AUTH_DB_PATH)
            c = conn.cursor()
            c.execute("SELECT permissions FROM user WHERE lower(email) = ? LIMIT 1", (email,))
            row = c.fetchone()
            conn.close()
            if not row:
                return ({}, False)
            return parse_permission_profile(row[0])
        except Exception as e:
            append_error_log("get_user_permissions_info_betterauth", e, {"email": email})
            return ({}, False)
    rec = get_auth_user_records().get(username)
    if not rec:
        return ({}, False)
    raw = rec.get("permissions", None)
    return parse_permission_profile(raw)

def user_has_permission(username: Optional[str], module: str, action: str) -> bool:
    if not username:
        return False
    if is_admin_user(username):
        return True
    module = s(module).lower()
    action = s(action).lower()
    role = get_user_role(username)
    role_permissions = {
        "admin": {mod: set(PERMISSION_ACTIONS) for mod in PERMISSION_MODULES},
        "manager": {
            "dashboard": {"view"},
            "api_wrapper": {"view", "execute"},
            "payments": {"view", "export", "submit", "edit", "update"},
            "sppd": {"view", "generate", "download"},
            "finance": {"view", "approve", "export"},
            "principles": {"view"},
            "summary": {"view", "export"},
            "validator": {"view", "download"},
        },
        "finance": {
            "dashboard": {"view"},
            "payments": {"view", "export"},
            "sppd": {"view", "download"},
            "finance": {"view", "approve", "transfer", "upload_proof", "post_accurate", "retry_post", "export", "update", "edit"},
            "principles": {"view"},
        },
        "staff": {
            "dashboard": {"view"},
            "payments": {"view", "create", "edit", "update", "upload", "submit"},
            "sppd": {"view", "generate", "download"},
            "principles": {"view"},
            "summary": {"view", "upload", "generate", "export", "edit", "update"},
            "validator": {"view", "upload", "run", "download", "edit"},
        },
        "viewer": {
            "dashboard": {"view"},
            "validator": {"view"},
            "summary": {"view"},
            "payments": {"view"},
            "sppd": {"view"},
            "finance": {"view"},
        },
    }
    perms, defined = get_user_permissions_info(username)
    if defined:
        allowed = perms.get(module, set())
        return action in allowed
    if role in role_permissions:
        return action in role_permissions.get(role, {}).get(module, set())
    if module == "finance":
        return is_finance_user(username)
    return True

def format_permissions(perms: Optional[Dict[str, Set[str]]], defined: bool) -> str:
    if not defined:
        return "default"
    if not perms:
        return "none"
    parts = []
    for mod in PERMISSION_MODULES:
        acts = sorted(perms.get(mod, set()))
        if acts:
            parts.append(f"{mod}:{','.join(acts)}")
    return "; ".join(parts) if parts else "none"

def append_audit_log(user: str, action: str, entity: str, details: Optional[Dict[str, Any]] = None) -> None:
    if not AUDIT_LOG_PATH:
        return
    entry = {
        "ts": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": s(user),
        "action": s(action),
        "entity": s(entity),
        "details": details or {},
    }
    try:
        os.makedirs(os.path.dirname(AUDIT_LOG_PATH), exist_ok=True)
        with open(AUDIT_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=True) + "\n")
    except Exception:
        pass


def append_error_log(where: str, err: Exception, context: Optional[Dict[str, Any]] = None) -> None:
    if not ERROR_LOG_PATH:
        return
    entry = {
        "ts": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "where": s(where),
        "error": s(str(err)),
        "context": context or {},
        "trace": traceback.format_exc(),
    }
    try:
        os.makedirs(os.path.dirname(ERROR_LOG_PATH), exist_ok=True)
        with open(ERROR_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=True) + "\n")
    except Exception:
        pass

# ---------------------------
# Payments (LPB) helpers
# ---------------------------
_PAYMENTS_DB_CACHE: Optional[Dict[str, Any]] = None
_PAYMENTS_DB_MTIME: float = 0.0
_PAYMENTS_DB_EMPTY: Dict[str, Any] = {"lpb": {}, "submissions": {}, "drafts": {}, "finance_mappings": {}, "proofs": {}, "sppd_settings": {}}
# ponytail: satu lock global cukup karena payments.json adalah satu file tunggal.
# Ceiling: serializes semua write — tidak ada parallelism untuk mutation routes.
# Upgrade path: per-resource lock jika ada multiple JSON stores.
_PAYMENTS_DB_LOCK: asyncio.Lock = asyncio.Lock()

def load_payments_db() -> Dict[str, Any]:
    global _PAYMENTS_DB_CACHE, _PAYMENTS_DB_MTIME
    if not PAYMENTS_DB_PATH or not os.path.exists(PAYMENTS_DB_PATH):
        return dict(_PAYMENTS_DB_EMPTY)
    try:
        mtime = os.path.getmtime(PAYMENTS_DB_PATH)
        if _PAYMENTS_DB_CACHE is not None and mtime == _PAYMENTS_DB_MTIME:
            return _PAYMENTS_DB_CACHE
        with open(PAYMENTS_DB_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return dict(_PAYMENTS_DB_EMPTY)
    if not isinstance(data, dict):
        return dict(_PAYMENTS_DB_EMPTY)
    data.setdefault("lpb", {})
    data.setdefault("submissions", {})
    data.setdefault("drafts", {})
    data.setdefault("finance_mappings", {})
    data.setdefault("proofs", {})
    data.setdefault("sppd_settings", {})
    _PAYMENTS_DB_CACHE = data
    _PAYMENTS_DB_MTIME = mtime
    return data

def save_payments_db(data: Dict[str, Any]) -> None:
    global _PAYMENTS_DB_CACHE, _PAYMENTS_DB_MTIME
    if not PAYMENTS_DB_PATH:
        return
    os.makedirs(os.path.dirname(PAYMENTS_DB_PATH), exist_ok=True)
    tmp_path = PAYMENTS_DB_PATH + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=True, indent=2)
    os.replace(tmp_path, PAYMENTS_DB_PATH)
    _PAYMENTS_DB_CACHE = data
    _PAYMENTS_DB_MTIME = os.path.getmtime(PAYMENTS_DB_PATH)

async def load_and_lock_payments_db():
    """Acquire _PAYMENTS_DB_LOCK lalu load payments.json.

    Gunakan sebagai async context manager di semua route yang melakukan
    read-modify-write agar tidak ada dua request concurrent yang menulis
    snapshot berbeda ke file yang sama (last-writer-wins silent data loss).

    Contoh pemakaian:
        async with load_and_lock_payments_db() as db:
            db["lpb"][key] = rec
            save_payments_db(db)
    """
    # ponytail: yield inside asyncio.Lock context — satu lock serializes semua
    # mutation. Ceiling: throughput mutation routes turun jadi sequential.
    # Upgrade: per-key optimistic locking jika contention terbukti tinggi.
    class _Ctx:
        def __init__(self):
            self._db: Dict[str, Any] = {}
        async def __aenter__(self) -> Dict[str, Any]:
            await _PAYMENTS_DB_LOCK.acquire()
            self._db = load_payments_db()
            return self._db
        async def __aexit__(self, exc_type, exc, tb):
            _PAYMENTS_DB_LOCK.release()
    return _Ctx()

def empty_payments_db_preserving_config(db: Dict[str, Any]) -> Dict[str, Any]:
    db = db if isinstance(db, dict) else {}
    cleared: Dict[str, Any] = {
        "lpb": {},
        "submissions": {},
        "drafts": {},
        "finance_mappings": db.get("finance_mappings", {}),
        "proofs": {},
        "sppd_settings": db.get("sppd_settings", {}),
    }
    if "sppd_seq" in db:
        cleared["sppd_seq"] = db.get("sppd_seq")
    return cleared

def normalize_lpb_no(no_lpb: str) -> str:
    return s(no_lpb).upper()


def normalize_pengajuan_type(raw: str) -> str:
    t = s(raw).upper()
    if t in ["LPB", "CBD", "NON_LPB"]:
        return t
    return "LPB"


def make_payment_record_id(prefix: str = "RID") -> str:
    base = re.sub(r"[^A-Z0-9_]+", "", s(prefix).upper()) or "RID"
    return f"{base}_{str(uuid.uuid4())[:8]}".upper()


def resolve_payment_record_key(db: Dict[str, Any], row_id: str) -> str:
    row_id = s(row_id)
    if not row_id:
        return ""
    if row_id in db.get("lpb", {}):
        return row_id
    row_norm = normalize_lpb_no(row_id)
    if row_norm in db.get("lpb", {}):
        return row_norm
    for k, r in db.get("lpb", {}).items():
        if normalize_lpb_no(s(r.get("no_lpb", ""))) == row_norm:
            return k
    return ""


def find_lpb_duplicate_key(db: Dict[str, Any], no_lpb: str, exclude_key: str = "") -> str:
    target = normalize_lpb_no(no_lpb)
    if not target:
        return ""
    exclude_key = s(exclude_key)
    for k, r in db.get("lpb", {}).items():
        if exclude_key and k == exclude_key:
            continue
        existing = normalize_lpb_no(s(r.get("no_lpb", "")) or s(k))
        if existing == target:
            return k
    return ""


def has_submitted_duplicate_payment(db: Dict[str, Any], current_key: str, rec: Dict[str, Any]) -> bool:
    principle = s(rec.get("principle", "")).upper()
    invoice_no = s(rec.get("invoice_no", "")).upper()
    nilai_invoice = parse_number_id(rec.get("nilai_invoice", rec.get("nilai_principle", 0)))
    for k, other in db.get("lpb", {}).items():
        if s(k) == s(current_key):
            continue
        if not s(other.get("submission_id", "")):
            continue
        if s(other.get("status_pembayaran", "")).lower() == "ajukan ulang":
            continue
        if normalize_pengajuan_type(other.get("tipe_pengajuan", "LPB")) != "CBD":
            continue
        if s(other.get("principle", "")).upper() != principle:
            continue
        other_invoice = s(other.get("invoice_no", "")).upper()
        if invoice_no and other_invoice and invoice_no == other_invoice:
            return True
        other_paid = parse_number_id(other.get("nilai_pembayaran", other.get("nilai_invoice", other.get("nilai_principle", 0))))
        if (not invoice_no or not other_invoice) and abs(float(other_paid or 0.0) - float(nilai_invoice or 0.0)) <= 1.0:
            return True
    return False

def to_date_str(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return s(val)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return s(val)

def parse_sppd_date_ddmmyyyy(value) -> str:
    """Normalkan tanggal import Excel SPPD ke YYYY-MM-DD dengan asumsi urutan
    DD/MM/YYYY untuk string. Tolak tanggal mustahil/rusak (raise ValueError).

    - datetime/Timestamp asli (sel tanggal Excel) -> dipakai apa adanya (tak ambigu).
    - "YYYY/MM/DD" (tahun 4-digit di depan); tukar ke YYYY/DD/MM hanya bila bulan>12.
    - "DD/MM/YYYY" (tahun 4-digit di belakang); default day-first. Tukar ke MM/DD
      hanya bila tak ambigu (bulan-slot >12 & hari-slot <=12).
    - Ambigu (keduanya <=12) -> tetap DD/MM (tidak ditukar).
    - Pemisah `-`, `/`, atau `.`. Format/komponen lain -> ValueError.
    """
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    from datetime import datetime as _dt, date as _date
    if isinstance(value, (_dt, _date)):
        return value.strftime("%Y-%m-%d")
    raw = s(value).strip()
    if not raw:
        return ""
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", raw)
    if m:
        y, a, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
        mo, da = a, b
        if a > 12 and b <= 12:  # YYYY/DD/MM -> tukar
            mo, da = b, a
    else:
        m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", raw)
        if not m:
            raise ValueError(f"'{raw}'")
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        da, mo = a, b  # default DD/MM
        if b > 12 and a <= 12:  # sebenarnya MM/DD -> tukar
            da, mo = b, a
    try:
        parsed = _date(y, mo, da)
    except ValueError:
        raise ValueError(f"'{raw}'")
    return parsed.strftime("%Y-%m-%d")

def parse_lpb_upload(content: bytes) -> List[Dict[str, Any]]:
    df = pd.read_excel(io.BytesIO(content))
    cols = {c.strip().upper(): c for c in df.columns}
    required = ["TGL. SETOR", "NO. LPB", "TGL. WIN", "TGL. J. TEMPO WIN", "PRINCIPLE", "NILAI WIN", "TGL TERIMA BARANG"]
    missing = [c for c in required if c not in cols]
    if missing:
        raise ValueError("Kolom wajib tidak lengkap: " + ", ".join(missing))

    out = []
    date_errors: List[str] = []

    def parse_lpb_upload_date(row, excel_row: int, *names: str) -> str:
        col = _col_lookup(cols, *names)
        if col is None:
            return ""
        try:
            return parse_sppd_date_ddmmyyyy(row[col])
        except ValueError as de:
            date_errors.append(f"kolom '{s(col)}' baris {excel_row} = {str(de)}")
            return ""

    for idx, (_, r) in enumerate(df.iterrows()):
        excel_row = idx + 2
        no_lpb = s(r[cols["NO. LPB"]])
        if not no_lpb:
            continue
        nilai_win = parse_number_id(r[cols["NILAI WIN"]])
        nilai_invoice = parse_number_id(_row_value(r, cols, "Nilai Invoice", "NILAI INVOICE", default=0))
        tgl_setor = parse_lpb_upload_date(r, excel_row, "TGL. SETOR")
        tgl_win = parse_lpb_upload_date(r, excel_row, "TGL. WIN")
        tgl_jtempo_win = parse_lpb_upload_date(r, excel_row, "TGL. J. TEMPO WIN")
        tgl_terima_barang = parse_lpb_upload_date(r, excel_row, "TGL TERIMA BARANG")
        tgl_invoice = parse_lpb_upload_date(r, excel_row, "Tgl Invoice", "TGL INVOICE", "TGL. INVOICE")
        jt_invoice = parse_lpb_upload_date(r, excel_row, "J.T Invoice", "J.T INVOICE", "JT INVOICE", "JATUH TEMPO INVOICE")
        actual_date = parse_lpb_upload_date(r, excel_row, "Actual Date", "ACTUAL DATE")
        tgl_pembayaran = parse_lpb_upload_date(r, excel_row, "Tgl Pembayaran", "TGL PEMBAYARAN")
        out.append({
            "no_lpb": no_lpb,
            "tgl_setor": tgl_setor,
            "tgl_win": tgl_win,
            "tgl_jtempo_win": tgl_jtempo_win,
            "principle": s(r[cols["PRINCIPLE"]]),
            "nilai_win": nilai_win,
            "tgl_terima_barang": tgl_terima_barang,
            "tgl_invoice": tgl_invoice,
            "invoice_no": s(_row_value(r, cols, "No Invoice", "NO INVOICE", "NO. INVOICE")),
            "nilai_invoice": nilai_invoice if nilai_invoice > 0 else "",
            "jt_invoice": jt_invoice,
            "actual_date": actual_date,
            "tgl_pembayaran": tgl_pembayaran,
            "gap_nilai": (nilai_win - nilai_invoice) if nilai_invoice > 0 else 0.0,
            "jenis_dokumen": s(_row_value(r, cols, "Jenis Dokumen", "JENIS DOKUMEN")),
            "nomor_dokumen": s(_row_value(r, cols, "Nomor Dokumen", "NOMOR DOKUMEN")),
            "keterangan": s(_row_value(r, cols, "Keterangan", "KETERANGAN")),
        })
    if date_errors:
        shown = "; ".join(date_errors[:5])
        extra = len(date_errors) - 5
        suffix = f"; dan {extra} lainnya" if extra > 0 else ""
        raise ValueError(
            f"Tanggal tidak valid (harus DD/MM/YYYY): {shown}{suffix}. Upload dibatalkan."
        )
    return out

def _col_lookup(cols: Dict[str, Any], *names: str) -> Optional[Any]:
    for name in names:
        key = str(name).strip().upper()
        if key in cols:
            return cols[key]
    return None

def to_datetime_str(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return s(val)
        if dt.hour or dt.minute or dt.second:
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return s(val)

def _row_value(row, cols: Dict[str, Any], *names: str, default: Any = "") -> Any:
    col = _col_lookup(cols, *names)
    if col is None:
        return default
    try:
        return row[col]
    except Exception:
        return default

def looks_like_payments_backup(cols: Dict[str, Any]) -> bool:
    required = ["RECORD ID", "TIPE PENGAJUAN", "PRINCIPLE", "NILAI INVOICE", "STATUS PEMBAYARAN"]
    return all(c in cols for c in required)

def parse_payments_backup_upload(content: bytes) -> List[Tuple[str, Dict[str, Any]]]:
    xl = pd.ExcelFile(io.BytesIO(content))
    sheet_name = "PAYMENTS" if "PAYMENTS" in xl.sheet_names else xl.sheet_names[0]
    df = xl.parse(sheet_name)
    cols = {str(c).strip().upper(): c for c in df.columns}
    if not looks_like_payments_backup(cols):
        raise ValueError("Format backup PAYMENTS tidak valid.")

    rows: List[Tuple[str, Dict[str, Any]]] = []
    date_errors: List[str] = []

    def parse_backup_date(r, excel_row: int, *names: str) -> str:
        col = _col_lookup(cols, *names)
        if col is None:
            return ""
        try:
            return parse_sppd_date_ddmmyyyy(r[col])
        except ValueError as de:
            date_errors.append(f"kolom '{s(col)}' baris {excel_row} = {str(de)}")
            return ""

    for idx, (_, r) in enumerate(df.iterrows()):
        excel_row = idx + 2
        raw_key = s(_row_value(r, cols, "Record ID"))
        tipe = normalize_pengajuan_type(_row_value(r, cols, "Tipe Pengajuan", default="LPB"))
        no_lpb = s(_row_value(r, cols, "No LPB"))
        principle = s(_row_value(r, cols, "Principle"))
        invoice_no = s(_row_value(r, cols, "No Invoice"))
        jenis_dokumen = s(_row_value(r, cols, "Jenis Dokumen"))
        nomor_dokumen = s(_row_value(r, cols, "Nomor Dokumen"))
        if not raw_key:
            raw_key = normalize_lpb_no(no_lpb) if no_lpb else make_payment_record_id(tipe)
        key = raw_key
        if not key:
            continue
        nilai_win = parse_number_id(_row_value(r, cols, "Nilai Sistem", "Nilai WIN", default=0))
        nilai_invoice = parse_number_id(_row_value(r, cols, "Nilai Invoice", default=0))
        potongan = parse_number_id(_row_value(r, cols, "Potongan", default=0))
        nilai_pembayaran = parse_number_id(_row_value(r, cols, "Nilai Pembayaran", default=0))
        gap_nilai = parse_number_id(_row_value(r, cols, "Gap Nilai", default=nilai_win - nilai_invoice))
        rec = {
            "record_id": key,
            "tipe_pengajuan": tipe,
            "no_lpb": no_lpb,
            "tgl_setor": parse_backup_date(r, excel_row, "Tgl Setor", "TGL. SETOR"),
            "tgl_win": parse_backup_date(r, excel_row, "Tgl Win", "TGL. WIN"),
            "tgl_jtempo_win": parse_backup_date(r, excel_row, "Tgl J.Tempo Win", "TGL. J. TEMPO WIN"),
            "principle": principle,
            "nilai_win": nilai_win,
            "tgl_terima_barang": parse_backup_date(r, excel_row, "Tgl Terima Barang", "TGL TERIMA BARANG"),
            "tgl_invoice": parse_backup_date(r, excel_row, "Tgl Invoice"),
            "jt_invoice": parse_backup_date(r, excel_row, "J.T Invoice"),
            "tgl_pembayaran": parse_backup_date(r, excel_row, "Tgl Pembayaran"),
            "actual_date": parse_backup_date(r, excel_row, "Actual Date"),
            "nilai_invoice": nilai_invoice,
            "gap_nilai": gap_nilai,
            "invoice_no": invoice_no,
            "status_pembayaran": s(_row_value(r, cols, "Status Pembayaran")),
            "payment_method": s(_row_value(r, cols, "Metode Pembayaran")),
            "submitted_at": to_datetime_str(_row_value(r, cols, "Submitted At")),
            "submitted_by": s(_row_value(r, cols, "Submitted By")),
            "submission_id": s(_row_value(r, cols, "Submission ID")),
            "draft_id": s(_row_value(r, cols, "Draft ID")),
            "potongan": potongan,
            "nilai_pembayaran": nilai_pembayaran,
            "target_payment_date": parse_backup_date(r, excel_row, "Tanggal Pengajuan Pembayaran"),
            "jenis_pembayaran": s(_row_value(r, cols, "Jenis Pembayaran")),
            "jenis_dokumen": jenis_dokumen,
            "nomor_dokumen": nomor_dokumen,
            "keterangan": s(_row_value(r, cols, "Keterangan")),
            "sppd_no": s(_row_value(r, cols, "SPPD No")),
            "created_at": to_datetime_str(_row_value(r, cols, "Created At")),
            "created_by": s(_row_value(r, cols, "Created By")),
        }
        rows.append((key, rec))
    if date_errors:
        shown = "; ".join(date_errors[:5])
        extra = len(date_errors) - 5
        suffix = f"; dan {extra} lainnya" if extra > 0 else ""
        raise ValueError(
            f"Tanggal tidak valid (harus DD/MM/YYYY): {shown}{suffix}. Restore backup dibatalkan."
        )
    return rows

def max_sppd_sequence_from_records(records: List[Dict[str, Any]]) -> int:
    max_seq = 0
    for rec in records:
        sppd_no = s(rec.get("sppd_no", ""))
        m = re.match(r"^\s*(\d+)\s*/", sppd_no)
        if not m:
            continue
        try:
            max_seq = max(max_seq, int(m.group(1)))
        except Exception:
            continue
    return max_seq

def rebuild_payment_submissions(db: Dict[str, Any]) -> None:
    submissions = dict(db.get("submissions", {}) or {})
    grouped: Dict[str, List[Tuple[str, Dict[str, Any]]]] = {}
    for key, rec in db.get("lpb", {}).items():
        submission_id = s(rec.get("submission_id", ""))
        if submission_id:
            grouped.setdefault(submission_id, []).append((s(key), rec))
    for submission_id, items in grouped.items():
        if submission_id in submissions:
            continue
        first = items[0][1]
        payment_method = s(first.get("payment_method", ""))
        method = "BANK_PANIN" if payment_method.lower() == "bank panin" else "NON_PANIN"
        submissions[submission_id] = {
            "id": submission_id,
            "created_at": s(first.get("submitted_at", "")),
            "created_by": s(first.get("submitted_by", "")),
            "draft_id": s(first.get("draft_id", "")),
            "method": method,
            "target_payment_date": s(first.get("target_payment_date", "")),
            "record_ids": [key for key, _ in items],
            "files": [],
            "sppd_file": "",
            "sppd_no": s(first.get("sppd_no", "")),
            "cart_items": {},
        }
    db["submissions"] = submissions

def validate_backup_restore_conflicts(db: Dict[str, Any], rows: List[Tuple[str, Dict[str, Any]]]) -> List[str]:
    conflicts: List[str] = []
    seen_keys: Set[str] = set()
    seen_lpb: Dict[str, str] = {}
    for key, rec in rows:
        if key in seen_keys:
            conflicts.append(f"Record ID duplikat di file: {key}")
        seen_keys.add(key)
        if key in db.get("lpb", {}):
            conflicts.append(f"Record ID sudah ada: {key}")
        no_lpb = s(rec.get("no_lpb", ""))
        if no_lpb:
            norm_no = normalize_lpb_no(no_lpb)
            if norm_no in seen_lpb and seen_lpb[norm_no] != key:
                conflicts.append(f"No LPB duplikat di file: {no_lpb}")
            seen_lpb[norm_no] = key
            dup_key = find_lpb_duplicate_key(db, no_lpb)
            if dup_key:
                conflicts.append(f"No LPB sudah ada di sistem: {no_lpb}")
        if len(conflicts) >= 10:
            break
    return conflicts

def slugify(text: str) -> str:
    t = re.sub(r"[^A-Za-z0-9]+", "-", s(text).strip())
    return t.strip("-").lower() or "item"

def format_idr(val: float) -> str:
    try:
        return f"{float(val):,.0f}".replace(",", ".")
    except Exception:
        return "0"

def format_idr_decimal(val: float) -> str:
    try:
        s = f"{float(val):,.2f}"
        return s.replace(",", "_").replace(".", ",").replace("_", ".")
    except Exception:
        return "0,00"

def format_id_date(dt: Optional[pd.Timestamp]) -> str:
    if dt is None or (isinstance(dt, float) and pd.isna(dt)):
        return ""
    try:
        d = pd.to_datetime(dt)
    except Exception:
        return s(dt)
    months = [
        "Januari","Februari","Maret","April","Mei","Juni",
        "Juli","Agustus","September","Oktober","November","Desember"
    ]
    return f"{d.day} {months[d.month-1]} {d.year}"

def roman_month(n: int) -> str:
    romans = ["I","II","III","IV","V","VI","VII","VIII","IX","X","XI","XII"]
    if n < 1 or n > 12:
        return ""
    return romans[n - 1]

def format_sppd_number(seq: int, dt: pd.Timestamp) -> str:
    return format_sppd_number_with_template(seq, dt, "{seq:03d}/SPA/PDSB/{roman_month}/{year}")

def format_sppd_number_with_template(seq: int, dt: pd.Timestamp, template: str) -> str:
    num = f"{int(seq):03d}"
    month = roman_month(int(dt.month))
    year = dt.year
    tpl = s(template) or "{seq:03d}/SPA/PDSB/{roman_month}/{year}"
    try:
        return tpl.format(
            seq=int(seq),
            seq3=num,
            roman_month=month,
            month=int(dt.month),
            year=year,
            yy=str(year)[-2:],
        )
    except Exception:
        return f"{num}/SPA/PDSB/{month}/{year}"

def default_sppd_settings(db: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    try:
        start_seq = int(os.getenv("SPPD_SEQ_START", "6"))
    except Exception:
        start_seq = 6
    legacy_last = start_seq - 1
    if isinstance(db, dict) and "sppd_seq" in db:
        try:
            legacy_last = int(db.get("sppd_seq", legacy_last))
        except Exception:
            pass
    return {
        "last_sequence": max(0, legacy_last),
        "number_template": "{seq:03d}/SPA/PDSB/{roman_month}/{year}",
        "fixed_jaminan_date": "2026-02-19",
        "maturity_months": 6,
        "items_per_page": max(1, int(os.getenv("SPPD_TRANSFER_ITEMS_PER_PAGE", "7"))),
    }

def normalize_sppd_settings(raw: Dict[str, Any], db: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    defaults = default_sppd_settings(db)
    raw = raw if isinstance(raw, dict) else {}
    settings = {**defaults, **raw}
    try:
        settings["last_sequence"] = max(0, int(settings.get("last_sequence", defaults["last_sequence"])))
    except Exception:
        settings["last_sequence"] = defaults["last_sequence"]
    settings["number_template"] = s(settings.get("number_template", defaults["number_template"])) or defaults["number_template"]
    fixed_date = _normalize_yyyy_mm_dd(s(settings.get("fixed_jaminan_date", defaults["fixed_jaminan_date"])))
    settings["fixed_jaminan_date"] = fixed_date or defaults["fixed_jaminan_date"]
    try:
        settings["maturity_months"] = max(1, min(24, int(settings.get("maturity_months", defaults["maturity_months"]))))
    except Exception:
        settings["maturity_months"] = defaults["maturity_months"]
    try:
        settings["items_per_page"] = max(1, min(20, int(settings.get("items_per_page", defaults["items_per_page"]))))
    except Exception:
        settings["items_per_page"] = defaults["items_per_page"]
    return settings

def get_sppd_settings(db: Dict[str, Any]) -> Dict[str, Any]:
    settings = normalize_sppd_settings(db.get("sppd_settings", {}), db)
    db["sppd_settings"] = settings
    return settings

def next_sppd_number(db: Dict[str, Any], dt: pd.Timestamp) -> Tuple[int, str, Dict[str, Any]]:
    settings = get_sppd_settings(db)
    next_seq = int(settings.get("last_sequence", 0)) + 1
    settings["last_sequence"] = next_seq
    db["sppd_seq"] = next_seq
    return next_seq, format_sppd_number_with_template(next_seq, dt, s(settings.get("number_template", ""))), settings

def _docx_paragraph_text(p: ET.Element, ns: Dict[str, str]) -> str:
    texts = []
    for t in p.findall(".//w:t", ns):
        texts.append(t.text or "")
    return "".join(texts)

def _docx_set_paragraph_text(p: ET.Element, text: str, ns: Dict[str, str]) -> None:
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return
    t_nodes[0].text = text
    for t in t_nodes[1:]:
        t.text = ""

def _docx_clear_highlights(element: ET.Element, ns: Dict[str, str]) -> None:
    for r_pr in element.findall(".//w:rPr", ns):
        for highlight in list(r_pr.findall("w:highlight", ns)):
            r_pr.remove(highlight)

def _docx_set_prefix_value(p: ET.Element, prefix: str, value: str, ns: Dict[str, str]) -> None:
    _docx_clear_highlights(p, ns)
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return
    if len(t_nodes) == 1:
        t_nodes[0].text = f"{prefix}{value}"
        return
    t_nodes[0].text = prefix
    t_nodes[1].text = value
    for t in t_nodes[2:]:
        t.text = ""

def _docx_set_field_value(p: ET.Element, value: str, ns: Dict[str, str], fallback_text: str) -> None:
    _docx_clear_highlights(p, ns)
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return
    colon_idx = None
    for idx, node in enumerate(t_nodes):
        if ":" in (node.text or ""):
            colon_idx = idx
    if colon_idx is None:
        _docx_set_paragraph_text(p, fallback_text, ns)
        return
    node = t_nodes[colon_idx]
    before = (node.text or "").split(":", 1)[0]
    node.text = f"{before}: {value}"
    for trailing in t_nodes[colon_idx + 1:]:
        trailing.text = ""

def _docx_replace_date_after_keyword(p: ET.Element, keyword: str, date_text: str, ns: Dict[str, str]) -> bool:
    _docx_clear_highlights(p, ns)
    t_nodes = p.findall(".//w:t", ns)
    if not t_nodes:
        return False
    chunks = []
    pos = 0
    full = ""
    for node in t_nodes:
        text = node.text or ""
        start = pos
        full += text
        pos += len(text)
        chunks.append((node, start, pos))
    keyword_idx = full.lower().find(keyword.lower())
    if keyword_idx < 0:
        return False
    match = re.search(r"\d{1,2}\s+[A-Za-zÀ-ÿ]+\s+\d{4}", full[keyword_idx:])
    if not match:
        return False
    start = keyword_idx + match.start()
    end = keyword_idx + match.end()
    first_node: Optional[ET.Element] = None
    for node, node_start, node_end in chunks:
        if node_end <= start or node_start >= end:
            continue
        before = (node.text or "")[:max(0, start - node_start)] if node_start <= start < node_end else ""
        after = (node.text or "")[max(0, end - node_start):] if node_start < end <= node_end else ""
        if first_node is None:
            node.text = f"{before}{date_text}{after}"
            first_node = node
        else:
            node.text = after
    return first_node is not None

def _docx_register_namespaces(xml_bytes: bytes) -> Dict[str, str]:
    namespace_map: Dict[str, str] = {}
    seen: Set[str] = set()
    for _, ns_def in ET.iterparse(io.BytesIO(xml_bytes), events=("start-ns",)):
        prefix, uri = ns_def
        namespace_map[prefix] = uri
        if prefix in seen:
            continue
        seen.add(prefix)
        try:
            ET.register_namespace(prefix, uri)
        except ValueError:
            continue
    return namespace_map

def _docx_preserve_namespace_declarations(xml_bytes: bytes, namespace_map: Dict[str, str]) -> bytes:
    text = xml_bytes.decode("utf-8")
    root_start = 0
    if text.startswith("<?xml"):
        root_start = text.find("<", text.find("?>") + 2)
    root_end = text.find(">", root_start)
    if root_start < 0 or root_end < 0:
        return xml_bytes
    root_open = text[root_start:root_end]
    additions = []
    for prefix, uri in namespace_map.items():
        if prefix == "xml":
            continue
        if prefix:
            marker = f"xmlns:{prefix}="
            attr = f' xmlns:{prefix}="{uri}"'
        else:
            marker = "xmlns="
            attr = f' xmlns="{uri}"'
        if marker not in root_open:
            additions.append(attr)
    if not additions:
        return xml_bytes
    return (text[:root_end] + "".join(additions) + text[root_end:]).encode("utf-8")

def _docx_page_break_paragraph(ns: Dict[str, str]) -> ET.Element:
    w_ns = ns["w"]
    p = ET.Element(f"{{{w_ns}}}p")
    r = ET.SubElement(p, f"{{{w_ns}}}r")
    br = ET.SubElement(r, f"{{{w_ns}}}br")
    br.set(f"{{{w_ns}}}type", "page")
    return p

def _label_prefix(text: str, default: str) -> str:
    if ":" in text:
        return text.split(":", 1)[0].strip()
    return default

def render_sppd_docx(
    template_path: str,
    out_path: str,
    submit_dt: pd.Timestamp,
    sppd_no: str,
    transfer_items: List[Dict[str, Any]],
    settings: Optional[Dict[str, Any]] = None,
) -> None:
    if not os.path.exists(template_path):
        raise FileNotFoundError("Template SPPD tidak ditemukan.")
    with zipfile.ZipFile(template_path) as z:
        data = {name: z.read(name) for name in z.namelist()}
    if "word/document.xml" not in data:
        raise ValueError("Template SPPD tidak valid.")

    namespace_map = _docx_register_namespaces(data["word/document.xml"])
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    root = ET.fromstring(data["word/document.xml"])
    body = root.find(".//w:body", ns)
    if body is None:
        raise ValueError("Template SPPD tidak valid.")
    settings = normalize_sppd_settings(settings or {})
    _docx_clear_highlights(root, ns)

    paras = list(body.findall("w:p", ns))
    texts = [_docx_paragraph_text(p, ns) for p in paras]

    # Update tanggal surat (Makassar, ...)
    tanggal_surat = format_id_date(submit_dt)
    for p, t in zip(paras, texts):
        if t.strip().startswith("Makassar"):
            _docx_set_prefix_value(p, "Makassar, ", tanggal_surat, ns)
            break

    # Update nomor surat
    for p, t in zip(paras, texts):
        if t.strip().startswith("Nomor") and "/SPA/PDSB/" in t:
            _docx_set_field_value(p, sppd_no, ns, f"Nomor : {sppd_no}")
            break

    # Update tanggal dokumen Jaminan dari konfigurasi SPPD.
    jaminan_date = format_id_date(pd.to_datetime(settings.get("fixed_jaminan_date", "2026-02-19")))
    for p, t in zip(paras, texts):
        if "Jaminan tertanggal" in t:
            _docx_replace_date_after_keyword(p, "Jaminan tertanggal", jaminan_date, ns)
            break

    # Update jatuh tempo Bank dari tanggal Makassar.
    maturity_months = int(settings.get("maturity_months", 6) or 6)
    jatuh_tempo = format_id_date(submit_dt + pd.DateOffset(months=maturity_months))
    for p, t in zip(paras, texts):
        if "jatuh tempo pada Bank yaitu tanggal" in t:
            _docx_replace_date_after_keyword(p, "jatuh tempo pada Bank yaitu tanggal", jatuh_tempo, ns)
            break

    # Replace daftar bank
    start_idx = None
    end_idx = None
    for i, t in enumerate(texts):
        if t.strip().startswith("Bank"):
            start_idx = i
            break
    for i, t in enumerate(texts):
        if "Kami menyatakan dan menyanggupi" in t:
            end_idx = i - 1
            break

    if start_idx is not None and end_idx is not None and end_idx >= start_idx:
        bank_tpl = paras[start_idx]
        norek_tpl = paras[start_idx + 1] if start_idx + 1 < len(paras) else bank_tpl
        atas_tpl = paras[start_idx + 2] if start_idx + 2 < len(paras) else bank_tpl
        jumlah_tpl = paras[start_idx + 3] if start_idx + 3 < len(paras) else bank_tpl

        sep_tpl = None
        for i in range(start_idx, min(start_idx + 8, len(paras))):
            if not texts[i].strip():
                sep_tpl = paras[i]
                break

        bank_label = _label_prefix(texts[start_idx], "Bank")
        norek_label = _label_prefix(texts[start_idx + 1], "Nomor Rekening")
        atas_label = _label_prefix(texts[start_idx + 2], "Atas Nama")
        jumlah_label = _label_prefix(texts[start_idx + 3], "Sejumlah")

        new_paras: List[ET.Element] = []
        items_per_page = max(1, int(settings.get("items_per_page", 7) or 7))
        for idx, item in enumerate(transfer_items):
            p_bank = copy.deepcopy(bank_tpl)
            bank_value = s(item.get("bank", ""))
            _docx_set_field_value(p_bank, bank_value, ns, f"{bank_label}: {bank_value}")
            new_paras.append(p_bank)

            p_norek = copy.deepcopy(norek_tpl)
            norek_value = s(item.get("rekening", ""))
            _docx_set_field_value(p_norek, norek_value, ns, f"{norek_label}: {norek_value}")
            new_paras.append(p_norek)

            p_atas = copy.deepcopy(atas_tpl)
            atas_value = s(item.get("penerima", ""))
            _docx_set_field_value(p_atas, atas_value, ns, f"{atas_label}: {atas_value}")
            new_paras.append(p_atas)

            p_jumlah = copy.deepcopy(jumlah_tpl)
            amount = float(item.get("amount", 0) or 0.0)
            jumlah_value = f"Rp {format_idr(amount)}"
            _docx_set_field_value(p_jumlah, jumlah_value, ns, f"{jumlah_label}: {jumlah_value}")
            new_paras.append(p_jumlah)

            if idx < len(transfer_items) - 1:
                if (idx + 1) % items_per_page == 0:
                    new_paras.append(_docx_page_break_paragraph(ns))
                elif sep_tpl is not None:
                    new_paras.append(copy.deepcopy(sep_tpl))

        for p in paras[start_idx:end_idx + 1]:
            body.remove(p)
        for offset, p in enumerate(new_paras):
            body.insert(start_idx + offset, p)

    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    new_xml = _docx_preserve_namespace_declarations(new_xml, namespace_map)
    data["word/document.xml"] = new_xml
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for name, content in data.items():
            z.writestr(name, content)

def load_bank_map() -> Dict[str, Dict[str, str]]:
    """
    Load data rekening dari Excel. Key = UPPERCASE principle name.
    Juga menyimpan normalized key untuk fuzzy matching.
    Nomor rekening disimpan sebagai string agar leading zeros tidak hilang.
    Data kosong (DF/VA tanpa nomor rekening) tetap disimpan apa adanya, tidak dikarang.
    """
    if not BANK_DATA_PATH or not os.path.exists(BANK_DATA_PATH):
        return {}
    # Try reading with header auto-detection: the Excel might have header at row 0 or row 2
    # First attempt: standard read
    try:
        df = pd.read_excel(BANK_DATA_PATH, dtype=str)
        cols = {c.strip().upper(): c for c in df.columns}
        required = ["PRINCIPLE", "NAMA BANK", "NOMOR REKENING", "NAMA PENERIMA"]
        missing = [c for c in required if c not in cols]
        if missing:
            # Try with header at row 2 (common format from rekprinciple.xlsx)
            df = pd.read_excel(BANK_DATA_PATH, header=2, dtype=str)
            cols = {c.strip().upper(): c for c in df.columns}
            missing = [c for c in required if c not in cols]
            if missing:
                return {}
    except Exception:
        return {}
    mp = {}
    for _, r in df.iterrows():
        raw_p = r.get(cols.get("PRINCIPLE", ""), None)
        p = s(raw_p) if raw_p is not None and not pd.isna(raw_p) else ""
        if not p:
            continue
        # Nomor rekening: pastikan string, jangan convert ke float
        raw_rek = r.get(cols.get("NOMOR REKENING", ""), None)
        if raw_rek is None or pd.isna(raw_rek):
            rek_str = ""
        else:
            rek_str = str(raw_rek).strip()
            # Hapus trailing .0 jika pandas membaca sebagai float
            if rek_str.endswith(".0"):
                rek_str = rek_str[:-2]

        raw_bank = r.get(cols.get("NAMA BANK", ""), None)
        bank_val = s(raw_bank) if raw_bank is not None and not pd.isna(raw_bank) else ""
        raw_pen = r.get(cols.get("NAMA PENERIMA", ""), None)
        penerima_val = s(raw_pen) if raw_pen is not None and not pd.isna(raw_pen) else ""

        # Jika bank = DF/VA dan rekening kosong, jangan mengarang
        # Simpan apa adanya
        mp[p.upper()] = {
            "principle": normalize_principle_name(p),
            "bank": bank_val,
            "rekening": rek_str,
            "penerima": penerima_val,
        }
    return mp


def load_bank_map_with_normalized_keys() -> Tuple[Dict[str, Dict[str, str]], Dict[str, str]]:
    """
    Load bank map + pre-build normalized keys untuk fuzzy matching.
    Returns: (bank_map, normalized_key_map)
    """
    bank_map = load_bank_map()
    norm_keys = build_normalized_key_map(bank_map)
    return bank_map, norm_keys

def write_invoice_excel(rows: List[Dict[str, Any]], out_path: str):
    cols = [
        "No",
        "Principle",
        "Tipe Pengajuan",
        "Nilai Invoice (Total)",
        "No. Invoice / Dokumen",
        "Potongan",
        "Nilai Pembayaran",
        "Jenis Pembayaran",
        "Keterangan",
    ]
    normalized_rows = []
    for row in rows or []:
        item = dict(row or {})
        if "No. Invoice / Dokumen" not in item:
            item["No. Invoice / Dokumen"] = s(item.get("No. Invoice", ""))
        if "Tipe Pengajuan" not in item:
            item["Tipe Pengajuan"] = s(item.get("Tipe", ""))
        normalized_rows.append(item)
    df = pd.DataFrame(normalized_rows, columns=cols)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="INVOICE")

def _normalize_yyyy_mm_dd(value: str) -> str:
    v = s(value)
    if not v:
        return ""
    try:
        dt = pd.to_datetime(v, errors="raise")
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""

def _style_excel_download_sheet(ws) -> None:
    money_tokens = ("NILAI", "AMOUNT", "TOTAL", "GAP", "POTONGAN", "PEMBAYARAN")
    date_tokens = ("TGL", "DATE", "J.T", "J TEMPO", "JATUH TEMPO")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1], start=1):
        header = s(cell.value).upper()
        width = max(12, min(34, len(header) + 3))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        if any(token in header for token in money_tokens):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = '#,##0.00'
                ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="right")
        elif any(token in header for token in date_tokens):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"

def _excel_download_response(rows: List[Dict[str, Any]], filename: str, sheet_name: str) -> Response:
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _style_excel_download_sheet(writer.book[sheet_name])
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

def safe_upload_filename(name: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(s(name)))
    return clean.strip("._") or "bukti-transfer"

def proof_public_url(file_name: str) -> str:
    return f"/payments/proofs/{os.path.basename(file_name)}"

def finance_mapping_key(principle: str) -> str:
    return s(principle).upper()

def get_finance_mapping(db: Dict[str, Any], principle: str) -> Dict[str, str]:
    mappings = db.setdefault("finance_mappings", {})
    item = mappings.get(finance_mapping_key(principle), {})
    if not isinstance(item, dict):
        return {}
    return {
        "principle": s(item.get("principle", principle)),
        "vendorNo": s(item.get("vendorNo", "")),
        "vendorName": s(item.get("vendorName", "")),
        "bankNo": s(item.get("bankNo", "")),
        "bankName": s(item.get("bankName", "")),
        "updated_at": s(item.get("updated_at", "")),
        "updated_by": s(item.get("updated_by", "")),
    }

def build_proof_metadata(proof_id: str, stored_name: str, original_name: str, content: bytes, uploaded_by: str) -> Dict[str, Any]:
    mime = mimetypes.guess_type(original_name)[0] or "application/octet-stream"
    return {
        "proof_id": proof_id,
        "original_filename": original_name,
        "stored_filename": stored_name,
        "mime": mime,
        "size": len(content),
        "sha256": hashlib.sha256(content).hexdigest(),
        "uploaded_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "uploaded_by": uploaded_by,
        "url": proof_public_url(stored_name),
    }

def xml_escape(text: str) -> str:
    return (
        s(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )

def render_docx_template(template_path: str, out_path: str, replacements: Dict[str, str], append_rows: str = "") -> None:
    if not os.path.exists(template_path):
        raise FileNotFoundError("Template SPPD tidak ditemukan.")
    with zipfile.ZipFile(template_path) as z:
        data = {name: z.read(name) for name in z.namelist()}
    if "word/document.xml" not in data:
        raise ValueError("Template SPPD tidak valid.")
    xml = data["word/document.xml"].decode("utf-8", errors="ignore")
    replaced = xml
    for k, v in replacements.items():
        replaced = replaced.replace("{{" + k + "}}", xml_escape(v))

    if "{{TRANSFER_ROWS}}" in replaced:
        replaced = replaced.replace("{{TRANSFER_ROWS}}", xml_escape(append_rows))
    elif append_rows:
        insert = (
            "<w:p><w:r><w:t>" + xml_escape(append_rows).replace("\n", "</w:t><w:br/><w:t>") + "</w:t></w:r></w:p>"
        )
        replaced = replaced.replace("</w:body>", insert + "</w:body>")

    data["word/document.xml"] = replaced.encode("utf-8")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for name, content in data.items():
            z.writestr(name, content)

def allocate_claim_by_mdstring(
    gross_amount: float,
    actual_discount: float,
    mdstring: str,
    rp_pabrik: float,
    pct_pabrik,
    pct_internal,
) -> Optional[Tuple[float, float, float]]:
    actual = float(actual_discount or 0.0)
    rp = float(rp_pabrik or 0.0)
    if actual <= 0:
        return (0.0, 0.0, 0.0)
    if actual < rp:
        return (actual, 0.0, 0.0)

    md_parts = normalize_pct_list(parse_mdstring_pct(mdstring))
    if not md_parts:
        return None

    p_parts = normalize_pct_list(pct_pabrik)
    i_parts = normalize_pct_list(pct_internal)

    base = max(float(gross_amount or 0.0) - rp, 0.0)
    rem = base
    p_idx = 0
    i_idx = 0
    p_pct_total = 0.0
    i_pct_total = 0.0
    matched_any = False

    for pct in md_parts:
        disc = rem * (pct / 100.0)
        p_next = p_parts[p_idx] if p_idx < len(p_parts) else None
        i_next = i_parts[i_idx] if i_idx < len(i_parts) else None
        p_match = (p_next is not None and abs(p_next - pct) < 1e-6)
        i_match = (i_next is not None and abs(i_next - pct) < 1e-6)

        if p_match and i_match:
            if (len(p_parts) - p_idx) >= (len(i_parts) - i_idx):
                p_pct_total += disc
                p_idx += 1
                matched_any = True
            else:
                i_pct_total += disc
                i_idx += 1
                matched_any = True
        elif p_match:
            p_pct_total += disc
            p_idx += 1
            matched_any = True
        elif i_match:
            i_pct_total += disc
            i_idx += 1
            matched_any = True
        else:
            return None

        rem -= disc

    if not matched_any:
        return None

    actual_pct_total = max(actual - rp, 0.0)
    pct_alloc = p_pct_total + i_pct_total
    if pct_alloc > 0:
        scale = min(actual_pct_total / pct_alloc, 1.0)
        p_pct_total *= scale
        i_pct_total *= scale

    p_total = rp + p_pct_total
    i_total = i_pct_total
    tanpat = max(actual - (p_total + i_total), 0.0)
    return (p_total, i_total, tanpat)

def sequential_pct_discount(base_amount: float, pct_parts: List[float]) -> float:
    remaining = float(base_amount or 0.0)
    total = 0.0
    for pct in pct_parts:
        disc = remaining * (pct / 100.0)
        total += disc
        remaining -= disc
    return total


def mdstring_is_bonus(md: str) -> bool:
    """Bonus line: MDSTRING represents 100% discount (free goods)."""
    parts = parse_mdstring_pct(md)
    if not parts:
        # sometimes users put plain "100" without %
        md_s = s(md).strip()
        return md_s == "100"
    mx = max(parts)
    if mx < 99.999:
        return False
    # allow 0% parts alongside 100%
    others = [p for p in parts if abs(p - mx) > 1e-6]
    return all(abs(p) < 1e-6 for p in others)

def pct_parts_full_discount(pct_parts: List[float]) -> bool:
    if not pct_parts:
        return False
    mx = max(pct_parts)
    if mx < 99.999:
        return False
    others = [p for p in pct_parts if abs(p - mx) > 1e-6]
    return all(abs(p) < 1e-6 for p in others)

def compute_expected_bonus_from_rules(
    sales_df: pd.DataFrame,
    rules_df: pd.DataFrame,
    channel_col: str,
    expected_col: str,
    debug_reason_prefix: str,
    program_name_col: Optional[str] = None,
):
    """
    Fill Expected bonus qty (PCS) per line.
    - Expected bonus is only written to rows that are 'bonus lines' (MDSTRING=100%).
    - For NON_GROUP: computed per line.
    - For grouped programs: computed per invoice+program key using invoice-level trigger metric,
      then allocated to bonus lines of the corresponding item(s).
    """
    if rules_df is None or rules_df.empty:
        sales_df.loc[:, expected_col] = 0.0
        return

    # normalize rules columns
    d = rules_df.copy()
    for c in ["KODE_BARANG","NAMA_BARANG","PROMO_LABEL","PROMO_GROUP_ID","PROMO_GROUP","TIER_NO","TRIGGER_QTY","TRIGGER_UNIT","BENEFIT_TYPE","BENEFIT_VALUE","BENEFIT_UNIT","PROMO_ACTIVE"]:
        if c not in d.columns:
            d.loc[:, c] = None
    d["KODE_BARANG"] = d["KODE_BARANG"].apply(s)
    d["NAMA_BARANG"] = d["NAMA_BARANG"].apply(s)
    d["_ALL_ITEM_RULE"] = (d["KODE_BARANG"] == "") & (d["NAMA_BARANG"] == "")

    # focus bonus programs only
    d = d[d["BENEFIT_TYPE"].astype(str).str.upper().isin(["BONUS_QTY","BONUSQTY","BONUS"])].copy()
    if d.empty:
        sales_df.loc[:, expected_col] = 0.0
        return

    # prep
    sales_df.loc[:, expected_col] = 0.0
    sales_df.loc[:, "_IS_BONUS_LINE"] = sales_df["MDSTRING"].apply(mdstring_is_bonus)

    # iterate invoices to keep it fast enough
    for inv, inv_df in sales_df.groupby("INVOICENO", sort=False):
        inv_idx = inv_df.index.tolist()

        # candidate rules for items present in this invoice
        items = set(inv_df["BRG"].astype(str))
        r0 = d[(d["KODE_BARANG"].astype(str).isin(items)) | (d["_ALL_ITEM_RULE"] == True)]
        if r0.empty:
            continue

        # group by program key
        for (pgid, plabel), pr in r0.groupby(["PROMO_GROUP_ID","PROMO_LABEL"], sort=False):
            pgid_s = norm(pgid)
            plabel_s = s(plabel)
            program_key = f"{pgid_s}||{plabel_s}"

            # channel filter (STD always allowed, otherwise must match effective channel)
            pr_ok = pr[pr.apply(
                lambda r: channel_ok(
                    r.get("PROMO_GROUP", ""),
                    inv_df.iloc[0].get(channel_col, ""),
                    r.get("PROMO_GROUP_ID", ""),
                    inv_df.iloc[0].get("SUB", "")
                ),
                axis=1
            )]
            if pr_ok.empty:
                continue

            # pick best tier (highest tier that passes trigger)
            # we decide trigger metric based on chosen trigger unit
            trigger_unit, _tu_source = choose_trigger_unit_satpam(pr_ok)
            if not trigger_unit:
                continue

            inv_non_bonus = inv_df[~inv_df["_IS_BONUS_LINE"]].copy()

            # Compute per-tier metrics to select highest satisfied tier
            best = None
            for tier_no in sorted(pr_ok["TIER_NO"].fillna(0).astype(int).unique(), reverse=True):
                tier_rows = pr_ok[pr_ok["TIER_NO"].fillna(0).astype(int) == int(tier_no)]
                if tier_rows.empty:
                    continue

                # trigger qty / benefit value (tier-level)
                tq = float(pd.to_numeric(tier_rows["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)
                if tq <= 0:
                    continue
                bq = float(pd.to_numeric(tier_rows["BENEFIT_VALUE"], errors="coerce").fillna(0).max() or 0.0)
                if bq <= 0:
                    continue

                if pgid_s == "NON_GROUP":
                    # non-group: per item using non-bonus lines for trigger
                    used_items: Set[str] = set()
                    items_tier = [x for x in tier_rows["KODE_BARANG"].astype(str).unique().tolist() if s(x)]
                    if bool(tier_rows["_ALL_ITEM_RULE"].any()):
                        items_tier = sorted(set(items_tier).union(set(inv_non_bonus["BRG"].astype(str).tolist())))
                    for item in items_tier:
                        if item in used_items:
                            continue
                        rows_item = tier_rows[tier_rows["KODE_BARANG"].astype(str) == str(item)]
                        if rows_item.empty:
                            rows_item = tier_rows[tier_rows["_ALL_ITEM_RULE"] == True]
                        if rows_item.empty:
                            continue

                        tq = float(pd.to_numeric(rows_item["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)
                        if tq <= 0:
                            continue
                        bq = float(pd.to_numeric(rows_item["BENEFIT_VALUE"], errors="coerce").fillna(0).max() or 0.0)
                        if bq <= 0:
                            continue

                        if trigger_unit == "GROSSAMOUNT":
                            metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str) == str(item)]["GROSSAMOUNT"].astype(float).fillna(0).sum())
                        else:
                            metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str) == str(item)].apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum())

                        if metric + 1e-9 < tq:
                            continue

                        total_exp = math.floor(metric / tq) * bq
                        if total_exp <= 0:
                            continue

                        bonus_idxs = [i for i in inv_idx if (str(sales_df.at[i, "BRG"]) == str(item)) and bool(sales_df.at[i, "_IS_BONUS_LINE"])]
                        if not bonus_idxs:
                            best = ("NON_GROUP_NO_BONUS_LINE", tier_no)
                            continue

                        if "Bonus_TriggerUnit" in sales_df.columns:
                            for i in bonus_idxs:
                                sales_df.at[i, "Bonus_TriggerUnit"] = trigger_unit
                                sales_df.at[i, "Bonus_TriggerQty"] = float(tq)
                                sales_df.at[i, "Bonus_TriggerMetric"] = float(metric)
                                sales_df.at[i, "Bonus_PickedTier"] = int(tier_no)

                        remaining = float(total_exp)
                        for i in bonus_idxs:
                            if remaining <= 0:
                                break
                            act_qty = float(qty_in_unit(sales_df.loc[i], "PCS"))
                            take = min(remaining, act_qty if act_qty > 0 else remaining)
                            sales_df.at[i, expected_col] = float(take)
                            remaining -= take
                            if program_name_col:
                                cur = s(sales_df.at[i, program_name_col])
                                sales_df.at[i, program_name_col] = add_list_item(cur, plabel_s)

                        used_items.add(item)
                        best = ("NON_GROUP", tier_no)
                    if used_items:
                        break

                # grouped program: invoice-level metric
                if bool(tier_rows["_ALL_ITEM_RULE"].any()):
                    elig_item_set = set(inv_non_bonus["BRG"].astype(str).tolist())
                else:
                    elig_item_set = set([x for x in tier_rows["KODE_BARANG"].astype(str).tolist() if s(x)])
                if not elig_item_set:
                    continue
                if trigger_unit == "GROSSAMOUNT":
                    metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str).isin(elig_item_set)]["GROSSAMOUNT"].astype(float).fillna(0).sum())
                else:
                    metric = float(inv_non_bonus[inv_non_bonus["BRG"].astype(str).isin(elig_item_set)].apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum())

                if metric + 1e-9 < tq:
                    continue

                total_exp = math.floor(metric / tq) * bq
                if total_exp <= 0:
                    continue

                # allocate to bonus lines for eligible items
                bonus_idxs = [i for i in inv_idx if (str(sales_df.at[i,"BRG"]) in elig_item_set) and bool(sales_df.at[i,"_IS_BONUS_LINE"])]
                if not bonus_idxs:
                    # no bonus line present -> still mark debug but can't allocate
                    best = ("GROUPED_NO_BONUS_LINE", tier_no)
                    break

                if "Bonus_TriggerUnit" in sales_df.columns:
                    for i in bonus_idxs:
                        sales_df.at[i, "Bonus_TriggerUnit"] = trigger_unit
                        sales_df.at[i, "Bonus_TriggerQty"] = float(tq)
                        sales_df.at[i, "Bonus_TriggerMetric"] = float(metric)
                        sales_df.at[i, "Bonus_PickedTier"] = int(tier_no)

                remaining = float(total_exp)
                for i in bonus_idxs:
                    if remaining <= 0:
                        break
                    # allocate up to actual bonus qty (pcs) on that line
                    act_qty = float(qty_in_unit(sales_df.loc[i], "PCS"))
                    take = min(remaining, act_qty if act_qty > 0 else remaining)
                    sales_df.at[i, expected_col] = float(take)
                    remaining -= take
                    if program_name_col:
                        cur = s(sales_df.at[i, program_name_col])
                        sales_df.at[i, program_name_col] = add_list_item(cur, plabel_s)

                best = ("GROUPED", tier_no)
                break

            # Debug: record if needed
            if best:
                for i in inv_idx:
                    cur = s(sales_df.at[i, "Debug_SkipReason"]) if "Debug_SkipReason" in sales_df.columns else ""
                    # keep it lightweight: only for bonus lines
                    if bool(sales_df.at[i, "_IS_BONUS_LINE"]):
                        msg = f"{debug_reason_prefix}:{program_key}:{best[0]}:tier{best[1]}"
                        sales_df.at[i, "Debug_SkipReason"] = (cur + (" | " if cur else "") + msg)

    # cleanup helper col
    if "_IS_BONUS_LINE" in sales_df.columns:
        sales_df.drop(columns=["_IS_BONUS_LINE"], inplace=True, errors="ignore")
def dedupe_grouped_benefits(tier_rows: pd.DataFrame) -> pd.DataFrame:
    if tier_rows.empty:
        return tier_rows
    cols = ["BENEFIT_TYPE", "BENEFIT_VALUE", "BENEFIT_UNIT", "TIER_NO", "TRIGGER_QTY", "TRIGGER_UNIT"]
    for c in cols:
        if c not in tier_rows.columns:
            tier_rows.loc[:, c] = ""
    return tier_rows.drop_duplicates(subset=cols).copy()


# ---------------------------
# Internal program helpers
# ---------------------------
def internal_channel_match(rule_channel: str, effective_sales_channel: str) -> bool:
    rc = norm(rule_channel)
    sc = norm(effective_sales_channel)
    if rc in ("", "ALL", "STD", "GLOBAL"):
        return True
    if sc == "":
        return False
    return (rc == sc) or (rc in sc) or (sc in rc)

def load_internal_rules(internal_bytes: Optional[bytes]) -> Optional[pd.DataFrame]:
    if not internal_bytes:
        return None
    df = pd.read_excel(io.BytesIO(internal_bytes))
    for c in ["CHANNEL","SUB","KODE_BARANG","RULE_TYPE","MIN_GROSS","MAX_GROSS","DISC_PCT","ACTIVE"]:
        ensure_col(df, c, "")
    df["CHANNEL"] = df["CHANNEL"].apply(s)
    df["SUB"] = df["SUB"].apply(s)
    df["KODE_BARANG"] = df["KODE_BARANG"].apply(s)
    df["RULE_TYPE"] = df["RULE_TYPE"].apply(lambda x: norm(x))
    df["DISC_PCT"] = df["DISC_PCT"].apply(s)
    df["MIN_GROSS"] = df["MIN_GROSS"].apply(parse_number_id)
    df["MAX_GROSS"] = df["MAX_GROSS"].apply(parse_number_id)

    def is_active(v):
        if pd.isna(v):
            return True
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in ["true","1","yes","y"]
    df = df[df["ACTIVE"].apply(is_active)].copy()
    return df

def pick_internal_rule(rules: pd.DataFrame, invoice_gross: float) -> pd.DataFrame:
    if rules is None or rules.empty:
        return rules.iloc[0:0].copy()
    ig = float(invoice_gross or 0.0)
    d = rules.copy()
    d = d[d["MIN_GROSS"].astype(float) <= ig]
    d = d[(d["MAX_GROSS"].astype(float) <= 0) | (ig <= d["MAX_GROSS"].astype(float))]
    if d.empty:
        return d
    d = d.sort_values(["MIN_GROSS","MAX_GROSS"], ascending=[False, True])
    top_min = float(d["MIN_GROSS"].iloc[0])
    return d[d["MIN_GROSS"].astype(float) == top_min].copy()

# ---------------------------
# Channel selection
# ---------------------------
def channel_ok(promo_group: str, effective_sales_channel: str, promo_group_id: str, sales_sub: str) -> bool:
    """
    promo_group is dataset column PROMO_GROUP
    effective_sales_channel is computed from channel map (SUB->CHANNEL); fallback to sales field
    """
    pg = norm(promo_group)
    sc = norm(effective_sales_channel)

    # STD = global (all channels)
    if pg == "STD":
        return True

    # Outlet special: require SUB match PROMO_GROUP_ID
    if pg == "OUTLET":
        return norm(sales_sub) != "" and norm(promo_group_id) != "" and norm(sales_sub) == norm(promo_group_id)

    # Online vs non-online should not cross-match
    if ("ONLINE" in pg or "ONLINE" in sc) and (("ONLINE" in pg) != ("ONLINE" in sc)):
        return False

    # Otherwise: match channel
    if pg != "" and sc != "" and (pg == sc or pg in sc or sc in pg):
        return True
    return False

# ---------------------------
# Tier picking
# ---------------------------
def pick_best_tier(program_rows: pd.DataFrame, metric_value: float) -> Tuple[int, pd.DataFrame]:
    if program_rows.empty:
        return (0, program_rows)

    d = program_rows.copy()
    d.loc[:, "TIER_NO"] = pd.to_numeric(d["TIER_NO"], errors="coerce").fillna(0).astype(int)
    d = d.sort_values(["TIER_NO", "TRIGGER_QTY"], ascending=[False, False])

    mv = float(metric_value or 0.0)
    picked = None
    for _, r in d.iterrows():
        if mv >= float(r["TRIGGER_QTY"]):
            picked = int(r["TIER_NO"])
            break
    if picked is None:
        return (0, d.iloc[0:0].copy())

    return (picked, d[d["TIER_NO"] == picked].copy())

# ---------------------------
# SATPAM trigger-unit (v9)
# ---------------------------
VALID_UNITS = {"GROSSAMOUNT", "PCS", "CTN", "KRT"}

def choose_trigger_unit_satpam(program_rows: pd.DataFrame) -> Tuple[Optional[str], str]:
    if "TRIGGER_UNIT" in program_rows.columns:
        tu = [norm(x) for x in program_rows["TRIGGER_UNIT"].tolist() if s(x)]
        tu = [x for x in tu if x in VALID_UNITS]
        if tu:
            if "GROSSAMOUNT" in tu:
                return ("GROSSAMOUNT", "TRIGGER_UNIT")
            if any(x in ["CTN", "KRT"] for x in tu):
                return ("CTN", "TRIGGER_UNIT")
            if "PCS" in tu:
                return ("PCS", "TRIGGER_UNIT")
            return (tu[0], "TRIGGER_UNIT")

    if "BENEFIT_UNIT" in program_rows.columns:
        bu = [norm(x) for x in program_rows["BENEFIT_UNIT"].tolist() if s(x)]
        bu = [x for x in bu if x in VALID_UNITS]
        if bu:
            if "GROSSAMOUNT" in bu:
                return ("GROSSAMOUNT", "BENEFIT_UNIT")
            if any(x in ["CTN", "KRT"] for x in bu):
                return ("CTN", "BENEFIT_UNIT")
            if "PCS" in bu:
                return ("PCS", "BENEFIT_UNIT")
            return (bu[0], "BENEFIT_UNIT")

    return (None, "UNKNOWN")

def add_reason(existing: str, reason: str) -> str:
    existing = s(existing)
    if not existing:
        return reason
    if reason in existing:
        return existing
    return existing + "; " + reason

def add_list_item(existing: str, item: str, sep: str = "; ") -> str:
    existing = s(existing)
    item = s(item)
    if not item:
        return existing
    if not existing:
        return item
    parts = [p.strip() for p in existing.split(";") if p.strip()]
    if item in parts:
        return existing
    return existing + sep + item


# ---------------------------
# UI (v16 FIXED) — safe raw HTML string (no JS leaked into Python)
# ---------------------------

UI_HTML = r"""<!doctype html>
<html lang="id">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <link rel="icon" href="/favicon.ico">
  <link rel="icon" type="image/png" href="/favicon.png">
  <title>Discount Validator</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    .pulse-soft { animation: pulseSoft 1.4s ease-in-out infinite; }
    @keyframes pulseSoft { 0%,100%{opacity:.55} 50%{opacity:1} }
  </style>
</head>

<body class="min-h-screen bg-[#EFEFEF] text-[#1D1F1E]">
  <div class="max-w-[1400px] mx-auto px-4 py-7">
    <div class="grid xl:grid-cols-12 gap-5">

      <!-- LEFT SIDEBAR -->
      <aside class="xl:col-span-3 rounded-2xl overflow-hidden shadow-lg border border-[#8A4703] bg-[#EFEFEF]">
        <div class="p-5 bg-[#BD7401] text-[#EFEFEF]">
          <div class="flex items-start justify-between gap-3">
            <div>
              <div class="text-xs uppercase tracking-widest text-[#EFEFEF]">Validator</div>
              <div class="text-xl font-semibold leading-tight">Discount & Claim</div>
              <div class="mt-1 text-xs text-[#EFEFEF]">Patch: __PATCH_VERSION__</div>
            </div>
            <div class="flex items-center gap-2">
              <a href="/"
                 class="shrink-0 text-xs px-3 py-1.5 rounded-full bg-[#F2DA82] hover:bg-[#BD7401] hover:text-[#EFEFEF] border border-[#8A4703] transition">
                Dashboard
              </a>
              <a href="/logout"
                 class="shrink-0 text-xs px-3 py-1.5 rounded-full bg-[#F2DA82] hover:bg-[#BD7401] hover:text-[#EFEFEF] border border-[#8A4703] transition">
                Logout
              </a>
            </div>
          </div>

          <div class="mt-4 grid grid-cols-3 gap-2">
            <div class="rounded-xl p-3 bg-[#F2DA82] border border-[#8A4703]">
              <div class="text-[11px] text-[#1D1F1E]">C</div>
              <div id="cntC" class="text-2xl font-semibold">0</div>
            </div>
            <div class="rounded-xl p-3 bg-[#F2DA82] border border-[#8A4703]">
              <div class="text-[11px] text-[#1D1F1E]">B</div>
              <div id="cntB" class="text-2xl font-semibold">0</div>
            </div>
            <div class="rounded-xl p-3 bg-[#F2DA82] border border-[#8A4703]">
              <div class="text-[11px] text-[#1D1F1E]">A</div>
              <div id="cntA" class="text-2xl font-semibold">0</div>
            </div>
          </div>

          <div class="mt-4 rounded-xl bg-[#F2DA82] border border-[#8A4703] p-3">
            <div class="text-xs text-[#1D1F1E]">API Status</div>
            <div id="apiStatus" class="mt-1 text-sm font-medium pulse-soft">checking...</div>
          </div>
        </div>

        <div class="p-5">
          <div class="flex items-center justify-between">
            <div class="text-sm font-semibold">Checklist Upload</div>
            <div id="readyPill"
                 class="text-[11px] px-2 py-1 rounded-full border border-[#8A4703] bg-[#EFEFEF] text-[#1D1F1E]">
              belum siap
            </div>
          </div>

          <div class="mt-4 space-y-3">
            <div class="flex items-center gap-3">
              <div id="dotSales" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Data Penjualan</div>
                <div id="nameSales" class="text-xs text-[#1D1F1E] truncate">belum diupload</div>
              </div>
            </div>

            <div class="flex items-center gap-3">
              <div id="dotPromo" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Dataset Diskon Pabrik</div>
                <div id="namePromo" class="text-xs text-[#1D1F1E] truncate">belum diupload</div>
              </div>
            </div>

            <div class="flex items-center gap-3">
              <div id="dotChannel" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Data Channel by SUB</div>
                <div id="nameChannel" class="text-xs text-[#1D1F1E] truncate">belum diupload</div>
              </div>
            </div>

            <div class="flex items-center gap-3">
              <div id="dotInternal" class="h-2.5 w-2.5 rounded-full bg-[#EFEFEF]"></div>
              <div class="min-w-0">
                <div class="text-sm font-medium">Dataset Diskon Internal <span class="text-xs text-[#1D1F1E]">(opsional)</span></div>
                <div id="nameInternal" class="text-xs text-[#1D1F1E] truncate">boleh kosong</div>
              </div>
            </div>
          </div>

          <div class="mt-5">
            <button id="downloadBtn" disabled
                    class="w-full rounded-xl px-4 py-3 text-sm font-semibold
                           bg-gradient-to-r from-[#BD7401] to-[#BD7401] text-[#EFEFEF]
                           hover:brightness-95 disabled:opacity-50 disabled:cursor-not-allowed transition">
              Download Result
            </button>
            <div class="mt-2 text-xs text-[#1D1F1E]">Setelah validate sukses, tombol ini aktif.</div>
          </div>

          <div class="mt-5 rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
            <div class="text-xs uppercase tracking-widest text-[#1D1F1E]">Tips</div>
            <div class="mt-2 text-sm text-[#1D1F1E]">
              Klik <b>Validate</b> kalau 3 file wajib sudah hijau.
              Kalau belum, nanti aku tegur manja dulu 😄
            </div>
          </div>
        </div>
      </aside>

      <!-- CENTER -->
      <main class="xl:col-span-6 rounded-2xl bg-[#EFEFEF] shadow-lg border border-[#8A4703] overflow-hidden">
        <div class="p-6 border-b border-[#8A4703]">
          <div class="flex items-start justify-between gap-4">
            <div>
              <h1 class="text-2xl md:text-3xl font-semibold tracking-tight">
                Validate Discount vs Program Dataset
              </h1>
              <p class="mt-2 text-[#1D1F1E]">__PATCH_TITLE__</p>
              <p class="mt-1 text-xs text-[#1D1F1E]">
                Required: Data Penjualan + Dataset Diskon Pabrik + Data Channel by SUB. Internal optional.
              </p>
            </div>
            <div class="hidden md:block rounded-2xl border border-[#8A4703] bg-[#EFEFEF] px-4 py-3">
              <div class="text-xs text-[#1D1F1E]">Mode</div>
              <div class="text-sm font-semibold text-[#1D1F1E]">Serius tapi ramah 😄</div>
            </div>
          </div>
        </div>

        <div class="p-6">
          <div id="toast" aria-live="polite" role="status" class="hidden mb-4 rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
            <div id="toastTitle" class="text-sm font-semibold text-[#1D1F1E]">Info</div>
            <div id="toastBody" class="mt-1 text-sm text-[#1D1F1E]"></div>
          </div>

          <form id="form" class="grid md:grid-cols-2 gap-4">
            <div class="md:col-span-2 rounded-xl border border-[#8A4703] bg-[#F8EFD1] p-3 text-xs text-[#1D1F1E]">
              <span class="font-semibold">Template Upload:</span>
              <a href="/validator/template/sales" class="underline ml-2">Data Penjualan</a>
              <span class="mx-1">|</span>
              <a href="/validator/template/promo" class="underline">Dataset Diskon Pabrik</a>
              <span class="mx-1">|</span>
              <a href="/validator/template/channel" class="underline">Data Channel by SUB</a>
            </div>
            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Data Penjualan <span class="text-xs text-[#1D1F1E]">(required)</span></div>
                <span id="badgeSales" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="sales" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Wajib. Tanpa ini, kita nggak bisa ngitung apa-apa.</div>
            </div>

            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Dataset Diskon Pabrik <span class="text-xs text-[#1D1F1E]">(required)</span></div>
                <span id="badgePromo" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="promo" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Program pabrik yang jadi patokan Expected_Pabrik.</div>
            </div>

            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Data Channel by SUB <span class="text-xs text-[#1D1F1E]">(required)</span></div>
                <span id="badgeChannel" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="channel" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Biar program by channel nggak salah pilih.</div>
            </div>

            <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] p-5">
              <div class="flex items-center justify-between">
                <div class="text-sm font-semibold">Dataset Diskon Internal <span class="text-xs text-[#1D1F1E]">(optional)</span></div>
                <span id="badgeInternal" class="hidden text-[11px] px-2 py-1 rounded-full bg-[#BD7401] text-[#EFEFEF] border border-[#8A4703]">OK</span>
              </div>
              <input id="internal" type="file" accept=".xlsx,.xls"
                     class="mt-4 block w-full text-sm file:mr-4 file:py-2 file:px-4 file:rounded-xl file:border-0
                            file:bg-[#BD7401] file:text-[#EFEFEF] hover:file:brightness-110"/>
              <div class="mt-2 text-xs text-[#1D1F1E]">Kalau diupload, Expected_Internal & FixClaim_Internal ikut dihitung.</div>
            </div>

            <div class="md:col-span-2">
              <button id="btn" type="submit" disabled
                      class="w-full rounded-2xl px-4 py-3 font-semibold text-[#EFEFEF]
                             bg-[#BD7401] hover:bg-[#8A4703] hover:text-[#EFEFEF] transition
                             disabled:opacity-50 disabled:cursor-not-allowed">
                Validate
              </button>

              <div class="mt-3 flex items-start gap-3">
                <div class="mt-0.5 h-2.5 w-2.5 rounded-full bg-[#EFEFEF]" id="readyDot"></div>
                <div class="text-sm text-[#1D1F1E]" id="msg">Upload 3 file wajib dulu, baru tombol Validate nyala.</div>
              </div>

              <div class="mt-3 grid md:grid-cols-3 gap-3">
                <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
                  <div class="text-xs text-[#1D1F1E]">Status A</div>
                  <div class="mt-1 text-sm font-semibold text-[#1D1F1E]">Tidak ada program cocok</div>
                  <div class="mt-1 text-xs text-[#1D1F1E]">Actual ada, expected 0</div>
                </div>
                <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
                  <div class="text-xs text-[#1D1F1E]">Status B</div>
                  <div class="mt-1 text-sm font-semibold text-[#1D1F1E]">Ada selisih</div>
                  <div class="mt-1 text-xs text-[#1D1F1E]">Selisih masuk TanpaTuan</div>
                </div>
                <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
                  <div class="text-xs text-[#1D1F1E]">Status C</div>
                  <div class="mt-1 text-sm font-semibold text-[#1D1F1E]">Sesuai</div>
                  <div class="mt-1 text-xs text-[#1D1F1E]">Dalam toleransi</div>
                </div>
              </div>
            </div>
          </form>
        </div>
      </main>

      <!-- RIGHT -->
      <section class="xl:col-span-3 space-y-5">
        __PATCH_NOTES_HTML__

        <div class="rounded-2xl bg-[#EFEFEF] border border-[#8A4703] shadow-lg overflow-hidden">
          <div class="p-5 border-b border-[#8A4703]">
            <div class="text-xs uppercase tracking-widest text-[#1D1F1E]">Insight</div>
            <div class="text-lg font-semibold text-[#1D1F1E]">Cara bacanya</div>
            <div class="mt-1 text-sm text-[#1D1F1E]">
              Goal: bikin kamu gampang ngecek “ini diskon beneran program atau nyasar”.
            </div>
          </div>

          <div class="p-5 space-y-3 text-sm text-[#1D1F1E]">
            <div class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
              <div class="font-semibold text-[#1D1F1E]">Urutan klaim (penting)</div>
              <div class="mt-1">Actual (LDISCAMOUNT) dibagi mengikuti urutan di <b>MDSTRING</b> untuk
                <b>Pabrik</b>/<b>Internal</b>, sisa masuk <b>TanpaTuan</b>.</div>
            </div>

            <details class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
              <summary class="cursor-pointer font-semibold text-[#1D1F1E]">Kenapa bisa “B”?</summary>
              <div class="mt-2 text-[#1D1F1E]">
                Karena expected (pabrik+internal) beda sama actual. Selisihnya otomatis ditaruh ke TanpaTuan,
                biar gampang investigasi.
              </div>
            </details>

            <details class="rounded-xl border border-[#8A4703] bg-[#EFEFEF] p-4">
              <summary class="cursor-pointer font-semibold text-[#1D1F1E]">Checklist sebelum validasi</summary>
              <ul class="mt-2 list-disc pl-5 space-y-1">
                <li>Pastikan 3 file wajib sudah hijau.</li>
                <li>Kalau internal belum ada, gap masuk TanpaTuan dulu.</li>
                <li>Kalau hasil “aneh”, cek kolom debug yang disediakan.</li>
              </ul>
            </details>
          </div>
        </div>
      </section>

    </div>
  </div>

<script>
const CSRF_TOKEN = "__CSRF_TOKEN__";
const CAN_VALIDATOR_EDIT = __CAN_VALIDATOR_EDIT__;
let downloadUrl = null;

function showToast(title, body){
  const t = document.getElementById('toast');
  document.getElementById('toastTitle').textContent = title;
  document.getElementById('toastBody').textContent = body;
  t.classList.remove('hidden');
  setTimeout(() => t.classList.add('hidden'), 5000);
}

async function checkHealth(){
  try{
    const r = await fetch('/health');
    const j = await r.json();
    document.getElementById('apiStatus').textContent = j.status || 'ok';
    document.getElementById('apiStatus').classList.remove('pulse-soft');
  }catch(e){
    document.getElementById('apiStatus').textContent = 'offline';
  }
}
checkHealth();

document.getElementById('downloadBtn').addEventListener('click', () => {
  if(downloadUrl) window.location.href = downloadUrl;
});

function setDot(id, ok, optional=false){
  const el = document.getElementById(id);
  el.className = "h-2.5 w-2.5 rounded-full " + (ok ? "bg-[#BD7401]" : "bg-[#EFEFEF] ring-1 ring-[#8A4703]");
}

function updateReadyState(){
  const sales = document.getElementById('sales').files[0];
  const promo = document.getElementById('promo').files[0];
  const channel = document.getElementById('channel').files[0];
  const internal = document.getElementById('internal').files[0];

  const ready = !!(sales && promo && channel);
  document.getElementById('btn').disabled = !ready;

  document.getElementById('readyDot').className = "mt-0.5 h-2.5 w-2.5 rounded-full " + (ready ? "bg-[#BD7401]" : "bg-[#EFEFEF] ring-1 ring-[#8A4703]");
  document.getElementById('readyPill').textContent = ready ? "siap" : "belum siap";
  document.getElementById('readyPill').className = "text-[11px] px-2 py-1 rounded-full border " + (ready ? "border-[#8A4703] bg-[#BD7401] text-[#EFEFEF]" : "border-[#8A4703] bg-[#EFEFEF] text-[#1D1F1E]");

  document.getElementById('msg').textContent = ready
    ? "Mantap. Kamu sudah bisa validate."
    : "Upload 3 file wajib dulu, baru tombol Validate nyala.";

  if(!CAN_VALIDATOR_EDIT){
    document.getElementById('btn').disabled = true;
    document.getElementById('msg').textContent = "Akses tidak diizinkan.";
  }

  document.getElementById('nameSales').textContent = sales ? sales.name : "belum diupload";
  document.getElementById('namePromo').textContent = promo ? promo.name : "belum diupload";
  document.getElementById('nameChannel').textContent = channel ? channel.name : "belum diupload";
  document.getElementById('nameInternal').textContent = internal ? internal.name : "boleh kosong";

  setDot('dotSales', !!sales);
  setDot('dotPromo', !!promo);
  setDot('dotChannel', !!channel);
  setDot('dotInternal', !!internal, true);

  document.getElementById('badgeSales').classList.toggle('hidden', !sales);
  document.getElementById('badgePromo').classList.toggle('hidden', !promo);
  document.getElementById('badgeChannel').classList.toggle('hidden', !channel);
  document.getElementById('badgeInternal').classList.toggle('hidden', !internal);
}

['sales','promo','channel','internal'].forEach(id => {
  document.getElementById(id).addEventListener('change', updateReadyState);
});
updateReadyState();

document.getElementById('form').addEventListener('submit', async (e) => {
  e.preventDefault();
  const btn = document.getElementById('btn');
  const dBtn = document.getElementById('downloadBtn');

  const sales = document.getElementById('sales').files[0];
  const promo = document.getElementById('promo').files[0];
  const channel = document.getElementById('channel').files[0];
  const internal = document.getElementById('internal').files[0];

  if(!(sales && promo && channel)){
    showToast("Eh belum bisa 😅", "__REQUIRED_MISSING_MSG__");
    return;
  }

  btn.disabled = true;
  btn.textContent = 'Validating... (lagi ngecek diskon satu-satu 😄)';
  dBtn.disabled = true;

  const fd = new FormData();
  fd.append('sales', sales);
  fd.append('promo', promo);
  fd.append('channel', channel);
  if(internal) fd.append('internal', internal);

  try{
    const res = await fetch('/validate_json', { method: 'POST', body: fd, headers: { 'X-CSRF-Token': CSRF_TOKEN } });
    const j = await res.json();
    if(!res.ok || !j.ok) throw new Error(j.error || ('HTTP ' + res.status));

    document.getElementById('cntA').textContent = j.counts.A;
    document.getElementById('cntB').textContent = j.counts.B;
    document.getElementById('cntC').textContent = j.counts.C;

    downloadUrl = j.download_url;
    dBtn.disabled = false;

    showToast("Selesai ✅", "Download hasilnya lewat tombol 'Download Result' di sidebar kiri.");
  }catch(err){
    showToast("Waduh error 😭", err.message || String(err));
  }finally{
    btn.disabled = false;
    btn.textContent = 'Validate';
  }
});
</script>
</body>
</html>"""












SPPD_EXCEL_FORBIDDEN_FIELDS = {
    "ajukan",
    "gap_nilai",
    "gap_nilai_display",
    "status_pembayaran",
    "status_tracking",
    "tracking_status",
    "submitted_at",
    "submitted_by",
    "submission_id",
    "draft_id",
    "sppd_no",
    "finance_status",
    "payment_proof",
    "proof_metadata",
    "accurate_posting",
}

SPPD_EXCEL_FIELD_ALIASES = {
    "recordid": "record_id",
    "record_id": "record_id",
    "id": "record_id",
    "tippepengajuan": "tipe_pengajuan",
    "tipepengajuan": "tipe_pengajuan",
    "tipe_pengajuan": "tipe_pengajuan",
    "nolpb": "no_lpb",
    "no_lpb": "no_lpb",
    "lpb": "no_lpb",
    "principle": "principle",
    "supplier": "principle",
    "vendor": "principle",
    "tglsetor": "tgl_setor",
    "tgl_setor": "tgl_setor",
    "tglwin": "tgl_win",
    "tgl_win": "tgl_win",
    "tgljtempowin": "tgl_jtempo_win",
    "tgl_jtempo_win": "tgl_jtempo_win",
    "tglterimabarang": "tgl_terima_barang",
    "tgl_terima_barang": "tgl_terima_barang",
    "tglinvoice": "tgl_invoice",
    "tgl_invoice": "tgl_invoice",
    "noinvoice": "invoice_no",
    "invoice": "invoice_no",
    "invoice_no": "invoice_no",
    "nomorinvoice": "invoice_no",
    "jenisdokumen": "jenis_dokumen",
    "jenis_dokumen": "jenis_dokumen",
    "nomordokumen": "nomor_dokumen",
    "nomor_dokumen": "nomor_dokumen",
    "nilaigiro": "nilai_invoice",
    "nilaiinvoice": "nilai_invoice",
    "nilai_invoice": "nilai_invoice",
    "nilaisistem": "nilai_win",
    "nilaiwin": "nilai_win",
    "nilai_win": "nilai_win",
    "potongan": "potongan",
    "nilaipembayaran": "nilai_pembayaran",
    "nilai_pembayaran": "nilai_pembayaran",
    "jtinvoice": "jt_invoice",
    "jt_invoice": "jt_invoice",
    "jatuhtempoinvoice": "jt_invoice",
    "actualdate": "actual_date",
    "actual_date": "actual_date",
    "tglpembayaran": "tgl_pembayaran",
    "tgl_pembayaran": "tgl_pembayaran",
    "tanggalpembayaran": "tgl_pembayaran",
    "tanggalpengajuanpembayaran": "target_payment_date",
    "targetpaymentdate": "target_payment_date",
    "target_payment_date": "target_payment_date",
    "metodepembayaran": "payment_method",
    "paymentmethod": "payment_method",
    "payment_method": "payment_method",
    "jenispembayaran": "jenis_pembayaran",
    "jenis_pembayaran": "jenis_pembayaran",
    "keterangan": "keterangan",
    "ajukan": "ajukan",
    "gap": "gap_nilai",
    "gapnilai": "gap_nilai",
    "gap_nilai": "gap_nilai",
    "status": "status_pembayaran",
    "statuspembayaran": "status_pembayaran",
    "status_pembayaran": "status_pembayaran",
    "statustracking": "status_tracking",
    "status_tracking": "status_tracking",
    "trackingstatus": "tracking_status",
    "tracking_status": "tracking_status",
    "submittedat": "submitted_at",
    "submitted_at": "submitted_at",
    "submittedby": "submitted_by",
    "submitted_by": "submitted_by",
    "submissionid": "submission_id",
    "submission_id": "submission_id",
    "draftid": "draft_id",
    "draft_id": "draft_id",
    "sppdno": "sppd_no",
    "sppd_no": "sppd_no",
}

SPPD_EXCEL_NUMERIC_FIELDS = {"nilai_invoice", "nilai_win", "potongan", "nilai_pembayaran"}
SPPD_EXCEL_DATE_FIELDS = {
    "tgl_setor",
    "tgl_win",
    "tgl_jtempo_win",
    "tgl_terima_barang",
    "tgl_invoice",
    "jt_invoice",
    "actual_date",
    "tgl_pembayaran",
    "target_payment_date",
}

def normalize_excel_field_name(value: Any) -> str:
    text = s(value).lower()
    text = text.replace(".", "")
    text = re.sub(r"[^a-z0-9_]+", "", text)
    return text

def normalize_sppd_excel_value(field: str, value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if field in SPPD_EXCEL_NUMERIC_FIELDS:
        return parse_number_id(value)
    if field in SPPD_EXCEL_DATE_FIELDS:
        return parse_sppd_date_ddmmyyyy(value)
    if field == "tipe_pengajuan":
        return normalize_pengajuan_type(value)
    return s(value)

def parse_sppd_excel_rows(content: bytes) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    df = pd.read_excel(io.BytesIO(content))
    if df.empty:
        return [], [], []
    columns: Dict[str, str] = {}
    ignored_columns: List[str] = []
    blocked_columns: List[str] = []
    for col in df.columns:
        normalized = normalize_excel_field_name(col)
        field = SPPD_EXCEL_FIELD_ALIASES.get(normalized)
        if field in SPPD_EXCEL_FORBIDDEN_FIELDS:
            blocked_columns.append(s(col))
            continue
        if field:
            columns[s(col)] = field
        elif normalized:
            ignored_columns.append(s(col))
    rows: List[Dict[str, Any]] = []
    date_errors: List[str] = []
    for idx, (_, row) in enumerate(df.iterrows()):
        excel_row = idx + 2  # header di baris 1
        item: Dict[str, Any] = {}
        for col, field in columns.items():
            try:
                value = normalize_sppd_excel_value(field, row.get(col))
            except ValueError as de:
                date_errors.append(f"kolom '{s(col)}' baris {excel_row} = {str(de)}")
                continue
            if value is None or s(value) == "":
                continue
            item[field] = value
        if item:
            rows.append(item)
    if date_errors:
        shown = "; ".join(date_errors[:5])
        extra = len(date_errors) - 5
        suffix = f"; dan {extra} lainnya" if extra > 0 else ""
        raise ValueError(
            f"Tanggal tidak valid (harus DD/MM/YYYY): {shown}{suffix}. Upload dibatalkan."
        )
    return rows, ignored_columns, blocked_columns

def _can_access_draft(user: str, draft: Dict[str, Any]) -> bool:
    if not user or not draft:
        return False
    if is_admin_user(user):
        return True
    return s(draft.get("created_by", "")) == s(user)

# ---------------------------
# Finalize
# ---------------------------
def finalize(sales_df: pd.DataFrame, sales_cols_fixed: List[str]) -> pd.DataFrame:
    # Ensure output columns exist before we reference them
    sales_df = sales_df.copy()
    for c, dv in [
        ("Expected_Pabrik", 0.0),
        ("Expected_Internal", 0.0),
        ("Expected_Total", 0.0),
        ("Actual_BonusQty", 0.0),
        ("Expected_Bonus_Pabrik", 0.0),
        ("Expected_Bonus_Internal", 0.0),
        ("Expected_Bonus_Total", 0.0),
        ("FixClaimBonus_Pabrik", 0.0),
        ("FixClaimBonus_Internal", 0.0),
        ("FixClaimBonus_TanpaTuan", 0.0),
        ("Bonus_Selisih", 0.0),
        ("FixClaim_Pabrik", 0.0),
        ("FixClaim_Internal", 0.0),
        ("FixClaim_TanpaTuan", 0.0),
        ("Fix Claim", 0.0),
        ("Selisih", 0.0),
        ("StatusValidasi", ""),
        ("PesanValidasi", ""),
        ("HAS_ANY_PROGRAM", False),
        ("HAS_ELIGIBLE_PROGRAM", False),
        ("Program_Pabrik_Eligible", ""),
        ("Program_Pabrik_Applied", ""),
        ("Program_Bonus_Applied", ""),
        ("Debug_TriggerQty", 0.0),
        ("Bonus_TriggerUnit", ""),
        ("Bonus_TriggerQty", 0.0),
        ("Bonus_TriggerMetric", 0.0),
        ("Bonus_PickedTier", 0),
        ("Penjelasan_Claim", ""),
        ("Penjelasan_Bonus", ""),
    ]:
        if c not in sales_df.columns:
            sales_df.loc[:, c] = dv

    actual = sales_df["LDISCAMOUNT"].astype(float).fillna(0.0)
    exp_p = sales_df["Expected_Pabrik"].astype(float).fillna(0.0)
    exp_i = sales_df["Expected_Internal"].astype(float).fillna(0.0)

    sales_df.loc[:, "Expected_Total"] = exp_p + exp_i

    use_mdstring_claim = all(
        c in sales_df.columns
        for c in ["_PCT_PARTS_PABRIK", "_PCT_PARTS_INTERNAL", "_RP_PABRIK"]
    )

    if use_mdstring_claim:
        def alloc_row(r):
            md_parts = normalize_pct_list(parse_mdstring_pct(r.get("MDSTRING", "")))
            if not md_parts:
                act = float(r.get("LDISCAMOUNT", 0.0) or 0.0)
                exp_p_row = float(r.get("Expected_Pabrik", 0.0) or 0.0)
                exp_i_row = float(r.get("Expected_Internal", 0.0) or 0.0)
                fix_p_row = min(act, exp_p_row)
                rem_row = max(act - fix_p_row, 0.0)
                fix_i_row = min(rem_row, exp_i_row)
                fix_t_row = max(rem_row - fix_i_row, 0.0)
                return pd.Series([fix_p_row, fix_i_row, fix_t_row])

            res = allocate_claim_by_mdstring(
                gross_amount=r.get("GROSSAMOUNT", 0.0),
                actual_discount=r.get("LDISCAMOUNT", 0.0),
                mdstring=r.get("MDSTRING", ""),
                rp_pabrik=r.get("_RP_PABRIK", 0.0),
                pct_pabrik=r.get("_PCT_PARTS_PABRIK", []),
                pct_internal=r.get("_PCT_PARTS_INTERNAL", []),
            )
            if res is None:
                act = float(r.get("LDISCAMOUNT", 0.0) or 0.0)
                exp_p_row = float(r.get("Expected_Pabrik", 0.0) or 0.0)
                exp_i_row = float(r.get("Expected_Internal", 0.0) or 0.0)
                fix_p_row = min(act, exp_p_row)
                rem_row = max(act - fix_p_row, 0.0)
                fix_i_row = min(rem_row, exp_i_row)
                fix_t_row = max(rem_row - fix_i_row, 0.0)
                return pd.Series([fix_p_row, fix_i_row, fix_t_row])

            fix_p_row, fix_i_row, fix_t_row = res
            return pd.Series([fix_p_row, fix_i_row, fix_t_row])

        alloc = sales_df.apply(alloc_row, axis=1)
        sales_df.loc[:, "FixClaim_Pabrik"] = alloc[0].values
        sales_df.loc[:, "FixClaim_Internal"] = alloc[1].values
        sales_df.loc[:, "FixClaim_TanpaTuan"] = alloc[2].values
        sales_df.loc[:, "Fix Claim"] = sales_df["FixClaim_Pabrik"] + sales_df["FixClaim_Internal"]
    else:
        # Allocate Actual -> Pabrik -> Internal -> TanpaTuan (legacy)
        fix_p = pd.concat([actual, exp_p], axis=1).min(axis=1)
        rem = (actual - fix_p).clip(lower=0.0)
        fix_i = pd.concat([rem, exp_i], axis=1).min(axis=1)
        rem2 = (rem - fix_i).clip(lower=0.0)

        sales_df.loc[:, "FixClaim_Pabrik"] = fix_p
        sales_df.loc[:, "FixClaim_Internal"] = fix_i
        sales_df.loc[:, "FixClaim_TanpaTuan"] = rem2
        sales_df.loc[:, "Fix Claim"] = fix_p + fix_i

    # ---------------------------
    # Bonus allocation (PCS)
    # Allocate Actual_BonusQty -> Expected_Bonus_Pabrik -> Expected_Bonus_Internal -> TanpaTuan
    # ---------------------------
    b_actual = sales_df["Actual_BonusQty"].astype(float).fillna(0.0)
    b_exp_p = sales_df["Expected_Bonus_Pabrik"].astype(float).fillna(0.0)
    b_exp_i = sales_df["Expected_Bonus_Internal"].astype(float).fillna(0.0)
    sales_df.loc[:, "Expected_Bonus_Total"] = b_exp_p + b_exp_i

    b_fix_p = pd.concat([b_actual, b_exp_p], axis=1).min(axis=1)
    b_rem = (b_actual - b_fix_p).clip(lower=0.0)
    b_fix_i = pd.concat([b_rem, b_exp_i], axis=1).min(axis=1)
    b_rem2 = (b_rem - b_fix_i).clip(lower=0.0)

    sales_df.loc[:, "FixClaimBonus_Pabrik"] = b_fix_p
    sales_df.loc[:, "FixClaimBonus_Internal"] = b_fix_i
    sales_df.loc[:, "FixClaimBonus_TanpaTuan"] = b_rem2
    sales_df.loc[:, "Bonus_Selisih"] = sales_df["Expected_Bonus_Total"].astype(float).fillna(0.0) - b_actual

    sales_df.loc[:, "Selisih"] = actual - (exp_p + exp_i)

    TOL = 1.0

    def status_row(r):
        act = float(r["LDISCAMOUNT"] or 0.0)
        exp = float(r["Expected_Total"] or 0.0)
        anyp = bool(r.get("HAS_ANY_PROGRAM", False))
        elig = bool(r.get("HAS_ELIGIBLE_PROGRAM", False))

        if act <= 0 and exp <= 0:
            return "C"
        if (not anyp) and act > 0:
            return "A"
        if anyp and (not elig) and act > 0 and exp <= 0:
            return "B"
        if abs(act - exp) <= TOL:
            return "C"
        return "B"

    sales_df.loc[:, "StatusValidasi"] = sales_df.apply(status_row, axis=1)

    def pesan_row(r):
        st = r["StatusValidasi"]
        act = float(r["LDISCAMOUNT"] or 0.0)
        exp = float(r["Expected_Total"] or 0.0)
        if st == "A":
            return "Tidak ada program yang cocok untuk item/outlet/channel ini (masuk TanpaTuan)"
        if st == "B":
            if bool(r.get("HAS_ANY_PROGRAM", False)) and (not bool(r.get("HAS_ELIGIBLE_PROGRAM", False))) and exp <= 0 and act > 0:
                return "Ada program di dataset, tapi faktur tidak memenuhi trigger/tier (expected 0). Masuk TanpaTuan."
            return f"Expected total {exp:.2f}, aktual {act:.2f}, selisih {(act-exp):.2f}. Selisih masuk TanpaTuan."
        return "Sesuai (toleransi)"

    sales_df.loc[:, "PesanValidasi"] = sales_df.apply(pesan_row, axis=1)

    def explain_claim_row(r):
        parts = []
        fix_p = float(r.get("FixClaim_Pabrik", 0.0) or 0.0)
        fix_i = float(r.get("FixClaim_Internal", 0.0) or 0.0)
        fix_t = float(r.get("FixClaim_TanpaTuan", 0.0) or 0.0)

        tu = s(r.get("Debug_TriggerUnit", ""))
        tq = float(r.get("Debug_TriggerQty", 0.0) or 0.0)
        metric = float(r.get("Debug_TriggerMetric", 0.0) or 0.0)
        tier = int(r.get("Debug_PickedTier", 0) or 0)

        trigger = ""
        if tu and tq > 0:
            trigger = f"trigger min {fmt_metric(tq, tu)} {tu}, tercapai {fmt_metric(metric, tu)} {tu}"
            if tier > 0:
                trigger += f", tier {tier}"

        md_parts = normalize_pct_list(parse_mdstring_pct(r.get("MDSTRING", "")))
        elig_parts = []
        elig_parts.extend(normalize_pct_list(r.get("_PCT_PARTS_PABRIK", [])))
        elig_parts.extend(normalize_pct_list(r.get("_PCT_PARTS_INTERNAL", [])))
        md_sum = sum(md_parts) if md_parts else 0.0
        elig_sum = sum(elig_parts) if elig_parts else 0.0
        leftover = max(md_sum - elig_sum, 0.0) if md_parts else 0.0

        if fix_p > 0:
            msg = "Pabrik: memenuhi trigger"
            if trigger:
                msg = f"Pabrik: hanya mencukupi {trigger}"
            if elig_sum > 0:
                msg += f" (disc {fmt_pct(elig_sum)}%)"
            parts.append(msg)

        if fix_i > 0:
            parts.append("Internal: rule match")

        if fix_t > 0:
            msg = "TanpaTuan: selisih diskon masuk TanpaTuan"
            if md_parts:
                md_s = pct_list_str(md_parts)
                if elig_sum > 0 and leftover > 0:
                    msg = f"TanpaTuan: sisa {fmt_pct(leftover)}% dari MDSTRING {fmt_pct(md_sum)}% masuk TanpaTuan"
                elif elig_sum <= 0:
                    msg = f"TanpaTuan: MDSTRING {md_s}% tidak ada program eligible; semua masuk TanpaTuan"
            parts.append(msg)

        return " | ".join([p for p in parts if p])

    def explain_bonus_row(r):
        actual = float(r.get("Actual_BonusQty", 0.0) or 0.0)
        expected = float(r.get("Expected_Bonus_Total", 0.0) or 0.0)
        if actual <= 0 and expected <= 0:
            return ""
        tanp = float(r.get("FixClaimBonus_TanpaTuan", 0.0) or 0.0)
        prog = s(r.get("Program_Bonus_Applied", "")) or s(r.get("Program_Pabrik_Applied", ""))
        tu = s(r.get("Bonus_TriggerUnit", ""))
        tq = float(r.get("Bonus_TriggerQty", 0.0) or 0.0)
        metric = float(r.get("Bonus_TriggerMetric", 0.0) or 0.0)
        tier = int(r.get("Bonus_PickedTier", 0) or 0)

        trigger = ""
        if tu and tq > 0:
            trigger = f"trigger min {fmt_metric(tq, tu)} {tu}, tercapai {fmt_metric(metric, tu)} {tu}"
            if tier > 0:
                trigger += f", tier {tier}"

        if tanp > 0:
            msg = f"Bonus eligible {fmt_metric(expected, 'PCS')} pcs, actual {fmt_metric(actual, 'PCS')} pcs; {fmt_metric(tanp, 'PCS')} pcs masuk TanpaTuan"
        elif expected > actual + 1e-9:
            diff = expected - actual
            msg = f"Bonus eligible {fmt_metric(expected, 'PCS')} pcs, actual {fmt_metric(actual, 'PCS')} pcs (kurang {fmt_metric(diff, 'PCS')} pcs)"
        else:
            msg = f"Bonus sesuai {fmt_metric(actual, 'PCS')} pcs"

        if trigger:
            msg += f" ({trigger})"
        return msg

    sales_df.loc[:, "Penjelasan_Claim"] = sales_df.apply(explain_claim_row, axis=1)
    sales_df.loc[:, "Penjelasan_Bonus"] = sales_df.apply(explain_bonus_row, axis=1)

    def program_name_row(r):
        name = s(r.get("Program_Pabrik_Applied", "")) or s(r.get("Program_Pabrik_Eligible", ""))
        if not name and bool(r.get("_IS_BONUS_LINE_MAIN", False)):
            name = s(r.get("Program_Bonus_Applied", ""))
        return name

    sales_df.loc[:, "Nama_Program_Pabrik"] = sales_df.apply(program_name_row, axis=1)

    out_cols = sales_cols_fixed + [
        "EffectiveChannel",
        "Nama_Program_Pabrik",
        "FixClaim_Pabrik", "FixClaim_Internal", "FixClaim_TanpaTuan",
        "StatusValidasi",
        "Penjelasan_Claim", "Penjelasan_Bonus",
    ]
    return sales_df[out_cols].copy()



def write_excel(out_df: pd.DataFrame, out_path: str):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="RESULT")

def accel_or_file_response(out_path: str, download_name: str):
    if not os.path.exists(out_path):
        return JSONResponse(status_code=404, content={"detail": "File not found"})

    if USE_X_ACCEL:
        prefix = X_ACCEL_PREFIX.rstrip("/") or "/protected-downloads"
        accel_path = f"{prefix}/{os.path.basename(out_path)}"
        headers = {
            "X-Accel-Redirect": accel_path,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f'attachment; filename="{download_name}"',
        }
        return Response(status_code=200, headers=headers)

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name
    )

# ---------------------------
# Load channel mapping
# ---------------------------
def load_channel_map(channel_bytes: bytes) -> Dict[str, str]:
    df = pd.read_excel(io.BytesIO(channel_bytes))
    # accept either column names:
    # - "KODE OUTLET" + "CHANNEL" (your current template)
    # - "SUB" + "CHANNEL_CODE"
    cols = {c.strip().upper(): c for c in df.columns}
    sub_col = cols.get("KODE OUTLET") or cols.get("SUB")
    ch_col = cols.get("CHANNEL") or cols.get("CHANNEL_CODE") or cols.get("CHANNELNAME") or cols.get("CHANNEL_NAME")
    if not sub_col or not ch_col:
        raise ValueError("Data Channel by SUB harus punya kolom 'KODE OUTLET'/'SUB' dan 'CHANNEL' (contoh: KODE OUTLET, CHANNEL).")
    df[sub_col] = df[sub_col].apply(s)
    df[ch_col] = df[ch_col].apply(s)
    mp = {}
    for _, r in df.iterrows():
        if s(r[sub_col]) and s(r[ch_col]):
            mp[s(r[sub_col])] = s(r[ch_col])
    return mp

# ---------------------------
# Engine
# ---------------------------
def run_engine(sales_bytes: bytes, promo_bytes: bytes, channel_bytes: bytes, internal_bytes: Optional[bytes]) -> pd.DataFrame:
    sales_df = pd.read_excel(io.BytesIO(sales_bytes))
    promo_df = pd.read_excel(io.BytesIO(promo_bytes))

    # Load channel map (required)
    channel_map = load_channel_map(channel_bytes)

    # Load internal (optional; used in v15)
    internal_rules = load_internal_rules(internal_bytes)
    if internal_bytes:
        try:
            internal_df = pd.read_excel(io.BytesIO(internal_bytes))
        except:
            internal_df = None

    sales_cols_fixed = [
        "SUBNAME", "SUB", "SUBADDRESS", "SUBKOTA", "JPSNAMA_CHANNEL",
        "BRG", "BRGNAME", "UNIT", "PACKAGING", "INVOICEDATE", "INVOICENO",
        "SALESORDER", "SALESMAN", "QTYSTR", "QTYPCS", "QTY", "QTYBONUSSTR",
        "PRICE", "GROSSAMOUNT", "MDSTRING", "LDISCAMOUNT", "DPP", "TAXVAL", "NETAMOUNT"
    ]
    for c in sales_cols_fixed:
        if c in ["QTYPCS", "QTY", "QTYBONUSSTR", "PRICE", "GROSSAMOUNT", "LDISCAMOUNT", "DPP", "TAXVAL", "NETAMOUNT"]:
            ensure_col(sales_df, c, 0.0)
        else:
            ensure_col(sales_df, c, "")

    for c in ["SUBNAME", "SUB", "SUBADDRESS", "SUBKOTA", "JPSNAMA_CHANNEL", "BRG", "BRGNAME",
              "UNIT", "PACKAGING", "INVOICENO", "SALESORDER", "SALESMAN", "QTYSTR", "MDSTRING"]:
        sales_df[c] = sales_df[c].apply(s)

    sales_df = to_float_series(
        sales_df,
        ["QTYPCS", "QTY", "QTYBONUSSTR", "PRICE", "GROSSAMOUNT", "LDISCAMOUNT", "DPP", "TAXVAL", "NETAMOUNT"]
    )
    sales_df = sales_df.copy()
    sales_df["_ROWID"] = range(len(sales_df))
    sales_df["_IS_BONUS_LINE_MAIN"] = sales_df["MDSTRING"].apply(mdstring_is_bonus)

    # Effective channel (from SUB map)
    def eff_channel(row):
        sub = s(row.get("SUB", ""))
        if sub in channel_map and channel_map[sub]:
            return channel_map[sub]
        return s(row.get("JPSNAMA_CHANNEL", ""))

    sales_df["EffectiveChannel"] = sales_df.apply(eff_channel, axis=1)

    promo_needed = [
        "KODE_BARANG", "NAMA_BARANG", "PROMO_LABEL", "PROMO_GROUP_ID", "PROMO_GROUP", "PROMO_ACTIVE",
        "TIER_NO", "TRIGGER_QTY", "TRIGGER_UNIT", "BENEFIT_TYPE", "BENEFIT_VALUE", "BENEFIT_UNIT"
    ]
    for c in promo_needed:
        ensure_col(promo_df, c, "")

    promo_df["KODE_BARANG"] = promo_df["KODE_BARANG"].apply(s)
    promo_df["NAMA_BARANG"] = promo_df["NAMA_BARANG"].apply(s)
    promo_df["PROMO_LABEL"] = promo_df["PROMO_LABEL"].apply(s)
    promo_df["PROMO_GROUP_ID"] = promo_df["PROMO_GROUP_ID"].apply(s)
    promo_df["PROMO_GROUP"] = promo_df["PROMO_GROUP"].apply(s)
    promo_df["TRIGGER_UNIT"] = promo_df["TRIGGER_UNIT"].apply(s)

    promo_df["TRIGGER_QTY"] = promo_df["TRIGGER_QTY"].apply(parse_number_id)
    promo_df["TIER_NO"] = pd.to_numeric(promo_df["TIER_NO"], errors="coerce").fillna(0).astype(int)

    promo_df["BENEFIT_UNIT"] = promo_df["BENEFIT_UNIT"].apply(s)
    promo_df["BENEFIT_TYPE"] = promo_df["BENEFIT_TYPE"].apply(s).str.upper()
    promo_df["BENEFIT_VALUE"] = promo_df["BENEFIT_VALUE"].apply(s)
    promo_df["_ALL_ITEM_RULE"] = (promo_df["KODE_BARANG"].apply(s) == "") & (promo_df["NAMA_BARANG"].apply(s) == "")

    if "PROMO_ACTIVE" in promo_df.columns:
        def is_active(v):
            if pd.isna(v):
                return True
            if isinstance(v, bool):
                return v
            return str(v).strip().lower() in ["true", "1", "yes", "y"]
        promo_df = promo_df[promo_df["PROMO_ACTIVE"].apply(is_active)].copy()

    # outputs + debug
    sales_df.loc[:, "HAS_ANY_PROGRAM"] = False
    sales_df.loc[:, "HAS_ELIGIBLE_PROGRAM"] = False
    sales_df.loc[:, "Expected_Pabrik"] = 0.0

    sales_df.loc[:, "Debug_ProgramKeysEligible"] = ""
    sales_df.loc[:, "Debug_ProgramKeysApplied"] = ""
    sales_df.loc[:, "Debug_TriggerUnit"] = ""
    sales_df.loc[:, "Debug_TriggerUnitSource"] = ""
    sales_df.loc[:, "Debug_TriggerMetric"] = 0.0
    sales_df.loc[:, "Debug_TriggerQty"] = 0.0
    sales_df.loc[:, "Debug_PickedTier"] = 0
    sales_df.loc[:, "Debug_SkipReason"] = ""
    sales_df.loc[:, "Debug_ProgramLock_ClaimedPct"] = ""
    sales_df.loc[:, "Debug_ProgramLock_SelectedPctPrograms"] = ""
    sales_df.loc[:, "Program_Pabrik_Eligible"] = ""
    sales_df.loc[:, "Program_Pabrik_Applied"] = ""
    sales_df.loc[:, "Program_Bonus_Applied"] = ""
    sales_df.loc[:, "_INTERNAL_BONUS_MATCH"] = False
    sales_df.loc[:, "Bonus_TriggerUnit"] = ""
    sales_df.loc[:, "Bonus_TriggerQty"] = 0.0
    sales_df.loc[:, "Bonus_TriggerMetric"] = 0.0
    sales_df.loc[:, "Bonus_PickedTier"] = 0
    sales_df.loc[:, "_PCT_PARTS_PABRIK"] = [list() for _ in range(len(sales_df))]
    sales_df.loc[:, "_PCT_PARTS_INTERNAL"] = [list() for _ in range(len(sales_df))]
    sales_df.loc[:, "_RP_PABRIK"] = 0.0

    def scope_program_rows(df: pd.DataFrame, pgid: str, plabel: str, brgs: List[str]) -> pd.DataFrame:
        r = df[
            (df["PROMO_GROUP_ID"].apply(s) == pgid) &
            (df["PROMO_LABEL"].apply(s) == plabel)
        ].copy()
        if r.empty:
            return r
        return r[(r["KODE_BARANG"].apply(s).isin(brgs)) | (r["_ALL_ITEM_RULE"] == True)].copy()

    def pick_rows_for_item(rows: pd.DataFrame, brg: str) -> pd.DataFrame:
        exact = rows[rows["KODE_BARANG"].apply(s) == brg].copy()
        if not exact.empty:
            return exact
        return rows[rows["_ALL_ITEM_RULE"] == True].copy()

    # Build candidate rows by channel (use EffectiveChannel!)
    promo_exact = promo_df[promo_df["_ALL_ITEM_RULE"] == False].copy()
    merged_exact = sales_df.merge(promo_exact, left_on="BRG", right_on="KODE_BARANG", how="left", suffixes=("", "_P"))
    merged_all = pd.DataFrame(columns=merged_exact.columns)
    promo_all = promo_df[promo_df["_ALL_ITEM_RULE"] == True].copy()
    if not promo_all.empty:
        merged_all = (
            sales_df.assign(_J=1)
            .merge(promo_all.assign(_J=1), on="_J", how="inner", suffixes=("", "_P"))
            .drop(columns=["_J"])
        )
    merged = pd.concat([merged_exact, merged_all], ignore_index=True, sort=False)
    merged["CHANNEL_OK"] = merged.apply(
        lambda r: channel_ok(r.get("PROMO_GROUP", ""), r.get("EffectiveChannel", ""), r.get("PROMO_GROUP_ID", ""), r.get("SUB", "")),
        axis=1
    )
    cand = merged[(merged["KODE_BARANG"].notna()) & (merged["CHANNEL_OK"] == True)].copy()
    if cand.empty:
        return finalize(sales_df, sales_cols_fixed)

    any_rowids = cand["_ROWID"].dropna().astype(int).unique().tolist()
    sales_df.loc[sales_df["_ROWID"].isin(any_rowids), "HAS_ANY_PROGRAM"] = True

    cand["PGID"] = cand["PROMO_GROUP_ID"].apply(s)
    cand["PLABEL"] = cand["PROMO_LABEL"].apply(s)
    cand["PROGRAM_KEY"] = cand["PGID"] + "||" + cand["PLABEL"]
    cand["IS_NON_GROUP"] = cand["PGID"].str.upper().eq("NON_GROUP")

    # Program scope per invoice+program_key
    program_scopes: Dict[Tuple[str, str], List[int]] = {}
    program_defs: Dict[Tuple[str, str], Dict] = {}
    for (inv, program_key), g in cand.groupby(["INVOICENO", "PROGRAM_KEY"], sort=True):
        inv = s(inv); program_key = s(program_key)
        rowids = g["_ROWID"].dropna().astype(int).unique().tolist()
        idxs = sales_df.index[sales_df["_ROWID"].isin(rowids)].tolist()
        program_scopes[(inv, program_key)] = idxs
        program_defs[(inv, program_key)] = {
            "pgid": s(g["PGID"].iloc[0]),
            "plabel": s(g["PLABEL"].iloc[0]),
            "is_non_group": bool(g["IS_NON_GROUP"].iloc[0])
        }
        for idx in idxs:
            cur = sales_df.at[idx, "Debug_ProgramKeysEligible"]
            sales_df.at[idx, "Debug_ProgramKeysEligible"] = (cur + (";" if cur else "") + program_key)
            sales_df.at[idx, "Program_Pabrik_Eligible"] = add_list_item(
                sales_df.at[idx, "Program_Pabrik_Eligible"], s(g["PLABEL"].iloc[0])
            )

    # PROGRAM LOCK (DISC_PCT grouped only) — uses mdstring per invoice
    inv_claimed_pct: Dict[str, Set[float]] = {}
    for inv, g in sales_df.groupby("INVOICENO"):
        claimed = []
        for md in g["MDSTRING"].tolist():
            claimed.extend(parse_mdstring_pct(md))
        claimed = [round(x, 4) for x in claimed if abs(x) > 1e-9]
        inv_claimed_pct[s(inv)] = set(claimed)

    inv_selected_pct_programs: Dict[str, Set[str]] = {}
    for inv, g in cand.groupby("INVOICENO"):
        inv = s(inv)
        claimed = inv_claimed_pct.get(inv, set())
        if not claimed:
            inv_selected_pct_programs[inv] = set()
            continue

        scores: Dict[str, int] = {}
        for program_key in g["PROGRAM_KEY"].unique().tolist():
            program_key = s(program_key)
            if program_defs.get((inv, program_key), {}).get("is_non_group", False):
                continue
            idxs = program_scopes.get((inv, program_key), [])
            if not idxs:
                continue
            scope = sales_df.loc[idxs]
            sub = s(scope["SUB"].iloc[0])
            eff_channel = s(scope["EffectiveChannel"].iloc[0])
            pgid = program_defs[(inv, program_key)]["pgid"]
            plabel = program_defs[(inv, program_key)]["plabel"]

            brgs = scope["BRG"].apply(s).unique().tolist()
            prog_rows = scope_program_rows(promo_df, pgid, plabel, brgs)
            if prog_rows.empty:
                continue

            prog_rows = prog_rows[prog_rows.apply(lambda r: channel_ok(r.get("PROMO_GROUP", ""), eff_channel, r.get("PROMO_GROUP_ID", ""), sub), axis=1)].copy()
            if prog_rows.empty:
                continue

            pct_rows = prog_rows[prog_rows["BENEFIT_TYPE"].str.upper() == "DISC_PCT"]
            if pct_rows.empty:
                continue

            program_pcts: Set[float] = set()
            for v in pct_rows["BENEFIT_VALUE"].tolist():
                for p in parse_pct_chain(s(v)):
                    if abs(p) > 1e-9:
                        program_pcts.add(round(p, 4))
            if not program_pcts:
                continue

            score = len(program_pcts.intersection(claimed))
            scores[program_key] = score

        if not scores:
            inv_selected_pct_programs[inv] = set()
            continue

        max_score = max(scores.values())
        if max_score <= 0:
            inv_selected_pct_programs[inv] = set()
        else:
            inv_selected_pct_programs[inv] = {k for k, sc in scores.items() if sc == max_score}

    for idx, r in sales_df.iterrows():
        inv = s(r["INVOICENO"])
        claimed = sorted(inv_claimed_pct.get(inv, set()))
        selected = sorted(inv_selected_pct_programs.get(inv, set()))
        sales_df.at[idx, "Debug_ProgramLock_ClaimedPct"] = ",".join([str(x).rstrip("0").rstrip(".") for x in claimed])
        sales_df.at[idx, "Debug_ProgramLock_SelectedPctPrograms"] = ";".join(selected)

    # Compute expected
    rp_line = pd.Series(0.0, index=sales_df.index)
    pct_tasks: List[Dict] = []
    eligible_rowids = set()

    # PASS 1 - RP
    for (inv, program_key) in sorted(program_scopes.keys()):
        idxs = program_scopes[(inv, program_key)]
        if not idxs:
            continue

        scope = sales_df.loc[idxs].copy()
        sub = s(scope["SUB"].iloc[0])
        eff_channel = s(scope["EffectiveChannel"].iloc[0])
        pgid = program_defs[(inv, program_key)]["pgid"]
        plabel = program_defs[(inv, program_key)]["plabel"]
        is_non_group = program_defs[(inv, program_key)]["is_non_group"]

        brgs = scope["BRG"].apply(s).unique().tolist()
        prog_rows = scope_program_rows(promo_df, pgid, plabel, brgs)
        if prog_rows.empty:
            continue

        prog_rows = prog_rows[prog_rows.apply(lambda r: channel_ok(r.get("PROMO_GROUP", ""), eff_channel, r.get("PROMO_GROUP_ID", ""), sub), axis=1)].copy()
        if prog_rows.empty:
            continue

        trigger_unit, tu_source = choose_trigger_unit_satpam(prog_rows)
        if trigger_unit is None:
            for idx in idxs:
                sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (rp)")
                sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
            continue

        if not is_non_group:
            metric = float(scope["GROSSAMOUNT"].sum() or 0.0) if trigger_unit == "GROSSAMOUNT" else float(scope.apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum() or 0.0)
            picked_tier, tier_rows = pick_best_tier(prog_rows, metric)
            if tier_rows.empty:
                continue

            tier_rows = dedupe_grouped_benefits(tier_rows)
            tq_val = float(pd.to_numeric(tier_rows["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)

            for idx in idxs:
                eligible_rowids.add(int(sales_df.at[idx, "_ROWID"]))
                sales_df.at[idx, "Debug_TriggerUnit"] = trigger_unit
                sales_df.at[idx, "Debug_TriggerUnitSource"] = tu_source
                sales_df.at[idx, "Debug_TriggerMetric"] = metric
                sales_df.at[idx, "Debug_TriggerQty"] = tq_val
                sales_df.at[idx, "Debug_PickedTier"] = picked_tier

            disc_rp_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_RP"].copy()
            if disc_rp_rows.empty:
                continue

            line_gross = scope["GROSSAMOUNT"].astype(float); sum_gross = float(line_gross.sum() or 0.0)
            line_pcs = scope.apply(lambda r: qty_in_unit(r, "PCS"), axis=1).astype(float); sum_pcs = float(line_pcs.sum() or 0.0)
            line_ctn = scope.apply(lambda r: qty_in_unit(r, "CTN"), axis=1).astype(float); sum_ctn = float(line_ctn.sum() or 0.0)

            for _, tr in disc_rp_rows.iterrows():
                bunit = norm(tr.get("BENEFIT_UNIT", "PCS"))
                rp = parse_number_id(tr.get("BENEFIT_VALUE", "0"))
                rp_dpp = rp / (1 + PPN_RATE)

                if bunit == "GROSSAMOUNT":
                    total_rp = rp_dpp
                    add = (total_rp * (line_gross / sum_gross)) if sum_gross > 0 else 0.0
                elif bunit == "PCS":
                    total_rp = rp_dpp * sum_pcs
                    add = (total_rp * (line_pcs / sum_pcs)) if sum_pcs > 0 else 0.0
                elif bunit in ["CTN", "KRT"]:
                    total_rp = rp_dpp * sum_ctn
                    add = (total_rp * (line_ctn / sum_ctn)) if sum_ctn > 0 else 0.0
                else:
                    total_rp = rp_dpp * sum_pcs
                    add = (total_rp * (line_pcs / sum_pcs)) if sum_pcs > 0 else 0.0

                if hasattr(add, "index"):
                    for idx in idxs:
                        rp_line.loc[idx] += float(add.loc[idx])
                else:
                    for idx in idxs:
                        rp_line.loc[idx] += float(add)

                for idx in idxs:
                    cur = sales_df.at[idx, "Debug_ProgramKeysApplied"]
                    sales_df.at[idx, "Debug_ProgramKeysApplied"] = (cur + (";" if cur else "") + program_key)
                    sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                        sales_df.at[idx, "Program_Pabrik_Applied"], program_defs[(inv, program_key)]["plabel"]
                    )

        else:
            for idx in idxs:
                line = sales_df.loc[idx]
                brg = s(line["BRG"])
                rows_item = pick_rows_for_item(prog_rows, brg)
                if rows_item.empty:
                    continue

                trigger_unit_line, tu_source_line = choose_trigger_unit_satpam(rows_item)
                if trigger_unit_line is None:
                    sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (rp/non_group)")
                    sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
                    continue

                metric = float(line["GROSSAMOUNT"]) if trigger_unit_line == "GROSSAMOUNT" else qty_in_unit(line, trigger_unit_line)
                picked_tier, tier_rows = pick_best_tier(rows_item, metric)
                if tier_rows.empty:
                    continue
                tq_val = float(pd.to_numeric(tier_rows["TRIGGER_QTY"], errors="coerce").fillna(0).max() or 0.0)

                eligible_rowids.add(int(line["_ROWID"]))
                sales_df.at[idx, "Debug_TriggerUnit"] = trigger_unit_line
                sales_df.at[idx, "Debug_TriggerUnitSource"] = tu_source_line
                sales_df.at[idx, "Debug_TriggerMetric"] = metric
                sales_df.at[idx, "Debug_TriggerQty"] = tq_val
                sales_df.at[idx, "Debug_PickedTier"] = picked_tier

                disc_rp_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_RP"].copy()
                if disc_rp_rows.empty:
                    continue

                rp_add = 0.0
                for _, tr in disc_rp_rows.iterrows():
                    bunit = norm(tr.get("BENEFIT_UNIT", "PCS"))
                    rp = parse_number_id(tr.get("BENEFIT_VALUE", "0"))
                    rp_dpp = rp / (1 + PPN_RATE)
                    rp_add += rp_dpp * qty_in_unit(line, bunit)

                if rp_add:
                    rp_line.loc[idx] += rp_add
                    cur = sales_df.at[idx, "Debug_ProgramKeysApplied"]
                    sales_df.at[idx, "Debug_ProgramKeysApplied"] = (cur + (";" if cur else "") + program_key)
                    sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                        sales_df.at[idx, "Program_Pabrik_Applied"], program_defs[(inv, program_key)]["plabel"]
                    )

    sales_df.loc[:, "_RP_PABRIK"] = rp_line.values

    # PASS 2 - PCT collect (with lock)
    for (inv, program_key) in sorted(program_scopes.keys()):
        idxs = program_scopes[(inv, program_key)]
        if not idxs:
            continue
        scope = sales_df.loc[idxs].copy()
        sub = s(scope["SUB"].iloc[0])
        eff_channel = s(scope["EffectiveChannel"].iloc[0])
        pgid = program_defs[(inv, program_key)]["pgid"]
        plabel = program_defs[(inv, program_key)]["plabel"]
        is_non_group = program_defs[(inv, program_key)]["is_non_group"]

        if not is_non_group:
            selected = inv_selected_pct_programs.get(s(inv), set())
            if selected and (program_key not in selected):
                for idx in idxs:
                    sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], f"LOCK_SKIP: pct program {program_key} not matching MDSTRING")
                continue

        brgs = scope["BRG"].apply(s).unique().tolist()
        prog_rows = scope_program_rows(promo_df, pgid, plabel, brgs)
        if prog_rows.empty:
            continue
        prog_rows = prog_rows[prog_rows.apply(lambda r: channel_ok(r.get("PROMO_GROUP", ""), eff_channel, r.get("PROMO_GROUP_ID", ""), sub), axis=1)].copy()
        if prog_rows.empty:
            continue

        trigger_unit, tu_source = choose_trigger_unit_satpam(prog_rows)
        if trigger_unit is None:
            for idx in idxs:
                sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (pct)")
                sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
            continue

        if is_non_group:
            for idx in idxs:
                line = sales_df.loc[idx]
                brg = s(line["BRG"])
                rows_item = pick_rows_for_item(prog_rows, brg)
                if rows_item.empty:
                    continue

                trigger_unit_line, tu_source_line = choose_trigger_unit_satpam(rows_item)
                if trigger_unit_line is None:
                    sales_df.at[idx, "Debug_SkipReason"] = add_reason(sales_df.at[idx, "Debug_SkipReason"], "SKIP: trigger unit UNKNOWN (pct/non_group)")
                    sales_df.at[idx, "Debug_TriggerUnitSource"] = "UNKNOWN"
                    continue

                metric = float(line["GROSSAMOUNT"]) if trigger_unit_line == "GROSSAMOUNT" else qty_in_unit(line, trigger_unit_line)
                picked_tier, tier_rows = pick_best_tier(rows_item, metric)
                if tier_rows.empty:
                    continue
                disc_pct_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_PCT"].copy()
                if disc_pct_rows.empty:
                    continue
                pct_parts = []
                for _, tr in disc_pct_rows.iterrows():
                    pct_parts.extend(parse_pct_chain(s(tr.get("BENEFIT_VALUE", ""))))
                if pct_parts:
                    pct_tasks.append({"indices": [idx], "pct_parts": pct_parts, "program_key": program_key})
                    cur = sales_df.at[idx, "_PCT_PARTS_PABRIK"]
                    if not isinstance(cur, list):
                        cur = []
                    cur.extend(pct_parts)
                    sales_df.at[idx, "_PCT_PARTS_PABRIK"] = cur
        else:
            metric = float(scope["GROSSAMOUNT"].sum() or 0.0) if trigger_unit == "GROSSAMOUNT" else float(scope.apply(lambda r: qty_in_unit(r, trigger_unit), axis=1).sum() or 0.0)
            picked_tier, tier_rows = pick_best_tier(prog_rows, metric)
            if tier_rows.empty:
                continue

            tier_rows = dedupe_grouped_benefits(tier_rows)

            disc_pct_rows = tier_rows[tier_rows["BENEFIT_TYPE"].str.upper() == "DISC_PCT"].copy()
            if disc_pct_rows.empty:
                continue

            pct_parts_scope = []
            for _, tr in disc_pct_rows.iterrows():
                pct_parts_scope.extend(parse_pct_chain(s(tr.get("BENEFIT_VALUE", ""))))
            if pct_parts_scope:
                pct_tasks.append({"indices": idxs, "pct_parts": pct_parts_scope, "program_key": program_key})
                for idx in idxs:
                    cur = sales_df.at[idx, "_PCT_PARTS_PABRIK"]
                    if not isinstance(cur, list):
                        cur = []
                    cur.extend(pct_parts_scope)
                    sales_df.at[idx, "_PCT_PARTS_PABRIK"] = cur

    # Apply % after rp_line known
    expected = rp_line.copy()
    for task in pct_tasks:
        idxs = task["indices"]; pct_parts = task["pct_parts"]; program_key = task.get("program_key", "")
        if not idxs or not pct_parts:
            continue
        scope = sales_df.loc[idxs].copy()
        rem = (scope["GROSSAMOUNT"].astype(float) - rp_line.loc[idxs]).clip(lower=0.0)
        base = float(rem.sum() or 0.0)
        if base <= 0:
            continue
        pct_total = sequential_pct_discount(base, pct_parts)
        if pct_total <= 0:
            continue
        srem = float(rem.sum() or 0.0)
        shares = (rem / srem) if srem > 0 else 0.0
        if hasattr(shares, "iloc"):
            for i, idx in enumerate(scope.index):
                expected.loc[idx] += float(pct_total * shares.iloc[i])

        for idx in idxs:
            cur = sales_df.at[idx, "Debug_ProgramKeysApplied"]
            if program_key:
                sales_df.at[idx, "Debug_ProgramKeysApplied"] = (cur + (";" if cur else "") + program_key)
                label = program_key.split("||", 1)[1] if "||" in program_key else program_key
                sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                    sales_df.at[idx, "Program_Pabrik_Applied"], label
                )

    sales_df.loc[:, "Expected_Pabrik"] = expected.values
    sales_df.loc[sales_df["_ROWID"].isin(list(eligible_rowids)), "HAS_ELIGIBLE_PROGRAM"] = True

    # ---------------------------
    # INTERNAL expected (v15)
    # ---------------------------
    sales_df.loc[:, "Expected_Internal"] = 0.0
    if internal_rules is not None and (not internal_rules.empty):
        for inv, g in sales_df.groupby("INVOICENO", sort=False):
            inv = s(inv)
            idxs = g.index.tolist()
            if not idxs:
                continue

            invoice_gross = float(g["GROSSAMOUNT"].sum() or 0.0)
            eff_channel = s(g["EffectiveChannel"].iloc[0])
            sub = s(g["SUB"].iloc[0])

            r0 = internal_rules[internal_rules["CHANNEL"].apply(lambda x: internal_channel_match(x, eff_channel))].copy()
            r0 = r0[(r0["SUB"].apply(lambda x: (s(x) == "") or (norm(x) == norm(sub))))].copy()
            if r0.empty:
                continue

            rem_line = (g["GROSSAMOUNT"].astype(float) - g["Expected_Pabrik"].astype(float)).clip(lower=0.0)
            rem_total = float(rem_line.sum() or 0.0)
            if rem_total <= 0:
                continue

            bonus_mask = g["_IS_BONUS_LINE_MAIN"] if "_IS_BONUS_LINE_MAIN" in g.columns else g["MDSTRING"].apply(mdstring_is_bonus)

            def apply_internal_indices(idxs_apply: List[int], pct_parts: List[float], mark_bonus: bool) -> bool:
                if not idxs_apply:
                    return False
                rem = rem_line.loc[idxs_apply]
                base = float(rem.sum() or 0.0)
                if base <= 0:
                    return False
                disc = sequential_pct_discount(base, pct_parts)
                if disc <= 0:
                    return False
                for idx in idxs_apply:
                    cur = sales_df.at[idx, "_PCT_PARTS_INTERNAL"]
                    if not isinstance(cur, list):
                        cur = []
                    cur.extend(pct_parts)
                    sales_df.at[idx, "_PCT_PARTS_INTERNAL"] = cur
                    if mark_bonus:
                        sales_df.at[idx, "_INTERNAL_BONUS_MATCH"] = True
                shares = rem / float(rem.sum() or 1.0)
                for idx in idxs_apply:
                    sales_df.at[idx, "Expected_Internal"] += float(disc * shares.loc[idx])
                return True

            used_any_discount = False

            # Item-specific rules first (if any)
            for brg, gi in g.groupby("BRG", sort=False):
                brg = s(brg)
                rr_item = r0[(r0["KODE_BARANG"].apply(lambda x: norm(x) == norm(brg)))].copy()
                if rr_item.empty:
                    continue
                picked = pick_internal_rule(rr_item, invoice_gross)
                if picked.empty:
                    continue
                pct_parts = []
                for v in picked["DISC_PCT"].tolist():
                    pct_parts.extend(parse_pct_chain(s(v)))
                if not pct_parts:
                    continue

                gi_idxs = gi.index.tolist()
                gi_bonus = [i for i in gi_idxs if bool(bonus_mask.loc[i])]
                gi_non_bonus = [i for i in gi_idxs if not bool(bonus_mask.loc[i])]

                applied_any = False
                if gi_non_bonus:
                    applied_any |= apply_internal_indices(gi_non_bonus, pct_parts, False)
                if pct_parts_full_discount(pct_parts) and gi_bonus:
                    applied_any |= apply_internal_indices(gi_bonus, pct_parts, True)

                if applied_any:
                    used_any_discount = True

            # General rule if no item-specific matched
            if not used_any_discount:
                rr_gen = r0[r0["KODE_BARANG"].apply(lambda x: s(x) == "")].copy()
                picked = pick_internal_rule(rr_gen, invoice_gross)
                if picked.empty:
                    continue
                pct_parts = []
                for v in picked["DISC_PCT"].tolist():
                    pct_parts.extend(parse_pct_chain(s(v)))
                if not pct_parts:
                    continue
                gen_bonus = [i for i in idxs if bool(bonus_mask.loc[i])]
                gen_non_bonus = [i for i in idxs if not bool(bonus_mask.loc[i])]
                if gen_non_bonus:
                    apply_internal_indices(gen_non_bonus, pct_parts, False)
                if pct_parts_full_discount(pct_parts) and gen_bonus:
                    apply_internal_indices(gen_bonus, pct_parts, True)

    
# ---------------------------
    # BONUS satpam (v16+bonus)
    # Bonus definition:
    # - MDSTRING contains 100% (free goods) -> treated as BONUS line
    # - Expected bonus is computed from Dataset Diskon (pabrik) where BENEFIT_TYPE = BONUS_QTY/BONUS
    #   using the same Trigger/Tier rules as discount engine (NON_GROUP per-line; grouped per-invoice).
    # - Expected bonus is allocated to bonus lines (MDSTRING=100%) for the eligible item(s).
    # ---------------------------
    sales_df.loc[:, "Actual_BonusQty"] = sales_df.apply(
        lambda r: float(qty_in_unit(r, "PCS")) if mdstring_is_bonus(r.get("MDSTRING", "")) else 0.0,
        axis=1
    )
    compute_expected_bonus_from_rules(
        sales_df=sales_df,
        rules_df=promo_df,
        channel_col="EffectiveChannel",
        expected_col="Expected_Bonus_Pabrik",
        debug_reason_prefix="BONUS_PABRIK",
        program_name_col="Program_Bonus_Applied"
    )
    # Internal bonus: template internal currently focuses on discount pct; keep 0 unless future internal rules add bonus columns.
    sales_df.loc[:, "Expected_Bonus_Internal"] = 0.0

    # Bonus discount allocation:
    # - Internal only if internal rule is 100% (marked by _INTERNAL_BONUS_MATCH)
    # - Otherwise, if bonus program matched (Expected_Bonus_Pabrik > 0) -> discount goes to Pabrik
    # - Else -> TanpaTuan
    bonus_mask = sales_df["_IS_BONUS_LINE_MAIN"] if "_IS_BONUS_LINE_MAIN" in sales_df.columns else sales_df["MDSTRING"].apply(mdstring_is_bonus)
    if bonus_mask.any():
        sales_df.loc[bonus_mask, "Expected_Pabrik"] = 0.0
        sales_df.loc[bonus_mask, "Expected_Internal"] = 0.0
        unit_disc = pd.Series(0.0, index=sales_df.index)
        bonus_qty = sales_df.loc[bonus_mask, "Actual_BonusQty"].astype(float).fillna(0.0)
        denom = bonus_qty.replace(0, float("nan"))
        unit_disc.loc[bonus_mask] = sales_df.loc[bonus_mask, "LDISCAMOUNT"].astype(float).fillna(0.0).div(denom).fillna(0.0)

        bonus_internal = bonus_mask & (sales_df["_INTERNAL_BONUS_MATCH"] == True)
        if bonus_internal.any():
            sales_df.loc[bonus_internal, "Expected_Internal"] = unit_disc.loc[bonus_internal] * sales_df.loc[bonus_internal, "Actual_BonusQty"].astype(float).fillna(0.0)

        bonus_pabrik = bonus_mask & (~bonus_internal) & (sales_df["Expected_Bonus_Pabrik"].astype(float) > 0)
        if bonus_pabrik.any():
            sales_df.loc[bonus_pabrik, "Expected_Pabrik"] = unit_disc.loc[bonus_pabrik] * sales_df.loc[bonus_pabrik, "Expected_Bonus_Pabrik"].astype(float).fillna(0.0)
            for idx in sales_df.index[bonus_pabrik]:
                sales_df.at[idx, "Program_Pabrik_Applied"] = add_list_item(
                    sales_df.at[idx, "Program_Pabrik_Applied"],
                    sales_df.at[idx, "Program_Bonus_Applied"]
                )
    return finalize(sales_df, sales_cols_fixed)

import openpyxl  # added for master parsing + excel output

# Audit F8: cache manual dipersist ke disk (JSON) agar selamat dari restart dan
# konsisten bila uvicorn multi-worker. Interface tetap dict — 21 call-site tak berubah.
# Flush hanya terjadi saat WRITE (upload master / generate output = aksi jarang),
# read path tetap dict lookup murni (0 tambahan latensi).
class _PersistentDict(dict):
    def __init__(self, path: str):
        super().__init__()
        self._path = path
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    self.update(json.load(f))
        except Exception as e:
            print(f"[MANUAL CACHE] gagal load {path}: {e}")

    def _flush(self) -> None:
        # Persist gagal TIDAK boleh mematikan request — cache in-memory tetap benar.
        try:
            tmp = self._path + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(self, f, ensure_ascii=False, default=str)
            os.replace(tmp, self._path)
        except Exception as e:
            print(f"[MANUAL CACHE] gagal persist {self._path}: {e}")

    def __setitem__(self, key, value):
        super().__setitem__(key, value)
        self._flush()

    def __delitem__(self, key):
        super().__delitem__(key)
        self._flush()

    def pop(self, *args, **kwargs):
        result = super().pop(*args, **kwargs)
        self._flush()
        return result

_MANUAL_CACHE_DIR = os.path.join(BASE_DIR, "data", "manual_cache")
MANUAL_MASTER_CACHE: dict = _PersistentDict(os.path.join(_MANUAL_CACHE_DIR, "master_cache.json"))   # token -> {"kelompok": [...], "variant_map": {...}, "gramasi_map": {...}}
MANUAL_OUTPUTS: dict = _PersistentDict(os.path.join(_MANUAL_CACHE_DIR, "outputs.json"))        # file_id -> {"form": path, "dataset": path}

def _norm_col(x: object) -> str:
    return " ".join(str(x or "").strip().split()).upper()

def _parse_master_barang_xlsx(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {}

    for c, h in enumerate(headers, start=1):
        hn = _norm_col(h)

        # Kode / Nama barang
        if hn == "KODE BARANG":
            idx["kode"] = c
        if hn == "NAMA BARANG":
            idx["nama"] = c
            
        # Principle Name Target (e.g. Camellia, Bellagio)
        if hn == "NAMA BARANG PRINCIPLE" or hn == "PRINCIPLE":
            idx["principle"] = c

        # Kelompok fields
        if ("NAMA" in hn and "KLP" in hn and "SUB" not in hn) or hn == "NAMA KLP":
            idx["klp"] = c
        if ("NAMA" in hn and "SUB" in hn and "KLP 2" not in hn and "KLP2" not in hn) or hn == "NAMA SUB KLP":
            idx["sub1"] = c
        if ("NAMA" in hn and "SUB" in hn and ("KLP 2" in hn or "KLP2" in hn)) or hn == "NAMA SUB KLP 2":
            idx["sub2"] = c

        # Variant = Aroma/Rasa
        if ("AROMA" in hn) or ("RASA" in hn):
            idx["variant"] = c

        # Gramasi
        if ("GRAMASI" in hn) or (("PACK" in hn or "PCS" in hn) and ("CTN" in hn or "KARTON" in hn or "CARTON" in hn)):
            idx["gramasi"] = c

    missing = [k for k in ["kode", "nama", "klp", "sub1", "sub2", "variant", "gramasi"] if k not in idx]
    if missing:
        raise ValueError("Kolom master barang tidak ketemu: " + ", ".join(missing))

    variant_map = {}
    gramasi_map = {}
    items = []

    for r in range(2, ws.max_row + 1):
        kode_barang = str(ws.cell(r, idx["kode"]).value or "").strip()
        nama_barang = str(ws.cell(r, idx["nama"]).value or "").strip()
        principle = str(ws.cell(r, idx["principle"]).value or "").strip() if "principle" in idx else ""
        
        klp = str(ws.cell(r, idx["klp"]).value or "").strip()
        sub1 = str(ws.cell(r, idx["sub1"]).value or "").strip()
        sub2 = str(ws.cell(r, idx["sub2"]).value or "").strip()
        kelompok = " - ".join([x for x in [klp, sub1, sub2] if x]).strip()
        if not kelompok:
            continue

        v = str(ws.cell(r, idx["variant"]).value or "").strip()
        g = str(ws.cell(r, idx["gramasi"]).value or "").strip()

        if v:
            variant_map.setdefault(kelompok, set()).add(v)
        if g:
            gramasi_map.setdefault(kelompok, set()).add(g)

        if kode_barang or nama_barang:
            items.append({
                "kode_barang": kode_barang,
                "nama_barang": nama_barang,
                "principle": principle,
                "kelompok": kelompok,
                "variant": v,
                "gramasi": g,
            })

    kelompok_list = sorted(set(list(variant_map.keys()) + list(gramasi_map.keys())))
    variant_map = {k: sorted(list(v)) for k, v in variant_map.items()}
    gramasi_map = {k: sorted(list(v)) for k, v in gramasi_map.items()}
    return kelompok_list, variant_map, gramasi_map, items

def _parse_master_customer_xlsx(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {}
    for c, h in enumerate(headers, start=1):
        hn = str(h or "").strip().upper()
        if hn == "KODE CUSTOMER": idx["kode"] = c
        elif hn == "NAMA CUSTOMER": idx["nama"] = c
            
    if "kode" not in idx or "nama" not in idx:
        raise ValueError("Kolom Master Customer tidak valid. Pastikan ada KODE CUSTOMER dan NAMA CUSTOMER.")
        
    customers = []
    for r in range(2, ws.max_row + 1):
        kode = str(ws.cell(r, idx["kode"]).value or "").strip()
        nama = str(ws.cell(r, idx["nama"]).value or "").strip()
        if kode and nama:
            customers.append({"kode_customer": kode, "nama_customer": nama})
            
    return customers

def _ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)


def _apply_native_kelompok(rows_to_check, master_items):
    # URC GUARD (deterministic matcher branch): the structure-only prompt
    # (routers/summary.py, URC) emits rows carrying `item_description` -- a
    # marker field NEVER used by Priskila's rows (`group_item_text`) or the
    # legacy path, so URC and Priskila rows can never cross-route into each
    # other's matcher+master by accident. Checked BEFORE the Priskila guard.
    try:
        if any(r.get("item_description") for r in rows_to_check):
            from urc_pipeline import apply_urc_matching
            return apply_urc_matching(rows_to_check, master_items)
    except Exception as _urc_err:
        try:
            append_error_log("urc_pipeline_fallback", _urc_err,
                             {"n_rows": len(rows_to_check)})
        except Exception:
            pass

    # PRISKILA GUARD (deterministic matcher branch): the structure-only prompt
    # (routers/summary.py, Priskila) emits rows carrying `group_item_text` -- that
    # field is the unambiguous marker of the new path (the legacy prompt never
    # emits it), so we route on its PRESENCE rather than on a principle string
    # (the master's principle column is frequently blank). Those rows go to the
    # offline-tested priskila_pipeline (surat line -> master SKUs, deterministic,
    # renderer-ready). NON-Priskila rows fall straight through unchanged below.
    try:
        if any(r.get("group_item_text") for r in rows_to_check):
            from priskila_pipeline import apply_priskila_matching
            return apply_priskila_matching(rows_to_check, master_items)
    except Exception as _prisk_err:
        # Never let the Priskila branch break the shared endpoint; on any failure
        # fall back to legacy matching -- but LOG it (a silent revert to the old
        # buggy path would otherwise look like success at the PDF gate).
        try:
            append_error_log("priskila_pipeline_fallback", _prisk_err,
                             {"n_rows": len(rows_to_check)})
        except Exception:
            pass

    final_rows_out = []
    def norm(x: object) -> str:
        return " ".join(str(x or "").strip().split()).upper()

    def split_list(val: str) -> list:
        import re
        return [x.strip() for x in re.split(r'[,&]', str(val or "")) if x.strip()]

    # Kode master yg kelompoknya di exclude-list variant_mapping (mis. Spray Cologne GLASS).
    # Distrip dari kode_barangs di AWAL tiap baris -> tak lolos jalur manapun, termasuk fallback
    # "row tak ter-explode" (yg dulu mengembalikan r mentah ber-GLASS -> 6 GLASS bocor run live).
    _excluded_codes = {str(it.get("kode_barang", "")).strip() for it in master_items
                       if norm(it.get("kelompok")) in _EXCLUDED_KELOMPOKS} if _EXCLUDED_KELOMPOKS else set()

    for r in rows_to_check:
        if _excluded_codes:
            _kb = [k.strip() for k in str(r.get("kode_barangs", "")).split(",") if k.strip()]
            _kb2 = [k for k in _kb if k not in _excluded_codes]
            if len(_kb2) != len(_kb):
                r = dict(r)
                r["kode_barangs"] = ",".join(_kb2)
        matched_items = []
        # FASE 3b: resolusi varian data-driven (variant_mapping.json) DULU, sebelum LLM/
        # kode_barangs manapun -- kasus terbukti: "Spray Cologne Series" harus jadi White SR
        # + Black SR (bukan cuma GLASS/salah satu), "EDT Sport" harus 4 varian tertentu
        # (Azzuro/Bianco/Nero/Rosso), bukan tebakan LLM yg terbukti salah/tak konsisten.
        _variant_hit = resolve_variant(str(r.get("kelompok", "")), master_items, _VARIANT_MAPPING)
        if _variant_hit is not None:
            matched_items = _variant_hit
            klist = []  # skip jalur matching lama sepenuhnya utk baris ini
        else:
            klist = [k.strip() for k in str(r.get("kode_barangs", "")).split(",") if k.strip()]
        if klist:
            for it in master_items:
                if str(it.get("kode_barang", "")).strip() in klist:
                    matched_items.append(it)

        # "All Variant": LLM sering under-enumerate kode di channel belakangan (terbukti
        # debug_ai.txt 2026-07-13: STAR OUTLET "Marie Jose All Variant" cuma kirim 1 dari 8
        # kode -> 7 varian hilang diam-diam). Perbaiki dgn ANCHOR ke item yg sudah cocok (seed):
        # tarik SEMUA varian se-(kelompok, gramasi) dari master, non-banded. Dibatasi ke
        # kelompok+gramasi seed -> TIDAK meledak (beda dgn fallback string-match di bawah yg bisa
        # jatuh ke SELURUH master saat kelompok LLM tak persis cocok). Aturan surat: banded (BND)
        # dikecualikan dari klaim promo.
        # _variant_hit is None: JANGAN ekspansi hasil FASE 3b (variant_resolver sengaja membatasi
        # varian, mis. Regazza EDT Sport = HANYA 4 -- ekspansi se-kelompok akan merusak batasan itu).
        if matched_items and _variant_hit is None:
            _vlist_probe = split_list(r.get("variant", ""))
            _v_all_probe = (not _vlist_probe) or any(norm(x) == "ALL VARIANT" for x in _vlist_probe)
            if _v_all_probe:
                _seed_keys = {(norm(it.get("kelompok")), norm(it.get("gramasi"))) for it in matched_items}
                _seen_codes = {str(it.get("kode_barang", "")).strip() for it in matched_items}
                # #3 (2026-07-15): kelompok yg BERPASANGAN per variant_mapping -- mis. surat
                # "Casablanca Spray Cologne Series" = White SR + Black SR SEKALIGUS (dikonfirmasi
                # user; SR=Series, lihat variant_mapping CASABLANCA_SPRAY_COLOGNE_SERIES). Kalau
                # seed masuk salah satu kelompok di resolve_to_kelompok sebuah rule, tarik JUGA
                # kelompok pasangannya (di gramasi seed) MINUS exclude_kelompok (GLASS di-EXCLUDE).
                # Additive & tak menyentuh resolve_variant -> baris mega-merge multi-kelompok tak
                # kehilangan kelompok lain (beda dgn kalau resolve_variant yg fire).
                _seed_kels = {norm(it.get("kelompok")) for it in matched_items}
                _seed_grams = {norm(it.get("gramasi")) for it in matched_items}
                for _rule in _VARIANT_MAPPING.values():
                    _rk = {norm(k) for k in _rule.get("resolve_to_kelompok", [])}
                    if not _rk or not (_seed_kels & _rk):
                        continue
                    _excl = {norm(k) for k in _rule.get("exclude_kelompok", [])}
                    for _k in _rk - _excl:
                        for _g in _seed_grams:
                            _seed_keys.add((_k, _g))
                for it in master_items:
                    _kode = str(it.get("kode_barang", "")).strip()
                    if _kode in _seen_codes:
                        continue
                    if "BND" in norm(it.get("nama_barang")).split():
                        continue
                    if (norm(it.get("kelompok")), norm(it.get("gramasi"))) in _seed_keys:
                        matched_items.append(it)
                        _seen_codes.add(_kode)
                    
        # Fallback Match: String matching if AI failed to provide API keys or "ALL VARIANT"
        if not matched_items:            
            kelompok = str(r.get("kelompok","") or "").strip()
            vlist = split_list(r.get("variant",""))
            glist = split_list(r.get("gramasi",""))
            v_all = (not vlist) or any(norm(x) == "ALL VARIANT" for x in vlist)
            g_all = (not glist) or any(norm(x) == "ALL GRAMASI" for x in glist)

            # Treat generic placeholders as empty
            if kelompok and any(skip in kelompok.lower() for skip in ["- kelompok -", "bisa meleset"]):
                kelompok = ""
                
            # Filter pool initially based on Kelompok if provided
            pool = []
            if kelompok:
                pool = [it for it in master_items if norm(it.get("kelompok")) == norm(kelompok)]
            else:
                pool = master_items

            # If pool is empty (hallucinated kelompok), fallback to all items
            if not pool: pool = master_items

            # Now filter the pool based on variants and gramasi
            for it in pool:
                it_variant = norm(it.get("variant"))
                it_nama = norm(it.get("nama_barang"))
                variant_match = False
                if v_all:
                    variant_match = True
                else:
                    for v in [norm(x) for x in vlist]:
                        if "- variant -" in v.lower() or "all variant" in v.lower() or "bisa meleset" in v.lower():
                            variant_match = True; break
                            
                        # Fuzzy Token Matching: Resolve LLM Hallucinations like 'BELLAGIO EAU DE TOILETTE'
                        if v == it_nama or v == it_variant or (len(v) > 5 and v in it_nama):
                            variant_match = True; break
                            
                        # Token overlap with Brand/Variant Aliasing
                        ALIASES = {"BELLAGIO": "BLAGIO", "CASABLANCA": "CSBNCA", "TOILETTE": "EDT", "PARFUME": "EDP", "PARFUM": "EDP"}
                        v_tokens = []
                        for t in v.split():
                            if len(t) >= 4:
                                t_mapped = ALIASES.get(t, t)
                                v_tokens.append(t_mapped)
                        v_tokens = set(v_tokens)
                        
                        db_tokens = set([t for t in it_nama.split() if len(t) >= 4])
                        if v_tokens and len(v_tokens.intersection(db_tokens)) >= 1:
                            variant_match = True; break
                            
                if not variant_match: continue
                    
                it_gramasi = norm(it.get("gramasi"))
                gramasi_match = False
                if g_all:
                    gramasi_match = True
                else:
                    for g in [norm(x) for x in glist]:
                        if "- gramasi -" in g.lower() or "all gramasi" in g.lower() or "bisa meleset" in g.lower():
                            gramasi_match = True; break
                        if g == it_gramasi or (len(g) > 2 and g in it_nama):
                            gramasi_match = True; break
                            
                if not gramasi_match: continue

                # Produk BANDED (nama ".. BTL BND") dikecualikan dari ekspansi promo -- aturan
                # surat: "Untuk Klaim tidak berlaku Produk Banded". Master menaruh item banded
                # di kelompok+gramasi+variant yg SAMA dgn non-banded (mis. Marie Jose ...011),
                # jadi tanpa filter ini ekspansi All Variant akan keliru menambah SKU banded.
                if "BND" in it_nama.split(): continue

                matched_items.append(it)
                    
        # EXCLUDE global: buang item ber-kelompok di exclude_kelompok variant_mapping (mis.
        # Spray Cologne GLASS/GLAS) dari SEMUA jalur -- termasuk saat LLM mencantumkan kode GLASS
        # langsung di kode_barangs (terbukti run live 2026-07-15: 18 baris GLASS bocor). Aturan
        # surat: Spray Cologne Series = White SR + Black SR saja.
        if _EXCLUDED_KELOMPOKS:
            matched_items = [it for it in matched_items if norm(it.get("kelompok")) not in _EXCLUDED_KELOMPOKS]

        # Deterministic Kelompok String Builder strictly from Master DB
        # EXPLODE MAGIC: If a single AI row contains items from DIFFERENT Brand Prefixes (Nama KLPs),
        # we must split it into separate rows so the frontend and generator handle them cleanly!
        from collections import defaultdict
        groups_by_prefix = defaultdict(list)
        
        for it in matched_items:
            real_kelompok = str(it.get("kelompok", "")).strip()
            if real_kelompok and real_kelompok.lower() not in ("nan", "null", "none"):
                # Extract prefix to group strictly by Nama KLP
                prefix = real_kelompok.split(" - ")[0] if " - " in real_kelompok else real_kelompok
                groups_by_prefix[prefix].append(it)
                
        if not groups_by_prefix:
            # Fallback for completely unmatched row
            final_rows_out.append(r)
        else:
            # Explode into multiple rows based on the prefix
            import copy
            is_first = True
            for prefix, its in groups_by_prefix.items():
                new_row = r if is_first else copy.deepcopy(r)
                is_first = False
                
                # Assign this prefix's valid kelompoks & matched items
                k_list = []
                k_codes = []
                for it in its:
                    k_name = str(it.get("kelompok", "")).strip()
                    k_str = str(it.get("kode_barang", "")).strip()
                    if k_name and k_name not in k_list: k_list.append(k_name)
                    if k_str and k_str not in k_codes: k_codes.append(k_str)
                    
                new_row["kelompok"] = " & ".join(k_list)
                new_row["kode_barangs"] = ",".join(k_codes)
                new_row["_matched_items_cache"] = its
                
                # Let's ensure no ID collisions if exploded
                if not is_first:
                     import uuid
                     new_row["id"] = str(uuid.uuid4())
                     
                final_rows_out.append(new_row)
                
    return final_rows_out

# ======================================================================================
# "AI LEARNING" DARI KOREKSI MANUAL (tombol Laporkan Salah di grid)
# Bukan fine-tuning model -- few-shot memory: tiap koreksi user (before->after) disimpan,
# lalu di-inject balik ke prompt parse_pdf_ai supaya kesalahan yg sama tidak berulang.
# ======================================================================================
CORRECTIONS_PATH = os.path.join(BASE_DIR, "data", "parse_corrections.jsonl")
_CORRECTION_IGNORE_KEYS = {"id", "_matched_items_cache", "no"}

def _load_corrections(principle_name: str, limit: int = 15) -> List[Dict[str, Any]]:
    if not os.path.exists(CORRECTIONS_PATH):
        return []
    items = []
    try:
        with open(CORRECTIONS_PATH, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    items.append(json.loads(line))
                except Exception:
                    continue
    except Exception:
        return []
    pn = s(principle_name).upper()
    matched = [it for it in items if pn and pn in s(it.get("principle_name", "")).upper()]
    pool = matched if matched else items
    return pool[-limit:]

def _format_corrections_for_prompt(corrections: List[Dict[str, Any]]) -> str:
    blocks = []
    for c in corrections:
        before = c.get("before") or {}
        after = c.get("after") or {}
        diffs = []
        for k in set(before.keys()) | set(after.keys()):
            if k in _CORRECTION_IGNORE_KEYS:
                continue
            bv, av = before.get(k, ""), after.get(k, "")
            if s(bv) != s(av):
                diffs.append(f'    {k}: SALAH="{bv}" -> BENAR="{av}"')
        if diffs:
            note = s(c.get("note", ""))
            header = f"KOREKSI ({note}):" if note else "KOREKSI:"
            blocks.append(header + "\n" + "\n".join(diffs))
    if not blocks:
        return ""
    return ("\n\n=== KOREKSI DARI USER SEBELUMNYA (WAJIB DIPATUHI, JANGAN ULANGI KESALAHAN INI) ===\n"
            + "\n".join(blocks) + "\n=== AKHIR KOREKSI ===")

def process_summary_generation_job(job_id: str, token: str, rows: List[Dict[str, Any]], user: str):
    try:
        BACKGROUND_JOBS[job_id]["status"] = "processing"
        cache = MANUAL_MASTER_CACHE[token]
        items = cache.get("items", [])

        def norm(x: object) -> str:
            return " ".join(str(x or "").strip().split()).upper()

        def smart_normalize(text: str) -> str:
            t = norm(text)
            aliases = {
                "SHM": "SHAMPOO",
                "F.WASH": "FACIAL WASH",
                "COND.": "CONDITIONER",
                "COND ": "CONDITIONER ",
                "ALOVERA": "ALOE",
                "ALOEVERA": "ALOE",
                "ALOE VERA": "ALOE",
                "ALOE VERE": "ALOE",
                "HABBATS": "HABBATUSSAUDA",
                "ANTI DANDRUFF": "TEA TREE",
                "ANTIDANDRUFF": "TEA TREE",
                "VIT ": "VITAMIN ",
                "165ML": "160ML",
                "BRIGHT&DC": "BRIGHTENING",
                "CR&OC": "ACNE",
                "EXTRACT": "",
                "&": " ",
                "OIL": ""
            }
            for k, v in aliases.items():
                t = t.replace(k, v)
            t = re.sub(r'[^A-Z0-9\s]', ' ', t)
            return " ".join(t.split())

        def split_list(val: str) -> List[str]:
            return [x.strip() for x in str(val or "").split(",") if x.strip()]

        def unit_from_text(text: str) -> str:
            t = norm(text)
            if "CTN" in t or "KRT" in t:
                return "CTN"
            return "PCS"

        def has_number(text: str) -> bool:
            return bool(re.search(r"\d", str(text or "")))

        out_dir = os.path.join(BASE_DIR, "output", "summary_manual")
        _ensure_dir(out_dir)
        file_id = str(uuid.uuid4())

        form_path = os.path.join(out_dir, f"{file_id}_Form_Summary_Program.pdf")
        
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.units import inch
        from datetime import datetime
        
        # Get current date formatted for Indonesia
        now = datetime.now()
        months = ["JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
        print_date = f"{now.month}/{now.day}/{now.year}"
        dibuat_date = f"{now.day:02d} {months[now.month - 1]} {now.year}"

        # TAHAP 2: Native Master DB Mapping (Injects Kelompok perfectly)
        _apply_native_kelompok(rows, items)
            
        def my_canvas(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setFont('Helvetica-Bold', 7)
            # Top Left
            canvas_obj.drawString(doc_obj.leftMargin, landscape(A4)[1] - 0.5*cm, f"Di Print Tgl : {print_date}")
            # Top Right (Page Number)
            page_str = f"Hal {doc_obj.page}"
            canvas_obj.drawRightString(landscape(A4)[0] - doc_obj.rightMargin, landscape(A4)[1] - 0.5*cm, page_str)
            
            # Bottom of Header (Before Table Starts)
            canvas_obj.setFont('Helvetica', 8)
            canvas_obj.drawString(doc_obj.leftMargin, landscape(A4)[1] - 2*cm, f"Dibuat Tanggal : {dibuat_date}")
            canvas_obj.setFont('Helvetica-Bold', 8)
            canvas_obj.drawRightString(landscape(A4)[0] - doc_obj.rightMargin, landscape(A4)[1] - 2*cm, "(ON PRINCIPLE COKLAT)")
            
            canvas_obj.restoreState()

        # FASE 6: PDF reproducible (CreationDate/ModDate/doc-id tetap) -> byte-identik antar-run
        enable_pdf_determinism()
        # Update margins to give space for the custom canvas headers
        doc = SimpleDocTemplate(form_path, pagesize=landscape(A4), rightMargin=0.5*cm, leftMargin=0.5*cm, topMargin=2.2*cm, bottomMargin=0.5*cm)
        elements = []
        styles = getSampleStyleSheet()

        # Add Title and Subtitle properly centered
        title_style = styles["Heading3"].clone("TitleStyle")
        title_style.alignment = TA_CENTER
        title_style.fontSize = 12
        elements.append(Paragraph("<b>SUMMARY PROGRAM ON FAKTUR BEBAN PRINCIPLE (COKLAT)</b>", title_style))
        
        sub_style = styles["Normal"].clone("SubStyle")
        sub_style.alignment = TA_CENTER
        sub_style.fontSize = 11
        sub_style.fontName = 'Helvetica'
        
        # Determine Period globally from the first row if available
        global_period = rows[0].get("periode", "") if rows else ""
        period_text = f" PERIODE {global_period.upper()}" if global_period else ""
        elements.append(Paragraph(f"CV. SURYA PERKASA {period_text}", sub_style))
        elements.append(Spacer(1, 15))

        headers_str = ["No", "Surat Program", "Nama Program", "Channel", "Periode", 
                   "Kelompok Barang", "Variant", "Gramasi", "Ketentuan", "Benefit", "Syarat Claim", "Keterangan"]
        
        header_style = styles["Normal"].clone("HeaderStyle")
        header_style.fontSize = 7
        header_style.leading = 8
        header_style.fontName = 'Helvetica-Bold'
        header_style.alignment = TA_CENTER
        header_style.textColor = colors.whitesmoke
        
        table_data = [[Paragraph(h, header_style) for h in headers_str]]

        cell_style = styles["Normal"].clone("CellStyle")
        cell_style.fontSize = 6
        cell_style.leading = 7
        cell_style.alignment = TA_CENTER

        for r in rows:
            benefit_pdf_text = str(r.get("benefit",""))
            if str(r.get("benefit_type","")).upper() == "DISC_RP":
                if benefit_pdf_text and not benefit_pdf_text.lower().startswith("cut price"):
                    benefit_pdf_text = f"cut price {benefit_pdf_text}"
                    
            table_data.append([
                Paragraph(str(r.get("no","")), cell_style),
                Paragraph(str(r.get("surat_program","")), cell_style),
                Paragraph(str(r.get("nama_program","")), cell_style),
                Paragraph(str(r.get("channel_gtmt","")), cell_style),
                Paragraph(str(r.get("periode","")), cell_style),
                Paragraph(str(r.get("kelompok","")), cell_style),
                Paragraph(str(r.get("variant","")), cell_style),
                Paragraph(str(r.get("gramasi","")), cell_style),
                Paragraph(str(r.get("ketentuan","")), cell_style),
                Paragraph(benefit_pdf_text, cell_style),
                Paragraph(str(r.get("syarat_claim","")), cell_style),
                Paragraph(str(r.get("keterangan","")), cell_style),
            ])
            
        # Total A4 landscape width is ~842. Margins are 0.5cm each (approx 14 points each, total 28 pts margin)
        # Usable width = 842 - 28 = 814 points
        usable = landscape(A4)[0] - (1 * cm)
        cw = [
            usable * 0.03, # No
            usable * 0.12, # Surat Program
            usable * 0.12, # Nama Program
            usable * 0.05, # Channel
            usable * 0.08, # Periode
            usable * 0.09, # Kelompok
            usable * 0.14, # Variant
            usable * 0.08, # Gramasi
            usable * 0.08, # Ketentuan
            usable * 0.08, # Benefit
            usable * 0.07, # Syarat Claim
            usable * 0.06  # Keterangan
        ]
            
        t = Table(table_data, repeatRows=1, colWidths=cw)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9E7C85')), # Match the brownish pink header color
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        elements.append(t)
        
        # Add the Footer Signatures
        elements.append(Spacer(1, 25))
        
        footer_style_left = styles["Normal"].clone("FooterLeft")
        footer_style_left.fontSize = 8
        footer_style_left.fontName = 'Helvetica-Bold'
        
        footer_style_right = styles["Normal"].clone("FooterRight")
        footer_style_right.fontSize = 8
        footer_style_right.fontName = 'Helvetica-Bold'
        footer_style_right.alignment = TA_RIGHT

        sig_data = [
            [Paragraph(f"Makassar , {dibuat_date}", footer_style_left), ""],
            [Paragraph("Diajukan Oleh,", footer_style_left), Paragraph("Disetujui Oleh,", footer_style_right)],
            [Spacer(1, 40), Spacer(1, 40)], # Space for signature
            [Paragraph("SM<br/>(.................................................)", footer_style_left), 
             Paragraph("OPERATIONAL MANAGER<br/>(.................................................)", footer_style_right)]
        ]
        
        # Table takes up full usable width so left is left, right is right
        sig_table = Table(sig_data, colWidths=[usable/2.0, usable/2.0])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(sig_table)
        
        doc.build(elements, onFirstPage=my_canvas, onLaterPages=my_canvas)

        dataset_path = os.path.join(out_dir, f"{file_id}_Dataset_Diskon_With_Channel.xlsx")
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Dataset"
        headers2 = [
            "KODE_BARANG", "NAMA_BARANG", "PROMO_LABEL", "PROMO_GROUP_ID", "PROMO_GROUP",
            "PERIODE", "PROMO_ACTIVE", "TIER_NO", "TRIGGER_QTY", "TRIGGER_UNIT",
            "BENEFIT_TYPE", "BENEFIT_VALUE", "BENEFIT_UNIT", "BENEFIT_BEBAN"
        ]
        ws2.append(headers2)

        tier_counter: Dict[Tuple[str, str, str, str], int] = {}
        channel_group_counters: Dict[str, int] = {}
        kelompok_group_map: Dict[Tuple[str, str, str], str] = {}

        for r in rows:
            kelompok = str(r.get("kelompok","") or "").strip()
            vlist = split_list(r.get("variant",""))
            glist = split_list(r.get("gramasi",""))
            v_all = (not vlist) or any(norm(x) == "ALL VARIANT" for x in vlist)
            g_all = (not glist) or any(norm(x) == "ALL GRAMASI" for x in glist)

            ket = str(r.get("ketentuan","") or "").strip()
            trig_has_num = has_number(ket)
            trig_qty = parse_number_id(ket) if trig_has_num else ""
            trig_unit = unit_from_text(ket) if trig_has_num else ""
            benefit_text_raw = str(r.get("benefit","") or "").strip()
            benefit_type = str(r.get("benefit_type","") or "").strip()
            
            import re
            # Extract only the numbers, avoiding capturing standalone periods like in "Pot."
            nums = re.findall(r'\d[\d\.,]*', benefit_text_raw)
            benefit_text = nums[0].replace(',', '.') if nums else benefit_text_raw
                
            benefit_unit = unit_from_text(benefit_text_raw) if benefit_text_raw else ""
            benefit_beban = "PABRIK"

            promo_label = str(r.get("nama_program","") or "").strip()
            promo_group_id = str(r.get("promo_group_id","") or "").strip()
            promo_group = str(r.get("channel_gtmt","") or "").strip()
            periode = str(r.get("periode","") or "").strip()
            
            if not promo_group_id or promo_group_id.upper() == "NON_GROUP":
                promo_group_id = promo_group

            matched_items = r.get("_matched_items_cache", [])
            
            if not matched_items:
                key_unmatched = (promo_label, promo_group_id, promo_group, "UNMATCHED")
                tier_counter[key_unmatched] = tier_counter.get(key_unmatched, 0) + 1
                tier_no = tier_counter[key_unmatched]
                
                ws2.append([
                    "",
                    "",
                    promo_label,
                    promo_group_id,
                    promo_group,
                    periode,
                    True,
                    tier_no,
                    trig_qty,
                    trig_unit,
                    benefit_type,
                    benefit_text,
                    benefit_unit,
                    benefit_beban,
                ])
                continue

            for it in matched_items:
                kb = str(it.get("kode_barang", ""))
                real_kelompok = str(it.get("kelompok", "")).strip()
                
                # Determine native promo group id progressively based on original valid kelompok
                base_pg_id = promo_group_id
                map_key = (promo_label, base_pg_id, real_kelompok)
                
                if map_key not in kelompok_group_map:
                    channel_group_counters[base_pg_id] = channel_group_counters.get(base_pg_id, 0) + 1
                    kelompok_group_map[map_key] = f"{base_pg_id}_{channel_group_counters[base_pg_id]}"
                
                final_promo_group_id = kelompok_group_map[map_key]
                
                key_item = (promo_label, final_promo_group_id, promo_group, kb)
                tier_counter[key_item] = tier_counter.get(key_item, 0) + 1
                tier_no = tier_counter[key_item]
                
                ws2.append([
                    kb,
                    str(it.get("nama_barang", "")),
                    promo_label,
                    final_promo_group_id,
                    promo_group,
                    periode,
                    True,
                    tier_no,
                    trig_qty,
                    trig_unit,
                    benefit_type,
                    benefit_text,
                    benefit_unit,
                    benefit_beban,
                ])
        wb2.save(dataset_path)

        MANUAL_OUTPUTS[file_id] = {"form": form_path, "dataset": dataset_path}
        
        BACKGROUND_JOBS[job_id]["status"] = "done"
        BACKGROUND_JOBS[job_id]["result"] = {
            "file_id": file_id
        }
        
    except Exception as e:
        import traceback
        with open(os.path.join(BASE_DIR, "debug_traceback.txt"), "w") as f:
            traceback.print_exc(file=f)
        BACKGROUND_JOBS[job_id]["status"] = "error"
        BACKGROUND_JOBS[job_id]["error"] = str(e)
        if hasattr(e, '__traceback__'):
            append_error_log("background_summary_manual", e, {"user": user, "job_id": job_id})

PRINCIPLES_JSON_PATH = os.path.join(BASE_DIR, "data", "principles.json")
MASTERS_DIR = os.path.join(BASE_DIR, "data", "masters")

def _load_principles() -> dict:
    try:
        import sqlite3
        conn = sqlite3.connect("database.sqlite")
        c = conn.cursor()
        c.execute("SELECT id, name, filename, uploaded_by, created_at FROM principles")
        res = {}
        for row in c.fetchall():
            res[row[0]] = {
                "name": row[1],
                "filename": row[2],
                "uploaded_by": row[3],
                "created_at": row[4]
            }
        conn.close()
        return res
    except:
        return {}

def _save_principles(data: dict):
    try:
        import sqlite3
        conn = sqlite3.connect("database.sqlite")
        c = conn.cursor()
        c.execute("DELETE FROM principles")
        for uid, info in data.items():
            name = info.get("name", "")
            filename = info.get("filename", "")
            upb = info.get("uploaded_by", "")
            ca = info.get("created_at", "")
            c.execute("INSERT INTO principles (id, name, filename, uploaded_by, created_at) VALUES (?, ?, ?, ?, ?)", (uid, name, filename, upb, ca))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error saving principles to SQLite: {e}")

# Tujuan: FastAPI backend untuk validator diskon, summary, payments, finance, dan principle.
# Caller: Next.js dashboard, halaman HTML legacy FastAPI, dan workflow internal operasional.
# Dependensi: pandas/openpyxl, auth helpers, validator_engine, payments, SumoPod/OpenAI-compatible API.
# Main Functions: `app`, endpoint validator/summary/payments/finance/principles.
# Side Effects: DB/file read-write runtime, HTTP call AI/SMTP, generate Excel/download artifacts.
#
# Tujuan: Backend FastAPI utama untuk validator, summary, payments/SPPD, finance, dan principle.
# Caller: Next.js dashboard, halaman legacy FastAPI, dan API browser internal AccAPI.
# Dependensi: pandas/openpyxl, payments.py, validator_engine.py, payments.json, file output, Accurate session dari UI.
# Main Functions: app routes, load/save payments DB, upload/restore payments, render SPPD, finance proof/mapping/update APIs.
# Side Effects: HTTP response, JSON/file I/O, DOCX/XLSX generation, audit/error log writes.


# =======================================================================================================
# Laporan Harian per SPV/SM — proses FIX LAP PENJ (Paste Acc only, retur sudah minus) -> agregat & split.
# Dipanggil Next.js app/api/laporan-harian/upload. Tidak kirim email (email = gate terpisah di Next.js).
# =======================================================================================================
LH_RUNTIME_DIR = os.getenv(
    "LH_RUNTIME_DIR",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "runtime", "laporan-harian"),
)
