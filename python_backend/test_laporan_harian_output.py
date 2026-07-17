# Tujuan: Menjaga output SPV/SM dan parity Power Query Principal tetap benar.
# Caller: Developer/CI melalui eksekusi Python langsung.
# Dependensi: pandas, openpyxl, laporan_harian.resolve_report_groups, dan write_report_files.
# Main Functions: main() memeriksa nama customer, tanggal, Principal, dan kontrak tujuh kolom Stock.
# Side Effects: Membuat lalu menghapus workbook sementara.

import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

import laporan_harian as laporan


def main() -> None:
    assert laporan._normal_text(pd.NA) == ""
    assert laporan._normal_text(None) == ""

    raw_accurate = pd.DataFrame([{
        "NO_NOTA": "INV-TEST", "TANGGAL": "2026-07-16", "MATA_UANG": "IDR",
        "KODE PELANGGAN INDUK": "C-001", "CUSTOMER": "C-001",
        "Nama Pelanggan Faktur Penjualan": "TOKO MAJU",
        "PRINCIPLE": "TEST PRINCIPAL", "KODE_SALESMAN": "S-01", "KODE_BARANG": "SKU-01",
        "NILAI JUAL": 1000, "QTY": 1,
    }])
    prepared = laporan._prep_acc(raw_accurate, laporan.LookupTables({}, {}, {}, {}, {}))
    assert prepared.iloc[0]["CUSTOMER"] == "TOKO MAJU"

    dated = pd.DataFrame([
        {"TANGGAL": "2026-07-15", "JENIS_TRANSAKSI": laporan.PENJ_LABEL},
        {"TANGGAL": "2026-07-16", "JENIS_TRANSAKSI": laporan.PENJ_LABEL},
        {"TANGGAL": "2026-07-17", "JENIS_TRANSAKSI": laporan.RETUR_LABEL},
    ])
    assert laporan.latest_sales_date(dated) == "2026-07-16"

    source = pd.DataFrame([
        {"GOLONGAN": "DENNY", "NAMA_SM": "HENDRIK", "PRINCIPAL": "ENERGIZER INDONESIA, PT",
         "KODE_BARANG": "SKU-1", "SALESMAN": "EN1_TEST", "AO": 7, "EC": 9, "IA": 11, "Item Aktif": 11},
        {"GOLONGAN": "YUDI", "NAMA_SM": "HENDRIK", "PRINCIPAL": "MOTASA INDONESIA, PT",
         "KODE_BARANG": "SKU-2", "SALESMAN": "MS1_TEST", "AO": 1, "EC": 1, "IA": 1, "Item Aktif": 1},
        {"GOLONGAN": "YUDI", "NAMA_SM": "HENDRIK", "PRINCIPAL": "MOTASA INDONESIA, PT",
         "KODE_BARANG": "SKU-3", "SALESMAN": "MS6_TEST", "AO": 1, "EC": 1, "IA": 1, "Item Aktif": 1},
        {"GOLONGAN": "SOFYAN", "NAMA_SM": "ADNAN", "PRINCIPAL": "FONTERRA BRANDS INDONESIA, PT",
         "KODE_BARANG": "SKU-F1", "KODE_CUST": "C-OK001", "SALESMAN": "FT1_TEST"},
        {"GOLONGAN": "SOFYAN", "NAMA_SM": "ADNAN", "PRINCIPAL": "FONTERRA BRANDS INDONESIA, PT",
         "KODE_BARANG": "SKU-F2", "KODE_CUST": "C-TUN020", "SALESMAN": pd.NA},
        {"GOLONGAN": "DENNY", "NAMA_SM": "HENDRIK", "PRINCIPAL": "RECKITT BENCKISER, PT",
         "KODE_BARANG": "20010010100010", "SALESMAN": "RB1_TEST"},
        {"GOLONGAN": "DENNY", "NAMA_SM": "HENDRIK", "PRINCIPAL": "MUSTIKA RATUBUANA INTERNATIONAL",
         "KODE_BARANG": "39110060200000", "NAMA_BARANG": 'ITEM "1006067" TEST',
         "KODE_CUST": "C-GRA001", "CUSTOMER": "OUTLET TEST", "KODE_SALESMAN": "MR1",
         "SALESMAN": "MR1_TEST", "TANGGAL": "2026-07-15", "NO_NOTA": "INV-1",
         "JENIS_TRANSAKSI": "1. Penjualan Bruto", "QTY_SATUANKECIL": 2,
         "NILAI_JUAL": 100000, "JUMLAH": 111000, "POTONGAN": 1000,
         "REGION": "OTHER_SELATAN KOTA", "ALAMAT": "JL TEST", "KOTA": "MAKASSAR"},
    ])
    stock = pd.DataFrame([
        {"KODE_BARANG": "SKU-1", "QTY AKHIR": 25, "PRINCIPAL": "ENERGIZER INDONESIA, PT",
         "GOLONGAN": "DENNY", "NAMA_SM": "HENDRIK"},
        {"KODE_BARANG": "SKU-2", "QTY AKHIR": 10, "PRINCIPAL": "MOTASA INDONESIA, PT",
         "GOLONGAN": "YUDI", "NAMA_SM": "HENDRIK"},
        {"KODE_BARANG": "SKU-3", "QTY AKHIR": 12, "PRINCIPAL": "MOTASA INDONESIA, PT",
         "GOLONGAN": "YUDI", "NAMA_SM": "HENDRIK"},
        {"KODE_BARANG": "SKU-F1", "QTY AKHIR": 8, "Principal": "FONTERRA BRANDS INDONESIA, PT",
         "GOLONGAN": "SOFYAN", "NAMA_SM": "ADNAN", "Kode Gudang": "GD01"},
        {"KODE_BARANG": "SKU-F2", "QTY AKHIR": 9, "Principal": "FONTERRA BRANDS INDONESIA, PT",
         "GOLONGAN": "SOFYAN", "NAMA_SM": "ADNAN", "Kode Gudang": "GD02"},
    ])
    lookups = laporan.LookupTables({}, {}, {}, {}, {
        "ENERGIZER": {"group_type": "principal", "values": ["ENERGIZER INDONESIA, PT"]},
        "FONTERRA": {"group_type": "principal", "values": ["FONTERRA BRANDS INDONESIA, PT"]},
        "MOTASA MKS 1": {"group_type": "principal", "values": ["MOTASA INDONESIA, PT"]},
        "MOTASA MKS 2": {"group_type": "principal", "values": ["MOTASA INDONESIA, PT"]},
        "MUSTIKA RATU": {"group_type": "principal", "values": ["MUSTIKA RATUBUANA INTERNATIONAL"]},
        "RECKIT": {"group_type": "principal", "values": ["RECKITT BENCKISER, PT"]},
    })

    frame = laporan.build_report_frame(source)
    assert frame.columns.tolist() == laporan.REPORT_COLUMNS
    assert not frame.columns.duplicated().any()

    with tempfile.TemporaryDirectory() as temp_dir:
        written, unmatched = laporan.write_report_files(
            source, temp_dir, "2026-07-15",
            ["DENNY", "HENDRIK", "ENERGIZER", "FONTERRA", "MOTASA MKS 1", "MOTASA MKS 2",
             "MUSTIKA RATU", "RECKIT", "UNKNOWN"],
            lookups, stock,
        )
        assert [item["keyword"] for item in written] == [
            "DENNY", "HENDRIK", "ENERGIZER", "FONTERRA", "MOTASA MKS 1", "MOTASA MKS 2",
            "MUSTIKA RATU", "RECKIT",
        ]
        assert unmatched == ["UNKNOWN"]
        assert {item["groupType"] for item in written} == {"spv", "sm", "principal"}
        output = Path(next(item["path"] for item in written if item["keyword"] == "DENNY"))
        workbook = openpyxl.load_workbook(output, read_only=True, data_only=True)
        assert workbook.sheetnames == ["DENNY", "DENNY Stock"]
        sheet = workbook["DENNY"]
        headers = [cell.value for cell in sheet[1]]
        values = [cell.value for cell in sheet[2]]

        assert headers[:len(laporan.REPORT_COLUMNS)] == laporan.REPORT_COLUMNS
        assert all(value is None for value in headers[len(laporan.REPORT_COLUMNS):])
        assert all(value is None for value in values[len(laporan.REPORT_COLUMNS):])
        assert values[headers.index("AO")] == 7
        assert values[headers.index("EC")] == 9
        assert values[headers.index("Item Aktif")] == 11
        stock_sheet = workbook["DENNY Stock"]
        stock_headers = [cell.value for cell in stock_sheet[1]]
        stock_values = [cell.value for cell in stock_sheet[2]]
        assert stock_headers == laporan.build_principal_stock(stock).columns.tolist()
        assert stock_headers == [
            "Kode", "Nama Barang", "Kode Gudang", "Nama Gudang", "Satuan", "Principal", "Saldo Akhir",
        ]
        assert stock_values[stock_headers.index("Kode")] == "SKU-1"
        assert stock_values[stock_headers.index("Saldo Akhir")] == 25
        assert next(item["stockRows"] for item in written if item["keyword"] == "DENNY") == 1
        workbook.close()

        sm_workbook = openpyxl.load_workbook(
            next(item["path"] for item in written if item["keyword"] == "HENDRIK"),
            read_only=True, data_only=True,
        )
        assert [cell.value for cell in sm_workbook["HENDRIK Stock"][1]] == stock_headers
        sm_workbook.close()

        motasa_1 = openpyxl.load_workbook(
            next(item["path"] for item in written if item["keyword"] == "MOTASA MKS 1"),
            read_only=True, data_only=True,
        )
        assert motasa_1["MOTASA MKS 1"]["G2"].value == "MS1_TEST"
        motasa_1.close()

        motasa_2 = openpyxl.load_workbook(
            next(item["path"] for item in written if item["keyword"] == "MOTASA MKS 2"),
            read_only=True, data_only=True,
        )
        assert motasa_2["MOTASA MKS 2"]["G2"].value == "MS6_TEST"
        motasa_2.close()

        fonterra = openpyxl.load_workbook(
            next(item["path"] for item in written if item["keyword"] == "FONTERRA"),
            read_only=True, data_only=True,
        )
        assert [cell.value for cell in fonterra["FONTERRA Stock"][1]] == stock_headers
        assert fonterra["FONTERRA Stock"]["C2"].value == "GD01"
        fonterra.close()
        assert next(item["rows"] for item in written if item["keyword"] == "FONTERRA") == 1
        assert next(item["stockRows"] for item in written if item["keyword"] == "FONTERRA") == 1

        mustika = openpyxl.load_workbook(
            next(item["path"] for item in written if item["keyword"] == "MUSTIKA RATU"),
            read_only=True, data_only=True,
        )
        mustika_sheet = mustika["MUSTIKA RATU"]
        mustika_headers = [cell.value for cell in mustika_sheet[1]]
        mustika_values = [cell.value for cell in mustika_sheet[2]]
        assert mustika_headers == laporan.build_principal_report(
            "MUSTIKA RATU", source.tail(1), laporan.REPORT_COLUMNS,
        ).columns.tolist()
        assert mustika_values[mustika_headers.index("Kode Item")] == "1006067"
        assert mustika_values[mustika_headers.index("HET")] == 57400
        assert mustika_values[mustika_headers.index("BA / NON BA")] == "BA"
        mustika.close()

        reckitt = openpyxl.load_workbook(
            next(item["path"] for item in written if item["keyword"] == "RECKIT"),
            read_only=True, data_only=True,
        )
        reckitt_sheet = reckitt["RECKIT"]
        reckitt_headers = [cell.value for cell in reckitt_sheet[1]]
        reckitt_values = [cell.value for cell in reckitt_sheet[2]]
        assert reckitt_headers[-1] == "Devisi"
        assert reckitt_values[-1] == "RBI-HEALTH"
        reckitt.close()

    print("OK: parity Power Query Principal, SPV/SM, dan Stock sesuai")


if __name__ == "__main__":
    main()
